# -*- coding: utf-8 -*-
"""
==============================================================================
 IntelliBI - BATCH PLANNER INTELLIGENCE SYSTEM
==============================================================================
 Purpose : Analyse student enrolment, course structure, technology mapping,
           session schedules, technology dependencies and timing preferences
           to help a Batch Coordinator decide WHAT technology batch to start
           next, FOR WHICH students, and in WHICH timing slot.

 Output  : Batch_Planner_Report.xlsx  (IntelliBI-styled), uploaded to the
           configured Google Drive folder as ONE file (no splitting).
           Sheets:
             1. Student_Progress_Report  (one row per student)
             2. Batch_Planning_Report    (technology x timing)
             3. Batch_Recommendation     (ranked "what to start next")
             4. Bottlenecks_Insights     (demand / bottleneck / idle)
             5. Assumptions

 Inputs (already mirrored in the project folder):
   - CourseAndTechnologiesMappingDocument.xlsx
   - TechnologyMappingDocuments.xlsx
   - IntellBIAttendance.xlsx               (Sessions)
   - IntelliBIStudentInfo.xlsx             (Students, ClassLearnerTeacherEnrolled)

 Dependencies: pandas, numpy, openpyxl,
               google-api-python-client, google-auth   (for Drive upload)

------------------------------------------------------------------------------
 KEY BUSINESS LOGIC  (see also the Assumptions sheet)
------------------------------------------------------------------------------
 * Per-student technology status comes from ClassLearnerTeacherEnrolled
   (class_name=technology, class_subject="dd-Mon-YYYY To dd-Mon-YYYY").
       Completed   -> end date < today
       In Progress -> "To Current Date" or end date >= today
       Pending     -> no enrolment record for that technology
 * A student's required technologies = TechnologiesInSeq of the course found
   from batch_name (Shortform = batch_name minus trailing MMYY + optional 'E').
   Only Is_Course=Y AND IsActiveCourse=Y courses are planned.
 * DEPENDENCY-ONLY eligibility (per the latest instruction: "apart from the
   listed dependencies there is no single dependency; you can start any
   technology and a student can be eligible for the next BUNCH of
   technologies"):
       A pending technology T is ELIGIBLE for a student iff every prerequisite
       in the dependency table below is Completed.  There is NO strict
       sequence/no-skip gate, so a student can have MULTIPLE next-eligible
       technologies at once.
   Dependencies:
       PySpark        <- Python
       Power BI       <- SQL
       Tableau        <- SQL
       Azure Services <- SQL, Python, PySpark
       Azure Project  <- SQL, Python, PySpark
       Databricks     <- SQL, Python, PySpark
       AWS            <- SQL, Python, PySpark
 * Technology-name normalisation collapses the 3 vocabularies (course sequence
   tokens, TechnologyMapping names, taught class_name) to one canonical name
   using the WiseCourseNameCreationNote grouping rules.
 * Suggested batch start = max(today, latest tentative end of that technology's
   currently-ongoing batch) + START_BUFFER_DAYS (2-5 days).  Ongoing tentative
   end = last scheduled session date of the running batch on the LMS.
==============================================================================
"""

# --- IntelliBI Operations Automation portability bootstrap (auto-inserted) ---
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "common"))
import _bootstrap  # noqa: E402  (sys.path + env defaults + config.yaml)
from paths import CREDENTIALS_DIR, CONFIG_DIR, LOGS_DIR, CACHE_DIR as PROJECT_CACHE_DIR  # noqa: E402
# --- end bootstrap ---

import os
import re
import sys
import logging
from datetime import datetime, date, timedelta

import numpy as np
import pandas as pd


# =============================================================================
# CONFIG
# =============================================================================
class CONFIG:
    BASE_DIR = os.environ.get("BP_BASE_DIR") or os.path.dirname(os.path.abspath(__file__))

    COURSE_FILE      = "CourseAndTechnologiesMappingDocument.xlsx"
    TECH_FILE        = "TechnologyMappingDocuments.xlsx"
    ATTENDANCE_FILE  = "IntellBIAttendance.xlsx"
    STUDENT_FILE     = "IntelliBIStudentInfo.xlsx"

    COURSE_SHEET   = "CourseAndTechnologiesMappingDoc"
    TECH_SHEET     = "TechnologyMapping"
    SESSIONS_SHEET = "Sessions"
    STUDENTS_SHEET = "Students"
    ENROLL_SHEET   = "ClassLearnerTeacherEnrolled"

    OUTPUT_FILE = "Batch_Planner_Report.xlsx"

    # planning knobs
    START_BUFFER_DAYS = 3        # 2-5 day buffer after a running batch ends
    TODAY = None

    # scope
    ACTIVE_STUDENTS_ONLY = True
    DUMMY_SHORTFORMS = {"DF", "DS", "FS"}
    DUMMY_BATCH_LITERALS = {"dummy faculty", "dummy student", "free student"}

    # technologies never planned as standalone batches
    NON_PLANNABLE_KEYWORDS = (
        "placement", "project", "interview", "real-time domain",
        "learning journey", "sample course", "business communication",
    )

    # --- Google Drive upload (same mechanism as the Reports/* pipelines) ------
    GDRIVE_UPLOAD        = True
    GDRIVE_FOLDER_ID     = "1goPKLAlbL-cEC8r9x7P0a4xAeg1pvDmN"
    SERVICE_ACCOUNT_FILE = os.path.join(CREDENTIALS_DIR, "service_account.json")
    IMPERSONATE_USER     = "info@intellibiinnovationstechnologies.in"


# =============================================================================
# LOGGING
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("BatchPlanner")


# =============================================================================
# IntelliBI REPORT PALETTE  (matches Reports/pyAttendaceFeedbackReport.py)
# =============================================================================
C_NAV       = "1A2E5A"   # title text + column headers
C_WHITE     = "FFFFFF"
C_BLUE_MID  = "2E75B6"   # section banner
C_BLUE_LITE = "BDD7EE"
C_ROW_ALT   = "EBF4FB"   # alternating row tint
C_GREY_BD   = "9E9E9E"
C_GREEN     = "C8E6C9"   # completed
C_GREEN_TXT = "1B5E20"
C_AMBER     = "FFE0B2"   # in progress
C_AMBER_TXT = "BF360C"
C_RED_LITE  = "FFCDD2"   # pending / blocked
C_RED_TXT   = "B71C1C"
C_BLUE_PALE = "DDEEFF"


# =============================================================================
# 0. TECHNOLOGY NORMALISATION
# =============================================================================
_CANON_ALIASES = {
    "SQL": ["sql"],
    "Python": ["python", "advanced python", "core advanced python",
               "advanced python | pyspark", "advanced python |  pyspark"],
    "PySpark": ["pyspark"],
    "Databricks": ["databricks"],
    "Power BI": ["power bi", "powerbi", "power-bi"],
    "Tableau": ["tableau"],
    "Data Warehouse": ["data warehouse", "datawarehouse", "dwh"],
    "Airflow": ["airflow"],
    "AWS": ["aws", "aws services", "aws data engineering"],
    "AWS Project": ["aws project"],
    "Azure Services": [
        "azure services", "azure service", "azure data factory", "adf",
        "azure synapse", "synapse", "azure", "advanced azure",
    ],
    "Azure Project": ["azure project", "azure  project", "azure project - a",
                      "azure project - b", "azure  project - a", "azure project - a/b"],
    "AI (GenAI - Agentic AI - ML)": [
        "ai (genai - agentic ai - ml)", "ai (genai • agentic ai • ml)",
        "generative ai", "genai", "agentic ai", "machine learning (ml)",
        "machine learning", "ml", "data science - ai", "data science",
    ],
    "Advanced Excel | R": [
        "advanced excel | r", "advanced excel", "r overview", "r",
        "advanced excel | r | tableau",
    ],
    "AI for Data Analytics": ["ai for data analytics"],
    "AI for Data Engineers": ["ai for data engineers"],
    "AI for Data Engineers & Data Analytics": ["ai for data engineers & data analytics"],
    "GoogleBigQuery": ["googlebigquery", "google bigquery", "bigquery"],
    "Data Fusion": ["data fusion"],
    "GCP": ["gcp"],
    "Business Communication": ["business communication"],
}


def _squeeze(s):
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return ""
    s = str(s).strip().lower()
    s = s.replace("•", "-").replace("·", "-")
    s = re.sub(r"\s+", " ", s)
    return s


_ALIAS_TO_CANON = {}
for _canon, _aliases in _CANON_ALIASES.items():
    _ALIAS_TO_CANON[_squeeze(_canon)] = _canon
    for _a in _aliases:
        _ALIAS_TO_CANON[_squeeze(_a)] = _canon


def normalize_tech(raw):
    """Map any raw technology / class / sequence token to its canonical name."""
    key = _squeeze(raw)
    if not key:
        return None
    if key in _ALIAS_TO_CANON:
        return _ALIAS_TO_CANON[key]
    base = re.sub(r"\s*-\s*[ab]$", "", key)
    base = re.sub(r"\s+(interviews?|sessions?)$", "", base)
    if base in _ALIAS_TO_CANON:
        return _ALIAS_TO_CANON[base]
    return str(raw).strip()


def is_plannable_tech(canon_name):
    if not canon_name:
        return False
    low = canon_name.lower()
    return not any(k in low for k in CONFIG.NON_PLANNABLE_KEYWORDS)


# =============================================================================
# DEPENDENCY RULES (the ONLY prerequisites; no strict-sequence gate)
# =============================================================================
DEPENDENCIES = {
    "PySpark":        ["Python"],
    "Power BI":       ["SQL"],
    "Tableau":        ["SQL"],
    "Azure Services": ["SQL", "Python", "PySpark"],
    "Azure Project":  ["SQL", "Python", "PySpark"],
    "Databricks":     ["SQL", "Python", "PySpark"],
    "AWS":            ["SQL", "Python", "PySpark"],
    "AWS Project":    ["SQL", "Python", "PySpark", "AWS"],
}


# =============================================================================
# 1. load_data
# =============================================================================
def load_data():
    p = lambda f: os.path.join(CONFIG.BASE_DIR, f)
    log.info("Loading source workbooks ...")
    data = {
        "courses":  pd.read_excel(p(CONFIG.COURSE_FILE),  sheet_name=CONFIG.COURSE_SHEET),
        "tech":     pd.read_excel(p(CONFIG.TECH_FILE),    sheet_name=CONFIG.TECH_SHEET),
        "sessions": pd.read_excel(p(CONFIG.ATTENDANCE_FILE), sheet_name=CONFIG.SESSIONS_SHEET),
        "students": pd.read_excel(p(CONFIG.STUDENT_FILE), sheet_name=CONFIG.STUDENTS_SHEET),
        "enroll":   pd.read_excel(p(CONFIG.STUDENT_FILE), sheet_name=CONFIG.ENROLL_SHEET),
    }
    for k, v in data.items():
        log.info("  %-9s rows=%-6d cols=%d", k, len(v), v.shape[1])
    return data


# =============================================================================
# 2. parse_date_ranges
# =============================================================================
_DATE_RE = re.compile(r"(\d{1,2}\s*-\s*[A-Za-z]{3,}\s*-\s*\d{4})")


def _parse_one_date(token):
    token = re.sub(r"\s+", "", str(token)).strip()
    for fmt in ("%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def parse_date_range(text):
    """'dd-Mon-YYYY To dd-Mon-YYYY' / 'dd-Mon-YYYY To Current Date'
       -> (start, end, ongoing).  end None when ongoing/unparseable."""
    if text is None or (isinstance(text, float) and np.isnan(text)):
        return (None, None, False)
    s = str(text).strip()
    ongoing = "current date" in s.lower()
    dates = _DATE_RE.findall(s)
    start = _parse_one_date(dates[0]) if len(dates) >= 1 else None
    if ongoing:
        return (start, None, True)
    end = _parse_one_date(dates[1]) if len(dates) >= 2 else None
    return (start, end, False)


# =============================================================================
# 3. clean_and_standardize + build_relationship_map
# =============================================================================
def shortform_from_batch(batch_name):
    if batch_name is None or (isinstance(batch_name, float) and np.isnan(batch_name)):
        return None
    b = str(batch_name).strip()
    m = re.match(r"^(.*?)(\d{4})([A-Za-z]{1,2})?$", b)
    return m.group(1).strip() if m else b.strip()


def batch_timing_from_batch(batch_name, fallback):
    """Morning / Evening / Weekend from Batch_Timing; infer Evening from
    trailing 'E' on batch_name when blank; default Morning."""
    if isinstance(fallback, str) and fallback.strip():
        return fallback.strip().title()
    b = str(batch_name).strip()
    if re.search(r"\d{4}E$", b):
        return "Evening"
    return "Morning"


def build_course_map(courses):
    cmap = {}
    for _, r in courses.iterrows():
        sf = str(r["CourseName Shortform"]).strip() if pd.notna(r["CourseName Shortform"]) else ""
        if not sf:
            continue
        seq = []
        if pd.notna(r.get("TechnologiesInSeq")):
            for tok in str(r["TechnologiesInSeq"]).split("|"):
                c = normalize_tech(tok)
                if c and c not in seq:
                    seq.append(c)
        cmap[sf.upper()] = {
            "shortform": sf,
            "course_name": str(r.get("CourseName", "")).strip(),
            "sequence": seq,
            "is_course": str(r.get("Is_Course", "")).strip().upper() == "Y",
            "is_active": str(r.get("IsActiveCourse", "")).strip().upper() == "Y",
        }
    return cmap


def build_tech_hours(tech):
    hours = {}
    for _, r in tech.iterrows():
        c = normalize_tech(r["TechnologyName"])
        if c and pd.notna(r.get("AvgNumofHours")):
            hours.setdefault(c, str(r["AvgNumofHours"]).strip())
    return hours


def build_session_maps(sessions):
    """Return:
       last_by_batch : (tech, course_title) -> last scheduled session date
       ongoing_end   : tech -> latest tentative end among its ONGOING batches
    """
    s = sessions.copy()
    s["start_time_ist"] = pd.to_datetime(s["start_time_ist"], errors="coerce")
    s["tech"] = s["course_name"].apply(normalize_tech)
    s["tag"] = s["course_title"].astype(str).str.strip()
    s = s.dropna(subset=["start_time_ist"])
    last_by_batch = {k: v.date() for k, v in
                     s.groupby(["tech", "tag"])["start_time_ist"].max().items()}
    ongoing_end = {}
    for (tech, tag), d in last_by_batch.items():
        if "current date" in tag.lower():
            if tech not in ongoing_end or d > ongoing_end[tech]:
                ongoing_end[tech] = d
    return last_by_batch, ongoing_end


# =============================================================================
# 4-5. compute_student_status + dependency_engine
# =============================================================================
def project_ongoing_end(tech, tag, last_by_batch):
    return last_by_batch.get((tech, tag)) or CONFIG.TODAY


def compute_student_progress(data, course_map, last_by_batch):
    """Returns (sheet1_rows, student_meta).  One Sheet-1 row per student."""
    students = data["students"]
    enroll = data["enroll"].copy()
    enroll["tech"] = enroll["class_name"].apply(normalize_tech)
    parsed = enroll["class_subject"].apply(parse_date_range)
    enroll["start"]   = parsed.apply(lambda t: t[0])
    enroll["end"]     = parsed.apply(lambda t: t[1])
    enroll["ongoing"] = parsed.apply(lambda t: t[2])
    enroll["tag"]     = enroll["class_subject"].astype(str).str.strip()

    enr_by_student = {}
    for _, r in enroll.iterrows():
        sid, tech = r["student_id"], r["tech"]
        if pd.isna(sid) or not tech:
            continue
        rec = {"end": r["end"], "ongoing": bool(r["ongoing"]),
               "start": r["start"], "tag": r["tag"]}
        bucket = enr_by_student.setdefault(sid, {})
        prev = bucket.get(tech)
        if prev is None:
            bucket[tech] = rec
        else:
            def score(x):
                if x["end"] is not None:
                    return (2, x["end"].toordinal())
                if x["ongoing"]:
                    return (1, x["start"].toordinal() if x["start"] else 0)
                return (0, 0)
            if score(rec) > score(prev):
                bucket[tech] = rec

    today = CONFIG.TODAY
    sheet1_rows, student_meta = [], {}

    for _, st in students.iterrows():
        sid = st["student_id"]
        if pd.isna(sid):
            continue
        batch = st.get("batch_name")
        sf_key = (shortform_from_batch(batch) or "").upper()

        if CONFIG.ACTIVE_STUDENTS_ONLY:
            if str(st.get("Is_Deleted", "")).strip().upper() == "Y":
                continue
            if sf_key in CONFIG.DUMMY_SHORTFORMS:
                continue
            if str(batch).strip().lower() in CONFIG.DUMMY_BATCH_LITERALS:
                continue

        course = course_map.get(sf_key)
        if not course or not course["is_course"] or not course["is_active"]:
            continue
        sequence = course["sequence"]
        if not sequence:
            continue

        timing = batch_timing_from_batch(batch, st.get("Batch_Timing"))
        name = str(st.get("student_name", "")).strip()
        enr = enr_by_student.get(sid, {})

        status_by_tech, end_by_tech = {}, {}
        for tech in sequence:
            rec = enr.get(tech)
            if rec is None:
                status_by_tech[tech] = "Pending"
                end_by_tech[tech] = None
            elif rec["ongoing"] or (rec["end"] is not None and rec["end"] >= today):
                status_by_tech[tech] = "In Progress"
                end_by_tech[tech] = project_ongoing_end(tech, rec["tag"], last_by_batch)
            elif rec["end"] is not None and rec["end"] < today:
                status_by_tech[tech] = "Completed"
                end_by_tech[tech] = rec["end"]
            else:
                status_by_tech[tech] = "In Progress"
                end_by_tech[tech] = project_ongoing_end(tech, rec["tag"], last_by_batch)

        def is_completed(t):
            return status_by_tech.get(t) == "Completed"

        def unmet_deps(t):
            return [d for d in DEPENDENCIES.get(t, [])
                    if d in status_by_tech and not is_completed(d)]

        completed = [t for t in sequence if status_by_tech[t] == "Completed"]
        inprog    = [t for t in sequence if status_by_tech[t] == "In Progress"]
        pending   = [t for t in sequence if status_by_tech[t] == "Pending"]

        # DEPENDENCY-ONLY eligibility -> possibly MULTIPLE next-eligible techs
        next_eligible, blockers = [], []
        for t in pending:
            um = unmet_deps(t)
            if not um:
                next_eligible.append(t)
            else:
                blockers.append(f"{t} (needs {', '.join(um)})")

        # overall student status
        if not pending and not inprog and completed:
            overall = "Completed"
        elif completed or inprog:
            overall = "In Progress"
        else:
            overall = "Pending"
        all_done = "Yes" if (not pending and not inprog and completed) else "No"

        comp_dates = [end_by_tech[t] for t in completed if end_by_tech[t]]
        completion_date = max(comp_dates).isoformat() if comp_dates else ""

        sheet1_rows.append({
            "Student ID": sid,
            "Student Name": name,
            "Course": course["course_name"],
            "Batch Name": batch,
            "Batch Timing": timing,
            "Technologies in Sequence": " | ".join(sequence),
            "Completed Technology": ", ".join(completed),
            "In Progress Technology": ", ".join(inprog),
            "Pending Technology": ", ".join(pending),
            "Status": overall,
            "Completion Date": completion_date,
            "Dependency Blockers": "; ".join(blockers),
            "Next Eligible Technologies": ", ".join(next_eligible),
            "Whether All Technologies Completed": all_done,
        })

        student_meta[sid] = {
            "name": name, "course": course["course_name"], "timing": timing,
            "batch": batch, "sequence": sequence, "status_by_tech": status_by_tech,
            "next_eligible": next_eligible, "pending": pending,
            "completed": set(completed),
        }

    log.info("Computed progress for %d students.", len(sheet1_rows))
    return sheet1_rows, student_meta


# =============================================================================
# 6. batch_planner_engine
# =============================================================================
def batch_planner(student_meta, tech_hours, ongoing_end):
    today = CONFIG.TODAY
    buf = timedelta(days=CONFIG.START_BUFFER_DAYS)

    # tech -> timing -> {eligible:[names], blocked:[names]}
    plan = {}
    for sid, m in student_meta.items():
        timing, name = m["timing"], m["name"]
        completed = m["completed"]
        for t in m["pending"]:
            if not is_plannable_tech(t):
                continue
            node = plan.setdefault(t, {}).setdefault(
                timing, {"eligible": [], "blocked": []})
            unmet = [d for d in DEPENDENCIES.get(t, []) if d not in completed]
            (node["eligible"] if not unmet else node["blocked"]).append(name)

    plan_rows, rec_rows = [], []
    for tech, timings in plan.items():
        deps = DEPENDENCIES.get(tech, [])
        hours = tech_hours.get(tech, "")
        # suggested start considers the running batch's tentative end
        oe = ongoing_end.get(tech)
        base_start = max(today, oe) if oe else today
        start_str = (base_start + buf).isoformat()

        total_elig = sum(len(n["eligible"]) for n in timings.values())
        best_timing, best_count = "", -1
        for timing, node in timings.items():
            if len(node["eligible"]) > best_count:
                best_count, best_timing = len(node["eligible"]), timing

        for timing, node in sorted(timings.items(), key=lambda x: -len(x[1]["eligible"])):
            elig, blk = node["eligible"], node["blocked"]
            tips = []
            if elig:
                tips.append(f"{len(elig)} eligible now")
            if blk:
                tips.append(f"{len(blk)} blocked (need {', '.join(deps)})")
            if oe:
                tips.append(f"running batch ends ~{oe.isoformat()}; start after buffer")
            if timing == best_timing and total_elig:
                tips.append(f"{timing} has max demand for this technology")
            plan_rows.append({
                "Technology": tech,
                "Batch Timing": timing,
                "Eligible Student Count": len(elig),
                "Eligible Student List": ", ".join(sorted(elig)),
                "Blocked Student Count": len(blk),
                "Blocked Student List": ", ".join(sorted(blk)),
                "Dependency Requirements": ", ".join(deps) if deps else "None",
                "Avg Completion Hours": hours,
                "Suggested Start Date": start_str if elig else "Hold (no one eligible)",
                "Additional Suggestions": "; ".join(tips) if tips else "-",
            })

        rec_rows.append({
            "Technology": tech,
            "Total Eligible (all timings)": total_elig,
            "Top Demand Timing": best_timing,
            "Top Demand Count": max(best_count, 0),
            "Dependency Requirements": ", ".join(deps) if deps else "None",
            "Avg Completion Hours": hours,
            "Suggested Start Date": start_str,
        })

    plan_rows.sort(key=lambda r: (-r["Eligible Student Count"], r["Technology"]))
    rec_rows.sort(key=lambda r: (-r["Total Eligible (all timings)"], r["Technology"]))
    for i, r in enumerate(rec_rows, 1):
        r["Priority Rank"] = i
        if r["Total Eligible (all timings)"] == 0:
            r["Recommendation"] = "Hold - no eligible students yet."
        else:
            r["Recommendation"] = (
                f"Start a {r['Top Demand Timing']} batch "
                f"({r['Top Demand Count']} ready); "
                f"{r['Total Eligible (all timings)']} eligible across timings.")
    log.info("Planned %d technology x timing batch rows.", len(plan_rows))
    return plan_rows, rec_rows


# =============================================================================
# 7. insights
# =============================================================================
def build_insights(student_meta, plan_rows):
    rows = []
    for sid, m in student_meta.items():
        sbt = m["status_by_tech"]
        done = sum(1 for v in sbt.values() if v == "Completed")
        prog = sum(1 for v in sbt.values() if v == "In Progress")
        pend = sum(1 for v in sbt.values() if v == "Pending")
        if done == 0 and prog == 0:
            rows.append({"Insight Type": "Idle Student",
                         "Detail": f"{m['name']} ({m['course']}, {m['timing']})",
                         "Metric": f"{pend} pending / nothing started"})
    agg = {}
    for r in plan_rows:
        a = agg.setdefault(r["Technology"], {"e": 0, "b": 0})
        a["e"] += r["Eligible Student Count"]
        a["b"] += r["Blocked Student Count"]
    for tech, a in sorted(agg.items(), key=lambda x: -x[1]["e"]):
        if a["e"] >= 5:
            rows.append({"Insight Type": "High-Demand Technology", "Detail": tech,
                         "Metric": f"{a['e']} eligible students waiting"})
    for tech, a in sorted(agg.items(), key=lambda x: -x[1]["b"]):
        if a["b"] >= 5:
            rows.append({"Insight Type": "Bottleneck Technology", "Detail": tech,
                         "Metric": f"{a['b']} students blocked by dependencies"})
    return rows


# =============================================================================
# 8. export_excel_report  (IntelliBI styling)
# =============================================================================
ASSUMPTIONS = [
    (1, "Per-student technology status taken from ClassLearnerTeacherEnrolled "
        "(class_name=technology, class_subject=date range). Sessions sheet used "
        "only to project tentative end of ongoing technologies."),
    (2, "Required technologies = TechnologiesInSeq of the course derived from "
        "batch_name (Shortform = batch_name minus trailing MMYY + optional 'E')."),
    (3, "Only Is_Course=Y and IsActiveCourse=Y courses are planned."),
    (4, "Completed = end date < today; In Progress = 'To Current Date' or end >= "
        "today; Pending = no enrolment record for that technology."),
    (5, "DEPENDENCY-ONLY eligibility (per latest instruction): no strict-sequence "
        "gate. A pending technology is eligible once its listed prerequisites are "
        "Completed, so a student may have MULTIPLE next-eligible technologies."),
    (6, "Dependencies: PySpark<-Python; Power BI<-SQL; Tableau<-SQL; "
        "Azure Services/Azure Project/Databricks/AWS <- SQL,Python,PySpark."),
    (7, "Technology names normalised across the 3 vocabularies using "
        "WiseCourseNameCreationNote grouping (e.g. GenAI|Agentic|ML -> AI group; "
        "Azure Data Factory|Synapse|Services -> Azure Services)."),
    (8, f"Suggested start = max(today, latest tentative end of the technology's "
        f"ongoing LMS batch) + {CONFIG.START_BUFFER_DAYS}-day buffer (2-5 days)."),
    (9, "Scope = active real students only: Is_Deleted!=Y; dummy/free batches "
        "(DF, DS, FS, Dummy Faculty/Student, Free Student) excluded."),
    (10, "Placement/Project/Interview/Real-Time-Domain/soft-skill phases are NOT "
         "planned as standalone technology batches."),
    (11, "Student-level 'Status' = Completed if all techs done; Pending if nothing "
         "started; else In Progress. 'Completion Date' = latest completed-tech end."),
    (12, "Batch_Timing from Students sheet; if blank, inferred from trailing 'E' on "
         "batch_name (Evening) else Morning. Weekend supported when present."),
]


def _style_sheet(ws, df, title, tab_color, status_col=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color=C_GREY_BD)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncols = max(len(df.columns), 1)

    # Title banner (row1) + generated stamp (row2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=f"IntelliBI  |  {title}")
    t.font = Font(name="Arial", bold=True, size=14, color=C_NAV)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    g = ws.cell(row=2, column=1,
                value=f"Generated: {CONFIG.TODAY.strftime('%d-%b-%Y')}   |   "
                      f"Rows: {len(df)}")
    g.font = Font(name="Arial", bold=False, size=9, color="555555")
    g.alignment = Alignment(horizontal="left", vertical="center")

    header_row = 3
    for j, colname in enumerate(df.columns, 1):
        c = ws.cell(row=header_row, column=j, value=colname)
        c.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_NAV)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[header_row].height = 30

    status_fill = {
        "Completed": (C_GREEN, C_GREEN_TXT),
        "In Progress": (C_AMBER, C_AMBER_TXT),
        "Pending": (C_RED_LITE, C_RED_TXT),
    }
    for i, (_, rrow) in enumerate(df.iterrows()):
        excel_row = header_row + 1 + i
        bg = C_ROW_ALT if i % 2 else C_WHITE
        for j, colname in enumerate(df.columns, 1):
            val = rrow[colname]
            c = ws.cell(row=excel_row, column=j,
                        value="" if (val is None or (isinstance(val, float) and np.isnan(val))) else val)
            c.font = Font(name="Arial", size=10, color="000000")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=colname not in ("Student ID",))
            c.border = border
            if status_col and colname == status_col and val in status_fill:
                fbg, ftx = status_fill[val]
                c.fill = PatternFill("solid", fgColor=fbg)
                c.font = Font(name="Arial", bold=True, size=10, color=ftx)

    # column widths
    for j, colname in enumerate(df.columns, 1):
        series = df[colname].astype(str)
        w = max(len(str(colname)),
                int(series.str.len().quantile(0.9)) if len(series) else 10)
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 3, 12), 60)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_properties.tabColor = tab_color


def export_excel_report(sheet1_rows, plan_rows, rec_rows, insight_rows):
    out = os.path.join(CONFIG.BASE_DIR, CONFIG.OUTPUT_FILE)
    df1 = pd.DataFrame(sheet1_rows)
    df2 = pd.DataFrame(plan_rows)
    df3 = pd.DataFrame(rec_rows)
    df4 = pd.DataFrame(insight_rows) if insight_rows else pd.DataFrame(
        columns=["Insight Type", "Detail", "Metric"])
    df5 = pd.DataFrame(ASSUMPTIONS, columns=["#", "Assumption"])

    if not df3.empty:
        df3 = df3[["Priority Rank", "Technology", "Total Eligible (all timings)",
                   "Top Demand Timing", "Top Demand Count", "Dependency Requirements",
                   "Avg Completion Hours", "Suggested Start Date", "Recommendation"]]

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    specs = [
        ("Student_Progress_Report", df1, "Student Technology Progress Report", C_NAV, "Status"),
        ("Batch_Planning_Report",   df2, "Batch Planning Report",             C_BLUE_MID, None),
        ("Batch_Recommendation",    df3, "Batch Recommendation (Start Next)",  "2E7D32", None),
        ("Bottlenecks_Insights",    df4, "Bottlenecks & Insights",            "E65100", None),
        ("Assumptions",             df5, "Assumptions & Business Rules",       "555555", None),
    ]
    for name, df, title, color, status_col in specs:
        ws = wb.create_sheet(name)
        _style_sheet(ws, df, title, color, status_col)
    wb.save(out)
    log.info("Report written -> %s", out)
    return out


# =============================================================================
# 9. upload_to_gdrive  (same mechanism as Reports/pyAttendaceFeedbackReport.py)
# =============================================================================
def upload_to_gdrive(local_path):
    if not CONFIG.GDRIVE_UPLOAD:
        log.info("[Drive] Upload disabled.")
        return None
    try:
        from googleapiclient.discovery import build as gdrive_build
        from googleapiclient.http import MediaFileUpload
        from google.oauth2 import service_account
    except ImportError as e:
        log.warning("[Drive] libs missing (%s). pip install google-api-python-client google-auth", e)
        return None
    try:
        filename = os.path.basename(local_path)
        creds = service_account.Credentials.from_service_account_file(
            CONFIG.SERVICE_ACCOUNT_FILE,
            scopes=["https://www.googleapis.com/auth/drive"],
        ).with_subject(CONFIG.IMPERSONATE_USER)
        drive = gdrive_build("drive", "v3", credentials=creds)
        q = (f"'{CONFIG.GDRIVE_FOLDER_ID}' in parents and name='{filename}' "
             f"and trashed=false")
        existing = drive.files().list(
            q=q, fields="files(id)",
            supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        for f in existing.get("files", []):
            drive.files().delete(fileId=f["id"], supportsAllDrives=True).execute()
        media = MediaFileUpload(
            local_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True)
        up = drive.files().create(
            body={"name": filename, "parents": [CONFIG.GDRIVE_FOLDER_ID]},
            media_body=media, fields="id,webViewLink",
            supportsAllDrives=True).execute()
        log.info("[Drive] Uploaded %s", filename)
        log.info("[Drive] Link: %s", up.get("webViewLink", ""))
        return up.get("webViewLink", "")
    except Exception as e:
        log.warning("[Drive] Upload failed: %s", e)
        return None


# =============================================================================
# MAIN
# =============================================================================
def main():
    CONFIG.TODAY = datetime.now().date()
    log.info("=== IntelliBI Batch Planner === run date %s", CONFIG.TODAY)

    data = load_data()
    course_map = build_course_map(data["courses"])
    tech_hours = build_tech_hours(data["tech"])
    last_by_batch, ongoing_end = build_session_maps(data["sessions"])

    sheet1_rows, student_meta = compute_student_progress(data, course_map, last_by_batch)
    plan_rows, rec_rows = batch_planner(student_meta, tech_hours, ongoing_end)
    insight_rows = build_insights(student_meta, plan_rows)

    out = export_excel_report(sheet1_rows, plan_rows, rec_rows, insight_rows)
    log.info("Report ready: %s", out)
    upload_to_gdrive(out)
    log.info("DONE.")
    return out


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log.exception("Batch planner failed: %s", e)
        sys.exit(1)
