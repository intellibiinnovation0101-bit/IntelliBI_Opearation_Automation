"""
================================================================================
  IntelliBI Attendance & Feedback Report Generator
  Reads live data from Google Sheets → Builds styled Excel report → Emails via Gmail

  Sheets read:
    IntellBIAttendance  (Sessions, Attendance, Student_Feedback)
    IntelliBIStudentInfo (Students — for phone numbers)

  Report sheets produced:
    1. Session Summary    — per-session attendance metrics
    2. Student Detail     — per-student per-session breakdown
    3. Absent & At-Risk   — action-required list
    4. Feedback Rating    — session-level feedback stats

  Usage:
    Execution is controlled by the RUN CONFIGURATION variables near the top of
    this file (report_type / report_date / start_date / end_date / send_email),
    then simply run:

        python pyAttendaceFeedbackReport.py

    Examples (edit the variables, then run):
        report_type = "daily"                                  # daily (default)
        report_type = "weekly"                                 # weekly
        report_type = "fortnightly"                            # fortnightly
        report_type = "monthly"                                # monthly
        report_type = "quarterly"                              # quarterly
        report_type = "yearly"                                 # yearly
        report_type = "daily";  report_date = "2026-03-29"     # specific date
        report_type = "daily";  send_email  = False            # skip email
        report_type = "manual"; start_date = "01-Feb-2026"; end_date = "31-Mar-2026"
================================================================================
"""

# --- IntelliBI Operations Automation portability bootstrap (auto-inserted) ---
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "common"))
import _bootstrap  # noqa: E402  (sys.path + env defaults + config.yaml)
from paths import CREDENTIALS_DIR, CONFIG_DIR, LOGS_DIR, CACHE_DIR as PROJECT_CACHE_DIR  # noqa: E402
# --- end bootstrap ---

import sys
import os
import io
import calendar
import smtplib
from datetime import datetime, date, timedelta, time
from email.mime.multipart import MIMEMultipart
from email.mime.base     import MIMEBase
from email.mime.text     import MIMEText
from email               import encoders

# ── Resolve parent directory so utils.py and service_account.json are always
#    found regardless of what working directory PyCharm uses. ─────────────────
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR   = os.path.dirname(_SCRIPT_DIR)          # IntelliBI Automation/
if _BASE_DIR not in sys.path:
    sys.path.insert(0, _BASE_DIR)

import pandas as pd
import openpyxl
from openpyxl.styles    import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils     import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG  ← fill in GMAIL_SENDER and GMAIL_APP_PASS before first run
# ─────────────────────────────────────────────────────────────────────────────

SERVICE_ACCOUNT_FILE = os.path.join(CREDENTIALS_DIR, "service_account.json")
ATTENDANCE_SHEET_ID  = "1TqDjq4gAyo32eRNMbuLd6uu0eCNZb7h1j5YH-q68AhU"
STUDENTS_SHEET_ID    = "1Eq7Q3Gota7nYiaorm1L0NoouVfYtS7JkbBp4U5MWzVA"

# E-mail credentials are centralised — edit config_files/email_config.py to rotate.
from email_config import GMAIL_SENDER, GMAIL_APP_PASS
REPORT_TO      = ["info@intellibiinnovationstechnologies.in","intellibihropsb2ch@gmail.com"]
REPORT_CC      = []                            # optional: ["other@email.com"]

# ─────────────────────────────────────────────────────────────────────────────
#  RUN CONFIGURATION  (replaces the old command-line arguments)
#  Execution is controlled entirely by the variables below — edit them instead
#  of passing CLI flags.
#    report_type : "daily" (default) | "weekly" | "fortnightly" | "monthly"
#                  | "quarterly" | "yearly" | "manual"
#    report_date : specific reference date as "YYYY-MM-DD", or None = today.
#                  (Used by the non-manual report types; e.g. a back-dated daily run.)
#    start_date  : manual-range START date "DD-Mon-YYYY" (e.g. "01-Feb-2026").
#                  Required only when report_type == "manual".
#    end_date    : manual-range END date "DD-Mon-YYYY" (e.g. "31-Mar-2026").
#                  Required only when report_type == "manual".
#    send_email  : True  -> build the report AND email it (default).
#                  False -> build the report but skip sending email (old --no-email).
#    upload_to_local_directory : True  -> ALSO save the report to the local
#                  computer directory (LOCAL_UPLOAD_DIR). False -> skip that save
#                  only. The Google Drive upload is unaffected either way.
# ─────────────────────────────────────────────────────────────────────────────
report_type = "daily"
report_date = None
start_date  = None
end_date    = None
send_email  = True
upload_to_local_directory = False

# Base local computer directory used only when upload_to_local_directory = True.
# The report is saved under a per-report-type subfolder (e.g. "Daily") to mirror
# the Google Drive layout. Edit this path to point anywhere on the computer.
LOCAL_UPLOAD_DIR = os.path.join(_BASE_DIR, "Local_Reports")

# ─────────────────────────────────────────────────────────────────────────────
#  COLOUR & STYLE CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

C_NAV        = "1A2E5A"   # deep navy — title bar & column headers
C_NAV2       = "243F6B"   # slightly lighter navy for accents
C_WHITE      = "FFFFFF"
C_BLUE_MID   = "2E75B6"   # section header bg
C_BLUE_LITE  = "BDD7EE"   # totals row / sub-header
C_BLUE_PALE  = "DDEEFF"   # light blue tint
C_ROW_ALT    = "EBF4FB"   # alternating row tint
C_GREY_BD    = "9E9E9E"   # border colour
C_GREY_LITE  = "F5F5F5"   # very light grey
C_GREY_SUSP  = "D5D5D5"   # light grey for suspended students
C_GREY_NA    = "A6A6A6"   # dark grey — attendance Not Applicable students
C_GREY_CANCEL= "E0E0E0"   # light grey for cancelled sessions
C_LAVENDER   = "E8DEF8"   # soft lavender for scheduled (not yet conducted) sessions
C_LAVENDER_TXT = "5E35B1"  # deep purple text for scheduled sessions
C_YEST_HL    = "FCE4D6"   # light orange — Daily Report: highlights previous-day rows

# ── Attendance status colours ──
C_GREEN_DARK = "1B5E20"   # Excellent header text
C_GREEN_MED  = "2E7D32"   # Good header text
C_GREEN      = "C8E6C9"   # Excellent row bg
C_GREEN_PALE = "E8F5E9"   # Good row bg / teacher feedback ok bg

# ── At-Risk / Amber colours ──
C_AMBER_DARK = "BF360C"   # at-risk header text
C_AMBER      = "FFE0B2"   # at-risk row bg
C_AMBER_PALE = "FFF3E0"   # at-risk alt row

# ── Absent / Red colours ──
C_RED_DARK   = "B71C1C"   # absent section header bg
C_RED_MED    = "D32F2F"   # absent accent
C_RED_LITE   = "FFCDD2"   # absent row bg
C_RED_PALE   = "FFE8E8"   # absent alt row

C_ORANGE     = "E65100"   # at-risk section header bg
C_TEAL       = "00695C"   # teacher / feedback accent
C_TEAL_LITE  = "E0F2F1"   # teacher / feedback bg

# ── KPI card palettes [label_bg, value_bg, text_color] ──
KPI_BLUE   = ("1A2E5A", "DDEEFF", "1A2E5A")
KPI_GREEN  = ("2E7D32", "E8F5E9", "2E7D32")
KPI_RED    = ("B71C1C", "FFCDD2", "B71C1C")
KPI_AMBER  = ("E65100", "FFF3E0", "BF360C")


# ── Performance: cached group-index for repeated equality filters ────────────
# Several builders filter a base DataFrame by a loop key many times, e.g.
# _eq_group(att_f, "session_id", sid) once per session — a full-column scan every
# iteration (O(rows × keys)). _eq_group() computes a groupby index for a
# (DataFrame, column) once, caches it, and turns each lookup into O(1). It returns
# a fresh copy each call, exactly matching the copy semantics AND the row
# content/order of boolean indexing (groupby preserves within-group order), so
# results are identical — only faster. Only used for exact `== scalar` filters;
# `.isin(...)` and `.str.strip() ==` filters are left untouched.
_EQ_INDEX_CACHE = {}


def _eq_group(df, col, key):
    ck = (id(df), col)
    ent = _EQ_INDEX_CACHE.get(ck)
    if ent is None or ent[0] is not df:
        ent = (df, {k: g for k, g in df.groupby(col, sort=False)})
        _EQ_INDEX_CACHE[ck] = ent
    g = ent[1].get(key)
    return g.copy() if g is not None else df.iloc[0:0].copy()


def _side(style="thin"):
    return Side(style=style, color=C_GREY_BD)


def _border(left="thin", right="thin", top="thin", bottom="thin"):
    return Border(
        left   = _side(left)   if left   else None,
        right  = _side(right)  if right  else None,
        top    = _side(top)    if top    else None,
        bottom = _side(bottom) if bottom else None,
    )


# ─────────────────────────────────────────────────────────────────────────────
#  DATE RANGE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_date_range(report_type: str, ref: date):
    """Return (start_date, end_date, human_label) for the given report type."""
    if report_type == "daily":
        return ref, ref, ref.strftime("%d-%b-%Y")

    if report_type == "weekly":
        # Previous full week Mon–Sun (run every Monday)
        mon = ref - timedelta(days=ref.weekday() + 7)
        sun = mon + timedelta(days=6)
        return mon, sun, f"{mon.strftime('%d %b')} – {sun.strftime('%d %b %Y')}"

    if report_type == "fortnightly":
        # Previous 2 full weeks Mon–Sun (run every other Monday)
        sun = ref - timedelta(days=ref.weekday() + 1)   # last Sunday
        mon = sun - timedelta(days=13)                   # Monday 2 weeks before
        return mon, sun, f"{mon.strftime('%d %b')} – {sun.strftime('%d %b %Y')}"

    if report_type == "monthly":
        # Previous calendar month (run on 1st of current month)
        first_this = ref.replace(day=1)
        last_prev  = first_this - timedelta(days=1)
        first_prev = last_prev.replace(day=1)
        return first_prev, last_prev, first_prev.strftime("%B %Y")

    if report_type == "quarterly":
        # Previous calendar quarter
        q = (ref.month - 1) // 3           # 0=Q1,1=Q2,2=Q3,3=Q4
        if q == 0:
            yr   = ref.year - 1
            sm, em, lbl = 10, 12, f"Q4 {yr}"
        else:
            yr   = ref.year
            sm   = (q - 1) * 3 + 1
            em   = sm + 2
            lbl  = f"Q{q} {yr}"
        end_day = calendar.monthrange(yr, em)[1]
        return date(yr, sm, 1), date(yr, em, end_day), lbl

    if report_type == "yearly":
        yr = ref.year - 1
        return date(yr, 1, 1), date(yr, 12, 31), str(yr)

    raise ValueError(f"Unknown report type: {report_type!r}")


def _period_for_date(report_type: str, d):
    """Return (start, end) of the period of report_type that contains date d."""
    if report_type == "daily":
        return d, d
    if report_type == "weekly":
        mon = d - timedelta(days=d.weekday())
        return mon, mon + timedelta(days=6)
    if report_type == "fortnightly":
        mon = d - timedelta(days=d.weekday())
        return mon, mon + timedelta(days=13)
    if report_type == "monthly":
        first = d.replace(day=1)
        last  = d.replace(day=calendar.monthrange(d.year, d.month)[1])
        return first, last
    if report_type == "quarterly":
        q  = (d.month - 1) // 3
        sm = q * 3 + 1
        em = sm + 2
        return date(d.year, sm, 1), date(d.year, em, calendar.monthrange(d.year, em)[1])
    if report_type == "yearly":
        return date(d.year, 1, 1), date(d.year, 12, 31)
    raise ValueError(f"Unknown report type for _period_for_date: {report_type!r}")


def _find_prev_period_with_data(service, report_type: str, current_start):
    """Read Sessions sheet once (no date filter) and return (start, end) of the
    most recent period — of the same report_type — that:
      • ends before current_start, AND
      • contains at least one session row.
    Returns (None, None) when no such period exists."""
    try:
        sess_all = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Sessions")
    except Exception as _e:
        print(f"[Trend] Could not read Sessions sheet for prev-period search: {_e}")
        return None, None

    if sess_all.empty or "start_time_ist" not in sess_all.columns:
        return None, None

    sess_all["_date"] = pd.to_datetime(sess_all["start_time_ist"], errors="coerce").dt.date
    past_dates = sorted(
        {d for d in sess_all["_date"].dropna() if d is not None and d < current_start},
        reverse=True,
    )
    if not past_dates:
        return None, None

    seen_periods: set = set()
    for d in past_dates:
        try:
            p_start, p_end = _period_for_date(report_type, d)
        except Exception:
            continue
        if p_end >= current_start:
            continue
        if (p_start, p_end) in seen_periods:
            continue
        seen_periods.add((p_start, p_end))
        return p_start, p_end

    return None, None


def report_title(report_type: str, label: str, sheet_name: str) -> str:
    prefixes = {
        "daily":       "Daily",
        "weekly":      "Weekly",
        "fortnightly": "Fortnightly",
        "monthly":     "Monthly",
        "quarterly":   "Quarterly",
        "yearly":      "Yearly",
        "manual":      "Custom Period",
    }
    p = prefixes.get(report_type, report_type.title())
    period_word = {
        "daily": "Report Period", "weekly": "Week", "fortnightly": "Fortnight",
        "monthly": "Month", "quarterly": "Quarter", "yearly": "Year",
        "manual": "Period",
    }.get(report_type, "Period")
    return f"{p} {sheet_name}  |  {period_word}: {label}"


# ─────────────────────────────────────────────────────────────────────────────
#  GOOGLE SHEETS READER
# ─────────────────────────────────────────────────────────────────────────────

def read_sheet_df(service, spreadsheet_id: str, tab: str) -> pd.DataFrame:
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"{tab}!A:ZZ",
    ).execute()
    values = result.get("values", [])
    if len(values) < 2:
        return pd.DataFrame()
    header = [h.strip() for h in values[0]]          # strip header names
    rows   = [r + [""] * (len(header) - len(r)) for r in values[1:]]
    df     = pd.DataFrame(rows, columns=header)
    # De-duplicate column names (Google Sheets can have repeated headers)
    seen: dict[str, int] = {}
    new_cols = []
    for c in df.columns:
        if c in seen:
            seen[c] += 1
            new_cols.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            new_cols.append(c)
    df.columns = new_cols
    # Strip whitespace from every string column so groupby never splits on
    # invisible leading/trailing spaces (e.g. "Instructor 2 " vs "Instructor 2").
    # NOTE: we do NOT use select_dtypes(include=["object","string"]) here because
    # pandas 3.0 made the new "str" StringDtype the default for text columns, and
    # that filter can miss those columns — leaving stray spaces that silently
    # split groupby keys. Instead, strip any column that exposes the .str
    # accessor (string-like) and skip the rest. Works on pandas 2.x and 3.x.
    for col in df.columns:
        try:
            df[col] = df[col].str.strip()
        except (AttributeError, TypeError):
            pass  # non-string column (numeric/datetime) → nothing to strip
    return df


def load_all_data(service, start_date: date, end_date: date, time_window=None):
    """Pull all required data from Google Sheets and filter to the date window.

    time_window : optional (start_dt, end_dt) datetime pair. When supplied
    (Daily Report only), rows are filtered on the exact timestamp so the window
    can span part of a day — e.g. yesterday 12:00 PM → now. When None, the
    original inclusive calendar-date filter (start_date … end_date) is used,
    keeping every other report type unchanged."""
    print("[Data] Reading Sessions …")
    sess_df = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Sessions")

    print("[Data] Reading Attendance …")
    att_df  = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Attendance")

    print("[Data] Reading Student_Feedback …")
    fb_df   = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Student_Feedback")

    print("[Data] Reading Teacher_Feedback …")
    tf_df   = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Teacher_Feedback")

    print("[Data] Reading Students (phone lookup) …")
    stu_df  = read_sheet_df(service, STUDENTS_SHEET_ID, "Students")

    # Parse dates
    def add_date_col(df, col):
        if col in df.columns:
            df["_date"] = pd.to_datetime(df[col], errors="coerce").dt.date

    # Parse full timestamps too — only used by the Daily Report's precise window.
    def add_dt_col(df, col):
        if col in df.columns:
            df["_dt"] = pd.to_datetime(df[col], errors="coerce")

    add_date_col(sess_df, "start_time_ist")
    add_date_col(att_df,  "session_start_ist")
    add_date_col(fb_df,   "session_start_ist")
    add_date_col(tf_df,   "session_start_ist")

    add_dt_col(sess_df, "start_time_ist")
    add_dt_col(att_df,  "session_start_ist")
    add_dt_col(fb_df,   "session_start_ist")
    add_dt_col(tf_df,   "session_start_ist")

    # Filter to window
    def in_window(df):
        # Daily Report: precise datetime window (e.g. yesterday noon → now).
        if time_window is not None and "_dt" in df.columns:
            w_start, w_end = time_window
            mask = df["_dt"].apply(
                lambda x: x is not None and not pd.isna(x) and w_start <= x <= w_end
            )
            return df[mask].copy()
        # All other report types: original inclusive calendar-date filter.
        if "_date" not in df.columns:
            return df
        mask = df["_date"].apply(lambda d: d is not None and not pd.isna(d) and start_date <= d <= end_date)
        return df[mask].copy()

    sess_f = in_window(sess_df)
    att_f  = in_window(att_df)
    fb_f   = in_window(fb_df)
    tf_f   = in_window(tf_df)

    # Phone lookup: student_id → phone
    if "student_id" in stu_df.columns and "phone" in stu_df.columns:
        phone_map      = stu_df.set_index("student_id")["phone"].to_dict()
        att_f["phone"] = att_f["student_id"].map(phone_map).fillna("")
    else:
        att_f["phone"] = ""

    # Pre-compute numeric columns
    att_f["_pct_num"] = att_f["attendance_percent"].apply(_parse_pct)
    att_f["_dur_sec"] = pd.to_numeric(att_f["duration"], errors="coerce").fillna(0)
    att_f["_dur_min"] = (att_f["_dur_sec"] / 60).round(1)

    # ── Attendance-applicable flag (is_attendance_required) ───────────────────
    # Only students whose is_attendance_required == "Y"/"y" count toward attendance
    # calculations. Blank or "N" → NOT applicable (they told us they won't attend),
    # so they are still enrolled but excluded from Present/Absent/Att% everywhere.
    if "student_id" in stu_df.columns and "is_attendance_required" in stu_df.columns:
        attn_map = stu_df.drop_duplicates(subset=["student_id"]) \
                         .set_index("student_id")["is_attendance_required"].to_dict()
        att_f["_attn_req"] = att_f["student_id"].map(attn_map).fillna("")
    else:
        att_f["_attn_req"] = ""
    att_f["_attn_applicable"] = att_f["_attn_req"].astype(str).str.strip().str.upper().eq("Y")

    fb_f["_rating"]   = pd.to_numeric(fb_f["rating"], errors="coerce")

    # ── Separate suspended students ──────────────────────────────────────────
    # suspend_status column is written by the pipeline; "Suspended" = suspended.
    # Suspended students are excluded from all attendance calculations but kept
    # for record-keeping in the Student Detail sheet.
    if "suspend_status" in att_f.columns:
        _susp_mask   = att_f["suspend_status"].str.strip().str.lower() == "suspended"
        att_susp     = att_f[_susp_mask].copy()
        att_f        = att_f[~_susp_mask].copy()
        print(f"[Data] Suspended students excluded from calculations: {len(att_susp)} row(s)")
    else:
        att_susp = pd.DataFrame(columns=att_f.columns)

    # ── Backfill blank end_time_ist in Sessions from Attendance session_end_ist ─
    # Built in explicit, individually-guarded steps (no long method-chain) so a
    # missing/renamed column can never raise mid-chain. This mirrors the safe
    # column-existence pattern used elsewhere in this file.
    _BLANKS = ("", "nan", "NaT", "None", "NAN")
    _have_cols = (
        "end_time_ist" in sess_f.columns
        and "session_id" in sess_f.columns
        and "session_end_ist" in att_f.columns
        and "session_id" in att_f.columns
    )
    if _have_cols and not att_f.empty and not sess_f.empty:
        # 1) keep only attendance rows that actually have a usable session_end_ist
        _end_str  = att_f["session_end_ist"].astype(str).str.strip()
        _att_good = att_f.loc[~_end_str.isin(_BLANKS), ["session_id", "session_end_ist"]].copy()
        # 2) one row per session_id → Series mapping session_id → session_end_ist
        _att_good = _att_good.drop_duplicates(subset=["session_id"])
        _att_end  = dict(zip(_att_good["session_id"], _att_good["session_end_ist"]))
        # 3) fill only the blank end_time_ist cells in Sessions
        _blank_mask = sess_f["end_time_ist"].astype(str).str.strip().isin(_BLANKS)
        _n_blank    = int(_blank_mask.sum())
        if _n_blank > 0 and _att_end:
            sess_f.loc[_blank_mask, "end_time_ist"] = (
                sess_f.loc[_blank_mask, "session_id"].map(_att_end)
            )
            _still_blank = int(sess_f["end_time_ist"].astype(str).str.strip().isin(_BLANKS).sum())
            print(f"[Data] Backfilled end_time_ist for {_n_blank - _still_blank}/{_n_blank} session(s) from Attendance.")
    elif not _have_cols:
        print("[Data] Skipped end_time_ist backfill (required columns missing: "
              f"sessions={list(sess_f.columns)[:8]}… attendance={list(att_f.columns)[:8]}…)")

    print(f"[Data] Filtered → Sessions: {len(sess_f)} | Attendance: {len(att_f)} "
          f"(+{len(att_susp)} suspended) | Student FB: {len(fb_f)} | Teacher FB: {len(tf_f)}")
    return sess_f, att_f, fb_f, tf_f, att_susp


# ─────────────────────────────────────────────────────────────────────────────
#  CALCULATION HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_pct(val) -> float:
    try:
        return float(str(val).replace("%", "").strip())
    except Exception:
        return 0.0


def _session_dur_min(row) -> float:
    """Compute session duration in minutes from start_time_ist → end_time_ist.
    Uses pd.to_datetime so any format Google Sheets stores (ISO-T, space-sep,
    with/without timezone) is handled automatically."""
    try:
        s = pd.to_datetime(str(row.get("start_time_ist", "")), errors="coerce")
        e = pd.to_datetime(str(row.get("end_time_ist",   "")), errors="coerce")
        if pd.isna(s) or pd.isna(e):
            return 0.0
        return round((e - s).total_seconds() / 60, 1)
    except Exception:
        return 0.0


def _remarks(pct: float, status: str) -> str:
    if status == "Absent":
        return "■ Absent"
    if pct >= 95: return "■ Excellent"
    if pct >= 85: return "■ Good"
    if pct >= 75: return "■ Satisfactory"
    if pct >= 60: return "■ Low"
    return            "■ Critical"


def _action(pct: float) -> str:
    if pct < 60: return "⚡ Immediate counselling"
    if pct < 75: return "📞 Call & counsel"
    return ""


def _row_bg(pct: float, status: str, row_idx: int) -> str:
    if status == "Absent":   return C_RED_LITE
    if pct < 75:             return C_AMBER
    if pct >= 95:            return C_GREEN
    return C_ROW_ALT if row_idx % 2 == 0 else C_WHITE


# ─────────────────────────────────────────────────────────────────────────────
#  OPENPYXL STYLING HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _font(bold=False, size=10, color=None, italic=False, strike=False):
    return Font(name="Arial", bold=bold, size=size, color=color or "000000",
                italic=italic, strike=strike)


def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_title_row(ws, row: int, col_start: int, col_end: int, text: str, row_h=38):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,  end_column=col_end)
    c           = ws.cell(row=row, column=col_start)
    c.value     = text
    c.font      = _font(bold=True, size=14, color=C_NAV)
    c.alignment = _align("left", "center")
    ws.row_dimensions[row].height = row_h


def style_header_cell(ws, row: int, col: int, text: str):
    c           = ws.cell(row=row, column=col)
    c.value     = text
    c.font      = _font(bold=True, size=10, color=C_WHITE)
    c.fill      = _fill(C_NAV)
    c.alignment = _align("center", "center", wrap=True)
    c.border    = _border()
    return c


def write_header_row(ws, row: int, headers: list, height=32):
    for col, h in enumerate(headers, 1):
        style_header_cell(ws, row, col, h)
    ws.row_dimensions[row].height = height


def style_data_cell(ws, row: int, col: int, value, bg=C_WHITE,
                    bold=False, h_align="left", wrap=False, number_fmt=None):
    c           = ws.cell(row=row, column=col)
    c.value     = value
    c.font      = _font(bold=bold, size=10)
    c.fill      = _fill(bg)
    c.alignment = _align(h_align, "center", wrap=wrap)
    c.border    = _border()
    if number_fmt:
        c.number_format = number_fmt
    return c


def write_section_banner(ws, row: int, n_cols: int, text: str, bg: str, height=26, h_align="left"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c           = ws.cell(row=row, column=1)
    c.value     = text
    c.font      = _font(bold=True, size=11, color=C_WHITE)
    c.fill      = _fill(bg)
    c.alignment = _align(h_align, "center")
    c.border    = _border()
    ws.row_dimensions[row].height = height


def auto_col_width(ws, min_w=10, max_w=55):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(c.value or "").split("\n")[0]) for c in col_cells),
            default=0,
        )
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 3))


def _border_thick_outer(ws, min_row, min_col, max_row, max_col):
    """Apply a medium outer border around a rectangular range."""
    thick = Side(style="medium", color="1A2E5A")
    thin  = Side(style="thin",   color=C_GREY_BD)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            t = thick if cell.row == min_row    else thin
            b = thick if cell.row == max_row    else thin
            l = thick if cell.column == min_col else thin
            r = thick if cell.column == max_col else thin
            cell.border = Border(left=l, right=r, top=t, bottom=b)


def _att_bg(pct: float, status: str) -> str:
    """Return a fill colour for attendance %."""
    if status == "Absent": return C_RED_LITE
    if pct >= 95:          return C_GREEN
    if pct >= 75:          return C_GREEN_PALE
    if pct >= 60:          return C_AMBER
    return C_RED_LITE


def _att_icon(pct: float, status: str) -> str:
    """Emoji prefix for attendance display value."""
    if status == "Absent": return "❌"
    if pct >= 95:          return "🟢"
    if pct >= 75:          return "🟡"
    return                        "🔴"


def _rating_bg(avg: float) -> str:
    """Fill colour for feedback average rating."""
    if avg >= 8:   return C_GREEN
    if avg >= 6:   return C_GREEN_PALE
    if avg >= 4:   return C_AMBER
    return C_RED_LITE


def _write_kpi_strip(ws, row_lbl: int, row_val: int, kpis: list, n_cols: int):
    """
    kpis: list of (label, value, icon, palette) where palette = KPI_BLUE/GREEN/RED/AMBER
    Columns are distributed evenly across n_cols.
    """
    span = max(1, n_cols // len(kpis))
    for idx, (label, value, icon, pal) in enumerate(kpis):
        c_start = idx * span + 1
        c_end   = c_start + span - 1
        if idx == len(kpis) - 1:
            c_end = n_cols          # last KPI fills to edge

        lbl_bg, val_bg, txt_col = pal

        # Merge label row
        if c_start < c_end:
            ws.merge_cells(start_row=row_lbl, start_column=c_start,
                           end_row=row_lbl,   end_column=c_end)
        lc = ws.cell(row=row_lbl, column=c_start)
        lc.value     = f"{icon}  {label}" if icon else label
        lc.font      = _font(bold=True, size=9, color=C_WHITE)
        lc.fill      = _fill(lbl_bg)
        lc.alignment = _align("center", "center")
        lc.border    = _border()

        # Merge value row
        if c_start < c_end:
            ws.merge_cells(start_row=row_val, start_column=c_start,
                           end_row=row_val,   end_column=c_end)
        vc = ws.cell(row=row_val, column=c_start)
        vc.value     = value
        vc.font      = _font(bold=True, size=18, color=txt_col)
        vc.fill      = _fill(val_bg)
        vc.alignment = _align("center", "center")
        vc.border    = _border()

    ws.row_dimensions[row_lbl].height = 20
    ws.row_dimensions[row_val].height = 36


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 1 — SESSION SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def _prefer_instructor_name(sess_f: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of sess_f whose `tutor_name` column holds `Instructor_Name`
    (when present / non-blank), falling back to the original `tutor_name`. This lets
    the Instructor column show Instructor_Name without touching downstream display /
    grouping code. If Instructor_Name is absent, behaviour is unchanged."""
    if "Instructor_Name" not in sess_f.columns:
        return sess_f
    sess_f = sess_f.copy()
    inst = sess_f["Instructor_Name"].astype(str).str.strip()
    base = sess_f["tutor_name"].astype(str) if "tutor_name" in sess_f.columns else ""
    sess_f["tutor_name"] = inst.where(inst != "", base)
    return sess_f


def _applicable(df: pd.DataFrame) -> pd.DataFrame:
    """Subset of attendance rows that count toward attendance calculations
    (is_attendance_required == 'Y'). If the flag column is absent, returns df
    unchanged so existing behaviour is preserved."""
    if "_attn_applicable" in df.columns:
        return df[df["_attn_applicable"]]
    return df


def _na_count(df: pd.DataFrame) -> int:
    """Number of attendance-NOT-applicable rows (blank / N) in the subset."""
    if "_attn_applicable" in df.columns:
        return int((~df["_attn_applicable"]).sum())
    return 0


def build_session_summary(ws, sess_f: pd.DataFrame, att_f: pd.DataFrame,
                          report_type: str, label: str, fb_f: pd.DataFrame = None,
                          prev_trend: dict = None, yest_date=None):
    is_daily  = (report_type == "daily")
    sess_f    = _prefer_instructor_name(sess_f)   # Instructor column ← Instructor_Name
    title     = report_title(report_type, label, "Attendance Report")
    prev_trend = prev_trend or {}

    if is_daily:
        headers = [
            "Tech Name", "Duration", "Instructor", "Session Date",
            "Duration\n(Mins)", "Total\nEnrolled", "Att. N/A\nCount", "Present", "Absent",
            "Att %", "Avg Time\nin Session %", "Att %\nTrend",
        ]
    else:
        headers = [
            "Tech Name", "Duration", "Instructor",
            "Total\nSessions", "Total\nEnrolled", "Att. N/A\nCount",
            "Total\nPresent", "Total\nAbsent",
            "Att %", "Avg Duration\n(min)", "Avg\nRating", "Feedback\nRate %",
        ]
    N = len(headers)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    style_title_row(ws, 1, 1, N, title)
    ws.cell(row=1, column=1).alignment = _align("center", "center")

    # ── KPI strip (rows 3-4) ──────────────────────────────────────────────────
    n_sessions    = len(sess_f)
    # Deduplicate: one row per unique (session_id, student_id) pair to avoid
    # double-counting students who appear multiple times in the same session.
    if fb_f is None:
        fb_f = pd.DataFrame()
    att_dedup = att_f.drop_duplicates(subset=["session_id", "student_id"]) \
                if "session_id" in att_f.columns and "student_id" in att_f.columns \
                else att_f
    # Present/Absent/Att% count ONLY attendance-applicable students.
    att_dedup_app = _applicable(att_dedup)
    total_present = int((att_dedup_app["status"] == "Present").sum())

    if is_daily:
        # ── Daily ─────────────────────────────────────────────────────────────
        # Total Students = total attendance records (one per student per session).
        att_f_app       = _applicable(att_f)
        total_enrolled  = len(att_f)                 # all enrolled (incl. N/A)
        applicable_n    = len(att_f_app)             # attendance-applicable only
        total_present   = int((att_f_app["status"] == "Present").sum())
        total_absent    = int((att_f_app["status"] == "Absent").sum())
        overall_pct     = round(total_present / applicable_n * 100, 2) if applicable_n else 0.0
        kpis = [
            ("Total Sessions", n_sessions,            "📚", KPI_BLUE),
            ("Total Students", total_enrolled,         "👥", KPI_BLUE),
            ("Overall Att %",  f"{overall_pct:.2f}%", "📊",
             KPI_GREEN if overall_pct >= 75 else KPI_RED),
            ("Absent",         total_absent,           "❌", KPI_RED),
            ("Present",        total_present,          "✅", KPI_GREEN),
        ]
    else:
        # ── Weekly / Monthly / Quarterly / Yearly ─────────────────────────────
        # Use deduplicated records: unique (session_id, student_id) pairs so that
        # each student is counted once per session even if the source sheet has
        # duplicate rows for the same student+session combination.
        applicable_n      = len(att_dedup_app)       # attendance-applicable only
        total_absent      = int((att_dedup_app["status"] == "Absent").sum())
        overall_pct       = round(total_present / applicable_n * 100, 1) if applicable_n else 0.0
        n_unique_stu_kpi  = att_f["student_id"].nunique() if "student_id" in att_f.columns else len(att_dedup)
        kpis = [
            ("Total Sessions", n_sessions,            "📚", KPI_BLUE),
            ("Total Students", n_unique_stu_kpi,       "👥", KPI_BLUE),
            ("Overall Att %",  f"{overall_pct:.1f}%", "📊",
             KPI_GREEN if overall_pct >= 75 else KPI_RED),
            ("Absent",         total_absent,           "❌", KPI_RED),
            ("Present",        total_present,          "✅", KPI_GREEN),
        ]
    _write_kpi_strip(ws, row_lbl=3, row_val=4, kpis=kpis, n_cols=N)

    # ── Row 6: Column headers ─────────────────────────────────────────────────
    write_header_row(ws, 6, headers)

    data_start = 7
    row_num    = data_start

    if is_daily:
        # ── Per-session rows ──────────────────────────────────────────────────
        for _, sess in sess_f.iterrows():
            sid         = sess.get("session_id", "")
            sa          = _eq_group(att_f, "session_id", sid)
            sa_app      = _applicable(sa)                 # applicable-only for Present/Absent/Att%
            na_cnt      = _na_count(sa)                   # attendance-N/A students
            present     = sa_app[sa_app["status"] == "Present"]
            absent_n    = len(sa_app) - len(present)
            avg_att     = round(len(present) / len(sa_app) * 100, 1) if len(sa_app) else 0.0
            # Session duration: prefer timestamps; fall back to max student
            # attendance duration so the column always shows a real value.
            dur_min = _session_dur_min(sess)
            if dur_min == 0 and not sa.empty:
                _max_stu_dur = sa["_dur_min"].max()
                if _max_stu_dur and _max_stu_dur > 0:
                    dur_min = round(float(_max_stu_dur), 1)
            # Avg Time = mean(student_dur / session_dur) × 100
            if dur_min > 0 and len(present) > 0:
                avg_time = round(present["_dur_min"].mean() / dur_min * 100, 1)
            else:
                avg_time = 0.0

            # ── Detect cancelled / scheduled sessions ────────────────────────
            # A session is "normal" if students attended OR it has duration,
            # regardless of whether end_time_ist is present.
            # If no attendance/duration AND no end_time:
            #   - Future start time → "scheduled" (hasn't happened yet)
            #   - Past start time   → "cancelled" (should have happened)
            _end_raw = str(sess.get("end_time_ist", "")).strip()
            _end_blank = _end_raw in ("", "nan", "NaT", "None", "NAN")
            _has_attendance = len(sa) > 0
            _has_duration   = dur_min > 0
            _start_raw = str(sess.get("start_time_ist", "")).strip()
            try:
                _start_dt = datetime.strptime(_start_raw[:19], "%Y-%m-%d %H:%M:%S")
                _is_future = _start_dt >= datetime.now()
            except (ValueError, TypeError):
                _is_future = False
            if _has_attendance or _has_duration:
                _sess_type = "normal"
            elif _end_blank and _is_future:
                _sess_type = "scheduled"
            elif _end_blank:
                _sess_type = "cancelled"
            else:
                _sess_type = "scheduled"

            if _sess_type == "cancelled":
                row_bg = C_GREY_CANCEL
            elif _sess_type == "scheduled":
                row_bg = C_LAVENDER
            else:
                row_bg = C_ROW_ALT if row_num % 2 == 0 else C_WHITE
                # Daily Report: tint the previous day's sessions light orange.
                if yest_date is not None and sess.get("_date") == yest_date:
                    row_bg = C_YEST_HL

            # Session Date: show full date+time from start_time_ist (trimmed to 19 chars)
            _session_dt = str(sess.get("start_time_ist", ""))[:19]
            vals   = [sess.get("course_name",""), sess.get("course_title",""),
                      sess.get("tutor_name",""),  _session_dt,
                      dur_min if _sess_type == "normal" else "",
                      len(sa) if _sess_type == "normal" else "",          # Total Enrolled (all)
                      na_cnt if _sess_type == "normal" else "",           # Att. N/A Count
                      len(present) if _sess_type == "normal" else "",     # Present (applicable)
                      absent_n if _sess_type == "normal" else ""]         # Absent (applicable)
            aligns = ["left","left","left","center","center","center","center","center","center"]

            if _sess_type == "cancelled":
                # Cancelled: light grey bg, italic font, grey text
                for col, (v, al) in enumerate(zip(vals, aligns), 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = v if col <= 4 else "Cancelled" if col == 5 else ""
                    c.font      = _font(italic=True, size=10, color="999999")
                    c.fill      = _fill(C_GREY_CANCEL)
                    c.alignment = _align(al, "center")
                    c.border    = _border()
                # Att % and remaining columns
                for col in range(10, N + 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = ""
                    c.font      = _font(italic=True, size=10, color="999999")
                    c.fill      = _fill(C_GREY_CANCEL)
                    c.alignment = _align("center", "center")
                    c.border    = _border()

            elif _sess_type == "scheduled":
                # Scheduled: lavender bg, italic font, deep purple text
                for col, (v, al) in enumerate(zip(vals, aligns), 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = v if col <= 4 else "Scheduled" if col == 5 else ""
                    c.font      = _font(italic=True, size=10, color=C_LAVENDER_TXT)
                    c.fill      = _fill(C_LAVENDER)
                    c.alignment = _align(al, "center")
                    c.border    = _border()
                # Att % and remaining columns
                for col in range(10, N + 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = ""
                    c.font      = _font(italic=True, size=10, color=C_LAVENDER_TXT)
                    c.fill      = _fill(C_LAVENDER)
                    c.alignment = _align("center", "center")
                    c.border    = _border()

            else:
                # Normal session — existing styling
                for col, (v, al) in enumerate(zip(vals, aligns), 1):
                    style_data_cell(ws, row_num, col, v, bg=row_bg, h_align=al)

                ac = ws.cell(row=row_num, column=10)
                ac.value        = avg_att
                ac.font         = _font(bold=True, size=10,
                                        color=C_GREEN_DARK if avg_att >= 75 else C_RED_DARK)
                ac.fill         = _fill(_att_bg(avg_att, "Present"))
                ac.alignment    = _align("center", "center")
                ac.border       = _border()
                ac.number_format = '0.0"%"'
                tc = style_data_cell(ws, row_num, 11, avg_time, bg=row_bg, h_align="center")
                tc.number_format = '0.0"%"'

            # Att % Trend (col 12) — compare with same course/tutor on previous session day
            _tn   = str(sess.get("tutor_name", ""))
            _cn   = str(sess.get("course_name", ""))
            prev_att = prev_trend.get((_cn, _tn))
            if _sess_type == "normal":
                _write_trend_cell(ws, row_num, 12, avg_att, prev_att, bg=row_bg)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

    else:
        # ── Grouped by Course + Subject + Tutor (weekly / monthly / etc.) ─────
        grp_keys = ["course_name", "course_title", "tutor_name"]
        for k in grp_keys:
            if k not in sess_f.columns:
                sess_f[k] = ""

        for (cn, ct, tn), grp in sess_f.groupby(grp_keys, sort=True):
            sids      = grp["session_id"].tolist()
            # Deduplicate: count each student once per session (ignore duplicate rows)
            sa        = att_dedup[att_dedup["session_id"].isin(sids)]
            sa_app    = _applicable(sa)                 # applicable-only for Present/Absent/Att%
            cfb       = fb_f[fb_f["course_name"].str.strip() == cn.strip()] \
                        if not fb_f.empty and "course_name" in fb_f.columns else pd.DataFrame()
            n_sess    = len(sids)
            # Total Enrolled = unique (session_id, student_id) pairs for the period
            # (includes attendance-N/A students); Att. N/A Count is broken out and
            # Present/Absent/Att% consider only attendance-applicable students.
            tot_enr   = len(sa)
            na_cnt    = _na_count(sa)
            applic_n  = len(sa_app)
            tot_pres  = int((sa_app["status"] == "Present").sum())
            tot_abs   = int((sa_app["status"] == "Absent").sum())
            att_pct   = round(tot_pres / applic_n * 100, 1) if applic_n else 0.0
            # Avg session duration from session start/end timestamps
            _durs         = [_session_dur_min(r) for _, r in grp.iterrows()]
            _valid_durs   = [d for d in _durs if d > 0]
            avg_dur       = round(sum(_valid_durs) / len(_valid_durs), 1) if _valid_durs else 0.0
            # Feedback stats
            fb_count      = len(cfb)
            # Feedback Rate % = feedbacks / present students (not total enrolled)
            fb_rate       = round(fb_count / tot_pres * 100, 1) if tot_pres else 0.0
            avg_rating    = round(pd.to_numeric(cfb["rating"], errors="coerce").mean(), 2) \
                            if not cfb.empty else 0.0
            row_bg        = C_ROW_ALT if row_num % 2 == 0 else C_WHITE

            vals   = [cn, ct, tn, n_sess, tot_enr, na_cnt, tot_pres, tot_abs]
            aligns = ["left","left","left","center","center","center","center","center"]
            for col, (v, al) in enumerate(zip(vals, aligns), 1):
                style_data_cell(ws, row_num, col, v, bg=row_bg, h_align=al)

            ac = ws.cell(row=row_num, column=9)
            ac.value        = att_pct
            ac.font         = _font(bold=True, size=10,
                                    color=C_GREEN_DARK if att_pct >= 75 else C_RED_DARK)
            ac.fill         = _fill(_att_bg(att_pct, "Present"))
            ac.alignment    = _align("center", "center")
            ac.border       = _border()
            ac.number_format = '0.0"%"'
            style_data_cell(ws, row_num, 10, avg_dur,    bg=row_bg, h_align="center")
            style_data_cell(ws, row_num, 11, avg_rating, bg=row_bg, h_align="center")
            fc = ws.cell(row=row_num, column=12)
            fc.value = fb_rate; fc.font = _font(size=10); fc.fill = _fill(row_bg)
            fc.alignment = _align("center","center"); fc.border = _border()
            fc.number_format = '0.0"%"'
            ws.row_dimensions[row_num].height = 18
            row_num += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    if row_num > data_start:
        de  = row_num - 1
        CL  = get_column_letter
        for col in range(1, N + 1):
            c           = ws.cell(row=row_num, column=col)
            c.fill      = _fill(C_BLUE_LITE)
            c.font      = _font(bold=True, size=10, color=C_NAV)
            c.border    = _border()
            c.alignment = _align("center", "center")

        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        ws.cell(row=row_num, column=1).value     = "⬛  TOTAL / AVERAGE"
        ws.cell(row=row_num, column=1).alignment = _align("left", "center")

        if is_daily:
            # Cols: 6 Total Enrolled, 7 Att.N/A, 8 Present, 9 Absent (sums);
            #       10 Att %, 11 Avg Time (averages); 12 Trend blank.
            for _c in (6, 7, 8, 9):
                ws.cell(row=row_num, column=_c).value = f"=SUM({CL(_c)}{data_start}:{CL(_c)}{de})"
            ws.cell(row=row_num, column=10).value         = f"=IFERROR(AVERAGE({CL(10)}{data_start}:{CL(10)}{de}),0)"
            ws.cell(row=row_num, column=10).number_format = '0.00"%"'
            ws.cell(row=row_num, column=11).value         = f"=IFERROR(AVERAGE({CL(11)}{data_start}:{CL(11)}{de}),0)"
            ws.cell(row=row_num, column=11).number_format = '0.00"%"'
            # col 12 (Trend) — leave blank in totals
        else:
            # Cols: 4 Sessions, 5 Total Enrolled, 6 Att.N/A, 7 Present, 8 Absent (sums);
            #       9 Att %, 10 Avg Duration, 11 Avg Rating, 12 Feedback Rate (averages).
            for _c in (4, 5, 6, 7, 8):
                ws.cell(row=row_num, column=_c).value = f"=SUM({CL(_c)}{data_start}:{CL(_c)}{de})"
            for _c in (9, 10, 11, 12):
                ws.cell(row=row_num, column=_c).value = f"=IFERROR(AVERAGE({CL(_c)}{data_start}:{CL(_c)}{de}),0)"

        ws.row_dimensions[row_num].height = 22
        _border_thick_outer(ws, data_start, 1, row_num, N)

    ws.auto_filter.ref = f"A6:{get_column_letter(N)}{row_num}"
    ws.freeze_panes    = "A7"
    ws.sheet_properties.tabColor = "1A2E5A"
    auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 2 — STUDENT DETAIL
# ─────────────────────────────────────────────────────────────────────────────

def build_student_detail(ws, sess_f: pd.DataFrame, att_f: pd.DataFrame,
                         report_type: str, label: str, fb_f: pd.DataFrame = None,
                         prev_sd_lookup: dict = None, att_susp: pd.DataFrame = None,
                         yest_date=None):
    is_daily      = (report_type == "daily")
    title         = report_title(report_type, label, "Student Detail Report")
    if fb_f is None:
        fb_f = pd.DataFrame()
    if att_susp is None:
        att_susp = pd.DataFrame(columns=att_f.columns)
    prev_sd_lookup = prev_sd_lookup or {}

    if is_daily:
        headers = [
            "Tech Name", "Duration", "Student Name", "Phone", "Email",
            "Status", "Attendance %", "Duration\n(min)",
            "Joined At", "Left At", "Remarks", "Feedback\nGiven?",
            "Prev Session\nAttendance",
        ]
    else:
        headers = [
            "Tech Name", "Duration", "Student Name", "Phone", "Email",
            "Total\nSessions", "Present", "Absent", "Attendance %",
            "Remarks",
        ]
    N = len(headers)

    style_title_row(ws, 1, 1, N, title)
    ws.cell(row=1, column=1).alignment = _align("center", "center")
    write_header_row(ws, 3, headers)

    row_num = 4

    if is_daily:
        # ── Per-student-per-session rows (daily), grouped by Course + Subject ──
        _tmp = att_f.copy()
        _tmp["_sort_absent"] = (_tmp["status"] != "Absent").astype(int)
        sorted_att = _tmp.sort_values(
            ["course_name", "course_title", "_sort_absent", "_pct_num"],
            na_position="last"
        )
        # Build a set of (session_id, student_id) pairs that have given feedback
        _fb_pairs = set()
        if not fb_f.empty and "session_id" in fb_f.columns and "student_id" in fb_f.columns:
            for _, _fr in fb_f.iterrows():
                _fb_pairs.add((str(_fr.get("session_id","")), str(_fr.get("student_id",""))))

        _current_course_key = None
        _row_alt_i          = 0
        for _, r in sorted_att.iterrows():
            _cn  = r.get("course_name", "")
            _ct  = r.get("course_title", "")
            _key = (_cn, _ct)
            if _key != _current_course_key:
                _current_course_key = _key
                _row_alt_i          = 0
                _banner = f"  {_cn}  —  {_ct}" if _ct else f"  {_cn}"
                write_section_banner(ws, row_num, N, _banner, C_BLUE_MID, h_align="center")
                row_num += 1

            # ── Attendance Not Applicable: shown under the group, dark-grey +
            #    strike-through, Status = "Not Applicable", no attendance numbers. ─
            if "_attn_applicable" in r.index and not bool(r.get("_attn_applicable", True)):
                _na_vals = [r.get("course_name",""), r.get("course_title",""),
                            r.get("student_name",""), r.get("phone",""), r.get("email","")]
                for col in range(1, N + 1):
                    c = ws.cell(row=row_num, column=col)
                    if col <= 5:
                        c.value = _na_vals[col - 1]
                    elif col == 6:
                        c.value = "Not Applicable"
                    else:
                        c.value = "—"
                    c.font      = _font(size=10, color=C_WHITE, strike=True, bold=(col == 6))
                    c.fill      = _fill(C_GREY_NA)
                    c.alignment = _align("left" if col <= 5 else "center", "center")
                    c.border    = _border()
                ws.row_dimensions[row_num].height = 18
                row_num += 1
                continue

            status      = str(r.get("status", ""))
            pct         = r["_pct_num"]
            row_bg      = _row_bg(pct, status, _row_alt_i)
            # Daily Report: tint the previous day's rows light orange. The Status
            # and Attendance % cells keep their own semantic fills below, so the
            # present/absent signal is preserved.
            if yest_date is not None and r.get("_date") == yest_date:
                row_bg = C_YEST_HL
            _row_alt_i += 1
            icon        = _att_icon(pct, status)
            remark_text = _remarks(pct, status)
            status_disp = "❌  Absent" if status == "Absent" else "✅  Present"
            att_disp    = f"{icon}  Absent" if status == "Absent" else f"{icon}  {pct:.1f}%"

            for col, (v, al) in enumerate(zip(
                [r.get("course_name",""), r.get("course_title",""),
                 r.get("student_name",""), r.get("phone",""), r.get("email","")],
                ["left","left","left","left","left"]
            ), 1):
                style_data_cell(ws, row_num, col, v, bg=row_bg, h_align=al)

            sc = ws.cell(row=row_num, column=6)
            sc.value     = status_disp
            sc.font      = _font(bold=True, size=10,
                                 color=C_RED_DARK if status=="Absent" else C_GREEN_DARK)
            sc.fill      = _fill(C_RED_LITE if status=="Absent" else C_GREEN)
            sc.alignment = _align("center","center"); sc.border = _border()

            ac = ws.cell(row=row_num, column=7)
            ac.value     = att_disp
            ac.font      = _font(bold=(status=="Absent" or pct<75), size=10,
                                 color=C_RED_DARK if status=="Absent" or pct<75
                                 else C_GREEN_DARK)
            ac.fill      = _fill(_att_bg(pct, status))
            ac.alignment = _align("center","center"); ac.border = _border()

            style_data_cell(ws, row_num, 8,
                            r["_dur_min"] if r["_dur_min"]>0 else "", bg=row_bg, h_align="center")
            style_data_cell(ws, row_num, 9,
                            str(r.get("first_join_ist",""))[:19], bg=row_bg, h_align="center")
            style_data_cell(ws, row_num, 10,
                            str(r.get("last_leave_ist",""))[:19], bg=row_bg, h_align="center")

            rc = ws.cell(row=row_num, column=11)
            rc.value     = remark_text
            rc.font      = _font(size=10,
                                 color=C_RED_DARK if status=="Absent" or pct<75
                                 else C_GREEN_DARK if pct>=95 else "000000")
            rc.fill      = _fill(row_bg)
            rc.alignment = _align("left","center"); rc.border = _border()

            # Feedback Given? column (col 12)
            _sid_str = str(r.get("session_id",""))
            _stid_str = str(r.get("student_id",""))
            if status == "Absent":
                _fb_val   = "Absent"
                _fb_color = C_GREY_BD
                _fb_bg    = C_GREY_LITE
            elif (_sid_str, _stid_str) in _fb_pairs:
                _fb_val   = "✅  Yes"
                _fb_color = C_GREEN_DARK
                _fb_bg    = C_GREEN
            else:
                _fb_val   = "❌  No"
                _fb_color = C_RED_DARK
                _fb_bg    = C_RED_LITE
            fbc = ws.cell(row=row_num, column=12)
            fbc.value     = _fb_val
            fbc.font      = _font(bold=True, size=10, color=_fb_color)
            fbc.fill      = _fill(_fb_bg)
            fbc.alignment = _align("center", "center")
            fbc.border    = _border()

            # Prev Session Attendance (col 13) — Y / N / —
            _pcn   = str(r.get("course_name", ""))
            _pstid = str(r.get("student_id",  ""))
            _prev_key = (_pcn, _pstid)
            if _prev_key in prev_sd_lookup:
                if prev_sd_lookup[_prev_key]:
                    _prev_val   = "✅  Y"
                    _prev_color = C_GREEN_DARK
                    _prev_bg    = C_GREEN
                else:
                    _prev_val   = "❌  N"
                    _prev_color = C_RED_DARK
                    _prev_bg    = C_RED_LITE
            else:
                _prev_val   = "—"
                _prev_color = C_GREY_BD
                _prev_bg    = C_GREY_LITE
            psc = ws.cell(row=row_num, column=13)
            psc.value     = _prev_val
            psc.font      = _font(bold=True, size=10, color=_prev_color)
            psc.fill      = _fill(_prev_bg)
            psc.alignment = _align("center", "center")
            psc.border    = _border()

            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # ── Append suspended students (daily) ────────────────────────────────
        if not att_susp.empty:
            # Filter suspended rows to sessions in this date range
            susp_in_range = att_susp[att_susp["session_id"].isin(
                sess_f["session_id"].tolist() if "session_id" in sess_f.columns else []
            )] if not att_susp.empty else att_susp
            if not susp_in_range.empty:
                write_section_banner(ws, row_num, N,
                                     "  Suspended Students (excluded from calculations)",
                                     C_GREY_BD, h_align="center")
                row_num += 1
                for _, r in susp_in_range.iterrows():
                    for col, (v, al) in enumerate(zip(
                        [r.get("course_name",""), r.get("course_title",""),
                         r.get("student_name",""), r.get("phone",""), r.get("email","")],
                        ["left","left","left","left","left"]
                    ), 1):
                        c = ws.cell(row=row_num, column=col)
                        c.value     = v
                        c.font      = _font(italic=True, size=10, color="666666")
                        c.fill      = _fill(C_GREY_SUSP)
                        c.alignment = _align(al, "center")
                        c.border    = _border()
                    # Status column — "Suspended" in bold
                    sc = ws.cell(row=row_num, column=6)
                    sc.value     = "Suspended"
                    sc.font      = _font(bold=True, size=10, color="000000")
                    sc.fill      = _fill(C_GREY_SUSP)
                    sc.alignment = _align("center", "center")
                    sc.border    = _border()
                    # Fill remaining columns with grey
                    for col in range(7, N + 1):
                        c = ws.cell(row=row_num, column=col)
                        c.value     = "—"
                        c.font      = _font(italic=True, size=10, color="999999")
                        c.fill      = _fill(C_GREY_SUSP)
                        c.alignment = _align("center", "center")
                        c.border    = _border()
                    ws.row_dimensions[row_num].height = 18
                    row_num += 1

    else:
        # ── Grouped by Course + Subject + Student (non-daily) ─────────────────
        # Only count sessions that actually appear in sess_f for this period
        valid_sids = set(sess_f["session_id"].tolist()) if "session_id" in sess_f.columns else set()
        att_f_det  = att_f[att_f["session_id"].isin(valid_sids)] if valid_sids else att_f

        grp_keys = ["course_name", "course_title", "student_name"]
        records  = []
        for (cn, ct, sn), grp in att_f_det.groupby(grp_keys, sort=True):
            # Attendance-NOT-applicable student (is_attendance_required != Y): still
            # shown under the Tech group, but excluded from all attendance numbers.
            na    = ("_attn_applicable" in grp.columns and not bool(grp["_attn_applicable"].any()))
            # Use Attendance sheet records directly — one row per session per student
            tot   = grp["session_id"].nunique()
            pres  = int((grp["status"] == "Present").sum())
            abst  = int((grp["status"] == "Absent").sum())   # direct count, not tot - pres
            pct   = round(pres / tot * 100, 1) if tot else 0.0
            phone = grp["phone"].iloc[0] if "phone" in grp.columns else ""
            email = grp["email"].iloc[0]  if "email"  in grp.columns else ""
            records.append(dict(cn=cn, ct=ct, sn=sn, phone=phone, email=email,
                                tot=tot, pres=pres, abst=abst, pct=pct, na=na))

        # Sort: course → Not-Applicable last → fully absent first → att% ascending
        records.sort(key=lambda x: (x["cn"], 1 if x.get("na") else 0,
                                    0 if x["pres"]==0 else 1, x["pct"]))

        for i, r in enumerate(records):
            # ── Attendance Not Applicable: dark-grey, strike-through, Status label ─
            if r.get("na"):
                _na_cols = [r["cn"], r["ct"], r["sn"], r["phone"], r["email"],
                            "Not Applicable", "—", "—", "Not Applicable", "Attendance not applicable"]
                for col, v in enumerate(_na_cols, 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = v
                    c.font      = _font(size=10, color=C_WHITE, strike=True,
                                        bold=(col == 6))
                    c.fill      = _fill(C_GREY_NA)
                    c.alignment = _align("left" if col in (1,2,3,4,5,10) else "center", "center")
                    c.border    = _border()
                ws.row_dimensions[row_num].height = 18
                row_num += 1
                continue

            status      = "Absent" if r["pres"] == 0 else "Present"
            pct         = r["pct"]
            row_bg      = _row_bg(pct, status, i)
            icon        = _att_icon(pct, status)
            remark_text = _remarks(pct, status)
            att_disp    = f"{icon}  0.0% (Never)" if status=="Absent" else f"{icon}  {pct:.1f}%"

            for col, (v, al) in enumerate(zip(
                [r["cn"], r["ct"], r["sn"], r["phone"], r["email"]],
                ["left","left","left","left","left"]
            ), 1):
                style_data_cell(ws, row_num, col, v, bg=row_bg, h_align=al)

            for col, (v, al) in enumerate(zip(
                [r["tot"], r["pres"], r["abst"]],
                ["center","center","center"]
            ), 6):
                style_data_cell(ws, row_num, col, v, bg=row_bg, h_align=al)

            ac = ws.cell(row=row_num, column=9)
            ac.value     = att_disp
            ac.font      = _font(bold=(status=="Absent" or pct<75), size=10,
                                 color=C_RED_DARK if status=="Absent" or pct<75
                                 else C_GREEN_DARK)
            ac.fill      = _fill(_att_bg(pct, status))
            ac.alignment = _align("center","center"); ac.border = _border()

            rc = ws.cell(row=row_num, column=10)
            rc.value     = remark_text
            rc.font      = _font(size=10,
                                 color=C_RED_DARK if status=="Absent" or pct<75
                                 else C_GREEN_DARK if pct>=95 else "000000")
            rc.fill      = _fill(row_bg)
            rc.alignment = _align("left","center"); rc.border = _border()

            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # ── Append suspended students (non-daily) ────────────────────────────
        if not att_susp.empty:
            _susp_names = att_susp.drop_duplicates(
                subset=["student_id", "course_name"]
            ) if "student_id" in att_susp.columns else att_susp
            if not _susp_names.empty:
                write_section_banner(ws, row_num, N,
                                     "  Suspended Students (excluded from calculations)",
                                     C_GREY_BD, h_align="center")
                row_num += 1
                for _, r in _susp_names.iterrows():
                    phone = r.get("phone", "") if "phone" in _susp_names.columns else ""
                    email = r.get("email", "") if "email" in _susp_names.columns else ""
                    for col, (v, al) in enumerate(zip(
                        [r.get("course_name",""), r.get("course_title",""),
                         r.get("student_name",""), phone, email],
                        ["left","left","left","left","left"]
                    ), 1):
                        c = ws.cell(row=row_num, column=col)
                        c.value     = v
                        c.font      = _font(italic=True, size=10, color="666666")
                        c.fill      = _fill(C_GREY_SUSP)
                        c.alignment = _align(al, "center")
                        c.border    = _border()
                    # Sessions/Present/Absent columns — dashes
                    for col in range(6, 9):
                        c = ws.cell(row=row_num, column=col)
                        c.value     = "—"
                        c.font      = _font(italic=True, size=10, color="999999")
                        c.fill      = _fill(C_GREY_SUSP)
                        c.alignment = _align("center", "center")
                        c.border    = _border()
                    # Attendance % column
                    ac = ws.cell(row=row_num, column=9)
                    ac.value     = "Suspended"
                    ac.font      = _font(bold=True, size=10, color="000000")
                    ac.fill      = _fill(C_GREY_SUSP)
                    ac.alignment = _align("center", "center")
                    ac.border    = _border()
                    # Remarks column
                    rc = ws.cell(row=row_num, column=10)
                    rc.value     = "Suspended"
                    rc.font      = _font(bold=True, size=10, color="000000")
                    rc.fill      = _fill(C_GREY_SUSP)
                    rc.alignment = _align("left", "center")
                    rc.border    = _border()
                    ws.row_dimensions[row_num].height = 18
                    row_num += 1

    if row_num > 4:
        _border_thick_outer(ws, 4, 1, row_num - 1, N)

    ws.auto_filter.ref = f"A3:{get_column_letter(N)}{row_num - 1}"
    ws.freeze_panes    = "A4"
    ws.sheet_properties.tabColor = "2E75B6"
    auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 3 — ABSENT & AT-RISK
# ─────────────────────────────────────────────────────────────────────────────

def build_absent_atrisk(ws, att_f: pd.DataFrame, report_type: str, label: str):
    title    = report_title(report_type, label, "Absent & At-Risk Students — Action Required")
    N        = 6
    is_daily = (report_type == "daily")

    if is_daily:
        absent  = att_f[att_f["status"] == "Absent"].copy()
        atrisk  = att_f[(att_f["status"] == "Present") & (att_f["_pct_num"] < 75)].copy()
        # Build display records for absent
        absent_recs = [{"name": f"{r.get('course_name','')} | {r.get('course_title','')} | {r.get('student_name','')}",
                        "phone": r.get("phone",""), "email": r.get("email",""), "pct": r["_pct_num"]}
                       for _, r in absent.sort_values(["course_name","student_name"]).iterrows()]
        atrisk_recs = [{"name": f"{r.get('course_name','')} | {r.get('course_title','')} | {r.get('student_name','')}",
                        "phone": r.get("phone",""), "pct": r["_pct_num"],
                        "action": _action(r["_pct_num"])}
                       for _, r in atrisk.sort_values("_pct_num").iterrows()]
    else:
        # Cumulative grouping: aggregate by (course, title, student)
        all_recs = []
        for (cn, ct, sn), grp in att_f.groupby(
                ["course_name","course_title","student_name"], sort=True):
            tot  = len(grp)
            pres = int((grp["status"] == "Present").sum())
            pct  = round(pres / tot * 100, 1) if tot else 0.0
            ph   = grp["phone"].iloc[0] if "phone" in grp.columns else ""
            em   = grp["email"].iloc[0]  if "email"  in grp.columns else ""
            all_recs.append(dict(name=f"{cn} | {ct} | {sn}",
                                 phone=ph, email=em, pct=pct, pres=pres, tot=tot))

        absent_recs = [r for r in sorted(all_recs, key=lambda x: x["name"])
                       if r["pres"] == 0]
        atrisk_recs = [{"name": r["name"], "phone": r["phone"],
                        "pct":  r["pct"],  "action": _action(r["pct"])}
                       for r in sorted(all_recs, key=lambda x: x["pct"])
                       if r["pres"] > 0 and r["pct"] < 75]

    n_absent = len(absent_recs)
    n_risk   = len(atrisk_recs)
    n_action = n_absent + n_risk

    style_title_row(ws, 1, 1, N, title)

    # ── Summary band ──────────────────────────────────────────────────────────
    for col, txt in enumerate(["❌ Absent & At-Risk", "", "Total Absent", "At Risk (< 75 %)", "Action Needed", ""], 1):
        c = ws.cell(row=3, column=col)
        c.value     = txt
        c.font      = _font(bold=True, size=11, color=C_WHITE)
        c.fill      = _fill(C_RED_DARK)
        c.alignment = _align("center", "center")
        c.border    = _border()
    ws.row_dimensions[3].height = 26

    for col, v in [(3, n_absent), (4, n_risk), (5, n_action)]:
        style_data_cell(ws, 4, col, v, bg=C_RED_LITE, bold=True, h_align="center")
    for col in [1, 2, 6]:
        style_data_cell(ws, 4, col, "", bg=C_RED_LITE)
    ws.row_dimensions[4].height = 22

    row_num = 6

    # ── Absent section ─────────────────────────────────────────────────────────
    write_section_banner(ws, row_num, N, f"ABSENT STUDENTS ({n_absent})", C_RED_DARK)
    row_num += 1
    write_header_row(ws, row_num, ["#", "Course  |  Title  |  Student Name", "Phone", "Email", "Follow-Up Done?", ""])
    row_num += 1

    for i, r in enumerate(absent_recs, 1):
        bg = C_RED_LITE if i % 2 == 1 else "FFD0D0"
        for col, (v, al) in enumerate(
            zip([i, r["name"], r.get("phone",""), r.get("email",""), "", ""],
                ["center","left","left","left","center","center"]), 1
        ):
            style_data_cell(ws, row_num, col, v, bg=bg, h_align=al)
        row_num += 1

    row_num += 1  # spacer

    # ── At-Risk section ────────────────────────────────────────────────────────
    write_section_banner(ws, row_num, N,
                         f"AT-RISK STUDENTS — Low Attendance (< 75 %)  ({n_risk})", C_ORANGE)
    row_num += 1
    write_header_row(ws, row_num, ["#", "Course  |  Title  |  Student Name", "Phone", "Att %", "Action Required", ""])
    row_num += 1

    for i, r in enumerate(atrisk_recs, 1):
        bg = C_AMBER if i % 2 == 1 else "FFE8A0"
        for col, (v, al) in enumerate(
            zip([i, r["name"], r.get("phone",""), f"{r['pct']:.1f}%", r.get("action",""), ""],
                ["center","left","left","center","left","center"]), 1
        ):
            style_data_cell(ws, row_num, col, v, bg=bg, h_align=al)
        row_num += 1

    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "B71C1C"
    auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 4 — FEEDBACK RATING
# ─────────────────────────────────────────────────────────────────────────────

def build_feedback_rating(ws, sess_f: pd.DataFrame, att_f: pd.DataFrame,
                          fb_f: pd.DataFrame, report_type: str, label: str,
                          prev_trend: dict = None, yest_date=None):
    is_daily   = (report_type == "daily")
    sess_f     = _prefer_instructor_name(sess_f)   # Instructor column ← Instructor_Name
    title      = report_title(report_type, label, "Session Feedback Rating Report")
    prev_trend = prev_trend or {}

    if is_daily:
        headers = [
            "Tech Name", "Duration", "Instructor", "Session Date",
            "Total\nEnrolled", "Total\nPresent", "No. of\nFeedbacks", "Feedback\nRate %",
            "⭐ Avg Rating\n(/10)", "Min\nRating", "Max\nRating", "Avg Rating\nTrend",
        ]
    else:
        headers = [
            "Tech Name", "Duration", "Instructor",
            "Total\nSessions", "Total\nFeedbacks", "Avg Feedback\nRate %", "⭐ Avg Rating\n(/10)",
        ]
    N = len(headers)

    style_title_row(ws, 1, 1, N, title)
    ws.cell(row=1, column=1).alignment = _align("center", "center")

    # ── Pre-compute per-course avg feedback rates for accurate KPI ────────────
    # (Deduplicate by session_id + student_id to avoid double-counting)
    def _dedup_fb(subset_fb):
        if subset_fb.empty:
            return subset_fb
        dedup_cols = [c for c in ["session_id", "student_id"] if c in subset_fb.columns]
        return subset_fb.drop_duplicates(subset=dedup_cols) if dedup_cols else subset_fb

    _pre_rates = []
    if not is_daily:
        _pre_gkeys = ["course_name", "course_title", "tutor_name"]
        for k in _pre_gkeys:
            if k not in sess_f.columns:
                sess_f[k] = ""
        for (_pcn, _pct, _ptn), _pgrp in sess_f.groupby(_pre_gkeys, sort=True):
            _psids  = _pgrp["session_id"].tolist()
            _psa_raw= att_f[att_f["session_id"].isin(_psids)]
            _pdd_k  = [c for c in ["session_id","student_id"] if c in _psa_raw.columns]
            _psa    = _psa_raw.drop_duplicates(subset=_pdd_k) if _pdd_k else _psa_raw
            _psf    = _dedup_fb(fb_f[fb_f["session_id"].isin(_psids)] if not fb_f.empty else fb_f)
            # Use present count as denominator (attendance-applicable students only)
            _ppres  = int((_applicable(_psa)["status"] == "Present").sum())
            _pre_rates.append(round(len(_psf) / _ppres * 100, 1) if _ppres else 0.0)

    # ── KPI strip ─────────────────────────────────────────────────────────────
    _fb_deduped  = _dedup_fb(fb_f)
    total_fb     = len(_fb_deduped) if not _fb_deduped.empty else 0
    # Feedback Rate KPI = feedbacks / total present (dedup) students
    _dd_att_k    = [c for c in ["session_id","student_id"] if c in att_f.columns]
    _att_dd_kpi  = att_f.drop_duplicates(subset=_dd_att_k) if _dd_att_k else att_f
    _total_pres  = int((_applicable(_att_dd_kpi)["status"] == "Present").sum())
    # Feedback Rate KPI = average of per-course Avg Feedback Rate % values
    fb_rate      = round(sum(_pre_rates) / len(_pre_rates), 1) if _pre_rates else (
        round(total_fb / _total_pres * 100, 1) if _total_pres else 0.0
    )
    avg_rating  = round(_fb_deduped["_rating"].mean(), 2) if total_fb else 0.0
    no_fb_cnt   = len(sess_f) - (
        len(_fb_deduped["session_id"].dropna().unique()) if not _fb_deduped.empty else 0
    )
    kpis_fb = [
        ("Total Sessions",    len(sess_f),                        "📚", KPI_BLUE),
        ("Total Present",     _total_pres,                        "👥", KPI_BLUE),
        ("Feedback Received", total_fb,                           "💬", KPI_BLUE),
        ("Feedback Rate",     f"{fb_rate:.1f}%",                 "📈",
         KPI_GREEN if fb_rate >= 50 else KPI_AMBER),
        ("Avg Rating (/10)",  f"{avg_rating:.2f}" if total_fb else "—",
         "⭐", KPI_GREEN if avg_rating >= 7 else KPI_AMBER),
    ]
    _write_kpi_strip(ws, row_lbl=3, row_val=4, kpis=kpis_fb, n_cols=N)
    write_header_row(ws, 6, headers)

    data_start = 7
    row_num    = data_start

    def _write_rating_cell(row, col, avg_r):
        rating_str = f"⭐ {avg_r:.2f}" if avg_r is not None else "—"
        rc = ws.cell(row=row, column=col)
        rc.value     = rating_str
        rc.font      = _font(bold=True, size=10,
                             color=C_GREEN_DARK if avg_r and avg_r >= 7
                             else C_AMBER_DARK  if avg_r and avg_r >= 5
                             else C_RED_DARK)
        rc.fill      = _fill(_rating_bg(avg_r) if avg_r is not None else C_GREY_LITE)
        rc.alignment = _align("center", "center")
        rc.border    = _border()

    if is_daily:
        for _, sess in sess_f.iterrows():
            sid      = sess.get("session_id", "")
            sa       = _eq_group(att_f, "session_id", sid)
            sa_app   = _applicable(sa)                    # applicable-only for Present
            sf       = _dedup_fb(_eq_group(fb_f, "session_id", sid) if not fb_f.empty else pd.DataFrame())
            n_stu    = len(sa)                            # Total Enrolled (all)
            n_present = int((sa_app["status"] == "Present").sum())   # Total Present (applicable)
            n_fb     = len(sf)
            # Feedback Rate % = feedbacks / students actually PRESENT (not enrolled).
            fb_rt    = round(n_fb / n_present * 100, 1) if n_present else 0.0
            _ratings = sf["_rating"].dropna().tolist() if n_fb and "_rating" in sf.columns else []
            avg_r    = round(sum(_ratings) / len(_ratings), 2) if _ratings else None
            min_r    = int(min(_ratings)) if _ratings else "—"
            max_r    = int(max(_ratings)) if _ratings else "—"

            # ── Detect cancelled / scheduled ─────────────────────────────────
            dur_min = _session_dur_min(sess)
            _end_raw = str(sess.get("end_time_ist", "")).strip()
            _end_blank = _end_raw in ("", "nan", "NaT", "None", "NAN")
            _has_attendance = len(sa) > 0
            _has_duration   = dur_min > 0
            _start_raw = str(sess.get("start_time_ist", "")).strip()
            try:
                _start_dt = datetime.strptime(_start_raw[:19], "%Y-%m-%d %H:%M:%S")
                _is_future = _start_dt >= datetime.now()
            except (ValueError, TypeError):
                _is_future = False
            if _has_attendance or _has_duration:
                _sess_type = "normal"
            elif _end_blank and _is_future:
                _sess_type = "scheduled"
            elif _end_blank:
                _sess_type = "cancelled"
            else:
                _sess_type = "scheduled"

            # Session Date with time
            _session_dt = str(sess.get("start_time_ist", ""))[:19]

            if _sess_type == "cancelled":
                vals = [sess.get("course_name",""), sess.get("course_title",""),
                        sess.get("tutor_name",""), _session_dt]
                aligns = ["left","left","left","center"]
                for col, (v, al) in enumerate(zip(vals, aligns), 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = v
                    c.font      = _font(italic=True, size=10, color="999999")
                    c.fill      = _fill(C_GREY_CANCEL)
                    c.alignment = _align(al, "center")
                    c.border    = _border()
                for col in range(5, N + 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = "Cancelled" if col == 5 else ""
                    c.font      = _font(italic=True, size=10, color="999999")
                    c.fill      = _fill(C_GREY_CANCEL)
                    c.alignment = _align("center", "center")
                    c.border    = _border()
            elif _sess_type == "scheduled":
                vals = [sess.get("course_name",""), sess.get("course_title",""),
                        sess.get("tutor_name",""), _session_dt]
                aligns = ["left","left","left","center"]
                for col, (v, al) in enumerate(zip(vals, aligns), 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = v
                    c.font      = _font(italic=True, size=10, color=C_LAVENDER_TXT)
                    c.fill      = _fill(C_LAVENDER)
                    c.alignment = _align(al, "center")
                    c.border    = _border()
                for col in range(5, N + 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = "Scheduled" if col == 5 else ""
                    c.font      = _font(italic=True, size=10, color=C_LAVENDER_TXT)
                    c.fill      = _fill(C_LAVENDER)
                    c.alignment = _align("center", "center")
                    c.border    = _border()
            else:
                bg = C_ROW_ALT if row_num % 2 == 0 else C_WHITE
                # Daily Report: tint the previous day's sessions light orange.
                if yest_date is not None and sess.get("_date") == yest_date:
                    bg = C_YEST_HL
                for col, (v, al) in enumerate(zip(
                    [sess.get("course_name",""), sess.get("course_title",""),
                     sess.get("tutor_name",""),  _session_dt,
                     n_stu, n_present, n_fb, f"{fb_rt:.1f}%"],
                    ["left","left","left","center","center","center","center","center"]
                ), 1):
                    style_data_cell(ws, row_num, col, v, bg=bg, h_align=al)
                _write_rating_cell(row_num, 9, avg_r)
                style_data_cell(ws, row_num, 10, min_r, bg=bg, h_align="center")
                style_data_cell(ws, row_num, 11, max_r, bg=bg, h_align="center")
                # Avg Rating Trend (col 12) — compare with same course/tutor on prev session
                _cn_fb  = str(sess.get("course_name", ""))
                _tn_fb  = str(sess.get("tutor_name",  ""))
                prev_r  = prev_trend.get((_cn_fb, _tn_fb))
                _curr_pct = round((avg_r  / 10) * 100, 1) if avg_r  else None
                _prev_pct = round((prev_r / 10) * 100, 1) if prev_r else None
                _write_trend_cell(ws, row_num, 12, _curr_pct, _prev_pct, bg=bg)
            ws.row_dimensions[row_num].height = 18
            row_num += 1
    else:
        # ── Grouped by Course + Subject + Tutor ──────────────────────────────
        grp_keys = ["course_name", "course_title", "tutor_name"]
        for k in grp_keys:
            if k not in sess_f.columns:
                sess_f[k] = ""

        for (cn, ct, tn), grp in sess_f.groupby(grp_keys, sort=True):
            sids          = grp["session_id"].tolist()
            # Deduplicate: one row per (session_id, student_id) pair
            _sa_raw       = att_f[att_f["session_id"].isin(sids)]
            _sa_dd_keys   = [c for c in ["session_id","student_id"] if c in _sa_raw.columns]
            sa            = _sa_raw.drop_duplicates(subset=_sa_dd_keys) if _sa_dd_keys else _sa_raw
            # Deduplicate feedbacks: one feedback per student per session
            sf            = _dedup_fb(fb_f[fb_f["session_id"].isin(sids)] if not fb_f.empty else fb_f)
            n_sess        = len(sids)
            n_fb          = len(sf)
            # Feedback Rate % = feedbacks / students present (attendance-applicable only)
            n_present_fb  = int((_applicable(sa)["status"] == "Present").sum())
            avg_fb_rate   = round(n_fb / n_present_fb * 100, 1) if n_present_fb else 0.0
            avg_r         = round(sf["_rating"].mean(), 2) if n_fb else None
            bg            = C_ROW_ALT if row_num % 2 == 0 else C_WHITE

            for col, (v, al) in enumerate(zip(
                [cn, ct, tn, n_sess, n_fb],
                ["left","left","left","center","center"]
            ), 1):
                style_data_cell(ws, row_num, col, v, bg=bg, h_align=al)

            # Store as float with % format so the AVERAGE formula in totals row works
            fc = ws.cell(row=row_num, column=6)
            fc.value         = avg_fb_rate
            fc.number_format = '0.0"%"'
            fc.font          = _font(size=10)
            fc.fill          = _fill(bg)
            fc.alignment     = _align("center", "center")
            fc.border        = _border()

            _write_rating_cell(row_num, 7, avg_r)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

    # ── Totals ────────────────────────────────────────────────────────────────
    if row_num > data_start:
        de  = row_num - 1
        CL  = get_column_letter
        for col in range(1, N + 1):
            c           = ws.cell(row=row_num, column=col)
            c.fill      = _fill(C_BLUE_LITE)
            c.font      = _font(bold=True, size=10, color=C_NAV)
            c.border    = _border()
            c.alignment = _align("center", "center")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        ws.cell(row=row_num, column=1).value     = "⬛  TOTAL / AVERAGE"
        ws.cell(row=row_num, column=1).alignment = _align("left", "center")
        if is_daily:
            ws.cell(row=row_num, column=5).value = f"=SUM({CL(5)}{data_start}:{CL(5)}{de})"
            ws.cell(row=row_num, column=6).value = f"=SUM({CL(6)}{data_start}:{CL(6)}{de})"
        else:
            ws.cell(row=row_num, column=4).value = f"=SUM({CL(4)}{data_start}:{CL(4)}{de})"
            ws.cell(row=row_num, column=5).value = f"=SUM({CL(5)}{data_start}:{CL(5)}{de})"
            ws.cell(row=row_num, column=6).value = f"=IFERROR(AVERAGE({CL(6)}{data_start}:{CL(6)}{de}),0)"
        ws.row_dimensions[row_num].height = 22
        _border_thick_outer(ws, data_start, 1, row_num, N)

    ws.auto_filter.ref = f"A6:{get_column_letter(N)}{row_num}"
    ws.freeze_panes    = "A7"
    ws.sheet_properties.tabColor = "2E7D32"
    auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 5 — TEACHER NO FEEDBACK
# ─────────────────────────────────────────────────────────────────────────────

def build_teacher_no_feedback(ws, sess_f: pd.DataFrame, tf_f: pd.DataFrame,
                               report_type: str, label: str,
                               att_f: pd.DataFrame = None, yest_date=None):
    """Sessions (daily) or Course+Tutor groups (non-daily) where teacher has not submitted feedback."""
    is_daily = (report_type == "daily")
    title    = report_title(report_type, label, "Teacher – No Feedback Sessions")

    if is_daily:
        headers = [
            "Course Name", "Course Title", "Tutor Name",
            "Start Time (IST)", "End Time (IST)", "Remark",
        ]
    else:
        headers = [
            "Course Name", "Course Title", "Tutor Name",
            "Total Sessions", "Sessions w/ Feedback", "Sessions w/o Feedback", "Remark",
        ]
    N = len(headers)

    style_title_row(ws, 1, 1, N, title)
    ws.cell(row=1, column=1).alignment = _align("center", "center")
    write_header_row(ws, 3, headers)

    # Use TEACHER feedback to determine which sessions have teacher feedback submitted
    sessions_with_fb = set(tf_f["session_id"].dropna().unique()) if not tf_f.empty else set()
    row_num = 4

    if att_f is None:
        att_f = pd.DataFrame()

    if is_daily:
        # ── Per-session (daily) ───────────────────────────────────────────────
        no_fb = sess_f[~sess_f["session_id"].isin(sessions_with_fb)].copy()
        no_fb = no_fb.sort_values(["course_name", "start_time_ist"], na_position="last")
        for i, (_, r) in enumerate(no_fb.iterrows()):
            sid = r.get("session_id", "")

            # ── Detect cancelled / scheduled (same logic as Session_Summary) ──
            dur_min = _session_dur_min(r)
            sa_count = len(_eq_group(att_f, "session_id", sid)) if not att_f.empty and "session_id" in att_f.columns else 0
            _end_raw   = str(r.get("end_time_ist", "")).strip()
            _end_blank = _end_raw in ("", "nan", "NaT", "None", "NAN")
            _has_attendance = sa_count > 0
            _has_duration   = dur_min > 0
            _start_raw = str(r.get("start_time_ist", "")).strip()
            try:
                _start_dt = datetime.strptime(_start_raw[:19], "%Y-%m-%d %H:%M:%S")
                _is_future = _start_dt >= datetime.now()
            except (ValueError, TypeError):
                _is_future = False
            if _has_attendance or _has_duration:
                _sess_type = "normal"
            elif _end_blank and _is_future:
                _sess_type = "scheduled"
            elif _end_blank:
                _sess_type = "cancelled"
            else:
                _sess_type = "scheduled"

            vals   = [r.get("course_name",""), r.get("course_title",""), r.get("tutor_name",""),
                      str(r.get("start_time_ist",""))[:19], str(r.get("end_time_ist",""))[:19]]
            aligns = ["left","left","left","center","center","left"]

            if _sess_type == "cancelled":
                remark = "Cancelled"
                for col, (v, al) in enumerate(zip(vals + [remark], aligns), 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = v
                    c.font      = _font(italic=True, size=10, color="999999")
                    c.fill      = _fill(C_GREY_CANCEL)
                    c.alignment = _align(al, "center")
                    c.border    = _border()
            elif _sess_type == "scheduled":
                remark = "Scheduled"
                for col, (v, al) in enumerate(zip(vals + [remark], aligns), 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = v
                    c.font      = _font(italic=True, size=10, color=C_LAVENDER_TXT)
                    c.fill      = _fill(C_LAVENDER)
                    c.alignment = _align(al, "center")
                    c.border    = _border()
            else:
                bg = C_ROW_ALT if i % 2 == 0 else C_WHITE
                # Daily Report: tint the previous day's sessions light orange.
                if yest_date is not None and r.get("_date") == yest_date:
                    bg = C_YEST_HL
                remark = "No feedback received"
                for col, (v, al) in enumerate(zip(vals + [remark], aligns), 1):
                    style_data_cell(ws, row_num, col, v, bg=bg, h_align=al)

            ws.row_dimensions[row_num].height = 18
            row_num += 1

    else:
        # ── Grouped by Course + Subject + Tutor (non-daily) ───────────────────
        grp_keys = ["course_name", "course_title", "tutor_name"]
        for k in grp_keys:
            if k not in sess_f.columns:
                sess_f[k] = ""

        for i, ((cn, ct, tn), grp) in enumerate(sess_f.groupby(grp_keys, sort=True)):
            sids      = grp["session_id"].tolist()
            n_sess    = len(sids)
            n_with_fb = sum(1 for s in sids if s in sessions_with_fb)
            n_no_fb   = n_sess - n_with_fb

            if n_no_fb == 0:
                # All sessions have feedback
                bg     = C_GREEN_PALE
                remark = "✅ All sessions received feedback"
            elif n_with_fb > 0:
                # Some sessions missing feedback — list the specific dates
                missing_sids  = [s for s in sids if s not in sessions_with_fb]
                missing_dates = []
                for _msid in missing_sids:
                    _mrow = grp[grp["session_id"] == _msid]
                    if not _mrow.empty and "_date" in _mrow.columns:
                        _d = _mrow.iloc[0]["_date"]
                        if _d is not None and not pd.isna(_d):
                            _ds = _d.strftime("%d-%b") if hasattr(_d, "strftime") else str(_d)
                            missing_dates.append(_ds)
                bg     = C_AMBER
                remark = (f"Missing feedback on: {', '.join(sorted(missing_dates))}"
                          if missing_dates else f"⚠️ {n_no_fb} of {n_sess} sessions missing feedback")
            else:
                # No feedback received at all
                bg     = C_RED_LITE
                remark = "❌ No feedback received at all"

            vals   = [cn, ct, tn, n_sess, n_with_fb, n_no_fb, remark]
            aligns = ["left","left","left","center","center","center","left"]
            for col, (v, al) in enumerate(zip(vals, aligns), 1):
                c = ws.cell(row=row_num, column=col)
                c.value     = v
                c.font      = _font(size=10)
                c.fill      = _fill(bg)
                c.alignment = _align(al, "center")
                c.border    = _border()
            ws.row_dimensions[row_num].height = 18
            row_num += 1

    if row_num == 4:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=N)
        c = ws.cell(row=4, column=1)
        c.value     = "✅ All sessions received teachers feedback."
        c.font      = _font(bold=True, size=11, color="2E7D32")
        c.fill      = _fill("E8F5E9")
        c.alignment = _align("center", "center")
        c.border    = _border()

    if row_num > 4:
        _border_thick_outer(ws, 4, 1, row_num - 1, N)

    ws.auto_filter.ref = f"A3:{get_column_letter(N)}{row_num - 1}"
    ws.freeze_panes    = "A4"
    ws.sheet_properties.tabColor = "E65100"
    auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
#  GOOGLE DRIVE UPLOAD
# ─────────────────────────────────────────────────────────────────────────────

# Direct Google Drive folder IDs for each report type
GDRIVE_FOLDER_MAP = {
    "daily":       "1hcbC4VIg92JGNQjFyEKI8zngoaAZgbCE",
    "weekly":      "1-CU6nYZ0tQu_tkjOjVS4rYUd1I9B9wpq",
    "fortnightly": "1-u6UytopVpSU1f0Vfewo6EHMTyMGg1tF",
    "monthly":     "1wX2hnmL5FaZXrkf500TvhxtAZWuBiwm8",
    "quarterly":   "1DTJu4B6EW_XcHE2yymZ9kDeOTpg_UCGy",
    "yearly":      "1vVgn_iZcwI4epEMBcjkB_hxEjA6YSjAQ",
    "manual":      "1rtQj0aXOBNc4Qj9L1X99Ishec_lPuWhD"
}


def _gdrive_get_or_create_subfolder(drive_service, parent_id: str, name: str) -> str:
    """Return the folder ID of <name> inside <parent_id>, creating it if needed."""
    query = (
        f"'{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' "
        f"and name='{name}' and trashed=false"
    )
    res = drive_service.files().list(q=query, fields="files(id,name)").execute()
    files = res.get("files", [])
    if files:
        return files[0]["id"]
    # Create
    meta = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }
    folder = drive_service.files().create(body=meta, fields="id").execute()
    return folder["id"]


def upload_to_gdrive(service_account_file: str, report_type: str,
                     filename: str, file_buffer: io.BytesIO):
    """Upload the report Excel buffer directly to the designated Drive folder.
    No local file is created — the workbook is streamed from memory."""
    try:
        from googleapiclient.discovery import build as gdrive_build
        from googleapiclient.http      import MediaIoBaseUpload
        from google.oauth2             import service_account

        folder_id = GDRIVE_FOLDER_MAP.get(report_type.lower())
        if not folder_id:
            print(f"[Drive] ⚠ No Drive folder configured for report type '{report_type}' — skipping upload.")
            return

        creds = service_account.Credentials.from_service_account_file(
            service_account_file,
            scopes=["https://www.googleapis.com/auth/drive"],
        )
        # Impersonate the info account so files are owned by it and
        # count against its storage quota (not the service account's zero quota).
        creds = creds.with_subject("info@intellibiinnovationstechnologies.in")
        drive_service = gdrive_build("drive", "v3", credentials=creds)

        # Remove existing file with same name to avoid duplicates
        query = f"'{folder_id}' in parents and name='{filename}' and trashed=false"
        existing = drive_service.files().list(
            q=query,
            fields="files(id)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        for f in existing.get("files", []):
            drive_service.files().delete(
                fileId=f["id"],
                supportsAllDrives=True,
            ).execute()

        # Stream from the in-memory buffer — no local file needed
        file_buffer.seek(0)
        media = MediaIoBaseUpload(
            file_buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=False,
        )
        meta     = {"name": filename, "parents": [folder_id]}
        uploaded = drive_service.files().create(
            body=meta,
            media_body=media,
            fields="id,webViewLink",
            supportsAllDrives=True,
        ).execute()
        print(f"[Drive] ✓ Uploaded → {report_type.title()} folder / {filename}")
        print(f"[Drive]   Link: {uploaded.get('webViewLink', '')}")
    except Exception as e:
        print(f"[Drive] ⚠ Upload failed: {e}")


def save_to_local_directory(report_type: str, filename: str, file_bytes: bytes):
    """Save the report .xlsx to the configured local computer directory.

    Runs only when upload_to_local_directory = True (see the RUN CONFIGURATION
    block). Mirrors the Google Drive layout by writing into a per-report-type
    subfolder of LOCAL_UPLOAD_DIR. Does not affect the Drive upload or email."""
    try:
        target_dir = os.path.join(LOCAL_UPLOAD_DIR, report_type.title())
        os.makedirs(target_dir, exist_ok=True)
        target_path = os.path.join(target_dir, filename)
        with open(target_path, "wb") as f:
            f.write(file_bytes)
        print(f"[Local] ✓ Saved → {report_type.title()} folder / {filename}")
        print(f"[Local]   Path: {target_path}")
    except Exception as e:
        print(f"[Local] ⚠ Save failed: {e}")


# ─────────────────────────────────────────────────────────────────────────────
#  EMAIL
# ─────────────────────────────────────────────────────────────────────────────

def send_report(report_type: str, label: str, filename: str, file_bytes: bytes,
                n_sessions: int, n_students: int,
                n_absent: int, combined_att_pct: float):
    subject = f"IntelliBI {report_type.title()} Attendance & Feedback Report — {label}"
    body = f"""\
<html><body style="font-family:Arial,sans-serif;color:#1F3864;">
<h2 style="color:#1F3864;">📊 IntelliBI {report_type.title()} Attendance &amp; Feedback Report</h2>
<p>Please find the attached report for <strong>{label}</strong>.</p>
<table style="border-collapse:collapse;font-size:14px;">
  <tr>
    <td style="padding:8px 18px;background:#D6E4F0;font-weight:bold;border:1px solid #B0B0B0;">Total Sessions</td>
    <td style="padding:8px 18px;border:1px solid #B0B0B0;">{n_sessions}</td>
  </tr><tr>
    <td style="padding:8px 18px;background:#D6E4F0;font-weight:bold;border:1px solid #B0B0B0;">Total Students (records)</td>
    <td style="padding:8px 18px;border:1px solid #B0B0B0;">{n_students}</td>
  </tr><tr>
    <td style="padding:8px 18px;background:#FFE0E0;font-weight:bold;border:1px solid #B0B0B0;">❌ Absent</td>
    <td style="padding:8px 18px;border:1px solid #B0B0B0;"><strong>{n_absent}</strong></td>
  </tr><tr>
    <td style="padding:8px 18px;background:#E8F5E9;font-weight:bold;border:1px solid #B0B0B0;">📊 Combined Attendance %</td>
    <td style="padding:8px 18px;border:1px solid #B0B0B0;"><strong>{combined_att_pct:.1f}%</strong></td>
  </tr>
</table>
<br>
<p style="font-size:12px;color:#888;">This is an automated report from IntelliBI Innovations Technologies.</p>
</body></html>"""

    to_list  = REPORT_TO if isinstance(REPORT_TO, list) else [REPORT_TO]
    msg             = MIMEMultipart()
    msg["From"]     = GMAIL_SENDER
    msg["To"]       = ", ".join(to_list)
    msg["Subject"]  = subject
    if REPORT_CC:
        msg["Cc"]   = ", ".join(REPORT_CC)
    msg.attach(MIMEText(body, "html"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(file_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    recipients = to_list + REPORT_CC
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_SENDER, GMAIL_APP_PASS)
        server.sendmail(GMAIL_SENDER, recipients, msg.as_string())

    print(f"[Email] ✓ Report sent to {', '.join(to_list)}")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════════════════
#  PERIOD REPORTS — Weekly / Monthly / Quarterly / Yearly
#  Built from scratch. Daily functions above are NOT modified.
# ═══════════════════════════════════════════════════════════════════════════════

def _period_title(report_type: str, label: str, sheet_name: str) -> str:
    names = {"weekly": "Weekly", "monthly": "Monthly",
             "quarterly": "Quarterly", "yearly": "Yearly",
             "manual": "Custom Period"}
    return f"{names.get(report_type, report_type.title())} {sheet_name}  |  Period: {label}"


def _avg_dur_for_group(sess_grp: pd.DataFrame, att_f: pd.DataFrame) -> float:
    """Average session duration (minutes) across a session group.
    Uses timestamps from Sessions; falls back to max student attendance duration."""
    durs = []
    for _, sr in sess_grp.iterrows():
        sid = sr.get("session_id", "")
        sa  = _eq_group(att_f, "session_id", sid) if sid else pd.DataFrame()
        d   = _session_dur_min(sr)
        if d <= 0 and not sa.empty and "_dur_min" in sa.columns:
            mx = sa["_dur_min"].max()
            if mx and float(mx) > 0:
                d = round(float(mx), 1)
        if d > 0:
            durs.append(d)
    return round(sum(durs) / len(durs), 1) if durs else 0.0


# ─── Trend helpers ────────────────────────────────────────────────────────────

def _build_ss_trend(sess_f: pd.DataFrame, att_f: pd.DataFrame) -> dict:
    """Return {(course_name, course_title, tutor_name): prev_att_pct} from prev-period data."""
    result = {}
    if sess_f.empty:
        return result
    for k in ["course_name", "course_title", "tutor_name"]:
        if k not in sess_f.columns:
            sess_f[k] = ""
    for (cn, ct, tn), grp in sess_f.groupby(
            ["course_name", "course_title", "tutor_name"], sort=True):
        sids     = grp["session_id"].tolist()
        sa_raw   = att_f[att_f["session_id"].isin(sids)]
        dd_cols  = [c for c in ["session_id", "student_id"] if c in sa_raw.columns]
        sa       = sa_raw.drop_duplicates(subset=dd_cols) if dd_cols else sa_raw
        enrolled = len(sa)
        present  = int((sa["status"] == "Present").sum())
        result[(cn, ct, tn)] = round(present / enrolled * 100, 1) if enrolled else 0.0
    return result


def _build_daily_ss_trend(sess_f: pd.DataFrame, att_f: pd.DataFrame) -> dict:
    """Return {(course_name, course_title, tutor_name): avg_att_pct} from previous-day data."""
    result = {}
    if sess_f.empty:
        return result
    for k in ["course_name", "course_title", "tutor_name", "session_id"]:
        if k not in sess_f.columns:
            sess_f[k] = ""
    for _, sess in sess_f.iterrows():
        cn  = sess.get("course_name", "")
        ct  = sess.get("course_title", "")
        tn  = sess.get("tutor_name",  "")
        sid = sess.get("session_id",  "")
        sa  = _eq_group(att_f, "session_id", sid)
        pct = round(len(sa[sa["status"] == "Present"]) / len(sa) * 100, 1) if len(sa) else 0.0
        key = (cn, ct, tn)
        result[key] = round((result[key] + pct) / 2, 1) if key in result else pct
    return result


def _build_daily_fb_trend(sess_f: pd.DataFrame, fb_f: pd.DataFrame) -> dict:
    """Return {(course_name, course_title, tutor_name): avg_rating} from previous-day data."""
    result = {}
    if sess_f.empty:
        return result
    for k in ["course_name", "course_title", "tutor_name", "session_id"]:
        if k not in sess_f.columns:
            sess_f[k] = ""
    for _, sess in sess_f.iterrows():
        cn  = sess.get("course_name", "")
        ct  = sess.get("course_title", "")
        tn  = sess.get("tutor_name",  "")
        sid = sess.get("session_id",  "")
        sf  = _eq_group(fb_f, "session_id", sid) if not fb_f.empty else pd.DataFrame()
        if not sf.empty and "_rating" in sf.columns:
            ratings = sf["_rating"].dropna().tolist()
            if ratings:
                avg_r = round(sum(ratings) / len(ratings), 2)
                key   = (cn, ct, tn)
                result[key] = round((result[key] + avg_r) / 2, 2) if key in result else avg_r
    return result


def _build_daily_sd_prev(att_f: pd.DataFrame) -> dict:
    """Return {(course_name, course_title, student_id): was_present}
    from previous-day attendance data."""
    result = {}
    if att_f.empty:
        return result
    for k in ["course_name", "course_title", "student_id", "status"]:
        if k not in att_f.columns:
            att_f[k] = ""
    for _, r in att_f.iterrows():
        cn   = r.get("course_name",  "")
        ct   = r.get("course_title", "")
        stid = str(r.get("student_id", ""))
        key  = (cn, ct, stid)
        # Any Present record for this student+course on prev day → True
        if result.get(key) is not True:
            result[key] = (str(r.get("status", "")) == "Present")
    return result


def _build_sd_trend(att_f: pd.DataFrame) -> dict:
    """Return {(course_name, course_title, student_id): att_pct} from prev-period data."""
    result = {}
    if att_f.empty:
        return result
    for k in ["course_name", "course_title", "student_id", "session_id"]:
        if k not in att_f.columns:
            att_f[k] = ""
    dd_cols = [c for c in ["session_id", "student_id"] if c in att_f.columns]
    att_dd  = att_f.drop_duplicates(subset=dd_cols) if dd_cols else att_f
    for (cn, ct, stid), grp in att_dd.groupby(
            ["course_name", "course_title", "student_id"], sort=True):
        n_sess  = grp["session_id"].nunique()
        present = int((grp["status"] == "Present").sum())
        result[(cn, ct, str(stid))] = round(present / n_sess * 100, 1) if n_sess else 0.0
    return result


def _build_fb_trend(sess_f: pd.DataFrame, fb_f: pd.DataFrame) -> dict:
    """Return {(course_name, course_title, tutor_name): prev_avg_rating} from prev-period data."""
    result = {}
    if sess_f.empty:
        return result
    for k in ["course_name", "course_title", "tutor_name"]:
        if k not in sess_f.columns:
            sess_f[k] = ""
    for (cn, ct, tn), grp in sess_f.groupby(
            ["course_name", "course_title", "tutor_name"], sort=True):
        sids    = grp["session_id"].tolist()
        sf      = fb_f[fb_f["session_id"].isin(sids)] if not fb_f.empty else pd.DataFrame()
        ratings = sf["_rating"].dropna().tolist() \
                  if not sf.empty and "_rating" in sf.columns else []
        result[(cn, ct, tn)] = round(sum(ratings) / len(ratings), 2) if ratings else None
    return result


def _collect_daily_trends(service, current_start):
    """For each (course_name, tutor_name) find the most recent session strictly
    before current_start, compute att% / avg_rating / student presence.

    Returns (ss_trend, fb_trend, sd_trend):
      ss_trend : {(course_name, tutor_name): att_pct}
      fb_trend : {(course_name, tutor_name): avg_rating}
      sd_trend : {(course_name, student_id_str): was_present bool}
    Keys intentionally omit course_title (it can change between sessions of the
    same course, causing mismatches if included).
    """
    ss_trend: dict = {}
    fb_trend: dict = {}
    sd_trend: dict = {}
    try:
        sess_all = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Sessions")
        att_all  = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Attendance")
        fb_all   = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Student_Feedback")
    except Exception as _e:
        print(f"[Trend] Could not read sheets for daily trends: {_e}")
        return ss_trend, fb_trend, sd_trend

    if sess_all.empty:
        return ss_trend, fb_trend, sd_trend

    for k in ["course_name", "tutor_name", "session_id"]:
        if k not in sess_all.columns:
            sess_all[k] = ""

    sess_all["_date"] = pd.to_datetime(
        sess_all["start_time_ist"], errors="coerce").dt.date

    sess_past = sess_all[sess_all["_date"].apply(
        lambda d: d is not None and not pd.isna(d) and d < current_start
    )].copy()

    if sess_past.empty:
        return ss_trend, fb_trend, sd_trend

    # Convert date to ISO string so pandas groupby max works reliably on all
    # pandas versions (avoids idxmax() issues with object-dtype date columns).
    sess_past = sess_past.copy()
    sess_past["_date_str"] = sess_past["_date"].apply(
        lambda d: d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else "")

    # Per (course_name, tutor_name): keep ALL sessions on the most-recent day
    max_date_str = sess_past.groupby(
        ["course_name", "tutor_name"], sort=False
    )["_date_str"].transform("max")
    prev_sessions = sess_past[sess_past["_date_str"] == max_date_str]

    # Build: (cn, tn) → [session_ids] for all sessions on that most-recent day
    key_to_sids: dict = {}
    for _, row in prev_sessions.iterrows():
        key = (str(row["course_name"]), str(row["tutor_name"]))
        key_to_sids.setdefault(key, []).append(str(row["session_id"]))
    prev_sids = {sid for sids in key_to_sids.values() for sid in sids}
    print(f"[Trend] Per-course prev session groups: {len(key_to_sids)}")

    # ── Attendance ───────────────────────────────────────────────────────────
    if not att_all.empty and "session_id" in att_all.columns:
        for k in ["student_id", "status"]:
            if k not in att_all.columns:
                att_all[k] = ""
        att_all["session_id"] = att_all["session_id"].astype(str)
        att_prev = att_all[att_all["session_id"].isin(prev_sids)]

        for (cn, tn), sids in key_to_sids.items():
            sa = att_prev[att_prev["session_id"].isin(sids)]
            if len(sa):
                pct = round(len(sa[sa["status"] == "Present"]) / len(sa) * 100, 1)
                ss_trend[(cn, tn)] = pct
                for _, ar in sa.iterrows():
                    stid   = str(ar.get("student_id", ""))
                    sd_key = (cn, stid)
                    was_present = (str(ar.get("status", "")) == "Present")
                    # If present in ANY session on that day, mark True
                    if not sd_trend.get(sd_key):
                        sd_trend[sd_key] = was_present

    # ── Feedback ─────────────────────────────────────────────────────────────
    if not fb_all.empty and "session_id" in fb_all.columns:
        if "rating" in fb_all.columns:
            fb_all["_rating"] = pd.to_numeric(fb_all["rating"], errors="coerce")
        fb_all["session_id"] = fb_all["session_id"].astype(str)
        fb_prev = fb_all[fb_all["session_id"].isin(prev_sids)]
        for (cn, tn), sids in key_to_sids.items():
            sf = fb_prev[fb_prev["session_id"].isin(sids)]
            if not sf.empty and "_rating" in sf.columns:
                ratings = sf["_rating"].dropna().tolist()
                if ratings:
                    fb_trend[(cn, tn)] = round(sum(ratings) / len(ratings), 2)

    print(f"[Trend] Daily trends: SS={len(ss_trend)} FB={len(fb_trend)} SD={len(sd_trend)}")
    return ss_trend, fb_trend, sd_trend


def _collect_period_trends(service, report_type, current_start):
    """For each (course_name, tutor_name) find the most recent period of
    report_type strictly before current_start that had sessions, compute
    att% / avg_rating / student att%.

    Returns (ss_trend, fb_trend, sd_trend):
      ss_trend : {(course_name, tutor_name): att_pct}
      fb_trend : {(course_name, tutor_name): avg_rating}
      sd_trend : {(course_name, student_id_str): att_pct}
    Keys intentionally omit course_title for the same reason as _collect_daily_trends.
    """
    ss_trend: dict = {}
    fb_trend: dict = {}
    sd_trend: dict = {}
    try:
        sess_all = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Sessions")
        att_all  = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Attendance")
        fb_all   = read_sheet_df(service, ATTENDANCE_SHEET_ID, "Student_Feedback")
    except Exception as _e:
        print(f"[Trend] Could not read sheets for period trends: {_e}")
        return ss_trend, fb_trend, sd_trend

    if sess_all.empty:
        return ss_trend, fb_trend, sd_trend

    for k in ["course_name", "tutor_name", "session_id"]:
        if k not in sess_all.columns:
            sess_all[k] = ""

    sess_all["_date"] = pd.to_datetime(
        sess_all["start_time_ist"], errors="coerce").dt.date

    def _safe_pstart(d):
        try:
            return _period_for_date(report_type, d)[0] if d else None
        except Exception:
            return None

    def _safe_pend(d):
        try:
            return _period_for_date(report_type, d)[1] if d else None
        except Exception:
            return None

    sess_all["_pstart"] = sess_all["_date"].apply(_safe_pstart)
    sess_all["_pend"]   = sess_all["_date"].apply(_safe_pend)

    # Keep only sessions whose entire period ends before current_start
    sess_past = sess_all[sess_all["_pend"].apply(
        lambda e: e is not None and e < current_start
    )].copy()

    if sess_past.empty:
        return ss_trend, fb_trend, sd_trend

    # Convert _pstart to ISO string so groupby max works reliably on all
    # pandas versions (avoids idxmax() issues with object-dtype date columns).
    sess_past = sess_past.copy()
    sess_past["_pstart_str"] = sess_past["_pstart"].apply(
        lambda d: d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else "")

    # Per (course_name, tutor_name): keep ALL sessions in the most-recent period
    max_pstart_str = sess_past.groupby(
        ["course_name", "tutor_name"], sort=False
    )["_pstart_str"].transform("max")
    sess_best = sess_past[sess_past["_pstart_str"] == max_pstart_str]

    # Collect session_ids per (cn, tn) from that best period
    key_to_sids: dict = {}
    for _, row in sess_best.iterrows():
        cn_tn = (str(row["course_name"]), str(row["tutor_name"]))
        key_to_sids.setdefault(cn_tn, []).append(str(row["session_id"]))

    all_prev_sids = {sid for sids in key_to_sids.values() for sid in sids}
    print(f"[Trend] Per-course prev periods found: {len(key_to_sids)}")

    # ── Attendance ───────────────────────────────────────────────────────────
    if not att_all.empty and "session_id" in att_all.columns:
        for k in ["student_id", "status"]:
            if k not in att_all.columns:
                att_all[k] = ""
        att_all["session_id"] = att_all["session_id"].astype(str)
        att_prev = att_all[att_all["session_id"].isin(all_prev_sids)]
        dd_cols  = [c for c in ["session_id", "student_id"] if c in att_prev.columns]
        att_dd   = att_prev.drop_duplicates(subset=dd_cols) if dd_cols else att_prev

        for cn_tn, sids in key_to_sids.items():
            cn, tn = cn_tn
            sa = att_dd[att_dd["session_id"].isin(sids)]
            if len(sa):
                pct = round(len(sa[sa["status"] == "Present"]) / len(sa) * 100, 1)
                ss_trend[cn_tn] = pct
                # Per-student att% in this prev period
                for stid_val, stu_grp in sa.groupby("student_id"):
                    stid   = str(stid_val)
                    n_s    = stu_grp["session_id"].nunique()
                    pres   = int((stu_grp["status"] == "Present").sum())
                    sd_trend[(cn, stid)] = round(pres / n_s * 100, 1) if n_s else 0.0

    # ── Feedback ─────────────────────────────────────────────────────────────
    if not fb_all.empty and "session_id" in fb_all.columns:
        if "rating" in fb_all.columns:
            fb_all["_rating"] = pd.to_numeric(fb_all["rating"], errors="coerce")
        fb_all["session_id"] = fb_all["session_id"].astype(str)
        fb_prev = fb_all[fb_all["session_id"].isin(all_prev_sids)]
        for cn_tn, sids in key_to_sids.items():
            sf = fb_prev[fb_prev["session_id"].isin(sids)]
            if not sf.empty and "_rating" in sf.columns:
                ratings = sf["_rating"].dropna().tolist()
                if ratings:
                    fb_trend[cn_tn] = round(sum(ratings) / len(ratings), 2)

    print(f"[Trend] Period trends: SS={len(ss_trend)} FB={len(fb_trend)} SD={len(sd_trend)}")
    return ss_trend, fb_trend, sd_trend


def _write_trend_cell(ws, row: int, col: int, current, previous, bg: str = C_WHITE):
    """Write a coloured trend cell.

    For Att %  : current/previous are plain floats (percentage points).
                 diff shown as percentage-point change  e.g. ▲ +5.0%
    For Avg Rating: current/previous are raw floats (0-10 scale).
                 diff shown as percentage change relative to previous
                 e.g. ▲ +7.1%  (= 0.5 / 7.0 * 100)
    Pass mode='pct' for percentage-change; default 'pts' keeps absolute diff.
    """
    tc           = ws.cell(row=row, column=col)
    tc.border    = _border()
    tc.alignment = _align("center", "center")
    if previous is None or current is None:
        tc.value = "—"
        tc.font  = _font(size=10, color=C_GREY_BD)
        tc.fill  = _fill(bg)
        return
    diff = round(current - previous, 2)
    if diff > 0:
        tc.value = f"▲ +{diff:.2f}%"
        tc.font  = _font(bold=True, size=10, color=C_GREEN_DARK)
        tc.fill  = _fill(C_GREEN)
    elif diff < 0:
        tc.value = f"▼ {diff:.2f}%"
        tc.font  = _font(bold=True, size=10, color=C_RED_DARK)
        tc.fill  = _fill(C_RED_LITE)
    else:
        tc.value = "→ 0.00%"
        tc.font  = _font(size=10, color=C_GREY_BD)
        tc.fill  = _fill(bg)


# ─────────────────────────────────────────────────────────────────────────────
#  PERIOD DRILL-DOWN HELPERS  (weekly → per-day, monthly → per-week)
# ─────────────────────────────────────────────────────────────────────────────

def _monthly_week_ranges(start_date, end_date):
    """Return [(header_label, week_start, week_end), ...] for every Mon-Sun week
    that intersects [start_date, end_date].  Used for monthly breakdown columns."""
    cur_mon = start_date - timedelta(days=start_date.weekday())
    weeks   = []
    wk_num  = 1
    while cur_mon <= end_date:
        cur_sun = cur_mon + timedelta(days=6)
        ws_c    = max(cur_mon, start_date)
        we_c    = min(cur_sun, end_date)
        label   = (f"Wk{wk_num}\n"
                   f"{ws_c.day:02d}-{we_c.day:02d} {we_c.strftime('%b')}")
        weeks.append((label, ws_c, we_c))
        cur_mon += timedelta(days=7)
        wk_num  += 1
    return weeks


def _write_bkd_att_cell(ws, row, col, att_pct, bg=C_WHITE):
    """Drill-down Att% cell: green/red percent or '—' when no session."""
    c = ws.cell(row=row, column=col)
    c.border    = _border()
    c.alignment = _align("center", "center")
    if att_pct is None:
        c.value = "—"; c.font = _font(size=9, color=C_GREY_BD); c.fill = _fill(bg)
    else:
        c.value = f"{att_pct:.2f}%"
        c.font  = _font(bold=True, size=9,
                        color=C_GREEN_DARK if att_pct >= 75 else C_RED_DARK)
        c.fill  = _fill(C_GREEN_PALE if att_pct >= 75 else C_RED_LITE)


def _write_bkd_rating_cell(ws, row, col, avg_rat, bg=C_WHITE):
    """Drill-down Avg Rating cell: coloured value or '—'."""
    c = ws.cell(row=row, column=col)
    c.border    = _border()
    c.alignment = _align("center", "center")
    if not avg_rat:
        c.value = "—"; c.font = _font(size=9, color=C_GREY_BD); c.fill = _fill(bg)
    else:
        c.value = f"{avg_rat:.2f}"
        c.font  = _font(bold=True, size=9)
        c.fill  = _fill(_rating_bg(avg_rat))


def _write_bkd_status_cell(ws, row, col, status, bg=C_WHITE):
    """Daily-presence cell for Student Detail: P (green) / A (red) / — (grey)."""
    c = ws.cell(row=row, column=col)
    c.border    = _border()
    c.alignment = _align("center", "center")
    if status is None:
        c.value = "—"; c.font = _font(size=9, color=C_GREY_BD); c.fill = _fill(bg)
    elif status == "Present":
        c.value = "P"; c.font = _font(bold=True, size=9, color=C_GREEN_DARK)
        c.fill  = _fill(C_GREEN)
    else:
        c.value = "A"; c.font = _font(bold=True, size=9, color=C_RED_DARK)
        c.fill  = _fill(C_RED_LITE)


# ─── Sheet 1: Session Summary ─────────────────────────────────────────────────

def build_period_session_summary(ws, sess_f: pd.DataFrame, att_f: pd.DataFrame,
                                  fb_f: pd.DataFrame, report_type: str, label: str,
                                  prev_trend: dict = None,
                                  start_date=None, end_date=None):
    """Period Session Summary: one row per Course + Subject + Tutor."""
    sess_f     = _prefer_instructor_name(sess_f)   # Instructor column ← Instructor_Name
    title      = _period_title(report_type, label, "Attendance Report")
    prev_trend = prev_trend or {}

    # ── Drill-down breakdown periods ──────────────────────────────────────────
    # weekly  → one column per day  (Mon … Sun)
    # monthly → one column per week (Wk1, Wk2, …)
    breakdown_periods = []   # [(header_label, period_start, period_end), ...]
    if report_type == "weekly" and start_date:
        for i in range(7):
            d = start_date + timedelta(days=i)
            breakdown_periods.append((d.strftime("%a\n%d-%b"), d, d))
    elif report_type == "monthly" and start_date and end_date:
        breakdown_periods = _monthly_week_ranges(start_date, end_date)
    n_bkd = len(breakdown_periods)

    # Pre-compute per-breakdown-period att% for each (cn, ct, tn)
    # {(cn, ct, tn, bp_i): att_pct} — None means no session that period
    bkd_att: dict = {}
    if n_bkd and "_date" in sess_f.columns:
        for bp_i, (_, bp_s, bp_e) in enumerate(breakdown_periods):
            bp_sess = sess_f[sess_f["_date"].apply(
                lambda d: d is not None and not pd.isna(d) and bp_s <= d <= bp_e)]
            if bp_sess.empty:
                continue
            for (cn, ct, tn), grp in bp_sess.groupby(
                    ["course_name", "course_title", "tutor_name"]):
                sids    = grp["session_id"].tolist()
                _sa     = _applicable(att_f[att_f["session_id"].isin(sids)])
                # No deduplication — matches daily report att% calculation exactly.
                # Attendance-applicable students only.
                enrolled = len(_sa)
                present  = int((_sa["status"] == "Present").sum())
                bkd_att[(cn, ct, tn, bp_i)] = (
                    round(present / enrolled * 100, 2) if enrolled else 0.0)

    bkd_hdrs = [lbl for lbl, _, _ in breakdown_periods]
    headers = [
        "Tech Name", "Duration", "Instructor",
        "Sessions", "Cancelled\nSessions", "Total\nEnrolled", "Att. N/A\nCount",
        "Present", "Absent",
        "Att %", "Avg Duration\n(min)", "Avg\nRating", "Feedback\nRate %",
    ] + bkd_hdrs + ["Att %\nTrend"]
    N         = len(headers)          # 14 + n_bkd
    TREND_COL = 14 + n_bkd           # 1-based column index of Trend

    style_title_row(ws, 1, 1, N, title)
    ws.cell(row=1, column=1).alignment = _align("center", "center")

    # ── KPI strip ─────────────────────────────────────────────────────────────
    # Deduplicate attendance: count each student once per session.
    # Also drop phantom/corrupt rows (null student_id or session_id) so they
    # don't inflate the KPI counts — these rows aren't linked to any real session
    # and are excluded from the per-course totals as well.
    _dedup_cols = [c for c in ["session_id", "student_id"] if c in att_f.columns]
    att_dedup_kpi = att_f.drop_duplicates(subset=_dedup_cols) if _dedup_cols else att_f
    # Drop phantom rows — handles both NaN (local Excel) AND empty string ""
    # (Google Sheets API fills missing cells with "" not NaN).
    if "student_id" in att_dedup_kpi.columns:
        _sid = att_dedup_kpi["student_id"].astype(str).str.strip()
        att_dedup_kpi = att_dedup_kpi[~_sid.isin(["", "nan", "None", "NaN"])]
    if "session_id" in att_dedup_kpi.columns:
        _sesid = att_dedup_kpi["session_id"].astype(str).str.strip()
        att_dedup_kpi = att_dedup_kpi[~_sesid.isin(["", "nan", "None", "NaN"])]
    n_sess_kpi   = len(sess_f)
    n_stu_kpi    = att_f["student_id"].nunique() if "student_id" in att_f.columns else len(att_f)
    # Present/Absent/Att% count attendance-applicable students only.
    att_dedup_kpi_app = _applicable(att_dedup_kpi)
    tot_pres_kpi = int((att_dedup_kpi_app["status"] == "Present").sum())
    tot_abs_kpi  = int((att_dedup_kpi_app["status"] == "Absent").sum())
    tot_enr_kpi  = len(att_dedup_kpi_app)
    pct_kpi      = round(tot_pres_kpi / tot_enr_kpi * 100, 1) if tot_enr_kpi else 0.0
    kpis = [
        ("Total Sessions", n_sess_kpi,         "📚", KPI_BLUE),
        ("Total Students", n_stu_kpi,          "👥", KPI_BLUE),
        ("Overall Att %",  f"{pct_kpi:.1f}%",  "📊",
         KPI_GREEN if pct_kpi >= 75 else KPI_RED),
        ("Absent",         tot_abs_kpi,         "❌", KPI_RED),
        ("Present",        tot_pres_kpi,        "✅", KPI_GREEN),
    ]
    _write_kpi_strip(ws, row_lbl=2, row_val=3, kpis=kpis, n_cols=N)

    write_header_row(ws, 4, headers)
    ws.freeze_panes = "D5"
    data_start = row_num = 5

    for k in ["course_name", "course_title", "tutor_name"]:
        if k not in sess_f.columns:
            sess_f[k] = ""

    tot = {"n_sess": 0, "cancelled": 0, "enr": 0, "na": 0, "pres": 0, "abs": 0,
           "fb": 0, "exp_fb": 0, "ratings": []}

    # ── Collect all rows first so we can sort before writing ──────────────────
    ss_rows = []
    for (cn, ct, tn), grp in sess_f.groupby(
            ["course_name", "course_title", "tutor_name"], sort=True):
        sids     = grp["session_id"].tolist()
        # Deduplicate: one row per unique (session_id, student_id) pair
        _sa_raw  = att_f[att_f["session_id"].isin(sids)]
        _dd_cols = [c for c in ["session_id", "student_id"] if c in _sa_raw.columns]
        sa       = _sa_raw.drop_duplicates(subset=_dd_cols) if _dd_cols else _sa_raw
        sf       = fb_f[fb_f["session_id"].isin(sids)] if not fb_f.empty else pd.DataFrame()

        n_sess    = len(sids)
        # Cancelled = sessions with no end_time_ist AND no attendance AND no duration.
        # A session with students or duration is "normal" even if end_time_ist is blank.
        cancelled = 0
        _now = datetime.now()
        for _, _s_row in grp.iterrows():
            _s_id = _s_row.get("session_id", "")
            _s_end_raw = str(_s_row.get("end_time_ist", "")).strip()
            _s_end_blank = _s_end_raw in ("", "nan", "NaT", "None", "NAN") or pd.isna(_s_row.get("end_time_ist"))
            _s_dur = _session_dur_min(_s_row)
            _s_att = len(att_f[att_f["session_id"] == _s_id]) if not att_f.empty and "session_id" in att_f.columns else 0
            # Check if session start is in the future → scheduled, not cancelled
            _s_start_raw = str(_s_row.get("start_time_ist", "")).strip()
            try:
                _s_start_dt = datetime.strptime(_s_start_raw[:19], "%Y-%m-%d %H:%M:%S")
                _s_is_future = _s_start_dt >= _now
            except (ValueError, TypeError):
                _s_is_future = False
            if _s_end_blank and _s_att == 0 and _s_dur == 0 and not _s_is_future:
                cancelled += 1
        sa_app    = _applicable(sa)                # applicable-only for Present/Absent/Att%
        enrolled  = len(sa)                         # Total Enrolled (incl. attendance-N/A)
        na_cnt    = _na_count(sa)                   # Att. N/A Count
        applic_n  = len(sa_app)
        present   = int((sa_app["status"] == "Present").sum())
        absent    = int((sa_app["status"] == "Absent").sum())
        att_pct   = round(present / applic_n * 100, 1) if applic_n else 0.0
        avg_dur   = _avg_dur_for_group(grp, att_f)

        ratings  = sf["_rating"].dropna().tolist() \
                   if not sf.empty and "_rating" in sf.columns else []
        avg_rat  = round(sum(ratings) / len(ratings), 2) if ratings else 0.0
        fb_rate  = round(len(sf) / present * 100, 1) if present else 0.0

        ss_rows.append((n_sess, att_pct,
                        cn, ct, tn, cancelled, enrolled, na_cnt, present, absent,
                        avg_dur, avg_rat, fb_rate, ratings, len(sf)))

        tot["n_sess"]    += n_sess;    tot["cancelled"] += cancelled
        tot["enr"]       += enrolled;  tot["na"]        += na_cnt
        tot["pres"]      += present;   tot["abs"]       += absent
        tot["fb"]        += len(sf)
        tot["exp_fb"]    += present
        tot["ratings"].extend(ratings)

    # Sort: Sessions descending, then Att % ascending
    ss_rows.sort(key=lambda r: (-r[0], r[1]))

    for (n_sess, att_pct,
         cn, ct, tn, cancelled, enrolled, na_cnt, present, absent,
         avg_dur, avg_rat, fb_rate, ratings, _) in ss_rows:

        row_bg = C_ROW_ALT if row_num % 2 == 0 else C_WHITE
        for col, (v, al) in enumerate(zip(
            [cn, ct, tn, n_sess, cancelled, enrolled, na_cnt, present, absent],
            ["left", "left", "left", "center", "center", "center", "center", "center", "center"]
        ), 1):
            style_data_cell(ws, row_num, col, v, bg=row_bg, h_align=al)

        # Att % (col 10)
        ac = ws.cell(row=row_num, column=10)
        ac.value = att_pct; ac.number_format = '0.0"%"'
        ac.font  = _font(bold=True, size=10,
                         color=C_GREEN_DARK if att_pct >= 75 else C_RED_DARK)
        ac.fill  = _fill(_att_bg(att_pct, "Present"))
        ac.alignment = _align("center"); ac.border = _border()

        # Avg Duration (col 11)
        style_data_cell(ws, row_num, 11, avg_dur if avg_dur else "—",
                        bg=row_bg, h_align="center")

        # Avg Rating (col 12) — 2 decimal places
        rc = ws.cell(row=row_num, column=12)
        rc.value         = avg_rat if avg_rat else "—"
        if avg_rat:
            rc.number_format = '0.00'
        rc.fill          = _fill(_rating_bg(avg_rat))
        rc.alignment     = _align("center"); rc.border = _border()
        rc.font          = _font(bold=True, size=10)

        # Feedback Rate % (col 13)
        fc = ws.cell(row=row_num, column=13)
        fc.value = fb_rate; fc.number_format = '0.0"%"'
        fc.alignment = _align("center"); fc.border = _border()
        fc.font  = _font(bold=True, size=10,
                         color=C_GREEN_DARK if fb_rate >= 75 else C_RED_DARK)
        fc.fill  = _fill(C_GREEN_PALE if fb_rate >= 75 else C_AMBER)

        # Breakdown columns (daily for weekly, weekly for monthly)
        for bp_i in range(n_bkd):
            _write_bkd_att_cell(ws, row_num, 14 + bp_i,
                                 bkd_att.get((cn, ct, tn, bp_i)), bg=row_bg)

        # Att % Trend
        prev_att = prev_trend.get((cn, tn))
        _write_trend_cell(ws, row_num, TREND_COL, att_pct, prev_att, bg=row_bg)

        ws.row_dimensions[row_num].height = 18
        row_num += 1

    # Totals row
    if row_num > data_start:
        t   = tot
        bg  = C_BLUE_LITE
        _applic_tot = t["enr"] - t["na"]     # attendance-applicable total
        tp  = round(t["pres"]  / _applic_tot * 100, 1) if _applic_tot else 0.0
        tfr = round(t["fb"]    / t["exp_fb"] * 100, 1) if t["exp_fb"] else 0.0
        tar = round(sum(t["ratings"]) / len(t["ratings"]), 2) if t["ratings"] else 0.0
        tot_vals  = (["⬛  TOTAL / AVERAGE", "", "",
                      t["n_sess"], t["cancelled"], t["enr"], t["na"], t["pres"], t["abs"],
                      tp, "", tar, tfr]
                     + [""] * n_bkd + [""])
        tot_aligns = (["left","left","left",
                       "center","center","center","center","center","center",
                       "center","center","center","center"]
                      + ["center"] * n_bkd + ["center"])
        for col, (v, al) in enumerate(zip(tot_vals, tot_aligns), 1):
            style_data_cell(ws, row_num, col, v if v != "" else "",
                            bg=bg, h_align=al, bold=True)
        ws.row_dimensions[row_num].height = 20

    ws.auto_filter.ref = f"A4:{get_column_letter(N)}{max(row_num, data_start)}"
    auto_col_width(ws)
    _border_thick_outer(ws, 1, 1, max(row_num, data_start), N)


# ─── Sheet 2: Student Detail ──────────────────────────────────────────────────

def build_period_student_detail(ws, att_f: pd.DataFrame, fb_f: pd.DataFrame,
                                 report_type: str, label: str,
                                 prev_trend: dict = None,
                                 start_date=None, end_date=None,
                                 att_susp: pd.DataFrame = None):
    """Period Student Detail: one row per student per course, grouped by course."""
    if att_susp is None:
        att_susp = pd.DataFrame(columns=att_f.columns)
    title      = _period_title(report_type, label, "Student Detail")
    prev_trend = prev_trend or {}

    headers = [
        "Tech Name", "Duration", "Student Name",
        "Sessions", "Present", "Absent", "Att %",
        "Avg Time\nin Session (min)", "Total\nRatings", "Avg\nRating",
        "Att %\nTrend",
    ]
    N         = len(headers)   # 11
    TREND_COL = 11

    style_title_row(ws, 1, 1, N, title)
    ws.cell(row=1, column=1).alignment = _align("center", "center")

    # KPI strip
    # Deduplicate attendance so each student is counted once per session.
    # Also drop phantom/corrupt rows where student_id or course_name is null.
    _sd_cols  = [c for c in ["session_id", "student_id"] if c in att_f.columns]
    _att_dedup = att_f.drop_duplicates(subset=_sd_cols) if _sd_cols else att_f
    # Drop phantom rows — handles both NaN and empty string "" from Google Sheets.
    if "student_id" in _att_dedup.columns:
        _sid2 = _att_dedup["student_id"].astype(str).str.strip()
        _att_dedup = _att_dedup[~_sid2.isin(["", "nan", "None", "NaN"])]
    if "course_name" in _att_dedup.columns:
        _cn2 = _att_dedup["course_name"].astype(str).str.strip()
        _att_dedup = _att_dedup[~_cn2.isin(["", "nan", "None", "NaN"])]
    n_stu_kpi = _att_dedup["student_id"].nunique() if "student_id" in _att_dedup.columns else 0
    _all_pcts = []
    for (_, _, _stid), grp in _att_dedup.groupby(
            ["course_name", "course_title", "student_id"], sort=False):
        n = grp["session_id"].nunique()
        p = int((grp["status"] == "Present").sum())
        if n:
            _all_pcts.append(p / n * 100)
    avg_pct_kpi = round(sum(_all_pcts) / len(_all_pcts), 1) if _all_pcts else 0.0
    at_risk_kpi = sum(1 for x in _all_pcts if x < 75)

    kpis = [
        ("Total Students", n_stu_kpi,            "👥", KPI_BLUE),
        ("At Risk (<75%)", at_risk_kpi,           "⚠️",  KPI_RED),
        ("Avg Att %",      f"{avg_pct_kpi:.1f}%", "📊",
         KPI_GREEN if avg_pct_kpi >= 75 else KPI_RED),
    ]
    _write_kpi_strip(ws, row_lbl=2, row_val=3, kpis=kpis, n_cols=N)

    write_header_row(ws, 4, headers)
    ws.freeze_panes = "D5"
    data_start = row_num = 5

    for k in ["course_name", "course_title", "student_id", "student_name"]:
        if k not in att_f.columns:
            att_f[k] = ""

    # Strip whitespace from student_name to merge duplicates caused by trailing
    # spaces/tabs across sessions (e.g. 'Tushar Koli ' vs 'Tushar Koli').
    # We pick the canonical name as the stripped version of the most-frequent
    # (or first) name for each student_id.
    if "student_name" in _att_dedup.columns:
        _att_dedup = _att_dedup.copy()
        _att_dedup["student_name"] = _att_dedup["student_name"].astype(str).str.strip()

    for (cn, ct), course_grp in _att_dedup.groupby(["course_name", "course_title"], sort=True):
        write_section_banner(ws, row_num, N, f"  {cn}  —  {ct}", C_BLUE_MID, h_align="center")
        row_num += 1

        # Session IDs for this (course_name, course_title) group — used to scope
        # feedback lookup to only sessions in this course.
        course_sids = set(course_grp["session_id"].dropna().unique())

        # Collect per-student stats first so we can sort by Att % descending
        student_rows = []
        for stid, grp in course_grp.groupby("student_id", sort=True):
            # Pick the most-frequent name (after stripping) as the canonical name
            sn = (grp["student_name"].mode().iloc[0]
                  if not grp["student_name"].empty else stid)
            # Attendance-NOT-applicable student (is_attendance_required != Y):
            # still listed under the course, but excluded from all counts.
            na = ("_attn_applicable" in grp.columns and not bool(grp["_attn_applicable"].any()))
            # grp is already deduplicated — one row per session per student
            n_sess  = grp["session_id"].nunique()
            present = int((grp["status"] == "Present").sum())
            absent  = int((grp["status"] == "Absent").sum())
            att_pct = round(present / n_sess * 100, 1) if n_sess else 0.0

            pres_rows = grp[grp["status"] == "Present"]
            avg_dur   = 0.0
            if not pres_rows.empty and "_dur_min" in pres_rows.columns:
                mv = pres_rows["_dur_min"].mean()
                avg_dur = round(float(mv), 1) if mv and float(mv) > 0 else 0.0

            # Ratings scoped to this student + this course's sessions only
            n_ratings = 0
            avg_rat   = 0.0
            if not fb_f.empty and "student_id" in fb_f.columns and "_rating" in fb_f.columns:
                sf_stu = fb_f[
                    (fb_f["student_id"] == stid) &
                    (fb_f["session_id"].isin(course_sids))
                ]
                valid     = sf_stu["_rating"].dropna()
                n_ratings = len(valid)
                avg_rat   = round(float(valid.mean()), 2) if n_ratings else 0.0

            student_rows.append((att_pct, cn, ct, str(stid), sn or stid, n_sess,
                                  present, absent, avg_dur, n_ratings, avg_rat, na))

        # Sort: attendance-applicable first (Att % desc), Not-Applicable rows last.
        student_rows.sort(key=lambda r: (1 if r[11] else 0, -r[0]))

        for (att_pct, _cn, _ct, stid, sn, n_sess, present, absent,
             avg_dur, n_ratings, avg_rat, na) in student_rows:
            # ── Attendance Not Applicable: dark-grey, strike-through, no counts ──
            if na:
                _na_cells = [_cn, _ct, sn, "Not Applicable", "—", "—",
                             "Not Applicable", "—", "—", "—", "—"]
                for col, v in enumerate(_na_cells, 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = v
                    c.font      = _font(size=10, color=C_WHITE, strike=True,
                                        bold=(col in (4, 7)))
                    c.fill      = _fill(C_GREY_NA)
                    c.alignment = _align("left" if col <= 3 else "center", "center")
                    c.border    = _border()
                ws.row_dimensions[row_num].height = 18
                row_num += 1
                continue

            row_bg = C_ROW_ALT if row_num % 2 == 0 else C_WHITE
            for col, (v, al) in enumerate(zip(
                [_cn, _ct, sn, n_sess, present, absent],
                ["left", "left", "left", "center", "center", "center"]
            ), 1):
                style_data_cell(ws, row_num, col, v, bg=row_bg, h_align=al)

            ac = ws.cell(row=row_num, column=7)
            ac.value     = att_pct; ac.number_format = '0.0"%"'
            ac.font      = _font(bold=True, size=10,
                                 color=C_GREEN_DARK if att_pct >= 75 else C_RED_DARK)
            ac.fill      = _fill(_att_bg(att_pct, "Present"))
            ac.alignment = _align("center"); ac.border = _border()

            style_data_cell(ws, row_num, 8,
                            avg_dur if avg_dur else "—", bg=row_bg, h_align="center")

            # Total Ratings column (col 9)
            style_data_cell(ws, row_num, 9,
                            n_ratings if n_ratings else "—", bg=row_bg, h_align="center")

            # Avg Rating (col 10) — 2 decimal places
            rc = ws.cell(row=row_num, column=10)
            rc.value     = avg_rat if avg_rat else "—"
            if avg_rat:
                rc.number_format = '0.00'
            rc.alignment = _align("center"); rc.border = _border()
            rc.font      = _font(bold=True, size=10)
            rc.fill      = _fill(_rating_bg(avg_rat))

            # Att % Trend
            prev_att = prev_trend.get((_cn, stid))
            _write_trend_cell(ws, row_num, TREND_COL, att_pct, prev_att, bg=row_bg)

            ws.row_dimensions[row_num].height = 18
            row_num += 1

    # ── Append suspended students (period) ────────────────────────────────────
    if not att_susp.empty:
        _susp_dedup = att_susp.drop_duplicates(
            subset=["student_id", "course_name"]
        ) if "student_id" in att_susp.columns and "course_name" in att_susp.columns else att_susp
        if not _susp_dedup.empty:
            write_section_banner(ws, row_num, N,
                                 "  Suspended Students (excluded from calculations)",
                                 C_GREY_BD, h_align="center")
            row_num += 1
            for _, r in _susp_dedup.iterrows():
                for col, (v, al) in enumerate(zip(
                    [r.get("course_name",""), r.get("course_title",""),
                     r.get("student_name","")],
                    ["left", "left", "left"]
                ), 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = v
                    c.font      = _font(italic=True, size=10, color="666666")
                    c.fill      = _fill(C_GREY_SUSP)
                    c.alignment = _align(al, "center")
                    c.border    = _border()
                # Sessions/Present/Absent columns — dashes
                for col in range(4, 7):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = "—"
                    c.font      = _font(italic=True, size=10, color="999999")
                    c.fill      = _fill(C_GREY_SUSP)
                    c.alignment = _align("center", "center")
                    c.border    = _border()
                # Att % column (col 7) — "Suspended" bold
                ac = ws.cell(row=row_num, column=7)
                ac.value     = "Suspended"
                ac.font      = _font(bold=True, size=10, color="000000")
                ac.fill      = _fill(C_GREY_SUSP)
                ac.alignment = _align("center", "center")
                ac.border    = _border()
                # Remaining columns — grey dashes
                for col in range(8, N + 1):
                    c = ws.cell(row=row_num, column=col)
                    c.value     = "—"
                    c.font      = _font(italic=True, size=10, color="999999")
                    c.fill      = _fill(C_GREY_SUSP)
                    c.alignment = _align("center", "center")
                    c.border    = _border()
                ws.row_dimensions[row_num].height = 18
                row_num += 1

    ws.auto_filter.ref = f"A4:{get_column_letter(N)}{max(row_num - 1, data_start)}"
    auto_col_width(ws)
    _border_thick_outer(ws, 1, 1, max(row_num - 1, data_start), N)


# ─── Sheet 3: Feedback Rating ─────────────────────────────────────────────────

def build_period_feedback_rating(ws, sess_f: pd.DataFrame, att_f: pd.DataFrame,
                                  fb_f: pd.DataFrame, report_type: str, label: str,
                                  prev_trend: dict = None,
                                  start_date=None, end_date=None):
    """Period Feedback Rating: one row per Course + Subject + Tutor."""
    sess_f     = _prefer_instructor_name(sess_f)   # Instructor column ← Instructor_Name
    title      = _period_title(report_type, label, "Feedback Rating")
    prev_trend = prev_trend or {}

    headers = [
        "Tech Name", "Duration", "Instructor",
        "Sessions", "Unique\nStudents", "Expected\nFeedbacks",
        "Received", "Feedback\nRate %", "Avg\nRating", "Min\nRating", "Max\nRating",
        "Avg Rating\nTrend",
    ]
    N         = len(headers)   # 12
    TREND_COL = 12

    style_title_row(ws, 1, 1, N, title)
    ws.cell(row=1, column=1).alignment = _align("center", "center")

    # KPI strip — feedback rate = feedbacks / total present students (dedup)
    _fbr_dd_cols   = [c for c in ["session_id", "student_id"] if c in att_f.columns]
    _att_dedup_fbr = att_f.drop_duplicates(subset=_fbr_dd_cols) if _fbr_dd_cols else att_f
    # Drop phantom rows — handles both NaN and empty string "" from Google Sheets.
    if "student_id" in _att_dedup_fbr.columns:
        _fbr_sid = _att_dedup_fbr["student_id"].astype(str).str.strip()
        _att_dedup_fbr = _att_dedup_fbr[~_fbr_sid.isin(["", "nan", "None", "NaN"])]
    if "session_id" in _att_dedup_fbr.columns:
        _fbr_sesid = _att_dedup_fbr["session_id"].astype(str).str.strip()
        _att_dedup_fbr = _att_dedup_fbr[~_fbr_sesid.isin(["", "nan", "None", "NaN"])]
    _kpi_present   = int((_applicable(_att_dedup_fbr)["status"] == "Present").sum())
    all_rat_kpi    = fb_f["_rating"].dropna().tolist() \
                     if not fb_f.empty and "_rating" in fb_f.columns else []
    kpi_avg_rat    = round(sum(all_rat_kpi) / len(all_rat_kpi), 2) if all_rat_kpi else 0.0
    kpi_fb_rate    = round(len(fb_f) / _kpi_present * 100, 1) if _kpi_present else 0.0
    kpis = [
        ("Total Feedbacks", len(fb_f),                 "💬", KPI_BLUE),
        ("Avg Rating",      f"{kpi_avg_rat:.2f}/10",   "⭐",
         KPI_GREEN if kpi_avg_rat >= 7 else KPI_RED),
        ("Feedback Rate %", f"{kpi_fb_rate:.1f}%",    "📊",
         KPI_GREEN if kpi_fb_rate >= 75 else KPI_RED),
    ]
    _write_kpi_strip(ws, row_lbl=2, row_val=3, kpis=kpis, n_cols=N)

    write_header_row(ws, 4, headers)
    ws.freeze_panes = "D5"
    data_start = row_num = 5

    for k in ["course_name", "course_title", "tutor_name"]:
        if k not in sess_f.columns:
            sess_f[k] = ""

    tot = {"recv": 0, "exp": 0, "ratings": []}

    # ── Collect all rows first so we can sort before writing ──────────────────
    fbr_rows = []
    for (cn, ct, tn), grp in sess_f.groupby(
            ["course_name", "course_title", "tutor_name"], sort=True):
        sids    = grp["session_id"].tolist()
        # Deduplicate: one row per (session_id, student_id)
        _sa_raw = att_f[att_f["session_id"].isin(sids)]
        _sa_dd  = _sa_raw.drop_duplicates(subset=_fbr_dd_cols) if _fbr_dd_cols else _sa_raw
        sf      = fb_f[fb_f["session_id"].isin(sids)] if not fb_f.empty else pd.DataFrame()

        n_sess    = len(sids)
        n_present = int((_applicable(_sa_dd)["status"] == "Present").sum())
        recv      = len(sf)
        fb_rate   = round(recv / n_present * 100, 1) if n_present else 0.0

        ratings = sf["_rating"].dropna().tolist() \
                  if not sf.empty and "_rating" in sf.columns else []
        avg_rat = round(sum(ratings) / len(ratings), 2) if ratings else 0.0
        min_rat = int(min(ratings)) if ratings else "—"
        max_rat = int(max(ratings)) if ratings else "—"

        fbr_rows.append((n_sess, fb_rate,
                         cn, ct, tn, n_present, recv, avg_rat, min_rat, max_rat, ratings))

        tot["recv"] += recv; tot["exp"] += n_present
        tot["ratings"].extend(ratings)

    # Sort: Sessions descending, then Feedback Rate % ascending
    fbr_rows.sort(key=lambda r: (-r[0], r[1]))

    for (n_sess, fb_rate,
         cn, ct, tn, n_present, recv, avg_rat, min_rat, max_rat, _) in fbr_rows:

        row_bg = C_ROW_ALT if row_num % 2 == 0 else C_WHITE
        for col, (v, al) in enumerate(zip(
            [cn, ct, tn, n_sess, n_present, n_present, recv],
            ["left","left","left","center","center","center","center"]
        ), 1):
            style_data_cell(ws, row_num, col, v, bg=row_bg, h_align=al)

        fc = ws.cell(row=row_num, column=8)
        fc.value = fb_rate; fc.number_format = '0.0"%"'
        fc.alignment = _align("center"); fc.border = _border()
        fc.font  = _font(bold=True, size=10,
                         color=C_GREEN_DARK if fb_rate >= 75 else C_RED_DARK)
        fc.fill  = _fill(C_GREEN_PALE if fb_rate >= 75 else C_AMBER)

        rc = ws.cell(row=row_num, column=9)
        rc.value     = avg_rat if avg_rat else "—"
        if avg_rat:
            rc.number_format = '0.00'
        rc.alignment = _align("center"); rc.border = _border()
        rc.font      = _font(bold=True, size=10)
        rc.fill      = _fill(_rating_bg(avg_rat))

        style_data_cell(ws, row_num, 10, min_rat, bg=row_bg, h_align="center")
        style_data_cell(ws, row_num, 11, max_rat, bg=row_bg, h_align="center")

        # Avg Rating Trend — percentage change vs previous period
        prev_rat      = prev_trend.get((cn, tn))
        _prev_rat_pct = round((prev_rat / 10) * 100, 2) if prev_rat else None
        _curr_rat_pct = round((avg_rat  / 10) * 100, 2) if avg_rat  else None
        _write_trend_cell(ws, row_num, TREND_COL, _curr_rat_pct, _prev_rat_pct, bg=row_bg)

        ws.row_dimensions[row_num].height = 18
        row_num += 1

    if row_num > data_start:
        t   = tot
        bg  = C_BLUE_LITE
        tfr = round(t["recv"] / t["exp"] * 100, 1) if t["exp"] else 0.0
        tar = round(sum(t["ratings"]) / len(t["ratings"]), 2) if t["ratings"] else 0.0
        tot_vals   = ["⬛  TOTAL / AVERAGE","","","","","",
                      t["recv"], tfr, tar,"","", ""]
        tot_aligns = ["left","left","left","center","center","center",
                      "center","center","center","center","center","center"]
        for col, (v, al) in enumerate(zip(tot_vals, tot_aligns), 1):
            style_data_cell(ws, row_num, col, v if v != "" else "",
                            bg=bg, h_align=al, bold=True)
        ws.row_dimensions[row_num].height = 20

    ws.auto_filter.ref = f"A4:{get_column_letter(N)}{max(row_num, data_start)}"
    auto_col_width(ws)
    _border_thick_outer(ws, 1, 1, max(row_num, data_start), N)


# ─── Sheet 4: Teacher Feedback Status ────────────────────────────────────────

def build_period_teacher_no_feedback(ws, sess_f: pd.DataFrame, tf_f: pd.DataFrame,
                                      report_type: str, label: str):
    """Period Teacher Feedback Status: one row per Course + Subject + Tutor."""
    title   = _period_title(report_type, label, "Teacher Feedback Status")
    headers = [
        "Course Name", "Subject", "Tutor",
        "Total\nSessions", "With\nFeedback", "Without\nFeedback",
        "Coverage %", "Missing Session Dates", "Remark",
    ]
    N = len(headers)

    style_title_row(ws, 1, 1, N, title)
    ws.cell(row=1, column=1).alignment = _align("center", "center")
    write_header_row(ws, 2, headers)
    ws.freeze_panes = "D3"
    data_start = row_num = 3

    for k in ["course_name", "course_title", "tutor_name"]:
        if k not in sess_f.columns:
            sess_f[k] = ""

    tf_sids = set(tf_f["session_id"].tolist()) \
              if not tf_f.empty and "session_id" in tf_f.columns else set()

    # Collect all rows first so we can sort by Coverage % ascending
    teacher_rows = []
    for (cn, ct, tn), grp in sess_f.groupby(
            ["course_name", "course_title", "tutor_name"], sort=True):
        sids       = grp["session_id"].tolist()
        n_sess     = len(sids)
        with_fb    = [s for s in sids if s in tf_sids]
        without_fb = [s for s in sids if s not in tf_sids]
        n_with     = len(with_fb)
        n_without  = len(without_fb)
        coverage   = round(n_with / n_sess * 100, 1) if n_sess else 0.0

        missing_dates = []
        for msid in without_fb:
            mrow = grp[grp["session_id"] == msid]
            if not mrow.empty and "_date" in mrow.columns:
                d = mrow.iloc[0]["_date"]
                if d is not None and not (isinstance(d, float) and pd.isna(d)):
                    ds = d.strftime("%d-%b") if hasattr(d, "strftime") else str(d)[:10]
                    missing_dates.append(ds)
        missing_str = ", ".join(sorted(missing_dates)) if missing_dates else "—"

        if n_without == 0:
            row_bg = C_GREEN_PALE
            remark = "✅ All sessions received feedback"
        elif n_with > 0:
            row_bg = C_AMBER
            remark = f"⚠️ Missing: {missing_str}"
        else:
            row_bg = C_RED_LITE
            remark = "❌ No feedback received at all"

        teacher_rows.append((coverage, cn, ct, tn, n_sess, n_with, n_without,
                              coverage, missing_str, remark, row_bg))

    # Sort: Total Sessions descending, then Coverage % ascending
    teacher_rows.sort(key=lambda r: (-r[4], r[0]))

    for (_, cn, ct, tn, n_sess, n_with, n_without,
         coverage, missing_str, remark, row_bg) in teacher_rows:
        for col, (v, al) in enumerate(zip(
            [cn, ct, tn, n_sess, n_with, n_without],
            ["left","left","left","center","center","center"]
        ), 1):
            style_data_cell(ws, row_num, col, v, bg=row_bg, h_align=al)

        cc = ws.cell(row=row_num, column=7)
        cc.value     = coverage; cc.number_format = '0.0"%"'
        cc.alignment = _align("center"); cc.border = _border()
        cc.font      = _font(bold=True, size=10,
                             color=C_GREEN_DARK if coverage >= 75 else C_RED_DARK)
        cc.fill      = _fill(C_GREEN_PALE if coverage >= 75 else
                             (C_AMBER if coverage > 0 else C_RED_LITE))

        style_data_cell(ws, row_num, 8, missing_str, bg=row_bg, h_align="left")
        style_data_cell(ws, row_num, 9, remark,      bg=row_bg, h_align="left", bold=True)
        ws.row_dimensions[row_num].height = 22
        row_num += 1

    ws.auto_filter.ref = f"A2:{get_column_letter(N)}{max(row_num - 1, data_start)}"
    auto_col_width(ws)
    _border_thick_outer(ws, 1, 1, max(row_num - 1, data_start), N)


# ─── Period report workbook entry point ──────────────────────────────────────

def generate_period_report(report_type: str, label: str,
                            sess_f: pd.DataFrame, att_f: pd.DataFrame,
                            fb_f: pd.DataFrame, tf_f: pd.DataFrame,
                            prev_ss_trend: dict = None,
                            prev_fb_trend: dict = None,
                            prev_sd_trend: dict = None,
                            start_date=None, end_date=None,
                            att_susp: pd.DataFrame = None) -> openpyxl.Workbook:
    """Build the full 4-sheet period workbook and return it."""
    if att_susp is None:
        att_susp = pd.DataFrame(columns=att_f.columns)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("[Report] Sheet 1 — Session Summary …")
    build_period_session_summary(
        wb.create_sheet("Session Summary"), sess_f, att_f, fb_f, report_type, label,
        prev_trend=prev_ss_trend or {},
        start_date=start_date, end_date=end_date)

    print("[Report] Sheet 2 — Student Detail …")
    build_period_student_detail(
        wb.create_sheet("Student Detail"), att_f, fb_f, report_type, label,
        prev_trend=prev_sd_trend or {},
        start_date=start_date, end_date=end_date, att_susp=att_susp)

    print("[Report] Sheet 3 — Feedback Rating …")
    build_period_feedback_rating(
        wb.create_sheet("Feedback Rating"), sess_f, att_f, fb_f, report_type, label,
        prev_trend=prev_fb_trend or {},
        start_date=start_date, end_date=end_date)

    print("[Report] Sheet 4 — Teacher Feedback Status …")
    build_period_teacher_no_feedback(
        wb.create_sheet("Teacher_No_Feedback"), sess_f, tf_f, report_type, label)

    return wb


def main():
    # ── Execution is controlled by the RUN CONFIGURATION variables at the top of
    #    this file (report_type / report_date / start_date / end_date / send_email),
    #    replacing the previous command-line arguments. They are gathered into a
    #    lightweight namespace so the report-generation logic below is unchanged.
    #    NOTE: start_date / end_date are read via globals() because both names are
    #    also reused as local variables (the resolved date objects) further down.
    from types import SimpleNamespace
    _cfg = globals()
    args = SimpleNamespace(
        type=report_type,
        date=report_date,
        start=_cfg.get("start_date"),
        end=_cfg.get("end_date"),
        no_email=not send_email,
    )

    def _config_error(msg):
        """Report an invalid RUN CONFIGURATION and exit (mirrors old parser.error)."""
        print(f"[Config Error] {msg}", file=sys.stderr)
        sys.exit(2)

    if args.type not in ("daily", "weekly", "fortnightly", "monthly",
                         "quarterly", "yearly", "manual"):
        _config_error(
            f"report_type = {args.type!r} is not supported. Valid values: "
            "daily, weekly, fortnightly, monthly, quarterly, yearly, manual.")

    # ── Resolve date range ────────────────────────────────────────────────────
    if args.type == "manual":
        if not args.start or not args.end:
            _config_error('report_type = "manual" requires both start_date and end_date. '
                          'Example: start_date = "01-Feb-2026"  end_date = "31-Mar-2026"')
        try:
            from datetime import datetime as _dt
            start_date = _dt.strptime(args.start.strip(), "%d-%b-%Y").date()
            end_date   = _dt.strptime(args.end.strip(),   "%d-%b-%Y").date()
        except ValueError:
            _config_error("start_date / end_date must be in DD-Mon-YYYY format, "
                          'e.g. "01-Feb-2026" or "31-Mar-2026"')
        if start_date > end_date:
            _config_error(f"start_date ({args.start}) must not be after end_date ({args.end})")
        label = f"{args.start.strip()} To {args.end.strip()}"
    else:
        ref_date              = date.fromisoformat(args.date) if args.date else date.today()
        start_date, end_date, label = get_date_range(args.type, ref_date)

    # ── Daily Report: widen the window to "yesterday 12:00 PM → now" ───────────
    # The report is run every morning, so the previous day's afternoon/evening
    # batches finish AFTER the prior run and would otherwise be missed. Extending
    # the start back to yesterday noon captures those alongside today's morning
    # sessions. Applies to the Daily Report ONLY — all other types are untouched.
    daily_window = None      # (start_dt, end_dt) passed to load_all_data
    yest_date    = None      # rows on this date are highlighted in the report
    if args.type == "daily":
        win_start = datetime.combine(ref_date - timedelta(days=1), time(12, 0, 0))
        if ref_date == date.today():
            win_end = datetime.now()                                  # live run
        else:
            win_end = datetime.combine(ref_date, time(23, 59, 59))    # historic --date
        daily_window = (win_start, win_end)
        yest_date    = ref_date - timedelta(days=1)
        label        = (f"{win_start.strftime('%d-%b-%Y %I:%M %p')} - "
                        f"{win_end.strftime('%d-%b-%Y %I:%M %p')}")

    sep = "=" * 64
    print(f"\n{sep}\n  IntelliBI Report Generator")
    print(f"  Type  : {args.type.upper()}")
    print(f"  Period: {start_date} → {end_date}  ({label})")
    print(f"{sep}\n")

    # Load data
    from utils import get_sheets_service  # utils.py is in _BASE_DIR, added to sys.path above
    service               = get_sheets_service(SERVICE_ACCOUNT_FILE)
    sess_f, att_f, fb_f, tf_f, att_susp = load_all_data(service, start_date, end_date,
                                                        time_window=daily_window)

    if sess_f.empty:
        print("[Warn] No sessions found for this period — report will be blank.")

    # Build workbook — daily uses its own functions; all period types (including
    # manual custom range) use the shared generate_period_report() pipeline.
    if args.type == "daily":
        # Load previous day data for trend columns
        prev_daily_ss: dict = {}
        prev_daily_fb: dict = {}
        prev_daily_sd: dict = {}
        try:
            print("[Trend] Finding per-course most recent previous session …")
            prev_daily_ss, prev_daily_fb, prev_daily_sd = _collect_daily_trends(
                service, start_date)
        except Exception as _te:
            print(f"[Trend] Could not build daily trends: {_te} — trends will show '—'")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        print("[Report] Sheet 1 — Session Summary …")
        build_session_summary(wb.create_sheet("Session Summary"), sess_f, att_f, args.type, label,
                              fb_f=fb_f, prev_trend=prev_daily_ss, yest_date=yest_date)
        print("[Report] Sheet 2 — Student Detail …")
        build_student_detail(wb.create_sheet("Student Detail"), sess_f, att_f, args.type, label,
                             fb_f=fb_f, prev_sd_lookup=prev_daily_sd, att_susp=att_susp, yest_date=yest_date)
        print("[Report] Sheet 3 — Feedback Rating …")
        build_feedback_rating(wb.create_sheet("Feedback Rating"), sess_f, att_f, fb_f, args.type, label,
                              prev_trend=prev_daily_fb, yest_date=yest_date)
        print("[Report] Sheet 4 — Teacher No Feedback …")
        build_teacher_no_feedback(wb.create_sheet("Teacher_No_Feedback"), sess_f, tf_f, args.type, label,
                                  att_f=att_f, yest_date=yest_date)
    else:
        # For period reports (weekly/monthly/quarterly/yearly), load the previous
        # period data to compute Att% and Avg-Rating trend columns.
        prev_ss_trend: dict = {}
        prev_fb_trend: dict = {}
        prev_sd_trend: dict = {}
        if args.type in ("weekly", "monthly", "quarterly", "yearly"):
            try:
                print(f"[Trend] Finding per-course most recent previous {args.type} period …")
                prev_ss_trend, prev_fb_trend, prev_sd_trend = _collect_period_trends(
                    service, args.type, start_date)
            except Exception as _te:
                print(f"[Trend] Could not build period trends: {_te} — trends will show '—'")

        # Covers: weekly, monthly, quarterly, yearly, manual
        wb = generate_period_report(args.type, label, sess_f, att_f, fb_f, tf_f,
                                    prev_ss_trend=prev_ss_trend,
                                    prev_fb_trend=prev_fb_trend,
                                    prev_sd_trend=prev_sd_trend,
                                    start_date=start_date,
                                    end_date=end_date,
                                    att_susp=att_susp)

    # Build filename
    # ":" only appears in the Daily Report's time-range label; strip it so the
    # filename stays valid on Windows/Drive (colons are illegal in filenames).
    safe_label = label.replace(" ", "_").replace("–", "to").replace("/", "-").replace(":", ".")
    if args.type == "fortnightly":
        filename = f"IntelliBI_Fortnightly_Report_{safe_label}.xlsx"
    else:
        filename = f"IntelliBI_{args.type.title()}_Attendance_Feedback_Report_{safe_label}.xlsx"

    # Save workbook to an in-memory buffer — no local file created
    buf = io.BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()   # keep bytes for email attachment
    print(f"[Report] Built in memory → {filename} ({len(file_bytes):,} bytes)")

    # Stats (att_f already excludes suspended students)
    n_total  = len(att_f)
    n_absent = int((att_f["status"] == "Absent").sum())
    n_present= n_total - n_absent
    combined_att_pct = round(n_present / n_total * 100, 1) if n_total else 0.0
    n_suspended = len(att_susp)

    # Google Drive upload (streams directly from buffer)
    print("[Drive] Uploading report …")
    upload_to_gdrive(SERVICE_ACCOUNT_FILE, args.type, filename, buf)

    # Local computer directory save — controlled by upload_to_local_directory.
    #   True  -> also save the report to LOCAL_UPLOAD_DIR; False -> skip entirely.
    # Independent of the Drive upload above and the email below.
    if upload_to_local_directory:
        print("[Local] Saving report …")
        save_to_local_directory(args.type, filename, file_bytes)

    # Email
    if args.no_email:
        print("[Email] Skipped (--no-email flag)")
    else:
        send_report(args.type, label, filename, file_bytes,
                    len(sess_f), n_total, n_absent, combined_att_pct)

    print(f"\n{sep}")
    print(f"  Done.  Sessions: {len(sess_f)} | Students: {n_total} | Absent: {n_absent} | Suspended: {n_suspended} | Combined Att%: {combined_att_pct:.1f}%")
    print(f"{sep}\n")


if __name__ == "__main__":
    main()
