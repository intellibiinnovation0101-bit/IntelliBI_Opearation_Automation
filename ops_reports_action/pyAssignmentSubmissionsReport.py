"""
================================================================================
  IntelliBI Assignment Submissions Report Generator
  Reads live data from Google Sheets → Builds styled Excel report → Emails/Uploads

  Sheets read:
    IntelliBIAssessmentSubmission  (Submissions, Assessment_Assigned,
                                    Assessment_Not_Assigned)

  Report sheets produced:
    1. Assignment_Summary   — KPI strip + assignment activity table
    2. Class_Breakdown      — per-class submission metrics
    3. Student_Detail       — per-student per-assignment breakdown
    4. Not_Assigned         — classes with no assignments in the system

  Modes:
    AUTO (default):
      - Checks for trigger file created by pyAssignmentSubmissions.py
      - If trigger found → generates report for past 1 year, deadline <= today
      - If no trigger   → exits without generating (no changes detected)

    MANUAL:
      - User-specified date range; bypasses trigger check
      - Filters by submission_start_date within the given range + deadline <= today

  Usage:
    python pyAssignmentSubmissionsReport.py                                      # auto mode (trigger-based)
    python pyAssignmentSubmissionsReport.py --type auto                           # explicit auto mode
    python pyAssignmentSubmissionsReport.py --type manual --start 01-Feb-2026 --end 31-Mar-2026
    python pyAssignmentSubmissionsReport.py --type auto --no-email
    python pyAssignmentSubmissionsReport.py --type manual --start 01-Jan-2025 --end 31-Dec-2025 --no-email
================================================================================
"""

# --- IntelliBI Operations Automation portability bootstrap (auto-inserted) ---
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "common"))
import _bootstrap  # noqa: E402  (sys.path + env defaults + config.yaml)
from paths import CREDENTIALS_DIR, CONFIG_DIR, LOGS_DIR, CACHE_DIR as PROJECT_CACHE_DIR  # noqa: E402
# --- end bootstrap ---

import sys
import os
import io
import json
import argparse
import smtplib
from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base      import MIMEBase
from email.mime.text      import MIMEText
from email                import encoders

# ── Resolve parent directory so utils.py / service_account.json are always found
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR   = os.path.dirname(_SCRIPT_DIR)          # IntelliBI Automation/
if _BASE_DIR not in sys.path:
    sys.path.insert(0, _BASE_DIR)

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils  import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────────────────────

SERVICE_ACCOUNT_FILE = os.path.join(CREDENTIALS_DIR, "service_account.json")
SUBMISSION_SHEET_ID  = "1E_pOuZfw4BUhQ1bRMmuc8lPDDJRtoXuGkP-qtHC96HU"
ASSIGNMENT_CONFIG_SHEET_ID = "1NH27v-S9XbW5QV0I3Z41WGY8Z0RwHAJQU6Wdt-PlCn4"

# Hardcoded current date for class-over check in Not_Assigned sheet.
# ClassLastDate < this date → class is over.  Update as needed.
CLASS_CUTOFF_DATE = date(2026, 4, 18)

# E-mail credentials are centralised — edit config_files/email_config.py to rotate.
from email_config import GMAIL_SENDER, GMAIL_APP_PASS
REPORT_TO      = ["intellibihropsb2ch@gmail.com",'info@intellibiinnovationstechnologies.in']
REPORT_CC      = []

# Google Drive folder IDs per report type
GDRIVE_FOLDER_MAP = {
    "auto":    "13HUGo68ahptotkf3lHqVBPgaGTZ6pB3-",
    "manual":  "1DuL3A_AGMWfvJbiVb6U3eThbzBoJBI0R",
}

# Trigger file path (created by pyAssignmentSubmissions.py on assignment changes)
TRIGGER_FILE = os.path.join(CREDENTIALS_DIR, ".assignment_report_trigger.json")

# ─────────────────────────────────────────────────────────────────────────────
#  COLOUR & STYLE CONSTANTS  (matches pyAttendaceFeedbackReport.py palette)
# ─────────────────────────────────────────────────────────────────────────────

C_NAV        = "1A2E5A"
C_NAV2       = "243F6B"
C_WHITE      = "FFFFFF"
C_BLUE_MID   = "2E75B6"
C_BLUE_LITE  = "BDD7EE"
C_BLUE_PALE  = "DDEEFF"
C_ROW_ALT    = "EBF4FB"
C_GREY_BD    = "9E9E9E"
C_GREY_LITE  = "F5F5F5"

C_GREEN      = "C8E6C9"
C_GREEN_PALE = "E8F5E9"
C_GREEN_DARK = "1B5E20"
C_AMBER      = "FFE0B2"
C_AMBER_PALE = "FFF3E0"
C_AMBER_DARK = "BF360C"
C_RED_LITE   = "FFCDD2"
C_RED_PALE   = "FFE8E8"
C_RED_DARK   = "B71C1C"
C_ORANGE     = "E65100"
C_PURPLE_PALE = "F3E5F5"
C_PURPLE_DARK = "4A148C"
C_TEAL       = "00695C"
C_TEAL_LITE  = "E0F2F1"

KPI_BLUE   = ("1A2E5A", "DDEEFF", "1A2E5A")

# Change-tracking highlight colours
C_HIGHLIGHT_NEW      = "2E7D32"   # dark green  — new assignments/submissions
C_HIGHLIGHT_CHANGED  = "F9A825"   # dark yellow — changed assignments/submissions
C_HIGHLIGHT_NEW_TXT  = "FFFFFF"   # white text on dark green
C_HIGHLIGHT_CHG_TXT  = "000000"   # black text on dark yellow
C_HIGHLIGHT_NEW_DARK = "1B5E20"   # dark green text/banner for new
C_HIGHLIGHT_CHG_DARK = "E65100"   # dark amber text/banner for changed
C_HIGHLIGHT_PURPLE   = "7B1FA2"   # purple — updated submission rows
C_HIGHLIGHT_PURPLE_TXT = "FFFFFF" # white text on purple
C_HIGHLIGHT_SUB_PINK = "F8BBD0"   # pink — submission columns highlight
KPI_GREEN  = ("2E7D32", "E8F5E9", "2E7D32")
KPI_RED    = ("B71C1C", "FFCDD2", "B71C1C")
KPI_AMBER  = ("E65100", "FFF3E0", "BF360C")
KPI_TEAL   = ("00695C", "E0F2F1", "00695C")

# Last-30-days soft styling
C_RECENT_BG   = "F0F4FF"   # very soft blue-grey background
C_RECENT_TXT  = "5C6BC0"   # indigo-ish text for recent items


# ─────────────────────────────────────────────────────────────────────────────
#  BORDER / STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

# ── Performance: cached group-index for repeated equality filters ────────────
# Report builders below filter a base DataFrame by a loop key many times, e.g.
# subs_f[subs_f["assessment_id"] == aid] once per assignment — a full-column scan
# every iteration (O(rows × keys)). _eq_group() computes a groupby index for a
# (DataFrame, column) once, caches it, and turns each lookup into O(1). It returns
# a fresh copy each call, exactly matching the copy semantics AND the row
# content/order of boolean indexing (groupby preserves within-group order), so
# results are identical — only faster. Only used for exact `== scalar` filters;
# `.isin(...)` filters are left untouched (their row order would differ).
_EQ_INDEX_CACHE = {}


def _eq_group(df, col, key):
    ck = (id(df), col)
    ent = _EQ_INDEX_CACHE.get(ck)
    if ent is None or ent[0] is not df:
        ent = (df, {k: g for k, g in df.groupby(col, sort=False)})
        _EQ_INDEX_CACHE[ck] = ent
    g = ent[1].get(key)
    return g.copy() if g is not None else df.iloc[0:0].copy()


def _side(style="thin"):
    return Side(style=style, color=C_GREY_BD)

def _border(left="thin", right="thin", top="thin", bottom="thin"):
    return Border(
        left   = _side(left)   if left   else None,
        right  = _side(right)  if right  else None,
        top    = _side(top)    if top    else None,
        bottom = _side(bottom) if bottom else None,
    )

def _font(bold=False, size=10, color=None, italic=False):
    return Font(name="Arial", bold=bold, size=size,
                color=color or "000000", italic=italic)

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_title_row(ws, row, col_start, col_end, text, row_h=38):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c           = ws.cell(row=row, column=col_start)
    c.value     = text
    c.font      = _font(bold=True, size=14, color=C_NAV)
    c.alignment = _align("center", "center")
    ws.row_dimensions[row].height = row_h

def style_header_cell(ws, row, col, text):
    c           = ws.cell(row=row, column=col)
    c.value     = text
    c.font      = _font(bold=True, size=10, color=C_WHITE)
    c.fill      = _fill(C_NAV)
    c.alignment = _align("center", "center", wrap=True)
    c.border    = _border()
    return c

def write_header_row(ws, row, headers, height=32):
    for col, h in enumerate(headers, 1):
        style_header_cell(ws, row, col, h)
    ws.row_dimensions[row].height = height

def style_data_cell(ws, row, col, value, bg=C_WHITE,
                    bold=False, h_align="left", wrap=False, number_fmt=None):
    c           = ws.cell(row=row, column=col)
    c.value     = value
    c.font      = _font(bold=bold, size=10)
    c.fill      = _fill(bg)
    c.alignment = _align(h_align, "center", wrap=wrap)
    c.border    = _border()
    if number_fmt:
        c.number_format = number_fmt
    return c

def write_section_banner(ws, row, n_cols, text, bg, height=24, h_align="left"):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=n_cols)
    c           = ws.cell(row=row, column=1)
    c.value     = text
    c.font      = _font(bold=True, size=11, color=C_WHITE)
    c.fill      = _fill(bg)
    c.alignment = _align(h_align, "center")
    c.border    = _border()
    ws.row_dimensions[row].height = height

def auto_col_width(ws, min_w=10, max_w=55):
    for col_cells in ws.columns:
        ltr     = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value or "").split("\n")[0])
                       for c in col_cells), default=0)
        ws.column_dimensions[ltr].width = max(min_w, min(max_w, max_len + 3))

def _border_thick_outer(ws, min_row, min_col, max_row, max_col):
    thick = Side(style="medium", color="1A2E5A")
    thin  = Side(style="thin",   color=C_GREY_BD)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            t = thick if cell.row    == min_row else thin
            b = thick if cell.row    == max_row else thin
            l = thick if cell.column == min_col else thin
            r = thick if cell.column == max_col else thin
            cell.border = Border(left=l, right=r, top=t, bottom=b)

def _build_change_legend_text(change_info) -> str:
    """Build a compact legend string from change_info (no row written)."""
    if not change_info:
        return ""
    new_count = len(change_info.get("new", set()))
    chg_count = len(change_info.get("changed", {}))
    ns_count  = len(change_info.get("new_sub", set()))
    us_count  = len(change_info.get("updated_sub", set()))

    if not (new_count or chg_count or ns_count or us_count):
        return ""

    parts = []
    if new_count:
        parts.append(f"New Assignment: {new_count} \U0001f7e2")         # green dot
    if chg_count:
        parts.append(f"Updated Assignment: {chg_count} \U0001f7e0")     # orange dot
    if ns_count:
        parts.append(f"New Submission: {ns_count} \U0001f7e1")          # yellow dot
    if us_count:
        parts.append(f"Updated Submission: {us_count} \U0001f7e3")
    return "  |  ".join(parts)


def _write_change_legend(ws, row, n_cols, change_info):
    """Write a compact colour-coded legend row showing change counts."""
    if not change_info:
        return row
    legend_text = _build_change_legend_text(change_info)
    if not legend_text:
        return row  # nothing to show

    # Merge full width for legend
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1)
    c.value = "  " + legend_text
    c.font  = _font(bold=True, size=9, color="333333")
    c.fill  = _fill("F5F5F5")
    c.alignment = _align("left", "center")
    c.border    = _border()
    ws.row_dimensions[row].height = 20
    return row + 1


def _write_kpi_strip(ws, row_lbl, row_val, kpis, n_cols):
    span = max(1, n_cols // len(kpis))
    for idx, (label, value, icon, pal) in enumerate(kpis):
        c_start = idx * span + 1
        c_end   = c_start + span - 1
        if idx == len(kpis) - 1:
            c_end = n_cols
        lbl_bg, val_bg, txt_col = pal

        if c_start < c_end:
            ws.merge_cells(start_row=row_lbl, start_column=c_start,
                           end_row=row_lbl,   end_column=c_end)
        lc           = ws.cell(row=row_lbl, column=c_start)
        lc.value     = f"{icon}  {label}" if icon else label
        lc.font      = _font(bold=True, size=9, color=C_WHITE)
        lc.fill      = _fill(lbl_bg)
        lc.alignment = _align("center", "center")
        lc.border    = _border()

        if c_start < c_end:
            ws.merge_cells(start_row=row_val, start_column=c_start,
                           end_row=row_val,   end_column=c_end)
        vc           = ws.cell(row=row_val, column=c_start)
        vc.value     = value
        vc.font      = _font(bold=True, size=18, color=txt_col)
        vc.fill      = _fill(val_bg)
        vc.alignment = _align("center", "center")
        vc.border    = _border()

    ws.row_dimensions[row_lbl].height = 20
    ws.row_dimensions[row_val].height = 36


# ─────────────────────────────────────────────────────────────────────────────
#  DATE RANGE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_auto_date_range(ref: date):
    """Auto mode: 1 year back from today."""
    start = date(ref.year - 1, ref.month, ref.day)
    label = f"{start.strftime('%d %b %Y')} – {ref.strftime('%d %b %Y')}"
    return start, ref, label


def report_title(report_type: str, label: str, sheet_name: str) -> str:
    prefixes = {
        "auto": "Assignment", "manual": "Custom Period",
    }
    p = prefixes.get(report_type, report_type.title())
    return f"{p} {sheet_name}  |  Period: {label}"


# ─────────────────────────────────────────────────────────────────────────────
#  GOOGLE SHEETS READER
# ─────────────────────────────────────────────────────────────────────────────

def read_sheet_df(service, spreadsheet_id: str, tab: str) -> pd.DataFrame:
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"{tab}!A:ZZ",
    ).execute()
    values = result.get("values", [])
    if len(values) < 2:
        return pd.DataFrame()
    header = [h.strip() for h in values[0]]
    rows   = [r + [""] * (len(header) - len(r)) for r in values[1:]]
    df     = pd.DataFrame(rows, columns=header)
    str_cols = df.select_dtypes(include=["object", "str"]).columns
    df[str_cols] = df[str_cols].apply(lambda col: col.str.strip())
    return df


# ─────────────────────────────────────────────────────────────────────────────
#  DATE PARSING HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_ist_date(val) -> date | None:
    """Parse 'DD/MM/YYYY HH:MM:SS IST' (or similar) → date.  Returns None on failure."""
    if not val or str(val).strip() in ("", "nan", "None"):
        return None
    s = str(val).replace(" IST", "").strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, errors="coerce").date()
    except Exception:
        return None


def _parse_class_end_date(duration_str) -> date | None:
    """Parse ClassLastDate from 'DD-Mon-YYYY To DD-Mon-YYYY' → date."""
    if not duration_str or str(duration_str).strip() in ("", "nan", "None"):
        return None
    s = str(duration_str).strip()
    parts = s.split("To")
    if len(parts) < 2:
        parts = s.split("to")
    if len(parts) < 2:
        return None
    end_part = parts[-1].strip()
    # "Current Date" means class is ongoing — return None (not over)
    if end_part.lower() in ("current date", "current", "ongoing", "present"):
        return None
    for fmt in ("%d-%b-%Y", "%d/%b/%Y", "%d-%B-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(end_part, fmt).date()
        except ValueError:
            continue
    try:
        return pd.to_datetime(end_part, errors="coerce").date()
    except Exception:
        return None


def _fmt_date(val) -> str:
    """Strip time + IST suffix; return 'MM/DD/YYYY' display string."""
    d = _parse_ist_date(val)
    return d.strftime("%m/%d/%Y") if d else str(val or "")


def _fmt_date_mdy(val) -> str:
    """Alias for _fmt_date — returns 'MM/DD/YYYY'."""
    return _fmt_date(val)


def _safe_float(val) -> float | None:
    try:
        f = float(str(val).strip())
        return f if not pd.isna(f) else None
    except (ValueError, TypeError):
        return None


# ─────────────────────────────────────────────────────────────────────────────
#  DATA LOADER
# ─────────────────────────────────────────────────────────────────────────────

def load_all_data(service, start_date: date, end_date: date):
    """
    Read all three sheets and filter Assessment_Assigned:
      1. submission_start_date must be within [start_date, end_date]
      2. submission_deadline must have already passed (deadline <= today)
    Returns (assigned_f, subs_f, not_assigned_df).
    """
    today = date.today()
    print("[Data] Reading Assessment_Assigned …")
    assigned_df = read_sheet_df(service, SUBMISSION_SHEET_ID, "Assessment_Assigned")

    print("[Data] Reading Submissions …")
    subs_df = read_sheet_df(service, SUBMISSION_SHEET_ID, "Submissions")

    print("[Data] Reading Assessment_Not_Assigned …")
    not_assigned_df = read_sheet_df(service, SUBMISSION_SHEET_ID, "Assessment_Not_Assigned")

    # Filter: submission_start_date within range AND deadline already passed
    if not assigned_df.empty and "submission_start_date" in assigned_df.columns:
        assigned_df["_start_date"] = assigned_df["submission_start_date"].apply(_parse_ist_date)
        assigned_df["_deadline"]   = assigned_df["submission_deadline"].apply(_parse_ist_date)

        mask = assigned_df.apply(
            lambda r: (
                r["_start_date"] is not None
                and not pd.isna(r["_start_date"])
                and start_date <= r["_start_date"] <= end_date
                and r["_deadline"] is not None
                and not pd.isna(r["_deadline"])
                and r["_deadline"] <= today
            ),
            axis=1,
        )
        assigned_f = assigned_df[mask].copy()
    else:
        assigned_f = pd.DataFrame()

    # Filter Submissions to only assessments present in assigned_f
    if not assigned_f.empty and not subs_df.empty and "assessment_id" in subs_df.columns:
        valid_aids = set(assigned_f["assessment_id"].tolist())
        subs_f = subs_df[subs_df["assessment_id"].isin(valid_aids)].copy()
    else:
        subs_f = pd.DataFrame()

    print(f"[Data] Filtered → Assigned: {len(assigned_f)} (deadline <= {today}) | "
          f"Submissions: {len(subs_f)} | Not-Assigned: {len(not_assigned_df)}")
    return assigned_f, subs_f, not_assigned_df


# ─────────────────────────────────────────────────────────────────────────────
#  SUBMISSION RATE COLOUR
# ─────────────────────────────────────────────────────────────────────────────

def _sub_rate_bg(rate: float, alt: int) -> str:
    if rate >= 80:
        return C_GREEN if alt % 2 == 0 else C_GREEN_PALE
    if rate >= 50:
        return C_AMBER if alt % 2 == 0 else C_AMBER_PALE
    return C_RED_LITE if alt % 2 == 0 else C_RED_PALE

def _sub_rate_icon(rate: float) -> str:
    if rate >= 80: return "🟢"
    if rate >= 50: return "🟡"
    return "🔴"

def _status_bg(status: str, alt: int) -> str:
    s = str(status).lower()
    if s == "graded":      return C_GREEN if alt % 2 == 0 else C_GREEN_PALE
    if s == "submitted":   return C_BLUE_PALE if alt % 2 == 0 else C_BLUE_PALE
    return C_RED_LITE if alt % 2 == 0 else C_RED_PALE

def _status_icon(status: str) -> str:
    s = str(status).lower()
    if s == "graded":    return "✅"
    if s == "submitted": return "📤"
    return "❌"


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 1 — ASSIGNMENT SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def build_assignment_summary(ws, assigned_f: pd.DataFrame, subs_f: pd.DataFrame,
                             report_type: str, label: str, change_info: dict = None):
    title = report_title(report_type, label, "Assignment Submission Report")

    HEADERS = [
        "#", "Class Name", "Class Duration", "Assessment Title",
        "Start Date", "Last Modified", "Deadline", "Max\nMarks",
        "Enrolled", "Submitted", "Pending", "Sub Rate %", "Avg Marks",
    ]
    N = len(HEADERS)

    # ── Sort assigned by Start Date DESCENDING (latest first) ─────────────────
    if not assigned_f.empty and "_start_date" in assigned_f.columns:
        assigned_f = assigned_f.sort_values("_start_date", ascending=False).reset_index(drop=True)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    style_title_row(ws, 1, 1, N, title)

    # ── KPI strip (rows 2–3) ─────────────────────────────────────────────────
    n_assignments = len(assigned_f) if not assigned_f.empty else 0
    n_classes     = assigned_f["class_id"].nunique() if not assigned_f.empty else 0

    total_enrolled  = 0
    total_submitted = 0
    avg_marks_vals  = []

    if not subs_f.empty:
        for aid, grp in subs_f.groupby("assessment_id"):
            total_enrolled  += len(grp)
            total_submitted += int((grp["submission_status"] != "Not Submitted").sum())
            for _, r in grp.iterrows():
                m = _safe_float(r.get("evaluation_marks", ""))
                if m is not None:
                    avg_marks_vals.append(m)

    overall_rate = round(total_submitted / total_enrolled * 100, 1) if total_enrolled else 0.0
    avg_marks    = round(sum(avg_marks_vals) / len(avg_marks_vals), 1) if avg_marks_vals else None
    avg_marks_str = f"{avg_marks}" if avg_marks is not None else "N/A"

    kpis = [
        ("Total Assignments",  n_assignments,     "📋", KPI_BLUE),
        ("Classes Assigned",   n_classes,          "🏫", KPI_TEAL),
        ("Overall Sub Rate",   f"{overall_rate}%", "📊",
         KPI_GREEN if overall_rate >= 80 else (KPI_AMBER if overall_rate >= 50 else KPI_RED)),
        ("Avg Marks Scored",   avg_marks_str,      "⭐", KPI_BLUE),
    ]
    _write_kpi_strip(ws, 2, 3, kpis, N)

    # ── Section banner (with change legend appended) ────────────────────────────
    banner_text = "  📋  Assignment Activity"
    legend_text = _build_change_legend_text(change_info)
    if legend_text:
        banner_text += " | " + legend_text
    write_section_banner(ws, 4, N, banner_text, C_NAV2)

    # ── Column headers ────────────────────────────────────────────────────────
    hdr_row = 5
    write_header_row(ws, hdr_row, HEADERS)

    # ── Data rows ─────────────────────────────────────────────────────────────
    row_num = hdr_row + 1
    tot_enrolled = tot_submitted = tot_pending = 0
    today = date.today()
    thirty_days_ago = today - timedelta(days=30)

    if assigned_f.empty:
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=N)
        c = ws.cell(row=row_num, column=1)
        c.value     = "✅ No assignments found for this period."
        c.font      = _font(bold=True, size=11, color="2E7D32")
        c.fill      = _fill(C_GREEN_PALE)
        c.alignment = _align("center", "center")
        c.border    = _border()
        row_num += 1
    else:
        for seq, (_, asgn) in enumerate(assigned_f.iterrows(), 1):
            aid   = str(asgn.get("assessment_id", ""))
            # Compute submission stats from Submissions tab
            if not subs_f.empty and aid:
                grp       = _eq_group(subs_f, "assessment_id", aid)
                enrolled  = len(grp)
                submitted = int((grp["submission_status"] != "Not Submitted").sum())
                pending   = enrolled - submitted
                mvals     = [_safe_float(r.get("evaluation_marks", ""))
                             for _, r in grp.iterrows()]
                mvals     = [v for v in mvals if v is not None]
                avg_m     = round(sum(mvals) / len(mvals), 1) if mvals else None
            else:
                enrolled = submitted = pending = 0
                avg_m = None

            sub_rate = round(submitted / enrolled * 100, 1) if enrolled else 0.0

            # Check if this assignment started in last 30 days
            start_dt = asgn.get("_start_date", None)
            is_recent = (start_dt is not None and not pd.isna(start_dt)
                         and start_dt >= thirty_days_ago)

            # Detect change type — only ASSIGNMENT-level changes highlight rows here
            is_new     = change_info and aid in change_info.get("new", set())
            is_changed = change_info and aid in change_info.get("changed", {})
            has_highlight = is_new or is_changed

            # Background: rate-based default
            bg = _sub_rate_bg(sub_rate, seq)

            # Override bg for new/updated ASSIGNMENT rows only
            if is_new:
                bg = C_HIGHLIGHT_NEW           # dark green 🟢
                txt_color = C_HIGHLIGHT_NEW_TXT
            elif is_changed:
                bg = C_HIGHLIGHT_CHANGED       # dark yellow 🟠
                txt_color = C_HIGHLIGHT_CHG_TXT
            else:
                txt_color = None  # default black

            max_marks = str(asgn.get("maximum_marks", "") or "")

            vals = [
                seq,
                str(asgn.get("class_name", "")),
                str(asgn.get("class_subject", "")),
                str(asgn.get("assessment_title", "")),
                _fmt_date(asgn.get("submission_start_date", "")),
                _fmt_date(asgn.get("last_modified_date", "")),
                _fmt_date(asgn.get("submission_deadline", "")),
                max_marks,
                enrolled, submitted, pending,
                f"{_sub_rate_icon(sub_rate)} {sub_rate:.1f}%",
                f"{avg_m}" if avg_m is not None else "N/A",
            ]
            aligns = ["center","left","left","left","center","center","center",
                      "center","center","center","center","center","center"]

            row_bold = has_highlight
            use_italic = is_recent and not has_highlight
            for col, (v, al) in enumerate(zip(vals, aligns), 1):
                c = style_data_cell(ws, row_num, col, v, bg=bg, bold=row_bold,
                                h_align=al, wrap=(col in (2, 3, 4)))
                if has_highlight:
                    c.font = _font(bold=True, size=10, color=txt_color)
                elif use_italic:
                    c.font = _font(bold=True, size=10,
                                   color="000000", italic=True)
            ws.row_dimensions[row_num].height = 18

            tot_enrolled  += enrolled
            tot_submitted += submitted
            tot_pending   += pending
            row_num += 1

        # ── Totals row ────────────────────────────────────────────────────────
        if row_num > 8:
            tot_rate = round(tot_submitted / tot_enrolled * 100, 1) if tot_enrolled else 0.0
            totals   = ["", "TOTALS", "", "", "", "", "", "",
                        tot_enrolled, tot_submitted, tot_pending,
                        f"{tot_rate:.1f}%", ""]
            for col, v in enumerate(totals, 1):
                c           = ws.cell(row=row_num, column=col)
                c.value     = v
                c.font      = _font(bold=True, size=10, color=C_WHITE)
                c.fill      = _fill(C_NAV)
                c.alignment = _align("center" if col != 2 else "left", "center")
                c.border    = _border()
            ws.row_dimensions[row_num].height = 20
            row_num += 1

    if row_num > hdr_row + 1:
        _border_thick_outer(ws, hdr_row, 1, row_num - 1, N)

    ws.auto_filter.ref  = f"A{hdr_row}:{get_column_letter(N)}{row_num - 1}"
    ws.freeze_panes     = f"A{hdr_row + 1}"
    ws.sheet_properties.tabColor = "1A2E5A"
    auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 2 — CLASS BREAKDOWN
# ─────────────────────────────────────────────────────────────────────────────

def _parse_class_start_date(class_subject: str) -> date | None:
    """
    Parse the class start date from 'Class Duration' field.
    Handles all observed formats:
      '21-Mar-2026 To Current Date'   — dd-Mon-yyyy
      '14 Sep 2025 To 13 Oct 2025'    — dd Mon yyyy
      '13-Oct 2025 To 11-Dec-2025'    — dd-Mon yyyy
      '15-Jan-2026 To 5-March-2026'   — dd-Mon-yyyy (full month)
      '29-03-2026 To Current Date'    — dd-mm-yyyy
    Returns the start date portion, or None for unparseable values.
    """
    import re
    s = str(class_subject or "").strip()
    if not s:
        return None

    # Split on ' To ' and take the start portion
    start_part = s.split(" To ")[0].split("+")[0].strip()
    if not start_part:
        return None

    # Normalise separators: replace dashes between date parts with spaces
    # '21-Mar-2026' → '21 Mar 2026', '29-03-2026' → '29 03 2026'
    normalised = re.sub(r"(\d)[- ](\w)[- ]?", r"\1 \2", start_part)
    # More aggressive: replace all dashes with spaces for date parsing
    normalised_sp = start_part.replace("-", " ").strip()

    # Try multiple date formats on both the original and normalised versions
    for candidate in [start_part, normalised, normalised_sp]:
        candidate = candidate.strip()
        for fmt in ("%d-%b-%Y", "%d %b %Y", "%d-%B-%Y", "%d %B %Y",
                    "%d-%b %Y", "%d-%m-%Y", "%d %m %Y"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue

    return None


def _is_live_batch(class_subject: str) -> bool:
    """Return True if the class duration ends with 'Current Date'."""
    return "current date" in str(class_subject or "").lower()


def build_class_breakdown(ws, assigned_f: pd.DataFrame, subs_f: pd.DataFrame,
                          report_type: str, label: str, change_info: dict = None):
    title = report_title(report_type, label, "Class-wise Breakdown")

    HEADERS = [
        "#", "Class Name", "Class Duration", "Sessions\nConducted",
        "Assignments\nGiven", "Total\nEnrolled",
        "Total\nSubmitted", "Total\nPending",
        "Sub Rate %", "Avg Marks",
    ]
    N = len(HEADERS)

    style_title_row(ws, 1, 1, N, title)

    # ── Section banner (with change legend appended) ────────────────────────────
    banner_text = "  🏫  Per-Class Submission Metrics"
    legend_text = _build_change_legend_text(change_info)
    if legend_text:
        banner_text += " | " + legend_text
    write_section_banner(ws, 2, N, banner_text, C_BLUE_MID)
    hdr_row = 3
    write_header_row(ws, hdr_row, HEADERS)

    row_num = hdr_row + 1
    tot_asgns = tot_enr = tot_sub = tot_pend = 0
    today = date.today()
    thirty_days_ago = today - timedelta(days=30)

    if assigned_f.empty:
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=N)
        c = ws.cell(row=row_num, column=1)
        c.value     = "✅ No assignments found for this period."
        c.font      = _font(bold=True, size=11, color="2E7D32")
        c.fill      = _fill(C_GREEN_PALE)
        c.alignment = _align("center", "center")
        c.border    = _border()
        row_num += 1
    else:
        # Group by class_id to aggregate, then sort by class start date descending
        class_groups = assigned_f.groupby("class_id", sort=False)
        class_data = []
        for cid, cls_asgn in class_groups:
            class_subject = str(cls_asgn["class_subject"].iloc[0])
            cls_start_dt  = _parse_class_start_date(class_subject)
            class_data.append((cid, cls_asgn, cls_start_dt))

        # Sort: parseable dates by start date desc (live batches first),
        #        unparseable formats at end
        def _has_valid_date(d):
            if d is None:
                return False
            try:
                return not pd.isna(d)
            except (TypeError, ValueError):
                return True

        class_data_parsed   = [x for x in class_data if _has_valid_date(x[2])]
        class_data_unparsed = [x for x in class_data if not _has_valid_date(x[2])]

        class_data_parsed.sort(
            key=lambda x: (
                x[2],
                1 if _is_live_batch(str(x[1]["class_subject"].iloc[0])) else 0,
            ),
            reverse=True,
        )
        class_data = class_data_parsed + class_data_unparsed

        seq = 0
        for cid, cls_asgn, cls_start_dt in class_data:
            seq += 1
            class_name    = str(cls_asgn["class_name"].iloc[0])
            class_subject = str(cls_asgn["class_subject"].iloc[0])
            sessions_str  = str(cls_asgn["sessions_since_start"].iloc[0])
            n_asgns       = len(cls_asgn)

            aids = cls_asgn["assessment_id"].tolist()
            if not subs_f.empty:
                cls_subs  = subs_f[subs_f["assessment_id"].isin(aids)]
                enrolled  = len(cls_subs)
                submitted = int((cls_subs["submission_status"] != "Not Submitted").sum())
                pending   = enrolled - submitted
                mvals     = [_safe_float(r.get("evaluation_marks", ""))
                             for _, r in cls_subs.iterrows()]
                mvals     = [v for v in mvals if v is not None]
                avg_m     = round(sum(mvals) / len(mvals), 1) if mvals else None
            else:
                enrolled = submitted = pending = 0
                avg_m = None

            sub_rate = round(submitted / enrolled * 100, 1) if enrolled else 0.0

            # Check if class started in last 30 days
            try:
                is_recent = (cls_start_dt is not None and not pd.isna(cls_start_dt)
                             and cls_start_dt >= thirty_days_ago)
            except (TypeError, ValueError):
                is_recent = False

            # Detect change type (assignment or submission level)
            # Only ASSIGNMENT-level changes highlight rows in Class_Breakdown
            class_has_new = False
            class_has_chg = False
            if change_info:
                class_has_new = any(a in change_info.get("new", set()) for a in aids)
                class_has_chg = any(a in change_info.get("changed", {}) for a in aids)
            has_highlight = class_has_new or class_has_chg

            # Background: rate-based default
            bg = _sub_rate_bg(sub_rate, seq)
            if class_has_new:
                bg = C_HIGHLIGHT_NEW
                txt_color = C_HIGHLIGHT_NEW_TXT
            elif class_has_chg:
                bg = C_HIGHLIGHT_CHANGED
                txt_color = C_HIGHLIGHT_CHG_TXT
            else:
                txt_color = None

            vals = [
                seq, class_name, class_subject, sessions_str,
                n_asgns, enrolled, submitted, pending,
                f"{_sub_rate_icon(sub_rate)} {sub_rate:.1f}%",
                f"{avg_m}" if avg_m is not None else "N/A",
            ]
            aligns = ["center","left","left","center","center","center",
                      "center","center","center","center"]

            row_bold = has_highlight
            use_italic = is_recent and not has_highlight
            for col, (v, al) in enumerate(zip(vals, aligns), 1):
                c = style_data_cell(ws, row_num, col, v, bg=bg, bold=row_bold,
                                h_align=al, wrap=(col in (2, 3)))
                if has_highlight:
                    c.font = _font(bold=True, size=10, color=txt_color)
                elif use_italic:
                    c.font = _font(bold=True, size=10,
                                   color="000000", italic=True)
            ws.row_dimensions[row_num].height = 18

            tot_asgns += n_asgns
            tot_enr   += enrolled
            tot_sub   += submitted
            tot_pend  += pending
            row_num   += 1

        # Totals
        if row_num > 5:
            tot_rate = round(tot_sub / tot_enr * 100, 1) if tot_enr else 0.0
            totals   = ["", "TOTALS", "", "", tot_asgns, tot_enr,
                        tot_sub, tot_pend, f"{tot_rate:.1f}%", ""]
            for col, v in enumerate(totals, 1):
                c           = ws.cell(row=row_num, column=col)
                c.value     = v
                c.font      = _font(bold=True, size=10, color=C_WHITE)
                c.fill      = _fill(C_NAV)
                c.alignment = _align("center" if col != 2 else "left", "center")
                c.border    = _border()
            ws.row_dimensions[row_num].height = 20
            row_num += 1

    if row_num > hdr_row + 1:
        _border_thick_outer(ws, hdr_row, 1, row_num - 1, N)

    ws.auto_filter.ref  = f"A{hdr_row}:{get_column_letter(N)}{row_num - 1}"
    ws.freeze_panes     = f"A{hdr_row + 1}"
    ws.sheet_properties.tabColor = "2E75B6"
    auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 3 — STUDENT DETAIL
# ─────────────────────────────────────────────────────────────────────────────

def build_student_detail(ws, assigned_f: pd.DataFrame, subs_f: pd.DataFrame,
                         report_type: str, label: str, change_info: dict = None):
    title = report_title(report_type, label, "Student Submission Detail")

    HEADERS = [
        "#", "Student Name", "Email", "Phone",
        "Submitted At", "Status",
        "Marks\nObtained", "Max\nMarks", "Feedback",
    ]
    N = len(HEADERS)

    style_title_row(ws, 1, 1, N, title)

    # ── Change legend ─────────────────────────────────────────────────────────
    row_num = 2
    row_num = _write_change_legend(ws, row_num, N, change_info)

    # ── Sort assigned by Start Date descending for Student Detail ──────────────
    if not assigned_f.empty and "_start_date" in assigned_f.columns:
        assigned_f = assigned_f.sort_values("_start_date", ascending=False).reset_index(drop=True)

    if assigned_f.empty:
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=N)
        c = ws.cell(row=row_num, column=1)
        c.value     = "✅ No assignments found for this period."
        c.font      = _font(bold=True, size=11, color="2E7D32")
        c.fill      = _fill(C_GREEN_PALE)
        c.alignment = _align("center", "center")
        c.border    = _border()
        ws.row_dimensions[row_num].height = 24
    else:
        for _, asgn in assigned_f.iterrows():
            aid       = str(asgn.get("assessment_id", ""))
            a_title   = str(asgn.get("assessment_title", "") or aid)
            cls_name  = str(asgn.get("class_name", ""))
            max_marks = str(asgn.get("maximum_marks", "") or "")
            deadline  = _fmt_date(asgn.get("submission_deadline", ""))
            start_dt  = _fmt_date(asgn.get("submission_start_date", ""))

            # ── Assessment banner ─────────────────────────────────────────────
            is_new     = change_info and aid in change_info.get("new", set())
            is_changed = change_info and aid in change_info.get("changed", {})
            is_new_sub = change_info and aid in change_info.get("new_sub", set())
            is_upd_sub = change_info and aid in change_info.get("updated_sub", set())
            has_any_new    = is_new or is_new_sub
            has_any_update = is_changed or is_upd_sub
            change_tag = ""
            banner_bg  = C_NAV2
            if is_new or is_new_sub:
                parts = []
                if is_new:     parts.append("NEW ASSIGNMENT")
                if is_new_sub: parts.append("NEW SUBMISSION")
                change_tag = "  ✅ " + " + ".join(parts)
                banner_bg  = C_HIGHLIGHT_NEW_DARK
            elif is_changed:
                parts = ["UPDATED ASSIGNMENT"]
                if is_upd_sub: parts.append("UPDATED SUBMISSION")
                change_tag = "  ✏️ " + " + ".join(parts)
                banner_bg  = C_HIGHLIGHT_CHG_DARK
            elif is_upd_sub:
                change_tag = "  🟣 UPDATED SUBMISSION"
                banner_bg  = C_HIGHLIGHT_PURPLE

            banner_text = (f"  📝  {a_title}  |  Class: {cls_name}  |  "
                           f"Start: {start_dt}  |  Deadline: {deadline}"
                           + (f"  |  Max Marks: {max_marks}" if max_marks else "")
                           + change_tag)
            write_section_banner(ws, row_num, N, banner_text, banner_bg, height=26)
            row_num += 1

            # ── Column headers for this group ─────────────────────────────────
            write_header_row(ws, row_num, HEADERS, height=26)
            first_data_row = row_num + 1
            row_num += 1

            # ── Student rows ──────────────────────────────────────────────────
            if not subs_f.empty:
                grp = _eq_group(subs_f, "assessment_id", aid).copy()
                # Sort: submitted/graded first, then not-submitted
                grp["_sort"] = grp["submission_status"].apply(
                    lambda s: 0 if str(s).lower() != "not submitted" else 1
                )
                grp = grp.sort_values(["_sort", "submitted_at"])
            else:
                grp = pd.DataFrame()

            # Submission-level change highlighting (Student_Detail only)
            # New Submission 🟡 → dark green bg + white bold on submission cols
            # Updated Submission 🟣 → purple bg + white bold on submission cols
            SUB_COLS = {5, 6, 7, 8, 9}  # Submitted At, Status, Marks, Max Marks, Feedback
            if is_new_sub:
                sub_col_bg  = C_HIGHLIGHT_NEW          # dark green
                sub_col_txt = C_HIGHLIGHT_NEW_TXT      # white
            elif is_upd_sub:
                sub_col_bg  = C_HIGHLIGHT_PURPLE       # purple
                sub_col_txt = C_HIGHLIGHT_PURPLE_TXT   # white
            else:
                sub_col_bg  = None
                sub_col_txt = None

            if grp.empty:
                ws.merge_cells(start_row=row_num, start_column=1,
                               end_row=row_num, end_column=N)
                c           = ws.cell(row=row_num, column=1)
                c.value     = "No submission records found."
                c.font      = _font(italic=True, size=10, color="888888")
                c.fill      = _fill(C_GREY_LITE)
                c.alignment = _align("center", "center")
                c.border    = _border()
                ws.row_dimensions[row_num].height = 18
                row_num += 1
            else:
                for seq_s, (_, sub) in enumerate(grp.iterrows(), 1):
                    status   = str(sub.get("submission_status", "Not Submitted"))
                    bg       = _status_bg(status, seq_s)
                    eval_m   = str(sub.get("evaluation_marks", "") or "")
                    feedback = str(sub.get("evaluation_feedback", "") or "")
                    subm_at  = _fmt_date(sub.get("submitted_at", ""))

                    vals   = [
                        seq_s,
                        str(sub.get("student_name",  "") or ""),
                        str(sub.get("student_email", "") or ""),
                        str(sub.get("student_phone", "") or ""),
                        subm_at,
                        f"{_status_icon(status)} {status}",
                        eval_m,
                        max_marks,
                        feedback,
                    ]
                    aligns = ["center","left","left","center","center",
                              "left","center","center","left"]

                    for col, (v, al) in enumerate(zip(vals, aligns), 1):
                        # Highlight submission columns for new/updated submissions
                        if sub_col_bg and col in SUB_COLS:
                            c = style_data_cell(ws, row_num, col, v, bg=sub_col_bg,
                                                h_align=al, wrap=(col == 9))
                            c.font = _font(bold=True, size=10, color=sub_col_txt)
                        else:
                            style_data_cell(ws, row_num, col, v, bg=bg,
                                            h_align=al, wrap=(col == 9))
                    ws.row_dimensions[row_num].height = 18
                    row_num += 1

                _border_thick_outer(ws, first_data_row - 1, 1, row_num - 1, N)


    ws.freeze_panes     = "A2"
    ws.sheet_properties.tabColor = "2E7D32"
    auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 4 — NOT ASSIGNED CLASSES
# ─────────────────────────────────────────────────────────────────────────────

_NOT_ASSIGNED_EXCLUDE_KEYWORDS = [
    "interview", "ci/cd", "domain", "real time",
    "chatgpt", "business communication", "journey",
    "consultations", "mastery",
]


def _filter_not_assigned(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove rows whose class_name (trimmed, case-insensitive) contains any of
    the exclude keywords.  Returns filtered copy.
    """
    if df.empty:
        return df
    def _should_keep(name):
        lower = str(name).strip().lower()
        return not any(kw in lower for kw in _NOT_ASSIGNED_EXCLUDE_KEYWORDS)
    mask = df["class_name"].apply(_should_keep)
    return df[mask].reset_index(drop=True)


def _load_assignment_config(service) -> list:
    """
    Read the Assignment Upload config Google Sheet and return a list of dicts:
        [{technology, assessment_title, session_start, session_end}, ...]
    session_start/session_end parsed from 'Upload Duration In Sessions From
    Module Start Date' column (format: '10-15 Sessions').
    """
    try:
        config_df = read_sheet_df(service, ASSIGNMENT_CONFIG_SHEET_ID,
                                  "Assignment Upload")
    except Exception as e:
        print(f"[Config] Warning — could not read assignment config: {e}")
        return []

    if config_df.empty:
        return []

    configs = []
    for _, row in config_df.iterrows():
        tech  = str(row.get("Technology", "") or "").strip()
        title = str(row.get("Assessment Title", "") or "").strip()
        dur   = str(row.get("Upload Duration In Sessions From Module Start Date",
                            "") or "").strip()
        if not tech or not title:
            continue

        # Parse "10-15 Sessions" → (10, 15)
        s_start, s_end = 0, 0
        dur_clean = dur.lower().replace("sessions", "").strip()
        if "-" in dur_clean:
            parts = dur_clean.split("-", 1)
            try:
                s_start = int(parts[0].strip())
                s_end   = int(parts[1].strip())
            except (ValueError, IndexError):
                pass
        elif dur_clean.isdigit():
            s_start = s_end = int(dur_clean)

        configs.append({
            "technology":       tech,
            "assessment_title": title,
            "session_start":    s_start,
            "session_end":      s_end,
        })
    return configs


def _match_technology(class_name: str, tech: str) -> bool:
    """Case-insensitive contains match between class name and technology."""
    return tech.strip().lower() in class_name.strip().lower()


def _assignment_status(sessions_conducted: int, s_start: int, s_end: int):
    """
    Determine assignment status and colour based on session count vs range.
    Returns (status_text, bg_colour, font_colour).
    """
    if s_start == 0 and s_end == 0:
        return ("No Config", C_GREY_LITE, C_GREY_BD)
    if sessions_conducted < s_start:
        return ("Upcoming", C_GREEN_PALE, C_GREEN_DARK)
    elif sessions_conducted <= s_end:
        return ("Due", C_AMBER_PALE, C_AMBER_DARK)
    else:
        return ("Overdue", C_RED_PALE, C_RED_DARK)


def build_not_assigned(ws, not_assigned_df: pd.DataFrame,
                       report_type: str, label: str,
                       assignment_config: list):
    title = report_title(report_type, label, "Classes Without Assignments")

    # ── Filter out excluded class names ─────────────────────────────────────
    not_assigned_df = _filter_not_assigned(not_assigned_df)

    HEADERS = [
        "#", "Class Name", "Class Duration", "Students\nEnrolled",
        "Course Start Date", "Sessions\nConducted",
        "Expected In\nSessions", "Assignment Title", "Assignment\nStatus",
    ]
    N = len(HEADERS)

    style_title_row(ws, 1, 1, N, title)

    # ── KPI strip (rows 2–3) ─────────────────────────────────────────────────
    n_not_assigned = len(not_assigned_df) if not not_assigned_df.empty else 0
    kpis = [
        ("Classes Without Assignments", n_not_assigned, "🚫", KPI_RED),
    ]
    _write_kpi_strip(ws, 2, 3, kpis, N)

    write_section_banner(ws, 4, N, "  🚫  Batches With No Assignments — Assignment Tracking", C_ORANGE)
    write_header_row(ws, 5, HEADERS)

    row_num = 6
    hdr_row = 5

    today = date.today()
    thirty_days_ago = today - timedelta(days=30)

    # Sort: Course Start Date descending (latest first)
    if not not_assigned_df.empty:
        not_assigned_df = not_assigned_df.copy()
        not_assigned_df["_sessions_num"] = pd.to_numeric(
            not_assigned_df["sessions_since_start"], errors="coerce"
        ).fillna(0).astype(int)
        not_assigned_df["_start_dt"] = not_assigned_df["course_start_date"].apply(
            _parse_ist_date
        )
        not_assigned_df = not_assigned_df.sort_values(
            ["_start_dt"], ascending=[False]
        ).reset_index(drop=True)

    if not_assigned_df.empty or n_not_assigned == 0:
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=N)
        c           = ws.cell(row=row_num, column=1)
        c.value     = "✅ All classes have at least one assignment."
        c.font      = _font(bold=True, size=11, color="2E7D32")
        c.fill      = _fill(C_GREEN_PALE)
        c.alignment = _align("center", "center")
        c.border    = _border()
        ws.row_dimensions[row_num].height = 24
        row_num += 1
    else:
        seq = 0
        for _, cls_row in not_assigned_df.iterrows():
            class_name  = str(cls_row.get("class_name", "") or "")
            class_subj  = str(cls_row.get("class_subject", "") or "")
            sessions    = int(cls_row.get("_sessions_num", 0))
            start_date_str = _fmt_date(cls_row.get("course_start_date", ""))

            _enrolled_raw = cls_row.get("students_enrolled", "") or "0"
            try:
                enrolled_val = int(float(str(_enrolled_raw)))
            except (ValueError, TypeError):
                enrolled_val = 0

            # Find all matching config entries for this class
            matched_configs = [
                cfg for cfg in assignment_config
                if _match_technology(class_name, cfg["technology"])
            ]

            # Check if class started in last 30 days
            cls_start_dt = cls_row.get("_start_dt", None)
            is_recent = (cls_start_dt is not None and not pd.isna(cls_start_dt)
                         and cls_start_dt >= thirty_days_ago)

            # Parse ClassLastDate from "Class Duration" to check if class is over
            cls_end_dt = _parse_class_end_date(class_subj)
            class_is_over = (cls_end_dt is not None and cls_end_dt < CLASS_CUTOFF_DATE)

            if not matched_configs:
                # No config match — show single row with "No Config"
                seq += 1
                bg = C_GREY_LITE
                vals = [
                    seq, class_name, class_subj, enrolled_val,
                    start_date_str, sessions, "—", "—", "No Config",
                ]
                aligns = ["center","left","left","center",
                          "center","center","center","left","center"]
                for col, (v, al) in enumerate(zip(vals, aligns), 1):
                    c = style_data_cell(ws, row_num, col, v, bg=bg, h_align=al,
                                        wrap=(col in (2, 8)))
                    c.font = _font(size=10, color=C_GREY_BD,
                                   bold=is_recent, italic=is_recent)
                ws.row_dimensions[row_num].height = 18
                row_num += 1
            else:
                # One row per expected assignment
                first_row_of_class = row_num
                for ci, cfg in enumerate(matched_configs):
                    seq += 1

                    # If class is over (end date < today):
                    #   sessions < session_start → "Not Applicable" (class ended too early)
                    #   sessions >= session_start → "Not Assigned"  (should have been uploaded)
                    if class_is_over:
                        if sessions < cfg["session_start"]:
                            status_bg = C_GREY_LITE
                            status_fg = C_GREY_BD
                            status_display = f"Not Applicable ({sessions}/{cfg['session_start']})"
                        else:
                            status_bg = C_PURPLE_PALE
                            status_fg = C_PURPLE_DARK
                            status_display = f"Not Assigned ({sessions}/{cfg['session_end']})"
                    else:
                        status_text, status_bg, status_fg = _assignment_status(
                            sessions, cfg["session_start"], cfg["session_end"]
                        )
                        # Append session count for clarity: "Overdue (32/15)"
                        s_end = cfg["session_end"]
                        status_display = f"{status_text} ({sessions}/{s_end})"

                    # Session range display: "10-15"
                    session_range = (f"{cfg['session_start']}-{cfg['session_end']}"
                                     if cfg["session_start"] or cfg["session_end"]
                                     else "—")

                    # Use status colour for entire row
                    bg = status_bg

                    vals = [
                        seq,
                        class_name if ci == 0 else "",
                        class_subj if ci == 0 else "",
                        enrolled_val if ci == 0 else "",
                        start_date_str if ci == 0 else "",
                        sessions if ci == 0 else "",
                        session_range,
                        cfg["assessment_title"],
                        status_display,
                    ]
                    aligns = ["center","left","left","center",
                              "center","center","center","left","center"]
                    for col, (v, al) in enumerate(zip(vals, aligns), 1):
                        c = style_data_cell(ws, row_num, col, v, bg=bg, h_align=al,
                                            wrap=(col in (2, 8)))
                        c.font = _font(size=10, color=status_fg,
                                        bold=(is_recent or col == 9),
                                        italic=is_recent)
                    ws.row_dimensions[row_num].height = 18
                    row_num += 1

                # Merge class-level cells (cols 2-6) if multiple config rows
                if len(matched_configs) > 1:
                    for merge_col in [2, 3, 4, 5, 6]:
                        ws.merge_cells(
                            start_row=first_row_of_class, start_column=merge_col,
                            end_row=row_num - 1, end_column=merge_col,
                        )
                        mc = ws.cell(row=first_row_of_class, column=merge_col)
                        mc.alignment = _align(
                            "left" if merge_col in (2, 3) else "center",
                            "center",
                        )

    if row_num > hdr_row + 1:
        _border_thick_outer(ws, hdr_row, 1, row_num - 1, N)

    ws.auto_filter.ref  = f"A{hdr_row}:{get_column_letter(N)}{row_num - 1}"
    ws.freeze_panes     = "A6"
    ws.sheet_properties.tabColor = "E65100"
    auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
#  GOOGLE DRIVE UPLOAD
# ─────────────────────────────────────────────────────────────────────────────

def upload_to_gdrive(report_type: str, filename: str, file_buffer: io.BytesIO):
    folder_id = GDRIVE_FOLDER_MAP.get(report_type.lower(), "")
    if not folder_id:
        print(f"[Drive] ⚠ No folder ID configured for '{report_type}' — skipping upload.")
        return
    try:
        from googleapiclient.discovery import build as gdrive_build
        from googleapiclient.http      import MediaIoBaseUpload
        from google.oauth2             import service_account as _sa

        creds = _sa.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE,
            scopes=["https://www.googleapis.com/auth/drive"],
        )
        creds = creds.with_subject("info@intellibiinnovationstechnologies.in")
        drive = gdrive_build("drive", "v3", credentials=creds)

        # Remove existing file with same name
        for f in drive.files().list(
            q=f"'{folder_id}' in parents and name='{filename}' and trashed=false",
            fields="files(id)",
        ).execute().get("files", []):
            drive.files().delete(fileId=f["id"]).execute()

        file_buffer.seek(0)
        media = MediaIoBaseUpload(
            file_buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        up = drive.files().create(
            body={"name": filename, "parents": [folder_id]},
            media_body=media,
            fields="id,webViewLink",
        ).execute()
        print(f"[Drive] ✓ Uploaded: {filename}")
        print(f"[Drive]   Link: {up.get('webViewLink', '')}")
    except Exception as e:
        print(f"[Drive] ⚠ Upload failed: {e}")


# ─────────────────────────────────────────────────────────────────────────────
#  EMAIL
# ─────────────────────────────────────────────────────────────────────────────

def send_report(report_type: str, label: str, filename: str, file_bytes: bytes,
                n_assignments: int, n_classes: int,
                overall_rate: float, n_not_assigned: int):
    subject = (f"IntelliBI {report_type.title()} Assignment Submission Report — {label}")
    body = f"""\
<html><body style="font-family:Arial,sans-serif;color:#1F3864;">
<h2 style="color:#1F3864;">📋 IntelliBI {report_type.title()} Assignment Submission Report</h2>
<p>Please find the attached report for <strong>{label}</strong>.</p>
<table style="border-collapse:collapse;font-size:14px;">
  <tr>
    <td style="padding:8px 18px;background:#D6E4F0;font-weight:bold;border:1px solid #B0B0B0;">Total Assignments</td>
    <td style="padding:8px 18px;border:1px solid #B0B0B0;">{n_assignments}</td>
  </tr><tr>
    <td style="padding:8px 18px;background:#D6E4F0;font-weight:bold;border:1px solid #B0B0B0;">Classes with Assignments</td>
    <td style="padding:8px 18px;border:1px solid #B0B0B0;">{n_classes}</td>
  </tr><tr>
    <td style="padding:8px 18px;background:#E8F5E9;font-weight:bold;border:1px solid #B0B0B0;">📊 Overall Submission Rate</td>
    <td style="padding:8px 18px;border:1px solid #B0B0B0;"><strong>{overall_rate:.1f}%</strong></td>
  </tr><tr>
    <td style="padding:8px 18px;background:#FFE0E0;font-weight:bold;border:1px solid #B0B0B0;">🚫 Classes Without Assignments</td>
    <td style="padding:8px 18px;border:1px solid #B0B0B0;"><strong>{n_not_assigned}</strong></td>
  </tr>
</table>
<br>
<p style="font-size:12px;color:#888;">This is an automated report from IntelliBI Innovations Technologies.</p>
</body></html>"""

    to_list = REPORT_TO if isinstance(REPORT_TO, list) else [REPORT_TO]
    msg             = MIMEMultipart()
    msg["From"]     = GMAIL_SENDER
    msg["To"]       = ", ".join(to_list)
    msg["Subject"]  = subject
    if REPORT_CC:
        msg["Cc"]   = ", ".join(REPORT_CC)
    msg.attach(MIMEText(body, "html"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(file_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    recipients = to_list + REPORT_CC
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_SENDER, GMAIL_APP_PASS)
        server.sendmail(GMAIL_SENDER, recipients, msg.as_string())
    print(f"[Email] ✓ Report sent to {', '.join(to_list)}")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="IntelliBI Assignment Submissions Report")
    parser.add_argument("--type",    default="auto",
                        choices=["auto", "manual"],
                        help="Report mode: 'auto' (trigger-based, 1 year) or 'manual' (custom date range)")
    parser.add_argument("--start",   default=None,
                        help="Manual period start (DD-Mon-YYYY or DD-MM-YYYY)")
    parser.add_argument("--end",     default=None,
                        help="Manual period end   (DD-Mon-YYYY or DD-MM-YYYY)")
    parser.add_argument("--no-email", action="store_true",
                        help="Skip sending email (still generates the file)")
    args = parser.parse_args()

    # ── Trigger check ──────────────────────────────────────────────────────────
    #  Auto mode : trigger file required (exit if missing)
    #  Manual mode: trigger file optional (highlight changes if available)
    trigger_info = None
    if args.type == "auto" and not os.path.isfile(TRIGGER_FILE):
        print("[Trigger] No trigger file found — no assignment changes detected.")
        print("[Trigger] Report generation skipped. Use --type manual to force.")
        return

    if os.path.isfile(TRIGGER_FILE):
        try:
            with open(TRIGGER_FILE, "r", encoding="utf-8") as f:
                trigger_info = json.load(f)
            print(f"[Trigger] \u2713 Trigger file found:")
            print(f"  Triggered at : {trigger_info.get('triggered_at', '?')}")
            print(f"  New assignments    : {trigger_info.get('new_count', 0)}")
            print(f"  Changed assignments: {trigger_info.get('changed_count', 0)}")
            n_ns = len(trigger_info.get("new_submission_aids", []))
            n_us = len(trigger_info.get("updated_submission_aids", []))
            if n_ns or n_us:
                print(f"  New submissions    : {n_ns}")
                print(f"  Updated submissions: {n_us}")
        except Exception as e:
            print(f"[Trigger] Warning \u2014 could not read trigger file: {e}")
            trigger_info = {}
    else:
        print("[Trigger] No trigger file found — change highlighting disabled.")

    # ── Resolve date range ──────────────────────────────────────────────────────
    if args.type == "manual":
        if not args.start or not args.end:
            print("[Error] --start and --end are required for --type manual")
            sys.exit(1)
        def _parse_arg_date(s):
            for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try: return datetime.strptime(s.strip(), fmt).date()
                except ValueError: pass
            raise ValueError(f"Cannot parse date: {s!r}")
        start_date = _parse_arg_date(args.start)
        end_date   = _parse_arg_date(args.end)
        label      = f"{start_date.strftime('%d %b %Y')} – {end_date.strftime('%d %b %Y')}"
    else:
        # Auto mode: 1 year from current date
        start_date, end_date, label = get_auto_date_range(date.today())

    sep = "=" * 64
    print(f"\n{sep}")
    print(f"  IntelliBI Assignment Submissions Report")
    print(f"  Type    : {args.type.title()}")
    print(f"  Period  : {label}  ({start_date} → {end_date})")
    print(f"{sep}\n")

    # ── Auth & data ─────────────────────────────────────────────────────────
    from utils import get_sheets_service
    service = get_sheets_service(SERVICE_ACCOUNT_FILE)

    assigned_f, subs_f, not_assigned_df = load_all_data(service, start_date, end_date)

    # ── Sort assigned by Start Date descending (latest first) ──────────────────
    if not assigned_f.empty and "_start_date" in assigned_f.columns:
        assigned_f = assigned_f.sort_values("_start_date", ascending=False).reset_index(drop=True)
        print("[Data] Sorted assignments by Start Date (descending — latest first)")

    # ── Build change_info from trigger data ────────────────────────────────
    change_info = {
        "new": set(),              # new assignment aids
        "changed": {},             # TRUE assignment-level changes only (field list)
        "new_sub": set(),          # aids with new submissions
        "updated_sub": set(),      # aids with updated submissions
    }
    if trigger_info:
        for a in trigger_info.get("new_assignments", []):
            aid = a.get("assessment_id", "")
            if aid:
                change_info["new"].add(aid)
        for a in trigger_info.get("changed_assignments", []):
            aid = a.get("assessment_id", "")
            fields = a.get("changed_fields", [])
            if not aid:
                continue
            # Separate: submission-only changes go to updated_sub, not changed
            if fields == ["submissions_updated"]:
                change_info["updated_sub"].add(aid)
            else:
                change_info["changed"][aid] = fields
        for aid in trigger_info.get("new_submission_aids", []):
            if aid:
                change_info["new_sub"].add(aid)
        for aid in trigger_info.get("updated_submission_aids", []):
            if aid:
                change_info["updated_sub"].add(aid)
        n_new = len(change_info["new"])
        n_chg = len(change_info["changed"])
        n_ns  = len(change_info["new_sub"])
        n_us  = len(change_info["updated_sub"])
        if n_new or n_chg or n_ns or n_us:
            print(f"[Highlight] Assignments: {n_new} new + {n_chg} changed | "
                  f"Submissions: {n_ns} new + {n_us} updated")
        else:
            print("[Highlight] No changes to highlight")

    # ── Build workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)    # remove default empty sheet

    print("[Report] Building Assignment_Summary …")
    ws1 = wb.create_sheet("Assignment_Summary")
    build_assignment_summary(ws1, assigned_f, subs_f, args.type, label, change_info)

    print("[Report] Building Class_Breakdown …")
    ws2 = wb.create_sheet("Class_Breakdown")
    build_class_breakdown(ws2, assigned_f, subs_f, args.type, label, change_info)

    print("[Report] Building Student_Detail …")
    ws3 = wb.create_sheet("Student_Detail")
    build_student_detail(ws3, assigned_f, subs_f, args.type, label, change_info)

    print("[Config] Loading assignment config …")
    assignment_config = _load_assignment_config(service)
    print(f"[Config] {len(assignment_config)} assignment entries loaded")

    print("[Report] Building Not_Assigned …")
    ws4 = wb.create_sheet("Not_Assigned")
    build_not_assigned(ws4, not_assigned_df, args.type, label, assignment_config)

    # ── Filename ──────────────────────────────────────────────────────────
    safe_label  = label.replace(" ", "_").replace("–", "-").replace("/", "-")
    type_prefix = {"auto": "Auto_", "manual": "Custom_"}
    filename    = f"AssignmentReport_{type_prefix.get(args.type, '')}{safe_label}.xlsx"

    # ── Save to in-memory buffer ──────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.getvalue()

    # ── Summary stats for email ─────────────────────────────────────────────
    n_assignments  = len(assigned_f) if not assigned_f.empty else 0
    n_classes      = assigned_f["class_id"].nunique() if not assigned_f.empty else 0
    n_not_assigned = len(not_assigned_df) if not not_assigned_df.empty else 0

    total_enrolled = total_submitted = 0
    if not subs_f.empty:
        total_enrolled  = len(subs_f)
        total_submitted = int((subs_f["submission_status"] != "Not Submitted").sum())
    overall_rate = round(total_submitted / total_enrolled * 100, 1) if total_enrolled else 0.0

    print(f"\n{sep}")
    print(f"  Report complete.")
    print(f"  Assignments     : {n_assignments}")
    print(f"  Classes covered : {n_classes}")
    print(f"  Overall sub rate: {overall_rate:.1f}%")
    print(f"  Not-assigned    : {n_not_assigned}")
    print(f"  File            : {filename}")
    print(f"{sep}\n")

    # ── Drive upload ────────────────────────────────────────────────────────
    upload_to_gdrive(args.type, filename, io.BytesIO(file_bytes))

    # ── Email ───────────────────────────────────────────────────────────
    if not args.no_email:
        try:
            send_report(args.type, label, filename, file_bytes,
                        n_assignments, n_classes, overall_rate, n_not_assigned)
        except Exception as e:
            print(f"[Email] ⚠ Failed to send: {e}")
    else:
        print("[Email] Skipped (--no-email flag set).")

    # ── Local saving DISABLED — report is delivered to Google Drive only ────────
    #    (Previously wrote a local .xlsx copy here for inspection.)
    print("[File] Local copy skipped — report delivered to Google Drive only.")

    # ── Clean up trigger file (auto mode) ──────────────────────────────────────
    if args.type == "auto" and os.path.isfile(TRIGGER_FILE):
        try:
            os.remove(TRIGGER_FILE)
            print("[Trigger] \u2713 Trigger file deleted after successful report generation.")
        except Exception as e:
            print(f"[Trigger] \u26a0 Could not delete trigger file: {e}")


if __name__ == "__main__":
    main()
