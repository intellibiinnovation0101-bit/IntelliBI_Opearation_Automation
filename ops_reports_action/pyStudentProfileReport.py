"""
================================================================================
  IntelliBI — Unified Student Profile Report Generator
  ------------------------------------------------------------------------------
  Produces ONE individual PDF per student that combines, end-to-end:

      1. Attendance performance        (from the Attendance pipeline sheets)
      2. Assignment submission status  (from the Assignment Submissions sheets)
      3. Interview feedback & scores   (from the Consolidated Interview sheet)

  It also produces a single MASTER Student Progress Report (Excel) consolidating
  every student for the CEO/admin, emailed only to the admin address.

  The three existing per-batch / per-cohort reports are the data + styling
  reference:
      Reports/pyAttendaceFeedbackReport.py     (attendance + feedback)
      Reports/pyAssignmentSubmissionsReport.py (assignment submissions)
      Reports/pyInterviewFeedbackReport.py     (interview feedback)
      Reports/pyInterviewConsolidatedReport.py (interview aggregation)

  PERIOD MODEL  ("Cumulative + month highlight")
  ----------------------------------------------
  Each PDF shows the student's full, all-time history AND a "This Month"
  highlight strip for the most-recent completed calendar month.

  OUTPUT FOLDER
  -------------
  Individual PDFs + the master workbook are written to a month-wise folder in
  the project root, e.g.  StudentProgressReport_May_2026/.

  FILE NAMING (per student)
  -------------------------
      FirstNameLastName_BatchName_EnrollmentDate.pdf
  e.g.  AnamikaThakur_ADA_14-Sep-2025.pdf
  (BatchName = the `batch_name` tag from the Students sheet; EnrollmentDate =
   `joined_on`.)

  DELIVERY
  --------
  - Each student PDF is emailed to the student's OWN registered email.
  - The master workbook is emailed ONLY to the admin (info@...).
  Use --admin-only for a safe first run (all student PDFs to admin), and
  --no-email to only write files to disk.

  SCHEDULING
  ----------
  Designed to run monthly on the 1st.

  USAGE
  -----
      python Reports/pyStudentProfileReport.py                 # live: students + master
      python Reports/pyStudentProfileReport.py --no-email      # build files only
      python Reports/pyStudentProfileReport.py --admin-only    # review run (all to admin)
      python Reports/pyStudentProfileReport.py --only-email someone@x.com
      python Reports/pyStudentProfileReport.py --no-email --limit 5
      python Reports/pyStudentProfileReport.py --no-master     # skip master workbook
      python Reports/pyStudentProfileReport.py --sample-xlsx "/path/sample.xlsx" --no-email
================================================================================
"""
from __future__ import annotations

# --- IntelliBI Operations Automation portability bootstrap (auto-inserted) ---
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "common"))
import _bootstrap  # noqa: E402  (sys.path + env defaults + config.yaml)
from paths import CREDENTIALS_DIR, CONFIG_DIR, LOGS_DIR, CACHE_DIR as PROJECT_CACHE_DIR  # noqa: E402
# --- end bootstrap ---

import argparse
import io
import json
import os
import re
import smtplib
import sys
import threading
import traceback
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta

# ── Resolve parent dir so utils.py + service_account.json are always found ────
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR   = os.path.dirname(_SCRIPT_DIR)            # IntelliBI Automation/
for _p in (_BASE_DIR, _SCRIPT_DIR):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from email.mime.multipart import MIMEMultipart
from email.mime.base      import MIMEBase
from email.mime.text      import MIMEText
from email                import encoders

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG — sheet IDs taken verbatim from the three reference reports
# ─────────────────────────────────────────────────────────────────────────────
SERVICE_ACCOUNT_FILE = os.path.join(CREDENTIALS_DIR, "service_account.json")

# Students roster (name / email / phone / batch_name tag / joined_on)
STUDENTS_SHEET_ID    = "1Eq7Q3Gota7nYiaorm1L0NoouVfYtS7JkbBp4U5MWzVA"
STUDENTS_TAB         = "Students"

# Attendance pipeline
ATTENDANCE_SHEET_ID  = "1TqDjq4gAyo32eRNMbuLd6uu0eCNZb7h1j5YH-q68AhU"
ATT_TAB              = "Attendance"
SESSIONS_TAB         = "Sessions"
STUDENT_FEEDBACK_TAB = "Student_Feedback"

# Assignment submissions pipeline
SUBMISSION_SHEET_ID  = "1E_pOuZfw4BUhQ1bRMmuc8lPDDJRtoXuGkP-qtHC96HU"
ASSIGNED_TAB         = "Assessment_Assigned"
SUBMISSIONS_TAB      = "Submissions"

# Consolidated interview report (header on row 2, data from row 3)
INTERVIEW_SHEET_ID   = "16IQtgrlvYZpEpsmtzWhyaS9ZgRHYB_jzQBF--DckCPg"
INTERVIEW_HEADER_ROW = 2
INTERVIEW_DATA_START = 3

# ── E-mail (Gmail SMTP — same sender used by the other IntelliBI reports) ─────
# E-mail credentials are centralised — edit config_files/email_config.py to rotate.
from email_config import (GMAIL_SENDER_DIGITAL as GMAIL_SENDER,
                                       GMAIL_APP_PASS_DIGITAL as GMAIL_APP_PASS)
ADMIN_EMAIL    = "info@intellibiinnovationstechnologies.in"

# Google Drive (uploads use the service account impersonating the info@ user,
# matching the other IntelliBI reports). Reports go into a month-wise subfolder
# created under this root folder.
DRIVE_DELEGATED_USER = "info@intellibiinnovationstechnologies.in"
DRIVE_ROOT_FOLDER_ID = "12hCSpu7uUYNDQeTIqr8IybKUjHOP6EWo"

LOCAL_PDF_DIR  = os.path.join(_SCRIPT_DIR, "_pdf_reports")

# ─────────────────────────────────────────────────────────────────────────────
#  REPORT GENERATION CONFIGURATION
#  The script is designed to run DAILY. What gets generated on a given run:
#    • Till-Date report  → ALWAYS (every day), overwriting the previous copy.
#    • Monthly report    → only on the last calendar day of the month
#                          (automatic), OR when MANUAL_MONTHLY_REPORT is True.
# ─────────────────────────────────────────────────────────────────────────────
# Manual monthly generation. When True, monthly reports are generated for
# REPORT_MONTH / REPORT_YEAR regardless of today's date. When False, the monthly
# report follows the automatic last-day-of-month rule.
MANUAL_MONTHLY_REPORT = False
REPORT_MONTH = 7          # 1–12, used only when MANUAL_MONTHLY_REPORT is True
REPORT_YEAR  = 2026       # e.g. 2026, used only when MANUAL_MONTHLY_REPORT is True

# Which students to generate reports for:
#   "ALL"         -> every student in the roster.
#   "ACTIVE_ONLY" -> only students whose active flag (is_active_student /
#                    is_candidate_active) reads truthy. Comparison is
#                    case-insensitive and blank / null values are treated as
#                    NOT active (so they are excluded under ACTIVE_ONLY).
REPORT_GENERATION_SCOPE = "ACTIVE_ONLY"

# Email each student their MONTHLY progress report.
#   True  -> send the monthly report to the student's registered email, but ONLY
#            on the last calendar day of the month. Report generation/upload is
#            never affected; an email failure is logged and processing continues.
#   False -> generate and upload reports but send no student emails.
SEND_STUDENT_REPORT_EMAIL = True

# Share the latest Till-Date report with each student (Google Drive, Viewer only).
#   True  -> after the Till-Date PDF is (over)written in the student's Drive folder,
#            share that file with the student's registered email as VIEWER. The file
#            is updated IN PLACE so the same link keeps working day-to-day and the
#            existing permission is retained (no daily re-share). Missing / invalid
#            emails and sharing errors are logged and skipped without interrupting
#            the other students.
#   False -> generate/overwrite the Till-Date report exactly as before; never share.
SHARE_TILL_DATE_REPORT_WITH_STUDENT = True

# ─────────────────────────────────────────────────────────────────────────────
#  BRAND PALETTE  (matches pyInterviewFeedbackReport / the sample workbook)
# ─────────────────────────────────────────────────────────────────────────────
BRAND_DARK_BLUE = "#1F3864"
BRAND_MID_BLUE  = "#2E5496"
BRAND_BANNER    = "#D6E4F0"
BRAND_SUBTLE    = "#F5F7FA"
BRAND_GROUP     = "#BDD7EE"
BRAND_GREY      = "#555555"

GREEN_BG, GREEN_FG = "#E8F5E9", "#1B5E20"
AMBER_BG, AMBER_FG = "#FFF3E0", "#E65100"
RED_BG,   RED_FG   = "#FFEBEE", "#B71C1C"
GREY_BG,  GREY_FG  = "#ECEFF1", "#455A64"

LOGO_PATH = os.path.join(_BASE_DIR, "intellibi_logo.png")


# ═════════════════════════════════════════════════════════════════════════════
#  DATA CONTRACT
# ═════════════════════════════════════════════════════════════════════════════
@dataclass
class AttendanceCourse:
    course:     str = "—"
    period:     str = "—"
    sessions:   int = 0
    present:    int = 0
    absent:     int = 0
    att_pct:    float | None = None
    avg_time:   float | None = None
    ratings:    int | None = None
    avg_rating: float | None = None


@dataclass
class AssignmentItem:
    seq:        str = ""
    title:      str = "—"
    class_name: str = "—"
    start:      str = "—"
    deadline:   str = "—"
    max_marks:  str = ""
    submitted_at: str = ""
    status:     str = "Not Submitted"
    marks:      str = ""
    feedback:   str = ""


@dataclass
class InterviewItem:
    no:        str = ""
    when:      str = "—"
    comm:      str = ""
    tech:      str = ""
    total:     str = ""
    avg:       str = ""
    comments:  str = ""
    zone:      str = "—"


@dataclass
class StudentProfile:
    name:            str = "—"
    email:           str = ""
    phone:           str = ""
    batch_name:      str = ""
    enrollment_date: str = ""
    student_id:      str = ""
    active_flag:     str = ""   # raw is_active_student / is_candidate_active value

    att_total_sessions: int = 0
    att_attended:       int = 0
    att_avg_pct:        float | None = None
    att_avg_dur_pct:    float | None = None
    att_courses:        list = field(default_factory=list)

    asg_total:      int = 0
    asg_submitted:  int = 0
    asg_pct:        float | None = None
    asg_items:      list = field(default_factory=list)

    itv_total:      int = 0
    itv_given:      int = 0
    itv_att_pct:    float | None = None
    itv_avg_score:  float | None = None
    itv_items:      list = field(default_factory=list)

    month_label:    str = ""
    month_sessions: int = 0
    month_present:  int = 0
    month_asg_sub:  int = 0
    month_itv:      int = 0


# ═════════════════════════════════════════════════════════════════════════════
#  SMALL HELPERS
# ═════════════════════════════════════════════════════════════════════════════
def sanitize_part(s: str) -> str:
    s = re.sub(r"[^0-9A-Za-z]+", "", str(s or "").strip())
    return s or "NA"


def to_float(v, default=None):
    try:
        s = str(v).strip().replace("%", "")
        if s in ("", "-", "—", "nan", "None", "NaT"):
            return default
        if "/" in s:
            s = s.split("/")[0].strip()
        return float(s)
    except Exception:
        return default


def to_int(v, default=0):
    f = to_float(v, None)
    return int(round(f)) if f is not None else default


# NOTE: IntelliBI assignment sheets use US MM/DD/YYYY, so month-first slash
# parsing is tried before day-first. ISO and DD-Mon-YYYY forms are unaffected.
_DATE_FORMATS = (
    "%Y-%m-%d", "%d-%b-%Y", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y",
    "%Y-%m-%d %H:%M:%S", "%d-%b-%Y %H:%M", "%d %b %Y", "%b %d, %Y",
    "%Y/%m/%d", "%d-%b-%y", "%Y-%m-%dT%H:%M:%S",
)


def parse_date(s):
    s = str(s or "").strip()
    if not s:
        return None
    s = s.split("T")[0].strip() if "T" in s else s
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s.split(" ")[0] if fmt in ("%Y-%m-%d", "%d-%b-%Y") else s, fmt).date()
        except Exception:
            continue
    try:
        import pandas as pd
        d = pd.to_datetime(s, errors="coerce", dayfirst=True)
        return None if pd.isna(d) else d.date()
    except Exception:
        return None


def fmt_date(s) -> str:
    d = parse_date(s)
    return d.strftime("%d-%b-%Y") if d else (str(s or "").strip() or "—")


def first_last(full_name: str) -> str:
    parts = [p for p in re.split(r"\s+", str(full_name or "").strip()) if p]
    if not parts:
        return "Student"
    if len(parts) == 1:
        return sanitize_part(parts[0])
    return sanitize_part(parts[0]) + sanitize_part(parts[-1])


def prev_month_label(today: date | None = None):
    today = today or date.today()
    y, m = today.year, today.month
    m -= 1
    if m == 0:
        m, y = 12, y - 1
    return date(y, m, 1).strftime("%b %Y"), y, m


def build_filename(p: StudentProfile) -> str:
    name = first_last(p.name)
    batch = sanitize_part(p.batch_name or "NoBatch")
    enr   = fmt_date(p.enrollment_date) if p.enrollment_date else "NoDate"
    enr   = sanitize_part(enr) if enr == "NoDate" else enr
    return f"{name}_{batch}_{enr}.pdf"


def _name_first_last(full_name: str) -> str:
    """'Pratik Patil' -> 'Pratik_Patil' (First_Last, sanitized)."""
    parts = [sanitize_part(x) for x in re.split(r"\s+", str(full_name or "").strip()) if x]
    if not parts:
        return "Student"
    if len(parts) == 1:
        return parts[0]
    return f"{parts[0]}_{parts[-1]}"


def student_folder_name(p: StudentProfile) -> str:
    """Per-student folder / file prefix: FirstName_LastName_BatchName."""
    return f"{_name_first_last(p.name)}_{sanitize_part(p.batch_name or 'NoBatch')}"


def monthly_report_filename(p: StudentProfile, month_long: str) -> str:
    """e.g. Pratik_Patil_ADA1224_Progress_Report_July_2026.pdf"""
    return f"{student_folder_name(p)}_Progress_Report_{month_long}.pdf"


def till_date_report_filename(p: StudentProfile) -> str:
    """Fixed name (NO timestamp) so the daily run overwrites the same file, e.g.
    Pratik_Patil_ADA1224_Progress_Report_Till_Date.pdf"""
    return f"{student_folder_name(p)}_Progress_Report_Till_Date.pdf"


def _is_active_flag(v) -> bool:
    """Case-insensitive, null-safe truthiness for the student active flag.
    Blank / null / unrecognised values are treated as NOT active."""
    return str(v or "").strip().lower() in ("y", "yes", "true", "1", "active", "a")


def _is_last_day_of_month(d: date | None = None) -> bool:
    """True when `d` (default today) is the last calendar day of its month."""
    d = d or date.today()
    nxt = d.replace(day=28) + timedelta(days=4)      # always lands in next month
    last = (nxt - timedelta(days=nxt.day)).day        # last day of d's month
    return d.day == last


def zone_colors(zone: str):
    z = str(zone or "").lower()
    if "strong" in z: return GREEN_BG, GREEN_FG
    if "good"   in z: return AMBER_BG, AMBER_FG
    if "weak"   in z or "poor" in z: return RED_BG, RED_FG
    return GREY_BG, GREY_FG


def att_pct_colors(pct):
    v = to_float(pct, 0) or 0
    if v >= 75: return GREEN_BG, GREEN_FG
    if v >= 50: return AMBER_BG, AMBER_FG
    return RED_BG, RED_FG


def status_colors(status: str):
    s = str(status or "").lower()
    if "graded" in s:    return GREEN_BG, GREEN_FG
    if "submit" in s and "not" not in s: return BRAND_BANNER, BRAND_DARK_BLUE
    return RED_BG, RED_FG


# ═════════════════════════════════════════════════════════════════════════════
#  PER-STUDENT PDF BUILDER  (source-agnostic)
# ═════════════════════════════════════════════════════════════════════════════
def build_student_pdf(p: StudentProfile, variant: str = "monthly"):
    # variant "monthly"  -> the existing month-wise report (unchanged format).
    # variant "till_date"-> consolidated report covering the student's full history
    #                       up to the execution timestamp (same layout; only the
    #                       header line is labelled differently).
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units     import mm
    from reportlab.lib           import colors
    from reportlab.platypus      import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, Image)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=12 * mm, bottomMargin=14 * mm,
        title=f"IntelliBI Student Profile Report — {p.name}",
        author="IntelliBI Innovations Technologies",
    )
    usable_w = doc.width

    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("H1", parent=styles["Title"], fontSize=17,
                        textColor=colors.HexColor(BRAND_DARK_BLUE),
                        spaceAfter=2, alignment=0)
    SUB = ParagraphStyle("SUB", parent=styles["Normal"], fontSize=9,
                         textColor=colors.HexColor(BRAND_GREY), spaceAfter=2)
    SEC = ParagraphStyle("SEC", parent=styles["Heading2"], fontSize=12.5,
                         textColor=colors.white, spaceBefore=2, spaceAfter=2,
                         leading=16)
    CELL = ParagraphStyle("CELL", parent=styles["Normal"], fontSize=8, leading=10)
    CELLC = ParagraphStyle("CELLC", parent=CELL, alignment=1)
    HEAD = ParagraphStyle("HEAD", parent=styles["Normal"], fontSize=8, leading=10,
                          textColor=colors.white, fontName="Helvetica-Bold",
                          alignment=1)
    NOTE = ParagraphStyle("NOTE", parent=styles["Normal"], fontSize=7.5,
                          textColor=colors.HexColor("#888888"))

    def section_banner(text: str):
        t = Table([[Paragraph(text, ParagraphStyle("sb", parent=SEC))]],
                  colWidths=[usable_w])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND_DARK_BLUE)),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    def kpi_strip(cards):
        n = len(cards)
        cw = usable_w / n
        top = [Paragraph(f'<font color="{fg}"><b>{lbl}</b></font>', CELLC)
               for (lbl, _v, _bg, fg) in cards]
        bot = [Paragraph(f'<font color="{fg}" size=13><b>{val}</b></font>', CELLC)
               for (_lbl, val, _bg, fg) in cards]
        t = Table([top, bot], colWidths=[cw] * n)
        cmds = [
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        for i, (_l, _v, bg, _fg) in enumerate(cards):
            cmds.append(("BACKGROUND", (i, 0), (i, -1), colors.HexColor(bg)))
        t.setStyle(TableStyle(cmds))
        return t

    story = []

    title_block = [
        Paragraph("IntelliBI Student Profile Report", H1),
        Paragraph("Comprehensive monthly performance summary &nbsp;·&nbsp; "
                  "Attendance &nbsp;|&nbsp; Assignments &nbsp;|&nbsp; Interviews",
                  SUB),
        Paragraph(
            (f"Progress Report (Till Date) &nbsp;·&nbsp; as of "
             f"{datetime.now().strftime('%d-%b-%Y %H:%M:%S')}"
             if variant == "till_date" else
             f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
             + (f" &nbsp;·&nbsp; Highlight month: <b>{p.month_label}</b>"
                if p.month_label else "")),
            SUB),
    ]
    if os.path.exists(LOGO_PATH):
        try:
            logo = Image(LOGO_PATH, width=26 * mm, height=26 * mm)
            head = Table([[title_block, logo]],
                         colWidths=[usable_w - 28 * mm, 28 * mm])
            head.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ]))
            story.append(head)
        except Exception:
            story.extend(title_block)
    else:
        story.extend(title_block)
    story.append(Spacer(1, 6))

    meta = [
        ["Student Name", p.name or "—", "Batch", p.batch_name or "—"],
        ["Email", p.email or "—", "Phone", p.phone or "—"],
        ["Enrollment Date", fmt_date(p.enrollment_date) if p.enrollment_date else "—",
         "Student ID", p.student_id or "—"],
    ]
    mt = Table(meta, colWidths=[28 * mm, usable_w / 2 - 28 * mm,
                                24 * mm, usable_w / 2 - 24 * mm])
    mt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(BRAND_DARK_BLUE)),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor(BRAND_DARK_BLUE)),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.white),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("BACKGROUND", (1, 0), (1, -1), colors.HexColor(BRAND_SUBTLE)),
        ("BACKGROUND", (3, 0), (3, -1), colors.HexColor(BRAND_SUBTLE)),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B0B0B0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(mt)
    story.append(Spacer(1, 8))

    if p.month_label:
        story.append(Paragraph(
            f'<b><font color="{BRAND_MID_BLUE}">This Month — {p.month_label}</font></b>',
            ParagraphStyle("mh", parent=styles["Normal"], fontSize=10,
                           spaceAfter=3)))
        mpct = (100.0 * p.month_present / p.month_sessions) if p.month_sessions else None
        story.append(kpi_strip([
            ("Sessions This Month", str(p.month_sessions), BRAND_BANNER, BRAND_DARK_BLUE),
            ("Present This Month", str(p.month_present), GREEN_BG, GREEN_FG),
            ("Month Attendance", f"{mpct:.0f}%" if mpct is not None else "—",
             *((GREEN_BG, GREEN_FG) if (mpct or 0) >= 75 else
               (AMBER_BG, AMBER_FG) if (mpct or 0) >= 50 else (RED_BG, RED_FG))),
            ("Assignments Submitted", str(p.month_asg_sub), BRAND_BANNER, BRAND_DARK_BLUE),
            ("Interviews", str(p.month_itv), BRAND_BANNER, BRAND_DARK_BLUE),
        ]))
        story.append(Spacer(1, 10))

    # ── SECTION 1 — ATTENDANCE ───────────────────────────────────────────────
    story.append(section_banner("1 &nbsp;·&nbsp; Attendance Performance"))
    story.append(Spacer(1, 4))
    story.append(kpi_strip([
        ("Total Sessions", str(p.att_total_sessions), BRAND_BANNER, BRAND_DARK_BLUE),
        ("Attended", str(p.att_attended), GREEN_BG, GREEN_FG),
        ("Avg Attendance",
         f"{p.att_avg_pct:.0f}%" if p.att_avg_pct is not None else "—",
         *att_pct_colors(p.att_avg_pct)),
        ("Avg Duration",
         f"{p.att_avg_dur_pct:.0f}%" if p.att_avg_dur_pct is not None else "—",
         BRAND_SUBTLE, BRAND_DARK_BLUE),
    ]))
    story.append(Spacer(1, 5))

    att_headers = ["Course", "Period", "Sess.", "Pres.", "Abs.", "Att %",
                   "Avg Min", "Ratings", "Avg Rating"]
    att_rows = [[Paragraph(h, HEAD) for h in att_headers]]
    att_bgs = []
    for c in p.att_courses:
        bg, _fg = att_pct_colors(c.att_pct)
        att_rows.append([
            Paragraph(c.course or "—", CELL),
            Paragraph(c.period or "—", CELL),
            Paragraph(str(c.sessions), CELLC),
            Paragraph(str(c.present), CELLC),
            Paragraph(str(c.absent), CELLC),
            Paragraph(f"{c.att_pct:.0f}%" if c.att_pct is not None else "—", CELLC),
            Paragraph(f"{c.avg_time:.0f}" if c.avg_time is not None else "—", CELLC),
            Paragraph(str(c.ratings) if c.ratings is not None else "—", CELLC),
            Paragraph(f"{c.avg_rating:.1f}" if c.avg_rating is not None else "—", CELLC),
        ])
        att_bgs.append(colors.HexColor(bg))
    if not p.att_courses:
        att_rows.append([Paragraph("No attendance records found.", CELL)]
                        + [Paragraph("", CELL)] * (len(att_headers) - 1))

    cw = [usable_w * w for w in
          (0.22, 0.22, 0.07, 0.07, 0.07, 0.09, 0.09, 0.08, 0.09)]
    t = Table(att_rows, colWidths=cw, repeatRows=1)
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_DARK_BLUE)),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]
    for i, bg in enumerate(att_bgs, start=1):
        cmds.append(("BACKGROUND", (5, i), (5, i), bg))
    t.setStyle(TableStyle(cmds))
    story.append(t)
    story.append(Spacer(1, 12))

    # ── SECTION 2 — ASSIGNMENTS ──────────────────────────────────────────────
    story.append(section_banner("2 &nbsp;·&nbsp; Assignment Submission Status"))
    story.append(Spacer(1, 4))
    story.append(kpi_strip([
        ("Assignments Given", str(p.asg_total), BRAND_BANNER, BRAND_DARK_BLUE),
        ("Submitted", str(p.asg_submitted), GREEN_BG, GREEN_FG),
        ("Submission Rate",
         f"{p.asg_pct:.0f}%" if p.asg_pct is not None else "—",
         *att_pct_colors(p.asg_pct)),
    ]))
    story.append(Spacer(1, 5))

    asg_headers = ["#", "Assignment", "Class", "Deadline", "Status",
                   "Marks", "Max", "Feedback"]
    asg_rows = [[Paragraph(h, HEAD) for h in asg_headers]]
    asg_bgs = []
    for a in p.asg_items:
        bg, _fg = status_colors(a.status)
        asg_rows.append([
            Paragraph(str(a.seq), CELLC),
            Paragraph(a.title or "—", CELL),
            Paragraph(a.class_name or "—", CELL),
            Paragraph(fmt_date(a.deadline), CELLC),
            Paragraph(a.status or "—", CELLC),
            Paragraph(a.marks or "—", CELLC),
            Paragraph(a.max_marks or "—", CELLC),
            Paragraph((a.feedback or "—").replace("\n", "<br/>"), CELL),
        ])
        asg_bgs.append(colors.HexColor(bg))
    if not p.asg_items:
        asg_rows.append([Paragraph("No assignments found.", CELL)]
                        + [Paragraph("", CELL)] * (len(asg_headers) - 1))

    cw = [usable_w * w for w in
          (0.05, 0.24, 0.16, 0.11, 0.13, 0.07, 0.06, 0.18)]
    t = Table(asg_rows, colWidths=cw, repeatRows=1)
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_DARK_BLUE)),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]
    for i, bg in enumerate(asg_bgs, start=1):
        cmds.append(("BACKGROUND", (4, i), (4, i), bg))
    t.setStyle(TableStyle(cmds))
    story.append(t)
    story.append(Spacer(1, 12))

    # ── SECTION 3 — INTERVIEWS ───────────────────────────────────────────────
    story.append(section_banner("3 &nbsp;·&nbsp; Interview Feedback & Performance"))
    story.append(Spacer(1, 4))
    story.append(kpi_strip([
        ("Total Interviews", str(p.itv_total), BRAND_BANNER, BRAND_DARK_BLUE),
        ("Attended", str(p.itv_given), GREEN_BG, GREEN_FG),
        ("Attendance",
         f"{p.itv_att_pct:.0f}%" if p.itv_att_pct is not None else "—",
         *att_pct_colors(p.itv_att_pct)),
        ("Avg Score",
         f"{p.itv_avg_score:.1f}" if p.itv_avg_score is not None else "—",
         BRAND_SUBTLE, BRAND_DARK_BLUE),
    ]))
    story.append(Spacer(1, 5))

    itv_headers = ["#", "Interview Time", "Comm.", "Tech Skills",
                   "Total", "Avg", "Comments", "Zone"]
    itv_rows = [[Paragraph(h, HEAD) for h in itv_headers]]
    itv_bgs = []
    for it in p.itv_items:
        bg, fg = zone_colors(it.zone)
        itv_rows.append([
            Paragraph(str(it.no), CELLC),
            Paragraph(it.when or "—", CELL),
            Paragraph(it.comm or "—", CELLC),
            Paragraph(it.tech or "—", CELL),
            Paragraph(it.total or "—", CELLC),
            Paragraph(it.avg or "—", CELLC),
            Paragraph((it.comments or "—").replace("\n", "<br/>"), CELL),
            Paragraph(f'<font color="{fg}"><b>{it.zone or "—"}</b></font>', CELLC),
        ])
        itv_bgs.append(colors.HexColor(bg))
    if not p.itv_items:
        itv_rows.append([Paragraph("No interview records available.", CELL)]
                        + [Paragraph("", CELL)] * (len(itv_headers) - 1))

    cw = [usable_w * w for w in
          (0.05, 0.22, 0.08, 0.16, 0.07, 0.06, 0.24, 0.12)]
    t = Table(itv_rows, colWidths=cw, repeatRows=1)
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_DARK_BLUE)),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]
    for i, bg in enumerate(itv_bgs, start=1):
        cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
    t.setStyle(TableStyle(cmds))
    story.append(t)
    story.append(Spacer(1, 14))

    story.append(Paragraph(
        "This is an automated comprehensive performance report from "
        "IntelliBI Innovations Technologies. Figures combine attendance, "
        "assignment submission and interview feedback data as of the generation "
        "date above.", NOTE))

    doc.build(story)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes, build_filename(p)


# ═════════════════════════════════════════════════════════════════════════════
#  LIVE DATA LOADERS  (Google Sheets)
# ═════════════════════════════════════════════════════════════════════════════
def _read_sheet_df(service, spreadsheet_id: str, tab: str,
                   header_row: int = 1, data_start=None):
    import pandas as pd
    if data_start is None:
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id, range=f"'{tab}'").execute()
        values = result.get("values", [])
        if not values:
            return pd.DataFrame()
        headers = [str(h).strip() for h in values[0]]
        rows = values[1:]
    else:
        hdr = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{tab}'!A{header_row}:ZZ{header_row}").execute()
        headers = [str(h).strip() for h in (hdr.get("values") or [[]])[0]]
        body = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{tab}'!A{data_start}:ZZ").execute()
        rows = body.get("values", [])
    width = len(headers)
    fixed = [(r + [""] * (width - len(r)))[:width] for r in rows]
    return pd.DataFrame(fixed, columns=headers)


def _first_tab(service, spreadsheet_id: str) -> str:
    meta = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id, fields="sheets.properties").execute()
    return meta["sheets"][0]["properties"]["title"]


def load_live_profiles(service, hl_year: int, hl_month: int, month_label: str):
    import pandas as pd

    print("[Data] Reading Students roster …")
    stu = _read_sheet_df(service, STUDENTS_SHEET_ID, STUDENTS_TAB)
    print("[Data] Reading Attendance …")
    att = _read_sheet_df(service, ATTENDANCE_SHEET_ID, ATT_TAB)
    print("[Data] Reading Student_Feedback …")
    fb  = _read_sheet_df(service, ATTENDANCE_SHEET_ID, STUDENT_FEEDBACK_TAB)
    print("[Data] Reading Assessment_Assigned …")
    assigned = _read_sheet_df(service, SUBMISSION_SHEET_ID, ASSIGNED_TAB)
    print("[Data] Reading Submissions …")
    subs = _read_sheet_df(service, SUBMISSION_SHEET_ID, SUBMISSIONS_TAB)
    print("[Data] Reading Consolidated Interview sheet …")
    itv = _read_sheet_df(service, INTERVIEW_SHEET_ID,
                         _first_tab(service, INTERVIEW_SHEET_ID),
                         header_row=INTERVIEW_HEADER_ROW,
                         data_start=INTERVIEW_DATA_START)

    def _dedup(df):
        df.columns = [str(c).strip() for c in df.columns]
        return df.loc[:, ~df.columns.duplicated()]
    stu, att, fb, assigned, subs, itv = (
        _dedup(stu), _dedup(att), _dedup(fb), _dedup(assigned), _dedup(subs), _dedup(itv))

    def col(df, *names):
        for n in names:
            if n in df.columns:
                return n
        low = {c.lower(): c for c in df.columns}
        for n in names:
            if n.lower() in low:
                return low[n.lower()]
        for n in names:
            for c in df.columns:
                if n.lower() in c.lower():
                    return c
        return None

    att_sid = col(att, "student_id")
    att_course = col(att, "course_name", "course_title")
    att_status = col(att, "status")
    att_pct_c  = col(att, "attendance_percent")
    att_dur_c  = col(att, "duration")
    att_date_c = col(att, "session_start_ist", "session_start")
    if att_sid:
        att["_pct"] = att[att_pct_c].map(lambda v: to_float(v, 0)) if att_pct_c else 0
        att["_durmin"] = att[att_dur_c].map(lambda v: (to_float(v, 0) or 0) / 60.0) if att_dur_c else 0
        att["_d"] = att[att_date_c].map(parse_date) if att_date_c else None

    fb_sid = col(fb, "student_id")
    fb_rate = col(fb, "rating")

    sub_email = col(subs, "student_email", "email")
    sub_aid   = col(subs, "assessment_id")
    sub_status = col(subs, "submission_status", "status")
    sub_marks  = col(subs, "evaluation_marks", "marks")
    sub_fb     = col(subs, "evaluation_feedback", "feedback")
    sub_at     = col(subs, "submitted_at")
    asg_id   = col(assigned, "assessment_id")
    asg_title = col(assigned, "assessment_title", "title")
    asg_class = col(assigned, "class_name")
    asg_max   = col(assigned, "maximum_marks", "max_marks")
    asg_start = col(assigned, "submission_start_date", "start")
    asg_dead  = col(assigned, "submission_deadline", "deadline")
    assigned_lookup = {}
    if asg_id:
        for _, a in assigned.iterrows():
            assigned_lookup[str(a[asg_id])] = a

    itv_email = col(itv, "email")
    itv_cand  = col(itv, "candidate name", "student name", "candidate", "student")
    itv_when  = col(itv, "slot", "interview time", "start time")
    itv_comm  = col(itv, "communication")
    itv_tech  = col(itv, "tech scores", "tech score", "tech skills")
    itv_total = col(itv, "total score")
    itv_avg   = col(itv, "average score", "avg score", "rating")
    itv_zone  = col(itv, "zone", "status", "result")
    itv_comm_txt = col(itv, "comments", "remark", "feedback")
    itv_date  = col(itv, "interview date", "date")

    s_id    = col(stu, "student_id")
    s_name  = col(stu, "student_name", "name")
    s_email = col(stu, "email")
    s_phone = col(stu, "phone")
    s_batch = col(stu, "batch_name")
    s_join  = col(stu, "joined_on", "joinedon", "joining")
    s_active = col(stu, "is_active_student", "is_candidate_active", "is_active")

    # ── Pre-index child tables by their join key ONCE ─────────────────────────
    # Previously every student triggered a FULL-table scan of attendance,
    # submissions and interviews — O(students × rows) — each rebuilding the same
    # .astype(str)/.str.strip()/.str.lower() Series from scratch. Grouping each
    # table by its join key a single time turns every per-student lookup into an
    # O(1) dict fetch. groupby preserves original row order within each group, so
    # each slice contains the EXACT same rows, in the same order, as the old
    # boolean-mask filters — results are identical, only faster.
    att_by_sid = {}
    if att_sid:
        for _k, _g in att.groupby(att[att_sid].astype(str), sort=False):
            att_by_sid[str(_k)] = _g

    _subs_empty = subs.iloc[0:0]
    subs_by_email = {}
    if sub_email and not subs.empty:
        for _k, _g in subs.groupby(
                subs[sub_email].astype(str).str.strip().str.lower(), sort=False):
            subs_by_email[str(_k)] = _g

    _itv_empty = itv.iloc[0:0]
    itv_by_email, itv_by_cand = {}, {}
    if not itv.empty:
        if itv_email:
            for _k, _g in itv.groupby(
                    itv[itv_email].astype(str).str.strip().str.lower(), sort=False):
                itv_by_email[str(_k)] = _g
        if itv_cand:
            for _k, _g in itv.groupby(
                    itv[itv_cand].astype(str).str.strip().str.lower(), sort=False):
                itv_by_cand[str(_k)] = _g

    profiles = []
    for _, srow in stu.iterrows():
        sid   = str(srow.get(s_id, "")) if s_id else ""
        name  = str(srow.get(s_name, "")) if s_name else ""
        email = str(srow.get(s_email, "")).strip().lower() if s_email else ""
        if not name and not email:
            continue
        p = StudentProfile(
            name=name or email, email=email,
            phone=str(srow.get(s_phone, "")) if s_phone else "",
            batch_name=str(srow.get(s_batch, "")) if s_batch else "",
            enrollment_date=str(srow.get(s_join, "")) if s_join else "",
            student_id=sid, month_label=month_label,
            active_flag=str(srow.get(s_active, "")) if s_active else "",
        )

        if att_sid and sid:
            mine = att_by_sid.get(sid)
            if mine is not None and not mine.empty:
                p.att_total_sessions = len(mine)
                present_mask = mine[att_status].astype(str).str.lower().ne("absent") \
                    if att_status else (mine["_pct"] > 0)
                p.att_attended = int(present_mask.sum())
                p.att_avg_pct = round(100.0 * p.att_attended / p.att_total_sessions, 1) \
                    if p.att_total_sessions else None
                p.att_avg_dur_pct = None
                if att_date_c:
                    mn = mine[mine["_d"].map(
                        lambda d: bool(d) and d.year == hl_year and d.month == hl_month)]
                    p.month_sessions = len(mn)
                    if att_status:
                        p.month_present = int(mn[att_status].astype(str).str.lower().ne("absent").sum())
                    else:
                        p.month_present = int((mn["_pct"] > 0).sum())
                if att_course:
                    for course, g in mine.groupby(mine[att_course].astype(str)):
                        sess = len(g)
                        if att_status:
                            pres = int(g[att_status].astype(str).str.lower().ne("absent").sum())
                        else:
                            pres = int((g["_pct"] > 0).sum())
                        attended_g = g.loc[(g[att_status].astype(str).str.lower().ne("absent"))
                                           if att_status else (g["_pct"] > 0)]
                        avgt = round(float(attended_g["_durmin"].mean()), 1) \
                            if len(attended_g) and attended_g["_durmin"].mean() == attended_g["_durmin"].mean() else None
                        dts = [d for d in g["_d"].tolist() if d] if att_date_c else []
                        period = (f"{min(dts).strftime('%d-%b-%Y')} To "
                                  f"{max(dts).strftime('%d-%b-%Y')}") if dts else "—"
                        p.att_courses.append(AttendanceCourse(
                            course=course or "—", period=period,
                            sessions=sess, present=pres, absent=sess - pres,
                            att_pct=round(100.0 * pres / sess, 1) if sess else None,
                            avg_time=avgt,
                        ))
                    p.att_courses.sort(key=lambda c: c.course.lower())

        if sub_email and email and not subs.empty:
            mine = subs_by_email.get(email, _subs_empty)
            p.asg_total = len(mine)

            def _is_sub(s):
                s = str(s).lower()
                return ("not" not in s) and ("submit" in s or "grad" in s)
            sub_flags = ([_is_sub(x) for x in mine[sub_status].tolist()]
                         if (sub_status and sub_status in mine.columns) else [False] * len(mine))
            p.asg_submitted = sum(1 for fl in sub_flags if fl)
            p.asg_pct = round(100.0 * p.asg_submitted / p.asg_total, 1) if p.asg_total else None
            if sub_at and sub_at in mine.columns:
                month_flags = [bool(_d) and _d.year == hl_year and _d.month == hl_month
                               for _d in (parse_date(v) for v in mine[sub_at].tolist())]
                p.month_asg_sub = sum(1 for fl, mm in zip(sub_flags, month_flags) if fl and mm)
            seq = 0
            for _, sub in mine.iterrows():
                seq += 1
                aid = str(sub.get(sub_aid, "")) if sub_aid else ""
                a = assigned_lookup.get(aid)
                p.asg_items.append(AssignmentItem(
                    seq=str(seq),
                    title=str(a[asg_title]) if (a is not None and asg_title) else "—",
                    class_name=str(a[asg_class]) if (a is not None and asg_class) else "—",
                    start=str(a[asg_start]) if (a is not None and asg_start) else "—",
                    deadline=str(a[asg_dead]) if (a is not None and asg_dead) else "—",
                    max_marks=str(a[asg_max]) if (a is not None and asg_max) else "",
                    submitted_at=str(sub.get(sub_at, "")) if sub_at else "",
                    status=str(sub.get(sub_status, "")) if sub_status else "Not Submitted",
                    marks=str(sub.get(sub_marks, "")) if sub_marks else "",
                    feedback=str(sub.get(sub_fb, "")) if sub_fb else "",
                ))

        if itv_email and email and not itv.empty:
            mine = itv_by_email.get(email, _itv_empty)
            if mine.empty and itv_cand and name:
                mine = itv_by_cand.get(name.strip().lower(), _itv_empty)
            p.itv_total = len(mine)
            scores, attended = [], 0
            no = 0
            for _, rr in mine.iterrows():
                no += 1
                zone = str(rr.get(itv_zone, "")) if itv_zone else ""
                avg = to_float(rr.get(itv_avg, "")) if itv_avg else None
                is_abs = "absent" in zone.lower()
                if not is_abs and (avg or to_float(rr.get(itv_total, ""), 0)):
                    attended += 1
                    if avg:
                        scores.append(avg)
                p.itv_items.append(InterviewItem(
                    no=str(no),
                    when=str(rr.get(itv_when, "")) if itv_when else "—",
                    comm=str(rr.get(itv_comm, "")) if itv_comm else "",
                    tech=str(rr.get(itv_tech, "")) if itv_tech else "",
                    total=str(rr.get(itv_total, "")) if itv_total else "",
                    avg=str(rr.get(itv_avg, "")) if itv_avg else "",
                    comments=str(rr.get(itv_comm_txt, "")) if itv_comm_txt else "",
                    zone=zone or "—",
                ))
                if itv_date:
                    d = parse_date(rr.get(itv_date, ""))
                    if d and d.year == hl_year and d.month == hl_month:
                        p.month_itv += 1
            p.itv_given = attended
            p.itv_att_pct = round(100.0 * attended / p.itv_total, 1) if p.itv_total else None
            p.itv_avg_score = round(sum(scores) / len(scores), 1) if scores else None

        profiles.append(p)

    print(f"[Data] Built {len(profiles)} student profile(s).")
    return profiles


# ═════════════════════════════════════════════════════════════════════════════
#  OFFLINE LOADER — build a profile from a per-student 3-sheet sample workbook
# ═════════════════════════════════════════════════════════════════════════════
def build_profile_from_sample_xlsx(path: str, month_label: str = "") -> StudentProfile:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    def sheet_by_kw(kw):
        for ws in wb.worksheets:
            if kw.lower() in ws.title.lower():
                return ws
        return None

    p = StudentProfile(month_label=month_label)

    aw = sheet_by_kw("attendance")
    if aw:
        rows = list(aw.iter_rows(values_only=True))
        title = str(rows[0][0] or "") if rows else ""
        p.name = title.replace("Attendance Report", "").strip() or p.name
        if len(rows) > 2:
            k = rows[2]
            p.att_total_sessions = to_int(k[0]) if len(k) > 0 else 0
            p.att_attended = to_int(k[1]) if len(k) > 1 else 0
            ap = to_float(k[2]) if len(k) > 2 else None
            p.att_avg_pct = round(ap * 100, 1) if ap is not None and ap <= 1.5 else ap
            ad = to_float(k[3]) if len(k) > 3 else None
            p.att_avg_dur_pct = round(ad * 100, 1) if ad is not None and ad <= 1.5 else ad
        for r in rows[4:]:
            if not r or all(c in (None, "") for c in r):
                continue
            if r[0] and len(str(r[0]).strip()) and (len(r) < 4 or r[3] in (None, "")):
                continue
            if len(r) >= 7 and r[0] not in (None, ""):
                p.att_courses.append(AttendanceCourse(
                    course=str(r[0]), period=str(r[1] or "—"),
                    sessions=to_int(r[3]), present=to_int(r[4]), absent=to_int(r[5]),
                    att_pct=to_float(r[6]), avg_time=to_float(r[7]) if len(r) > 7 else None,
                    ratings=to_int(r[8]) if len(r) > 8 and str(r[8]).strip() not in ("—", "") else None,
                    avg_rating=to_float(r[9]) if len(r) > 9 else None,
                ))

    gw = sheet_by_kw("assignment")
    if gw:
        rows = list(gw.iter_rows(values_only=True))
        if len(rows) > 2:
            k = rows[2]
            p.asg_total = to_int(k[0]) if len(k) > 0 else 0
            p.asg_submitted = to_int(k[3]) if len(k) > 3 else 0
            p.asg_pct = to_float(k[6]) if len(k) > 6 else None
        seq = 0
        i = 4
        while i < len(rows):
            r = rows[i]
            if r and r[1] and "|" in str(r[1]):
                meta = str(r[1])
                m_title = meta.split("|")[0].replace("📝", "").strip()
                m_class = ""; m_dead = ""; m_max = ""
                for seg in meta.split("|"):
                    seg = seg.strip()
                    if seg.lower().startswith("class:"):
                        m_class = seg.split(":", 1)[1].strip()
                    elif seg.lower().startswith("deadline:"):
                        m_dead = seg.split(":", 1)[1].strip()
                    elif seg.lower().startswith("max marks:"):
                        m_max = seg.split(":", 1)[1].strip()
                drow = rows[i + 1] if i + 1 < len(rows) else None
                status = "Not Submitted"; marks = ""; fbk = ""
                if drow:
                    for cell in drow:
                        cs = str(cell or "")
                        if any(t in cs for t in ("Submitted", "Not Submitted", "Graded",
                                                 "❌", "✅", "📤")):
                            status = cs
                            break
                    if len(drow) > 6 and drow[6] not in (None, ""):
                        marks = str(drow[6])
                    if len(drow) > 8 and drow[8] not in (None, ""):
                        fbk = str(drow[8])
                seq += 1
                p.asg_items.append(AssignmentItem(
                    seq=str(r[0] if r[0] not in (None, "") else seq),
                    title=m_title or "—", class_name=m_class or "—",
                    deadline=m_dead or "—", max_marks=m_max,
                    status=status, marks=marks, feedback=fbk,
                ))
                i += 2
            else:
                i += 1

    iw = sheet_by_kw("interview")
    if iw:
        rows = list(iw.iter_rows(values_only=True))
        if len(rows) > 2:
            k = rows[2]
            p.itv_total = to_int(k[0]) if len(k) > 0 else 0
            p.itv_given = to_int(k[2]) if len(k) > 2 else 0
            ap = to_float(k[4]) if len(k) > 4 else None
            p.itv_att_pct = round(ap, 1) if ap is not None and ap > 1.5 else (round(ap * 100, 1) if ap is not None else None)
            p.itv_avg_score = to_float(k[6]) if len(k) > 6 else None
        for r in rows[4:]:
            if not r or r[0] in (None, ""):
                continue
            p.itv_items.append(InterviewItem(
                no=str(r[0]), when=str(r[1] or "—"),
                comm=str(r[3]) if len(r) > 3 and r[3] not in (None, "") else "",
                tech=str(r[4]) if len(r) > 4 and r[4] not in (None, "") else "",
                total=str(r[5]) if len(r) > 5 and r[5] not in (None, "") else "",
                avg=str(r[6]) if len(r) > 6 and r[6] not in (None, "") else "",
                comments=str(r[7]) if len(r) > 7 and r[7] not in (None, "") else "",
                zone=str(r[8]) if len(r) > 8 and r[8] not in (None, "") else "—",
            ))
    return p


# ═════════════════════════════════════════════════════════════════════════════
#  EMAIL DELIVERY
# ═════════════════════════════════════════════════════════════════════════════
def send_student_report(to_email: str, p: StudentProfile, pdf_bytes: bytes,
                        filename: str):
    subject = f"Your IntelliBI Student Profile Report — {datetime.now().strftime('%b %Y')}"
    body = f"""\
<html><body style="font-family:Arial,sans-serif;color:#1F3864;">
<h2 style="color:#1F3864;">📘 Your IntelliBI Student Profile Report</h2>
<p>Hi {p.name or 'there'},</p>
<p>Please find attached your comprehensive performance report covering your
<strong>attendance</strong>, <strong>assignment submissions</strong> and
<strong>interview feedback</strong>.</p>
<table style="border-collapse:collapse;font-size:14px;">
  <tr><td style="padding:8px 18px;background:#D6E4F0;font-weight:bold;border:1px solid #B0B0B0;">Overall Attendance</td>
      <td style="padding:8px 18px;border:1px solid #B0B0B0;">{(f'{p.att_avg_pct:.0f}%' if p.att_avg_pct is not None else '—')}</td></tr>
  <tr><td style="padding:8px 18px;background:#D6E4F0;font-weight:bold;border:1px solid #B0B0B0;">Assignment Submission Rate</td>
      <td style="padding:8px 18px;border:1px solid #B0B0B0;">{(f'{p.asg_pct:.0f}%' if p.asg_pct is not None else '—')}</td></tr>
  <tr><td style="padding:8px 18px;background:#E8F5E9;font-weight:bold;border:1px solid #B0B0B0;">Interviews Attended</td>
      <td style="padding:8px 18px;border:1px solid #B0B0B0;">{p.itv_given} / {p.itv_total}</td></tr>
</table>
<br>
<p style="font-size:12px;color:#888;">This is an automated report from IntelliBI Innovations Technologies.
If any details look incorrect, please reply to this email.</p>
</body></html>"""

    msg = MIMEMultipart()
    msg["From"] = GMAIL_SENDER
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(pdf_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_SENDER, GMAIL_APP_PASS)
        server.sendmail(GMAIL_SENDER, [to_email], msg.as_string())
    _emit(f"   ↳ [Email] sent to {to_email}")


# ═════════════════════════════════════════════════════════════════════════════
#  MASTER STUDENT PROGRESS REPORT  (executive Excel workbook, admin-only)
# ═════════════════════════════════════════════════════════════════════════════
def _overall_status(att_pct, asg_pct, itv_att_pct) -> str:
    a = att_pct if att_pct is not None else 0
    s = asg_pct if asg_pct is not None else 0
    if a < 50 or s < 40:
        return "At Risk"
    if a >= 75 and s >= 75:
        return "Excellent"
    return "On Track"


def build_master_report(profiles: list, month_label: str, out_path: str = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

    NAV = "1F3864"; MID = "2E5496"; BANNER = "D6E4F0"; SUBTLE = "F5F7FA"
    WHITE = "FFFFFF"; GREEN = "C8E6C9"; AMBER = "FFE0B2"; RED = "FFCDD2"; FONT = "Arial"
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    blank = Border()

    def style(cell, *, bold=False, size=10, color="000000", bg=None,
              align="left", wrap=False, fmt=None, bd=True):
        cell.font = Font(name=FONT, bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cell.border = border if bd else blank
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if fmt:
            cell.number_format = fmt

    wb = Workbook()

    # ── Student Progress (data) ──────────────────────────────────────────────
    sp = wb.active
    sp.title = "Student Progress"
    headers = [
        "#", "Student Name", "Batch", "Email", "Enrollment Date",
        "Total Sessions", "Attended", "Attendance %", "This Month Att %",
        "Assignments Given", "Submitted", "Submission %",
        "Interviews", "Attended", "Avg Score", "Latest Zone", "Overall Status",
    ]
    NC = len(headers)
    sp.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
    style(sp.cell(1, 1, f"Student Progress Detail — {month_label}"),
          bold=True, size=14, color=WHITE, bg=NAV, align="center")
    sp.row_dimensions[1].height = 26
    for j, h in enumerate(headers, 1):
        style(sp.cell(2, j, h), bold=True, size=9, color=WHITE, bg=MID,
              align="center", wrap=True)
    sp.row_dimensions[2].height = 34

    profs = sorted(profiles, key=lambda p: ((p.batch_name or "~").lower(),
                                            (p.name or "").lower()))
    r = 3
    for i, p in enumerate(profs, 1):
        mpct = (100.0 * p.month_present / p.month_sessions) if p.month_sessions else None
        latest_zone = ""
        for it in p.itv_items:
            if it.zone and it.zone != "—":
                latest_zone = it.zone
        bg_row = SUBTLE if r % 2 else WHITE
        vals = [
            (i, "center"), (p.name or "—", "left"), (p.batch_name or "—", "left"),
            (p.email or "—", "left"),
            (fmt_date(p.enrollment_date) if p.enrollment_date else "—", "center"),
            (p.att_total_sessions, "center"), (p.att_attended, "center"),
        ]
        for col_idx, (v, al) in enumerate(vals, 1):
            style(sp.cell(r, col_idx, v), size=9, align=al, bg=bg_row)
        style(sp.cell(r, 8), size=9, align="center", bg=bg_row, fmt='0"%";;"—"')
        sp.cell(r, 8).value = f"=IF(F{r}=0,\"\",ROUND(G{r}/F{r}*100,0))"
        style(sp.cell(r, 9, (round(mpct) if mpct is not None else "—")), size=9,
              align="center", bg=bg_row, fmt=('0"%"' if mpct is not None else "General"))
        style(sp.cell(r, 10, p.asg_total), size=9, align="center", bg=bg_row)
        style(sp.cell(r, 11, p.asg_submitted), size=9, align="center", bg=bg_row)
        style(sp.cell(r, 12), size=9, align="center", bg=bg_row, fmt='0"%";;"—"')
        sp.cell(r, 12).value = f"=IF(J{r}=0,\"\",ROUND(K{r}/J{r}*100,0))"
        style(sp.cell(r, 13, p.itv_total), size=9, align="center", bg=bg_row)
        style(sp.cell(r, 14, p.itv_given), size=9, align="center", bg=bg_row)
        style(sp.cell(r, 15, (round(p.itv_avg_score, 1) if p.itv_avg_score is not None else "—")),
              size=9, align="center", bg=bg_row,
              fmt=('0.0' if p.itv_avg_score is not None else "General"))
        style(sp.cell(r, 16, latest_zone or "—"), size=9, align="center", bg=bg_row)
        st = _overall_status(p.att_avg_pct, p.asg_pct, p.itv_att_pct)
        style(sp.cell(r, 17, st), bold=True, size=9, align="center", bg=bg_row)
        r += 1
    last = r - 1

    if last >= 3:
        for colL in ("H", "I", "L"):
            sp.conditional_formatting.add(
                f"{colL}3:{colL}{last}",
                ColorScaleRule(start_type="num", start_value=0, start_color=RED,
                               mid_type="num", mid_value=50, mid_color=AMBER,
                               end_type="num", end_value=100, end_color=GREEN))
        for word, color in (("At Risk", RED), ("On Track", AMBER), ("Excellent", GREEN)):
            sp.conditional_formatting.add(
                f"Q3:Q{last}",
                CellIsRule(operator="equal", formula=[f'"{word}"'],
                           fill=PatternFill("solid", fgColor=color)))
    widths = [4, 22, 14, 28, 14, 10, 9, 11, 13, 12, 10, 11, 10, 9, 9, 12, 13]
    for j, w in enumerate(widths, 1):
        sp.column_dimensions[get_column_letter(j)].width = w
    sp.freeze_panes = "A3"
    sp.auto_filter.ref = f"A2:{get_column_letter(NC)}{max(last, 2)}"

    # ── Batch Summary ────────────────────────────────────────────────────────
    bs = wb.create_sheet("Batch Summary")
    bheaders = ["Batch", "Students", "Avg Attendance %", "Avg Submission %",
                "Interviews Attended", "Total Interviews", "Avg Interview Score",
                "At Risk", "Excellent"]
    bs.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(bheaders))
    style(bs.cell(1, 1, f"Batch-wise Performance Summary — {month_label}"),
          bold=True, size=14, color=WHITE, bg=NAV, align="center")
    bs.row_dimensions[1].height = 26
    for j, h in enumerate(bheaders, 1):
        style(bs.cell(2, j, h), bold=True, size=9, color=WHITE, bg=MID,
              align="center", wrap=True)
    bs.row_dimensions[2].height = 34

    batches = sorted({(p.batch_name or "—") for p in profs}, key=lambda x: x.lower())
    br = 3
    for b in batches:
        bl = b.replace('"', '""')
        crit = f"'Student Progress'!$C$3:$C${last}"
        style(bs.cell(br, 1, b), size=9, bg=(SUBTLE if br % 2 else WHITE))
        bs.cell(br, 2).value = f'=COUNTIF({crit},"{bl}")'
        bs.cell(br, 3).value = f"=IFERROR(ROUND(AVERAGEIF({crit},\"{bl}\",'Student Progress'!$H$3:$H${last}),0),\"—\")"
        bs.cell(br, 4).value = f"=IFERROR(ROUND(AVERAGEIF({crit},\"{bl}\",'Student Progress'!$L$3:$L${last}),0),\"—\")"
        bs.cell(br, 5).value = f"=SUMIF({crit},\"{bl}\",'Student Progress'!$N$3:$N${last})"
        bs.cell(br, 6).value = f"=SUMIF({crit},\"{bl}\",'Student Progress'!$M$3:$M${last})"
        bs.cell(br, 7).value = f"=IFERROR(ROUND(AVERAGEIF({crit},\"{bl}\",'Student Progress'!$O$3:$O${last}),1),\"—\")"
        bs.cell(br, 8).value = f'=COUNTIFS({crit},"{bl}",\'Student Progress\'!$Q$3:$Q${last},"At Risk")'
        bs.cell(br, 9).value = f'=COUNTIFS({crit},"{bl}",\'Student Progress\'!$Q$3:$Q${last},"Excellent")'
        for j in range(2, 10):
            fmt = '0"%"' if j in (3, 4) else ('0.0' if j == 7 else "0")
            style(bs.cell(br, j), size=9, align="center",
                  bg=(SUBTLE if br % 2 else WHITE), fmt=fmt)
        br += 1
    blast = br - 1
    style(bs.cell(br, 1, "ALL BATCHES"), bold=True, size=9, color=WHITE, bg=NAV)
    bs.cell(br, 2).value = f"=SUM(B3:B{blast})"
    bs.cell(br, 3).value = f"=IFERROR(ROUND(AVERAGE('Student Progress'!$H$3:$H${last}),0),\"—\")"
    bs.cell(br, 4).value = f"=IFERROR(ROUND(AVERAGE('Student Progress'!$L$3:$L${last}),0),\"—\")"
    bs.cell(br, 5).value = f"=SUM(E3:E{blast})"
    bs.cell(br, 6).value = f"=SUM(F3:F{blast})"
    bs.cell(br, 7).value = f"=IFERROR(ROUND(AVERAGE('Student Progress'!$O$3:$O${last}),1),\"—\")"
    bs.cell(br, 8).value = f"=SUM(H3:H{blast})"
    bs.cell(br, 9).value = f"=SUM(I3:I{blast})"
    for j in range(2, 10):
        fmt = '0"%"' if j in (3, 4) else ('0.0' if j == 7 else "0")
        style(bs.cell(br, j), bold=True, size=9, align="center", color=WHITE, bg=NAV, fmt=fmt)
    for j, w in enumerate([22, 10, 16, 16, 16, 14, 16, 9, 10], 1):
        bs.column_dimensions[get_column_letter(j)].width = w
    bs.freeze_panes = "A3"

    # ── Executive Summary (first tab) ────────────────────────────────────────
    ex = wb.create_sheet("Executive Summary", 0)
    ex.sheet_view.showGridLines = False
    ex.merge_cells("A1:H1")
    style(ex.cell(1, 1, "IntelliBI — Master Student Progress Report"),
          bold=True, size=16, color=NAV, align="left", bd=False)
    ex.row_dimensions[1].height = 28
    ex.merge_cells("A2:H2")
    style(ex.cell(2, 1, f"Executive overview  ·  Reporting period: {month_label}  ·  "
                        f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}"),
          size=10, color="555555", align="left", bd=False)

    kpis = [
        ("Total Students", f"=COUNT('Student Progress'!A3:A{last})", "0", BANNER, NAV),
        ("Total Batches", f"=COUNTA('Batch Summary'!A3:A{blast})", "0", BANNER, NAV),
        ("Avg Attendance", f"=IFERROR(ROUND(AVERAGE('Student Progress'!H3:H{last}),0),0)", '0"%"', GREEN, "1B5E20"),
        ("Avg Submission Rate", f"=IFERROR(ROUND(AVERAGE('Student Progress'!L3:L{last}),0),0)", '0"%"', AMBER, "E65100"),
        ("Interview Attendance", f"=IFERROR(ROUND(SUM('Student Progress'!N3:N{last})/SUM('Student Progress'!M3:M{last})*100,0),0)", '0"%"', BANNER, NAV),
        ("Avg Interview Score", f"=IFERROR(ROUND(AVERAGE('Student Progress'!O3:O{last}),1),0)", "0.0", BANNER, NAV),
        ("Students At Risk", f'=COUNTIF(\'Student Progress\'!Q3:Q{last},"At Risk")', "0", RED, "B71C1C"),
        ("Students Excellent", f'=COUNTIF(\'Student Progress\'!Q3:Q{last},"Excellent")', "0", GREEN, "1B5E20"),
    ]
    start_row = 4
    for idx, (label, formula, fmt, bg, fg) in enumerate(kpis):
        block = idx // 4
        col = 1 + (idx % 4) * 2
        rr = start_row + block * 3
        ex.merge_cells(start_row=rr, start_column=col, end_row=rr, end_column=col + 1)
        style(ex.cell(rr, col, label), bold=True, size=9, color=fg, bg=bg, align="center")
        ex.merge_cells(start_row=rr + 1, start_column=col, end_row=rr + 1, end_column=col + 1)
        c = ex.cell(rr + 1, col); c.value = formula
        style(c, bold=True, size=18, color=fg, bg=WHITE, align="center", fmt=fmt)
        ex.row_dimensions[rr + 1].height = 30

    note_row = start_row + 2 * 3 + 1
    ex.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    style(ex.cell(note_row, 1,
                  "Status key:  Excellent = attendance >=75% and submission >=75%   ·   "
                  "On Track = meeting minimums   ·   At Risk = attendance <50% or submission <40%."),
          size=9, color="555555", align="left", bd=False)
    ex.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=8)
    style(ex.cell(note_row + 1, 1,
                  "See the 'Batch Summary' tab for batch-level rollups and the "
                  "'Student Progress' tab for sortable/filterable per-student detail."),
          size=9, color="555555", align="left", bd=False)
    for j, w in enumerate([20, 12, 20, 12, 20, 12, 20, 12], 1):
        ex.column_dimensions[get_column_letter(j)].width = w

    # Drive-only policy: build the workbook in memory and return its bytes.
    # (out_path is accepted for backward-compat but intentionally NOT written.)
    import io as _io
    _buf = _io.BytesIO()
    wb.save(_buf)
    return _buf.getvalue()


def send_master_report(to_email: str, xlsx_bytes: bytes, filename: str,
                       month_label: str, n_students: int):
    subject = f"IntelliBI Master Student Progress Report — {month_label}"
    body = f"""\
<html><body style="font-family:Arial,sans-serif;color:#1F3864;">
<h2 style="color:#1F3864;">📊 IntelliBI Master Student Progress Report</h2>
<p>Attached is the consolidated executive report for <strong>{month_label}</strong>,
covering attendance, assignment submissions and interview performance across
<strong>{n_students}</strong> student(s).</p>
<p>The workbook contains three tabs: <b>Executive Summary</b> (KPIs),
<b>Batch Summary</b> (batch-level rollups), and <b>Student Progress</b>
(sortable per-student detail).</p>
<p style="font-size:12px;color:#888;">Automated report — IntelliBI Innovations Technologies.</p>
</body></html>"""
    data = xlsx_bytes
    msg = MIMEMultipart()
    msg["From"] = GMAIL_SENDER
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html"))
    part = MIMEBase("application",
                    "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(data)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_SENDER, GMAIL_APP_PASS)
        server.sendmail(GMAIL_SENDER, [to_email], msg.as_string())
    print(f"[Email] ✓ Master report sent to {to_email}")


# ═════════════════════════════════════════════════════════════════════════════
#  GOOGLE DRIVE UPLOAD
# ═════════════════════════════════════════════════════════════════════════════
# ═════════════════════════════════════════════════════════════════════════════
#  PERFORMANCE — parallel per-student uploads + persistent Drive-ID cache
# ═════════════════════════════════════════════════════════════════════════════
#  The per-student loop makes several independent Google Drive round-trips
#  (folder lookup, monthly upload, Till-Date upsert, permission share). Doing
#  them one student at a time is network-bound and slow. We now (a) run students
#  across a small thread pool so the network waits overlap, and (b) cache each
#  student's Drive folder-id and Till-Date file-id to a small JSON so later runs
#  skip the two lookup round-trips.
#
#  Behaviour is preserved EXACTLY: same files, same shares, same emails, same
#  final counts, same log lines in the same order. Safety details:
#    • each worker thread uses its OWN Drive client (httplib2 is not thread-safe);
#    • email sends are serialised through a lock (Gmail-friendly, order-neutral);
#    • every student's log lines (including those emitted inside the Drive
#      helpers) are buffered per-student and printed in the original order;
#    • a stale cached id is detected (HTTP 404) and the file/folder transparently
#      recreated, so the outcome equals the uncached path.
MAX_UPLOAD_WORKERS = 8

_thread_local = threading.local()
_email_lock   = threading.Lock()
_cache_lock   = threading.Lock()
_DRIVE_CACHE_PATH = os.path.join(PROJECT_CACHE_DIR or _SCRIPT_DIR,
                                 "pyStudentProfileReport_drive_cache.json")
_drive_id_cache = {"folders": {}, "till_files": {}}


def _emit(msg):
    """Thread-safe log sink: append to the current worker's buffer when one is
    set (so each student's lines print together, in order), else print now."""
    buf = getattr(_thread_local, "logbuf", None)
    if buf is not None:
        buf.append(msg)
    else:
        print(msg)


def _http_status(exc):
    """HTTP status code of a googleapiclient error, or None."""
    resp = getattr(exc, "resp", None)
    status = getattr(resp, "status", None)
    try:
        return int(status) if status is not None else None
    except (TypeError, ValueError):
        return None


def _thread_drive():
    """One Drive service per worker thread (never share a client across threads)."""
    d = getattr(_thread_local, "drive", None)
    if d is None:
        d = _get_drive_service()
        _thread_local.drive = d
    return d


def _load_drive_cache():
    global _drive_id_cache
    try:
        with open(_DRIVE_CACHE_PATH, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if isinstance(data, dict):
            _drive_id_cache = {"folders": dict(data.get("folders") or {}),
                               "till_files": dict(data.get("till_files") or {})}
    except Exception:
        _drive_id_cache = {"folders": {}, "till_files": {}}


def _save_drive_cache():
    try:
        os.makedirs(os.path.dirname(_DRIVE_CACHE_PATH), exist_ok=True)
        with _cache_lock:
            snapshot = {"folders": dict(_drive_id_cache["folders"]),
                        "till_files": dict(_drive_id_cache["till_files"])}
        with open(_DRIVE_CACHE_PATH, "w", encoding="utf-8") as fh:
            json.dump(snapshot, fh)
    except Exception as e:
        # The cache is a pure optimisation; never fail the run over it.
        print(f"[Cache] ⚠ could not save Drive-ID cache: {e}")


def _cached_folder_id(drive, parent_id, name):
    """Folder id from cache (skips a files.list round-trip) or a real
    get-or-create, remembering the result for next run."""
    with _cache_lock:
        cached = _drive_id_cache["folders"].get(name)
    if cached:
        return cached
    fid = _drive_get_or_create_subfolder(drive, parent_id, name)
    with _cache_lock:
        _drive_id_cache["folders"][name] = fid
    return fid


def _cached_upsert_till(drive, folder_id, filename, data):
    """Like _drive_upsert_file, but uses the cached file-id to update in place
    without the files.list lookup. On a cache miss — or a stale id whose file was
    removed (HTTP 404) — it falls back to the exact original upsert and then
    remembers the id. Result is identical to _drive_upsert_file."""
    from googleapiclient.http import MediaIoBaseUpload
    key = f"{folder_id}/{filename}"
    with _cache_lock:
        cached_id = _drive_id_cache["till_files"].get(key)
    if cached_id:
        try:
            media = MediaIoBaseUpload(io.BytesIO(data), mimetype="application/pdf",
                                      resumable=False)
            drive.files().update(fileId=cached_id, media_body=media,
                                 supportsAllDrives=True).execute()
            _emit(f"   ↳ [Drive] Till-Date PDF updated in place (id={cached_id}) — "
                  f"permissions retained")
            return cached_id
        except Exception as e:
            if _http_status(e) == 404:          # file removed → rebuild via slow path
                with _cache_lock:
                    _drive_id_cache["till_files"].pop(key, None)
            else:
                raise
    fid = _drive_upsert_file(drive, folder_id, filename, data, "application/pdf")
    with _cache_lock:
        _drive_id_cache["till_files"][key] = fid
    return fid


class _FolderRef:
    __slots__ = ("id",)

    def __init__(self, fid):
        self.id = fid


def _with_folder_recovery(drive, folder_name, ref, op):
    """Run op(folder_id). If it fails because a cached folder id is stale (the
    folder was removed — HTTP 404), recreate the folder once, refresh the cache,
    and retry so the outcome matches a fresh, uncached run. Any other error
    propagates unchanged to the caller's existing handler."""
    try:
        return op(ref.id)
    except Exception as e:
        if _http_status(e) != 404:
            raise
        with _cache_lock:
            _drive_id_cache["folders"].pop(folder_name, None)
            for k in [k for k in _drive_id_cache["till_files"]
                      if k.startswith(f"{ref.id}/")]:
                _drive_id_cache["till_files"].pop(k, None)
        new_id = _drive_get_or_create_subfolder(drive, DRIVE_ROOT_FOLDER_ID, folder_name)
        with _cache_lock:
            _drive_id_cache["folders"][folder_name] = new_id
        ref.id = new_id
        return op(new_id)


def _student_uploads(drive, p, folder, folder_id, monthly_name, monthly_bytes,
                     till_name, till_bytes, res):
    """Monthly + Till-Date upload/share for one student — identical logic and
    per-file error isolation to the original sequential block, with the Drive-ID
    cache and 404 folder-recovery layered underneath."""
    ref = _FolderRef(folder_id)
    # Monthly — existing behaviour (replace the same-named file).
    if monthly_name:
        try:
            _with_folder_recovery(drive, folder, ref, lambda fid:
                _drive_upload_bytes(drive, fid, monthly_name, monthly_bytes,
                                    "application/pdf"))
            res["uploaded"] += 1
        except Exception as e:
            _emit(f"   ↳ [Drive] ⚠ upload failed for {monthly_name}: {e}")

    # Till-Date — same switch on SHARE_TILL_DATE_REPORT_WITH_STUDENT as before.
    try:
        if SHARE_TILL_DATE_REPORT_WITH_STUDENT:
            till_id = _with_folder_recovery(drive, folder, ref, lambda fid:
                _cached_upsert_till(drive, fid, till_name, till_bytes))
            res["uploaded"] += 1
            if _drive_share_as_viewer(drive, till_id, p.email):
                res["shared"] += 1
            else:
                res["share_skipped"] += 1
        else:
            _with_folder_recovery(drive, folder, ref, lambda fid:
                _drive_upload_bytes(drive, fid, till_name, till_bytes,
                                    "application/pdf"))
            res["uploaded"] += 1
    except Exception as e:
        _emit(f"   ↳ [Drive] ⚠ Till-Date upload/share failed for "
              f"{till_name}: {e}")


def _get_drive_service():
    from googleapiclient.discovery import build
    from google.oauth2 import service_account
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=["https://www.googleapis.com/auth/drive"])
    creds = creds.with_subject(DRIVE_DELEGATED_USER)
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _drive_get_or_create_subfolder(drive, parent_id, name):
    safe = name.replace("'", "\\'")
    q = (f"'{parent_id}' in parents and "
         f"mimeType='application/vnd.google-apps.folder' and "
         f"name='{safe}' and trashed=false")
    res = drive.files().list(q=q, fields="files(id,name)",
                             supportsAllDrives=True,
                             includeItemsFromAllDrives=True).execute()
    files = res.get("files", [])
    if files:
        return files[0]["id"]
    meta = {"name": name, "mimeType": "application/vnd.google-apps.folder",
            "parents": [parent_id]}
    folder = drive.files().create(body=meta, fields="id",
                                  supportsAllDrives=True).execute()
    return folder["id"]


def _drive_upload_bytes(drive, folder_id, filename, data, mimetype):
    """Upload bytes to a Drive folder, replacing any existing same-named file."""
    from googleapiclient.http import MediaIoBaseUpload
    safe = filename.replace("'", "\\'")
    q = f"'{folder_id}' in parents and name='{safe}' and trashed=false"
    existing = drive.files().list(q=q, fields="files(id)",
                                  supportsAllDrives=True,
                                  includeItemsFromAllDrives=True).execute()
    for fobj in existing.get("files", []):
        drive.files().delete(fileId=fobj["id"], supportsAllDrives=True).execute()
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mimetype, resumable=False)
    meta = {"name": filename, "parents": [folder_id]}
    up = drive.files().create(body=meta, media_body=media,
                              fields="id,webViewLink",
                              supportsAllDrives=True).execute()
    return up.get("webViewLink", "")


def _drive_upsert_file(drive, folder_id, filename, data, mimetype):
    """Create the file, or UPDATE an existing same-named file's contents IN PLACE
    (keeping its file id and therefore any sharing permissions). Used for the
    Till-Date report so the daily overwrite retains the student's Viewer access and
    the shared link stays valid. Returns the file id."""
    from googleapiclient.http import MediaIoBaseUpload
    safe = filename.replace("'", "\\'")
    q = f"'{folder_id}' in parents and name='{safe}' and trashed=false"
    existing = drive.files().list(q=q, fields="files(id)",
                                  supportsAllDrives=True,
                                  includeItemsFromAllDrives=True).execute().get("files", [])
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mimetype, resumable=False)
    if existing:
        file_id = existing[0]["id"]
        # Remove any accidental duplicates so exactly one file (with its
        # permissions) survives.
        for extra in existing[1:]:
            try:
                drive.files().delete(fileId=extra["id"], supportsAllDrives=True).execute()
            except Exception:
                pass
        drive.files().update(fileId=file_id, media_body=media,
                             supportsAllDrives=True).execute()
        _emit(f"   ↳ [Drive] Till-Date PDF updated in place (id={file_id}) — "
              f"permissions retained")
        return file_id
    meta = {"name": filename, "parents": [folder_id]}
    up = drive.files().create(body=meta, media_body=media, fields="id",
                              supportsAllDrives=True).execute()
    _emit(f"   ↳ [Drive] Till-Date PDF created (id={up['id']})")
    return up["id"]


def _drive_share_as_viewer(drive, file_id, email):
    """Grant the student VIEWER (reader) access to `file_id`. Viewer only — never
    editor/commenter. Idempotent: if the student already has a permission it is
    left untouched. Returns True when the student has (or now has) access.
    Missing/invalid email or any API error is logged and returns False."""
    email_n = str(email or "").strip()
    if not email_n or "@" not in email_n:
        _emit(f"   ↳ [Share] ⚠ skipped — student email missing/invalid ({email!r})")
        return False
    try:
        perms = drive.permissions().list(
            fileId=file_id, fields="permissions(id,emailAddress,role,type)",
            supportsAllDrives=True).execute().get("permissions", [])
        for pm in perms:
            if str(pm.get("emailAddress", "")).strip().lower() == email_n.lower():
                _emit(f"   ↳ [Share] already shared with {email_n} "
                      f"(role={pm.get('role')}) — retained, no re-share needed")
                return True
        drive.permissions().create(
            fileId=file_id,
            body={"type": "user", "role": "reader", "emailAddress": email_n},
            sendNotificationEmail=False,
            supportsAllDrives=True,
            fields="id").execute()
        _emit(f"   ↳ [Share] ✓ Till-Date PDF shared with {email_n} as Viewer")
        return True
    except Exception as e:
        _emit(f"   ↳ [Share] ⚠ sharing failed for {email_n}: {e}")
        return False


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="Generate & email a unified Student Profile PDF per student, "
                    "plus a consolidated master Excel report for the admin.")
    parser.add_argument("--no-email", action="store_true",
                        help="Build files only; don't send any email.")
    parser.add_argument("--admin-only", action="store_true",
                        help="Mail every student PDF to the admin inbox for review "
                             "instead of to each student.")
    parser.add_argument("--only-email", metavar="EMAIL",
                        help="Process a single student by email.")
    parser.add_argument("--limit", type=int, default=0,
                        help="Process at most N students (smoke test).")
    parser.add_argument("--out-dir", default=None,
                        help="Override the output folder. Default is a month-wise "
                             "folder in the project root, e.g. "
                             "StudentProgressReport_May_2026/.")
    parser.add_argument("--no-master", action="store_true",
                        help="Skip building/sending the master Excel report.")
    parser.add_argument("--no-drive", action="store_true",
                        help="Skip uploading the reports to Google Drive.")
    parser.add_argument("--sample-xlsx", metavar="PATH",
                        help="OFFLINE: build a single PDF (+ master) from a "
                             "per-student 3-sheet sample workbook (no Google / email).")
    args = parser.parse_args()

    # ── Decide the monthly reporting period + whether to generate it ──────────
    #   MANUAL_MONTHLY_REPORT=True → generate for REPORT_MONTH/REPORT_YEAR (any day).
    #   else (automatic)           → generate ONLY on the last calendar day of the
    #                                month, covering the month that is ending today.
    #   The Till-Date report is generated every run regardless (see the loop below).
    today = date.today()
    last_day = _is_last_day_of_month(today)
    if MANUAL_MONTHLY_REPORT:
        hl_y, hl_m = int(REPORT_YEAR), int(REPORT_MONTH)
        generate_monthly = True
        monthly_reason = f"MANUAL_MONTHLY_REPORT — {date(hl_y, hl_m, 1).strftime('%b %Y')}"
    else:
        hl_y, hl_m = today.year, today.month
        generate_monthly = last_day
        monthly_reason = ("automatic — last day of month" if last_day
                          else "skipped — today is not the last day of the month")
    month_label = date(hl_y, hl_m, 1).strftime("%b %Y")
    month_long  = date(hl_y, hl_m, 1).strftime("%B_%Y")   # e.g. July_2026
    folder_name = "StudentProgressReport_" + month_label.replace(" ", "_")
    # Drive-only policy: reports are delivered to Google Drive, never written
    # to a local folder. out_dir is kept only as a label for log messages.
    out_dir = args.out_dir or folder_name
    master_name = f"IntelliBI_Master_Student_Progress_Report_{month_label.replace(' ', '_')}.xlsx"

    # ── Offline sample mode ──────────────────────────────────────────────────
    if args.sample_xlsx:
        print(f"[Sample] Building profile from {args.sample_xlsx}")
        p = build_profile_from_sample_xlsx(args.sample_xlsx, month_label)
        pdf_bytes, filename = build_student_pdf(p)
        print(f"[Sample] ✓ Built {filename} ({len(pdf_bytes):,} bytes; not saved locally)")
        if not args.no_master:
            master_bytes = build_master_report([p], month_label)
            print(f"[Sample] ✓ Built master {master_name} "
                  f"({len(master_bytes):,} bytes; not saved locally)")
        return

    # ── Live mode ────────────────────────────────────────────────────────────
    from utils import get_sheets_service
    service = get_sheets_service(SERVICE_ACCOUNT_FILE)
    profiles = load_live_profiles(service, hl_y, hl_m, month_label)

    # ── Student-selection scope (REPORT_GENERATION_SCOPE) ─────────────────────
    scope = str(REPORT_GENERATION_SCOPE or "ALL").strip().upper()
    if scope == "ACTIVE_ONLY":
        before = len(profiles)
        profiles = [p for p in profiles if _is_active_flag(p.active_flag)]
        print(f"[Scope] ACTIVE_ONLY → {len(profiles)}/{before} active student(s) kept.")
    else:
        print(f"[Scope] ALL → {len(profiles)} student(s).")

    if args.only_email:
        tgt = args.only_email.strip().lower()
        profiles = [p for p in profiles if p.email.strip().lower() == tgt]
        if not profiles:
            raise SystemExit(f"✗ No student with email '{args.only_email}'.")
    if args.limit:
        profiles = profiles[:args.limit]

    # ── Email gating: the monthly report is emailed ONLY on the last calendar day
    #    of the month, only when SEND_STUDENT_REPORT_EMAIL is on, never with
    #    --no-email, and only when a monthly report was actually generated.
    #    Report generation/upload is independent of email outcome.
    email_enabled = SEND_STUDENT_REPORT_EMAIL and (not args.no_email)
    will_email = email_enabled and generate_monthly and last_day
    if email_enabled and not will_email:
        print("[Email] SEND_STUDENT_REPORT_EMAIL is on, but monthly emails only go "
              "out on the last calendar day of the month → skipped this run.")

    print(f"[Monthly]  {'GENERATE' if generate_monthly else 'SKIP'} — {monthly_reason}")
    print(f"[Till-Date] GENERATE (every run, overwriting existing file)"
          + (" · share with student as Viewer"
             if SHARE_TILL_DATE_REPORT_WITH_STUDENT else " · sharing OFF"))
    print(f"→ Processing {len(profiles)} student(s) → Google Drive (per-student "
          f"folders). Monthly emails: {'ON' if will_email else 'off'}")

    # Google Drive: one subfolder per student (FirstName_LastName_BatchName) under
    # the configured root folder.
    drive = None
    if not args.no_drive:
        try:
            drive = _get_drive_service()
            print("[Drive] Uploading into per-student folders under the Drive root …")
        except Exception as e:
            drive = None
            print(f"[Drive] ⚠ Could not initialise Drive upload: {e}")

    # Persistent Drive-ID cache: load prior folder/file ids so this run can skip
    # the per-student lookup round-trips (self-healing on any stale id).
    _load_drive_cache()

    def _process_student(p):
        """Build + upload + (optionally) email one student's reports. Runs in a
        worker thread; buffers its log lines and returns per-student counters so
        the caller aggregates and prints them in the original order — identical
        output to the old sequential loop."""
        _thread_local.logbuf = []
        res = {"ok": 0, "fail": 0, "skipped": 0, "uploaded": 0,
               "emailed": 0, "shared": 0, "share_skipped": 0}
        try:
            folder = student_folder_name(p)

            # Per-student Drive subfolder (thread-local client + cached id)
            drv = _thread_drive() if drive else None
            student_folder_id = None
            if drv:
                try:
                    student_folder_id = _cached_folder_id(
                        drv, DRIVE_ROOT_FOLDER_ID, folder)
                except Exception as e:
                    _emit(f"   ↳ [Drive] ⚠ folder create failed for {folder}: {e}")

            monthly_bytes = monthly_name = None

            # (1) Monthly progress report — ONLY on the last day / manual run.
            #     Existing format & calculations unchanged.
            if generate_monthly:
                monthly_bytes, _ = build_student_pdf(p, variant="monthly")
                monthly_name = monthly_report_filename(p, month_long)

            # (2) Consolidated Till-Date report — EVERY run, fixed name so the
            #     same file is overwritten daily with the latest progress.
            till_bytes, _ = build_student_pdf(p, variant="till_date")
            till_name = till_date_report_filename(p)

            names = ([monthly_name] if monthly_name else []) + [till_name]
            _emit(f"• {folder}/  →  " + " | ".join(names))

            if drv and student_folder_id:
                _student_uploads(drv, p, folder, student_folder_id,
                                 monthly_name, monthly_bytes,
                                 till_name, till_bytes, res)

            # (3) Email the MONTHLY report (last day only, when generated).
            #     Failures are logged and never stop the remaining students.
            #     SMTP is serialised via _email_lock to stay Gmail-friendly.
            if will_email and monthly_bytes is not None:
                recipient = ADMIN_EMAIL if args.admin_only else p.email
                if recipient and "@" in recipient:
                    try:
                        with _email_lock:
                            send_student_report(recipient, p, monthly_bytes, monthly_name)
                        res["emailed"] += 1
                    except Exception as e:
                        _emit(f"   ↳ [Email] ⚠ failed for {recipient}: {e}")
                else:
                    res["skipped"] += 1
                    _emit(f"   ↳ [skip] no valid email for {p.name}")
            res["ok"] += 1
        except Exception as e:
            res["fail"] += 1
            _emit(f"   ✗ FAILED for {p.name}: {e}")
            _emit(traceback.format_exc().rstrip())
        res["logs"] = _thread_local.logbuf
        _thread_local.logbuf = None
        return res

    ok, fail, skipped, uploaded, emailed, shared, share_skipped = 0, 0, 0, 0, 0, 0, 0
    if profiles:
        _workers = max(1, min(MAX_UPLOAD_WORKERS, len(profiles)))
        with ThreadPoolExecutor(max_workers=_workers) as _ex:
            _results = list(_ex.map(_process_student, profiles))
        # Aggregate + print in the ORIGINAL profile order (map preserves order).
        for _res in _results:
            for _line in _res["logs"]:
                print(_line)
            ok            += _res["ok"]
            fail          += _res["fail"]
            skipped       += _res["skipped"]
            uploaded      += _res["uploaded"]
            emailed       += _res["emailed"]
            shared        += _res["shared"]
            share_skipped += _res["share_skipped"]

    # Persist any newly-learned folder/file ids for next run's fast path.
    _save_drive_cache()

    print(f"\n✓ Done. {ok} student(s) processed, {fail} failed, "
          f"{uploaded} file(s) uploaded to Drive, {emailed} email(s) sent, "
          f"{skipped} had no email"
          + (f", {shared} Till-Date report(s) shared as Viewer, "
             f"{share_skipped} share(s) skipped/failed"
             if SHARE_TILL_DATE_REPORT_WITH_STUDENT else "")
          + ". (Google Drive only — no local copies.)")

    # ── Master report (consolidated MONTHLY report, admin-only) ───────────────
    #    Tied to the monthly cadence: only built/sent when the monthly reports are
    #    generated (last day of month or manual run), not on every daily run.
    if not args.no_master and profiles and generate_monthly:
        try:
            master_bytes = build_master_report(profiles, month_label)
            print(f"✓ Master report built: {master_name} ({len(master_bytes):,} bytes)")
            if drive:
                try:
                    master_folder_id = _drive_get_or_create_subfolder(
                        drive, DRIVE_ROOT_FOLDER_ID, folder_name)
                    _drive_upload_bytes(
                        drive, master_folder_id, master_name, master_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    print(f"[Drive] ✓ Master uploaded → Drive root / {folder_name} / {master_name}")
                except Exception as e:
                    print(f"[Drive] ⚠ Master upload failed: {e}")
            if not args.no_email:
                send_master_report(ADMIN_EMAIL, master_bytes, master_name,
                                   month_label, len(profiles))
        except Exception as e:
            print(f"✗ Master report FAILED: {e}")
            traceback.print_exc()


if __name__ == "__main__":
    main()
