"""
================================================================================
  IntelliBI Assignment Submission — Email Reminder System
  Sends deadline-approaching reminder emails to students with pending submissions.

  FLOW:
    1. Read Submissions tab from IntelliBIAssessmentSubmission Google Sheet
    2. Identify students with submission_status = "Not Submitted"
    3. Check submission_deadline against today (ALL reminders enabled):
         - 2 days before deadline → 1st Reminder    (Gentle nudge)
         - 1 day  before deadline → 2nd Reminder    (Urgent warning)
         - Deadline day itself    → Final Reminder  (Last chance)
         - 1 day  AFTER  deadline → Missed Deadline (Overdue follow-up; sent
                                    ONCE, notes that 3 reminders were already
                                    sent and the assignment is now missed)
    4. CONSOLIDATE: If one student has multiple pending assignments,
       send ONE email listing all of them (most urgent level applied).
    5. Generate a BATCH-WISE PDF report listing reminder level + pending
       students per batch, upload to Google Drive.
    6. Print batch/subject-wise console summary with email counts.

  SCHEDULE: Run daily (e.g. via cron / Task Scheduler / run_all.py)

  Usage:
    Execution is controlled by the RUN CONFIGURATION variables near the top of
    this file (report_date / send_email / dry_run / generate_pdf /
    include_overdue_report / overdue_all), then simply run:

        python pyAssignmentSubmissionEmailReminder.py

    Examples (edit the variables, then run):
        send_email  = False           # skip email, show summary + still build PDF
        dry_run     = True            # preview emails + PDF (no send, no upload)
        report_date = "2026-03-29"    # override "today"
================================================================================
"""

# --- IntelliBI Operations Automation portability bootstrap (auto-inserted) ---
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "common"))
import _bootstrap  # noqa: E402  (sys.path + env defaults + config.yaml)
from paths import CREDENTIALS_DIR, CONFIG_DIR, LOGS_DIR, CACHE_DIR as PROJECT_CACHE_DIR  # noqa: E402
# --- end bootstrap ---

import sys
import os
import io
import time
import smtplib
from datetime import datetime, date, timedelta
from email.mime.multipart    import MIMEMultipart
from email.mime.text         import MIMEText
from email.mime.application  import MIMEApplication
from collections import defaultdict

# ── Resolve parent directory so utils.py / service_account.json are always found
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR   = os.path.dirname(_SCRIPT_DIR)          # IntelliBI Automation/
if _BASE_DIR not in sys.path:
    sys.path.insert(0, _BASE_DIR)

import pandas as pd

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────────────────────

SERVICE_ACCOUNT_FILE = os.path.join(CREDENTIALS_DIR, "service_account.json")
SUBMISSION_SHEET_ID  = "1E_pOuZfw4BUhQ1bRMmuc8lPDDJRtoXuGkP-qtHC96HU"
SUBMISSIONS_TAB      = "Submissions"

# Student active-status source — reminder emails go ONLY to active students.
# The 'Students' tab's is_candidate_active column is Y (active) / N (inactive or
# course completed). Matched to each reminder by student email (case-insensitive).
STUDENT_INFO_SHEET_ID = "1Eq7Q3Gota7nYiaorm1L0NoouVfYtS7JkbBp4U5MWzVA"
STUDENT_INFO_TAB      = "Students"
STUDENT_EMAIL_COLUMN  = "email"
ACTIVE_FLAG_COLUMN    = "is_candidate_active"

# E-mail credentials are centralised — edit config_files/email_config.py to rotate.
from email_config import GMAIL_SENDER, GMAIL_APP_PASS

CC_RECIPIENTS  = []

# Recipients of the end-of-run staff summary email (one combined mail after
# every student has been notified). Keeping these OFF each student email
# dramatically reduces Gmail's per-recipient send count.
STAFF_SUMMARY_RECIPIENTS = ["intellibihropsb2ch@gmail.com","info@intellibiinnovationstechnologies.in"]

# ── Send throttling (to stay safely under Gmail anti-spam heuristics) ────────
INTER_EMAIL_DELAY_SEC = 3     # wait N seconds between consecutive student emails
EMAIL_BATCH_SIZE      = 20    # after this many emails, pause longer
EMAIL_BATCH_PAUSE_SEC = 60    # pause (seconds) between batches

# Reminder schedule: days_before_deadline → (reminder_level, label, priority)
#   Priority is used to pick the most-urgent level when a student has
#   multiple pending assignments at different stages.
#   ALL reminder levels are enabled below. The "missed" level (days_remaining
#   = -1, i.e. ONE day AFTER the deadline) is a post-deadline follow-up sent
#   ONCE when a student has not submitted despite the 3 prior reminders.
REMINDER_SCHEDULE = {
     2: ("1st",    "1st Reminder",    1),
     1: ("2nd",    "2nd Reminder",    2),
     0: ("final",  "Final Reminder",  3),
    -1: ("missed", "Missed Deadline", 4),
}

LEVEL_PRIORITY = {"1st": 1, "2nd": 2, "final": 3, "missed": 4}
PRIORITY_TO_LEVEL = {v: k for k, v in LEVEL_PRIORITY.items()}
LEVEL_LABEL = {"1st": "1st Reminder", "2nd": "2nd Reminder",
               "final": "Final Reminder", "missed": "Missed Deadline"}
# Canonical level ordering — use everywhere counts/columns are built so adding
# a new level never causes a KeyError on the per-level dicts below.
LEVEL_ORDER = ["1st", "2nd", "final", "missed"]

# Google Drive folder for PDF upload
PDF_DRIVE_FOLDER_ID = "1suIqjRZxDc3QH3NiWje7XcjRcn5w6zfB"


# ─────────────────────────────────────────────────────────────────────────────
#  RUN CONFIGURATION  (replaces the old command-line arguments)
#  Execution is controlled entirely by the variables below — edit them instead
#  of passing CLI flags.
#    report_date : "YYYY-MM-DD" to run for a specific date, or None = today.
#    send_email  : True  -> build the report/PDF AND send the emails (default).
#                  False -> build everything exactly as before but skip sending
#                           email (old --no-email); the execution summary still
#                           prints as it does today.
#  The options below preserve the other former CLI flags so all existing
#  behaviour stays available (these defaults reproduce a normal live run):
#  dry_run                : True  -> preview emails + PDF, no send/upload (old --dry-run).
#  generate_pdf           : False -> skip PDF generation / Drive upload (old --no-pdf).
#  include_overdue_report : False -> skip the overdue-submissions Excel report
#                             (old --no-overdue-report).
#  overdue_all            : True  -> overdue report includes ALL not-submitted
#  rows, not just overdue (old --overdue-all).
# ─────────────────────────────────────────────────────────────────────────────
report_date            = None
send_email             = True
dry_run                = False
generate_pdf           = True
include_overdue_report = True
overdue_all            = False

# ─────────────────────────────────────────────────────────────────────────────
#  DATE PARSING
# ─────────────────────────────────────────────────────────────────────────────

def _parse_ist_date(val):
    """Parse 'DD/MM/YYYY HH:MM:SS IST' (or similar) → date."""
    if not val or str(val).strip() in ("", "nan", "None"):
        return None
    s = str(val).replace(" IST", "").strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, errors="coerce").date()
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────────────────────
#  GOOGLE SHEETS READER
# ─────────────────────────────────────────────────────────────────────────────

def read_sheet_df(service, spreadsheet_id: str, tab: str) -> pd.DataFrame:
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"{tab}!A:ZZ",
    ).execute()
    values = result.get("values", [])
    if len(values) < 2:
        return pd.DataFrame()
    header = [h.strip() for h in values[0]]
    rows   = [r + [""] * (len(header) - len(r)) for r in values[1:]]
    df     = pd.DataFrame(rows, columns=header)
    str_cols = df.select_dtypes(include=["object", "str"]).columns
    df[str_cols] = df[str_cols].apply(lambda col: col.str.strip())
    return df


# ─────────────────────────────────────────────────────────────────────────────
#  STUDENT ACTIVE-STATUS (restrict reminder emails to active students)
# ─────────────────────────────────────────────────────────────────────────────

def load_active_status_map(service) -> dict:
    """Build {student_email_lower: 'Y'/'N'/''} from the Students tab's
    is_candidate_active column, used to restrict reminder emails to active
    students only.

    Returns an empty dict if the sheet/columns cannot be read; the caller then
    treats every student as unmatched and skips sending (fail-safe)."""
    try:
        df = read_sheet_df(service, STUDENT_INFO_SHEET_ID, STUDENT_INFO_TAB)
    except Exception as e:                                    # noqa: BLE001
        print(f"[Active-Check] ⚠ Could not read '{STUDENT_INFO_TAB}' tab ({e}); "
              f"no reminder emails will be sent this run.")
        return {}
    if df.empty or STUDENT_EMAIL_COLUMN not in df.columns or ACTIVE_FLAG_COLUMN not in df.columns:
        print(f"[Active-Check] ⚠ '{STUDENT_INFO_TAB}' tab missing "
              f"'{STUDENT_EMAIL_COLUMN}' or '{ACTIVE_FLAG_COLUMN}' column; "
              f"no reminder emails will be sent this run.")
        return {}

    status_map: dict = {}
    for _, r in df.iterrows():
        em = str(r.get(STUDENT_EMAIL_COLUMN, "") or "").strip().lower()
        if not em:
            continue
        flag = str(r.get(ACTIVE_FLAG_COLUMN, "") or "").strip().upper()
        # If the same email appears more than once, an active ('Y') record wins
        # over a stale inactive one so a currently-active student is never
        # wrongly suppressed.
        if status_map.get(em) == "Y":
            continue
        status_map[em] = flag
    active_n = sum(1 for v in status_map.values() if v == "Y")
    print(f"[Active-Check] Loaded is_candidate_active for {len(status_map)} "
          f"student email(s) — {active_n} active.")
    return status_map


def is_student_active(student: dict, status_map: dict):
    """Decide whether a reminder email may be sent to `student`.

    Returns (True, "") when is_candidate_active == 'Y' (case/space-insensitive),
    otherwise (False, reason). Matching is by student email (trimmed, lowercased).
    Blank flag, missing student, or blank email all resolve to (False, reason)."""
    email = str(student.get("student_email", "") or "").strip().lower()
    if not email:
        return False, "student email is blank — cannot verify active status"
    flag = status_map.get(email)
    if flag is None:
        return False, (f"student not found in {STUDENT_INFO_TAB} tab "
                       f"({student.get('student_email', '')})")
    if flag == "Y":
        return True, ""
    if flag == "N":
        return False, "is_candidate_active = N (inactive / course completed)"
    if flag == "":
        return False, "is_candidate_active is blank/unavailable"
    return False, f"is_candidate_active = '{flag}' (not 'Y')"


# ─────────────────────────────────────────────────────────────────────────────
#  IDENTIFY PENDING SUBMISSIONS NEARING DEADLINE
# ─────────────────────────────────────────────────────────────────────────────

def find_pending_reminders(subs_df: pd.DataFrame, today: date) -> list:
    """
    Scan submissions for students who:
      - Have submission_status = "Not Submitted"
      - Have a deadline that is 2/1/0 days from today
    Returns a list of dicts: one entry per (student, assignment) row.
    """
    if subs_df.empty:
        return []

    records = []

    for _, row in subs_df.iterrows():
        status = str(row.get("submission_status", "")).strip()
        if status.lower() != "not submitted":
            continue

        email = str(row.get("student_email", "")).strip()
        if not email or "@" not in email:
            continue

        deadline = _parse_ist_date(row.get("submission_deadline", ""))
        if deadline is None:
            continue

        days_remaining = (deadline - today).days
        if days_remaining not in REMINDER_SCHEDULE:
            continue

        reminder_level, reminder_label, _ = REMINDER_SCHEDULE[days_remaining]

        # Assigned date — try submission_start_date first, fall back to alternates
        assigned_date = _parse_ist_date(row.get("submission_start_date", ""))
        if assigned_date is None:
            assigned_date = _parse_ist_date(row.get("assigned_date", ""))
        if assigned_date is None:
            assigned_date = _parse_ist_date(row.get("assignment_start_date", ""))

        if assigned_date is not None:
            assigned_date_str   = assigned_date.strftime("%d-%b-%Y")
            days_since_assigned = (today - assigned_date).days
        else:
            assigned_date_str   = ""
            days_since_assigned = None

        records.append({
            "student_name":        str(row.get("student_name", "") or "Student"),
            "student_email":       email,
            "student_phone":       str(row.get("student_phone", "") or ""),
            "class_name":          str(row.get("class_name", "") or ""),
            "class_subject":       str(row.get("class_subject", "") or ""),
            "assessment_title":    str(row.get("assessment_title", "") or ""),
            "assessment_id":       str(row.get("assessment_id", "") or ""),
            "enrolled_count":      str(row.get("enrolled_count", "") or ""),
            "maximum_marks":       str(row.get("maximum_marks", "") or ""),
            "assigned_date":       assigned_date,
            "assigned_date_str":   assigned_date_str,
            "days_since_assigned": days_since_assigned,
            "deadline":            deadline,
            "deadline_str":        deadline.strftime("%d-%b-%Y"),
            "days_remaining":      days_remaining,
            "reminder_level":      reminder_level,
            "reminder_label":      reminder_label,
        })

    return records


def _dedup_records(records: list) -> list:
    """Keep only one row per (student_email, assessment_id)."""
    seen = set()
    out  = []
    for r in records:
        key = (r["student_email"].lower(), r["assessment_id"])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def consolidate_by_student(records: list) -> list:
    """
    Group per-assignment records into one reminder per student.
    Each consolidated reminder contains:
      - student_name / student_email / student_phone
      - assignments: list of dicts (class_name, class_subject,
        assessment_title, assessment_id, maximum_marks, deadline_str,
        days_remaining, reminder_level, reminder_label)
      - reminder_level / reminder_label = most urgent level across assignments
      - most_urgent_deadline = earliest deadline (for sorting / subject line)
    """
    grouped = defaultdict(list)
    name_map = {}
    phone_map = {}

    for r in records:
        key = r["student_email"].lower()
        grouped[key].append(r)
        # Keep the most complete name / phone we find for this student
        if key not in name_map or (not name_map[key] and r["student_name"]):
            name_map[key] = r["student_name"]
        if key not in phone_map or (not phone_map[key] and r["student_phone"]):
            phone_map[key] = r["student_phone"]

    consolidated = []
    for email_key, items in grouped.items():
        # Most urgent level across all assignments
        top_pri = max(LEVEL_PRIORITY[i["reminder_level"]] for i in items)
        top_level = PRIORITY_TO_LEVEL[top_pri]
        top_label = LEVEL_LABEL[top_level]

        # Earliest deadline drives subject line
        items_sorted = sorted(items, key=lambda x: (x["deadline"], x["assessment_title"]))
        earliest = items_sorted[0]

        assignments = []
        for it in items_sorted:
            assignments.append({
                "class_name":          it["class_name"],
                "class_subject":       it["class_subject"],
                "assessment_title":    it["assessment_title"],
                "assessment_id":       it["assessment_id"],
                "maximum_marks":       it["maximum_marks"],
                "assigned_date":       it.get("assigned_date"),
                "assigned_date_str":   it.get("assigned_date_str", ""),
                "days_since_assigned": it.get("days_since_assigned"),
                "deadline":            it["deadline"],
                "deadline_str":        it["deadline_str"],
                "days_remaining":      it["days_remaining"],
                "reminder_level":      it["reminder_level"],
                "reminder_label":      it["reminder_label"],
            })

        consolidated.append({
            "student_name":        name_map.get(email_key) or "Student",
            "student_email":       items[0]["student_email"],
            "student_phone":       phone_map.get(email_key, ""),
            "assignments":         assignments,
            "assignment_count":    len(assignments),
            "reminder_level":      top_level,
            "reminder_label":      top_label,
            "earliest_deadline":   earliest["deadline"],
            "earliest_deadline_str": earliest["deadline_str"],
            "min_days_remaining":  earliest["days_remaining"],
            # Primary batch / subject for grouping in summary (earliest deadline row)
            "primary_class_name":   earliest["class_name"],
            "primary_class_subject": earliest["class_subject"],
        })

    # Sort: most urgent first, then earliest deadline
    consolidated.sort(key=lambda x: (-LEVEL_PRIORITY[x["reminder_level"]],
                                      x["earliest_deadline"],
                                      x["student_name"].lower()))
    return consolidated


# ─────────────────────────────────────────────────────────────────────────────
#  EMAIL TEMPLATE — ONE email per student listing ALL pending assignments
# ─────────────────────────────────────────────────────────────────────────────

def _level_theme(level: str) -> dict:
    """Visual theme for each reminder level."""
    if level == "1st":
        return {
            "accent":    "#2E75B6",
            "bg_banner": "#E3F2FD",
            "row_bg":    "#F5F5F5",
            "row_border": "#DDDDDD",
            "deadline_bg": "#FFF3E0",
            "deadline_fg": "#E65100",
            "icon":      "📋",
            "heading":   "Assignment Submission Reminder",
            "intro_color": "#1F1F1F",
            "urgency_color": "#1565C0",
            "footer_link": "#2E75B6",
        }
    elif level == "2nd":
        return {
            "accent":    "#E65100",
            "bg_banner": "#FFF3E0",
            "row_bg":    "#FBE9E7",
            "row_border": "#FFCCBC",
            "deadline_bg": "#FF6D00",
            "deadline_fg": "#FFFFFF",
            "icon":      "⚠️",
            "heading":   "URGENT: Assignment Deadline Approaching",
            "intro_color": "#BF360C",
            "urgency_color": "#BF360C",
            "footer_link": "#E65100",
        }
    elif level == "final":
        return {
            "accent":    "#B71C1C",
            "bg_banner": "#FFEBEE",
            "row_bg":    "#FFCDD2",
            "row_border": "#EF9A9A",
            "deadline_bg": "#D50000",
            "deadline_fg": "#FFFFFF",
            "icon":      "🚨",
            "heading":   "FINAL WARNING: Assignment Due TODAY",
            "intro_color": "#B71C1C",
            "urgency_color": "#B71C1C",
            "footer_link": "#B71C1C",
        }
    else:  # missed — post-deadline follow-up (darker maroon, distinct from Final)
        return {
            "accent":    "#791F1F",
            "bg_banner": "#FCEBEB",
            "row_bg":    "#F7C1C1",
            "row_border": "#F09595",
            "deadline_bg": "#A32D2D",
            "deadline_fg": "#FFFFFF",
            "icon":      "⛔",
            "heading":   "MISSED DEADLINE: Assignment Overdue",
            "intro_color": "#501313",
            "urgency_color": "#791F1F",
            "footer_link": "#791F1F",
        }


def _deadline_cell_text(days: int, deadline_str: str) -> str:
    if days is not None and days < 0:
        n = abs(days)
        unit = "day" if n == 1 else "days"
        return f"{deadline_str} — OVERDUE by {n} {unit}"
    if days == 0:
        return f"{deadline_str} — TODAY!"
    if days == 1:
        return f"{deadline_str} — TOMORROW"
    return f"{deadline_str} ({days} days left)"


def _assigned_cell_text(days_ago, assigned_str: str) -> str:
    if not assigned_str:
        return "—"
    if days_ago is None:
        return assigned_str
    if days_ago == 0:
        return f"{assigned_str} (today)"
    if days_ago == 1:
        return f"{assigned_str} (1 day ago)"
    if days_ago < 0:
        return f"{assigned_str} ({abs(days_ago)} day(s) in future)"
    return f"{assigned_str} ({days_ago} days ago)"


def _build_email_html(s: dict) -> str:
    """
    Build consolidated HTML email for a student. 's' is a consolidated reminder
    (see consolidate_by_student) — may contain multiple assignments.
    """
    name       = s["student_name"]
    level      = s["reminder_level"]
    n          = s["assignment_count"]
    min_days   = s["min_days_remaining"]
    theme      = _level_theme(level)

    # ── Intro line (plural-aware) ────────────────────────────────────────────
    if level == "1st":
        if n == 1:
            intro = ("This is a friendly reminder that you have a pending "
                     "assignment submission as detailed below.")
        else:
            intro = (f"This is a friendly reminder that you currently have "
                     f"<strong>{n} pending assignment submissions</strong> "
                     f"as detailed below.")
        closing = ("<p>Please plan your time and submit before the respective "
                   "deadlines. Early submissions are always appreciated!</p>")
        urgency_bar = ""
    elif level == "2nd":
        if n == 1:
            intro = (f"Your assignment submission is due in just <u>{min_days} day(s)</u>. "
                     "This is your <strong>second reminder</strong>.")
        else:
            intro = (f"You have <strong>{n} pending assignment submissions</strong> — "
                     f"the earliest is due in just <u>{min_days} day(s)</u>. "
                     "This is your <strong>second reminder</strong>.")
        closing = ("<p>Do not delay further. Missing a deadline may affect your "
                   "grades and progress. Please submit your work as soon as possible.</p>")
        urgency_bar = (
            '<div style="background:#FF6D00;color:white;padding:10px 18px;'
            'border-radius:6px;font-weight:bold;font-size:15px;text-align:center;'
            f'margin:12px 0;">⏰ ONLY {min_days} DAY(S) LEFT — Submit NOW!</div>'
        )
    elif level == "final":
        if n == 1:
            intro = ("This is your <u>FINAL reminder</u>. Your assignment is due "
                     "<strong>TODAY</strong>.")
        else:
            intro = (f"This is your <u>FINAL reminder</u>. You have "
                     f"<strong>{n} pending assignment submissions</strong> — "
                     f"at least one is due <strong>TODAY</strong>.")
        closing = ('<p style="color:#B71C1C;">If you do not submit before the end '
                   'of today, your assignment will be marked as <strong>missed</strong>. '
                   'This will directly impact your professional progress.</p>'
                   '<p style="font-weight:bold;">Please submit your assignment '
                   '<u>right now</u>. No further reminders will be sent.</p>')
        urgency_bar = (
            '<div style="background:#D50000;color:white;padding:14px 18px;'
            'border-radius:6px;font-weight:bold;font-size:16px;text-align:center;'
            'margin:12px 0;letter-spacing:0.5px;">'
            '🚨 DEADLINE IS TODAY — SUBMIT IMMEDIATELY 🚨</div>'
        )
    else:  # missed — post-deadline follow-up (deadline already passed)
        if n == 1:
            intro = ("Our records show that the deadline for the assignment below "
                     "has now <strong>passed</strong>, and we have not yet received "
                     "your submission.")
        else:
            intro = (f"Our records show that the deadline for "
                     f"<strong>{n} of your assignment submissions</strong> has now "
                     f"<strong>passed</strong>, and we have not yet received your work.")
        closing = (
            '<p style="color:#501313;">We sent you <strong>three reminders</strong> '
            'before the due date — but the assignment was not submitted in time.</p>'
            '<p style="color:#501313;">This assignment has now been marked as '
            '<strong>missed</strong>. Do not reply.</p>'
        )
        urgency_bar = (
            '<div style="background:#A32D2D;color:white;padding:14px 18px;'
            'border-radius:6px;font-weight:bold;font-size:15px;text-align:center;'
            'margin:12px 0;letter-spacing:0.3px;">'
            '⛔ DEADLINE MISSED — Despite 3 reminders, this assignment is still '
            'unsubmitted</div>'
        )

    # ── Assignment table rows ────────────────────────────────────────────────
    row_bg    = theme["row_bg"]
    row_bd    = theme["row_border"]
    dl_bg     = theme["deadline_bg"]
    dl_fg     = theme["deadline_fg"]

    header_html = (
        f'<tr>'
        f'<th style="padding:10px 12px;background:{theme["accent"]};color:white;'
        f'border:1px solid {row_bd};text-align:center;font-size:13px;">#</th>'
        f'<th style="padding:10px 12px;background:{theme["accent"]};color:white;'
        f'border:1px solid {row_bd};text-align:left;font-size:13px;">Course</th>'
        f'<th style="padding:10px 12px;background:{theme["accent"]};color:white;'
        f'border:1px solid {row_bd};text-align:center;font-size:13px;">Assigned Date</th>'
        f'<th style="padding:10px 12px;background:{theme["accent"]};color:white;'
        f'border:1px solid {row_bd};text-align:center;font-size:13px;">Deadline</th>'
        f'<th style="padding:10px 12px;background:{theme["accent"]};color:white;'
        f'border:1px solid {row_bd};text-align:center;font-size:13px;">Max Marks</th>'
        f'<th style="padding:10px 12px;background:{theme["accent"]};color:white;'
        f'border:1px solid {row_bd};text-align:center;font-size:13px;">Status</th>'
        f'</tr>'
    )

    rows_html = ""
    for i, a in enumerate(s["assignments"], 1):
        dl_txt = _deadline_cell_text(a["days_remaining"], a["deadline_str"])
        asg_txt = _assigned_cell_text(a.get("days_since_assigned"),
                                      a.get("assigned_date_str", ""))
        mm     = a["maximum_marks"] or "—"
        # Course cell shows the class name on line 1 and the assignment title
        # on line 2 (smaller) so all information stays visible even though
        # the Assignment column was removed.
        course_cell = (
            f'<div style="font-weight:bold;">{a["class_name"] or "—"}</div>'
            f'<div style="font-size:11px;color:#555555;margin-top:2px;">'
            f'📝 {a["assessment_title"] or "—"}</div>'
        )
        rows_html += (
            f'<tr>'
            f'<td style="padding:10px 12px;background:{row_bg};border:1px solid {row_bd};'
            f'font-weight:bold;text-align:center;">{i}</td>'
            f'<td style="padding:10px 12px;background:{row_bg};border:1px solid {row_bd};">'
            f'{course_cell}</td>'
            f'<td style="padding:10px 12px;background:{row_bg};border:1px solid {row_bd};'
            f'text-align:center;">{asg_txt}</td>'
            f'<td style="padding:10px 12px;background:{dl_bg};border:1px solid {row_bd};'
            f'color:{dl_fg};font-weight:bold;text-align:center;">{dl_txt}</td>'
            f'<td style="padding:10px 12px;background:{row_bg};border:1px solid {row_bd};'
            f'text-align:center;">{mm}</td>'
            f'<td style="padding:10px 12px;background:{row_bg};border:1px solid {row_bd};'
            f'text-align:center;color:#B71C1C;font-weight:bold;">{a["reminder_label"]}</td>'
            f'</tr>'
        )

    table_html = (
        '<table style="border-collapse:collapse;margin:16px 0;width:100%;font-size:13px;">'
        + header_html + rows_html + '</table>'
    )

    # ── Signature block (always the same) ─────────────────────────────────────
    signature_html = f"""
      <div style="margin-top:22px;padding-top:14px;border-top:1px dashed #BDBDBD;">
        <p style="margin:0;font-size:14px;color:#1F1F1F;">Thanks &amp; Regards,</p>
        <p style="margin:4px 0 0;font-size:16px;font-weight:bold;color:{theme['accent']};">
          IntelliBI
        </p>
        <p style="margin:2px 0 0;font-size:13px;color:#1F1F1F;">
          📞 <a href="tel:+917020629915" style="color:#1F1F1F;text-decoration:none;">7020629915</a>
        </p>
      </div>"""

    # ── Assemble full HTML ────────────────────────────────────────────────────
    html = f"""\
<html>
<body style="font-family:'Segoe UI',Arial,sans-serif;color:#1F1F1F;margin:0;padding:0;">
  <div style="max-width:720px;margin:20px auto;border:1px solid #E0E0E0;border-radius:8px;overflow:hidden;">

    <!-- Header Banner -->
    <div style="background:{theme['accent']};padding:20px 24px;text-align:center;">
      <span style="font-size:28px;">{theme['icon']}</span>
      <h1 style="color:white;margin:8px 0 0;font-size:20px;">{theme['heading']}</h1>
    </div>

    <!-- Body -->
    <div style="padding:24px 28px;background:{theme['bg_banner']};">
      <p style="font-size:15px;">Dear <strong>{name}</strong>,</p>
      <p style="color:{theme['intro_color']};">{intro}</p>
      {urgency_bar}
      {table_html}
      {closing}
      {signature_html}
    </div>

    <!-- Footer -->
    <div style="background:#F5F5F5;padding:16px 24px;border-top:1px solid #E0E0E0;">
      <p style="font-size:12px;color:#888;margin:0;">
        This is an automated reminder from <strong>IntelliBI Innovations Technologies</strong>.<br>
        If you have already submitted your assignment(s), please disregard this email.<br>
        For any queries, contact us at
        <a href="mailto:intellibiinnovation0101@gmail.com" style="color:{theme['footer_link']};">
        intellibiinnovation0101@gmail.com</a>
      </p>
    </div>
  </div>
</body>
</html>"""
    return html


def _build_subject_line(s: dict) -> str:
    """Build email subject line for a consolidated student reminder."""
    level = s["reminder_level"]
    n     = s["assignment_count"]
    course = s["primary_class_name"]
    deadline = s["earliest_deadline_str"]
    name = s["student_name"]

    if n == 1:
        a = s["assignments"][0]
        title = a["assessment_title"]
        if level == "1st":
            return f"Reminder: Assignment \"{title}\" pending — Due {deadline} | {course}"
        elif level == "2nd":
            return f"URGENT: Assignment \"{title}\" due TOMORROW — {course}"
        elif level == "final":
            return f"FINAL: Assignment \"{title}\" due TODAY — Submit NOW! | {course}"
        else:  # missed
            return f"IntelliBI {name} --- {title} Missed"
    else:
        if level == "1st":
            return f"Reminder: {n} Assignments pending — Earliest due {deadline}"
        elif level == "2nd":
            return f"URGENT: {n} Assignments pending — Earliest due TOMORROW ({deadline})"
        elif level == "final":
            return f"FINAL: {n} Assignments pending — Deadline TODAY ({deadline})"
        else:  # missed
            return f"IntelliBI {name} --- {n} Assignments Missed"


# ─────────────────────────────────────────────────────────────────────────────
#  EMAIL SENDER
# ─────────────────────────────────────────────────────────────────────────────

def send_reminder_email(s: dict, dry_run: bool = False) -> bool:
    """Send one consolidated email to a student. Returns True on success.

    NOTE: Student reminder emails are sent TO-only (no Cc / no Bcc) so that
    each send counts as exactly ONE recipient against Gmail's daily cap and
    anti-spam quota. Staff receive a single consolidated summary email at
    the end of the run instead of being cc'd on every student message.
    """
    to_email = s["student_email"]
    subject  = _build_subject_line(s)
    html     = _build_email_html(s)

    if dry_run:
        # Print email preview: header, subject, body
        sep = "─" * 76
        print(f"\n  {sep}")
        print(f"  [DRY-RUN] EMAIL PREVIEW")
        print(f"  {sep}")
        print(f"  From    : {GMAIL_SENDER}")
        print(f"  To      : {to_email}")
        print(f"  Cc      : (none — staff get a single summary email after the run)")
        print(f"  Subject : {subject}")
        print(f"  Level   : {s['reminder_label']}  ({s['assignment_count']} assignment(s))")
        print(f"  Student : {s['student_name']}")
        print(f"  {sep}")
        print(f"  --- Pending assignments ---")
        for i, a in enumerate(s["assignments"], 1):
            print(f"    {i}. {a['assessment_title']}  |  {a['class_name']}  "
                  f"|  {a['class_subject']}  |  Max: {a['maximum_marks'] or '—'}  "
                  f"|  Deadline: {a['deadline_str']} ({a['reminder_label']})")
        print(f"  {sep}")
        print(f"  --- HTML body (full) ---")
        for ln in html.splitlines():
            print(f"  {ln}")
        print(f"  {sep}\n")
        return True

    try:
        msg             = MIMEMultipart()
        msg["From"]     = GMAIL_SENDER
        msg["To"]       = to_email
        # Intentionally NO Cc / Bcc — student emails are 1-recipient sends.
        msg["Subject"]  = subject
        msg.attach(MIMEText(html, "html"))

        envelope_recipients = [to_email]
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_SENDER, GMAIL_APP_PASS)
            server.sendmail(GMAIL_SENDER, envelope_recipients, msg.as_string())
        return True
    except Exception as e:
        print(f"    [ERROR] Failed to send to {to_email}: {e}")
        return False


def _build_staff_summary_html(consolidated: list,
                              sent_ok: list,
                              sent_fail: list,
                              pdfs: list,
                              today: date) -> str:
    """Render the HTML body for the end-of-run staff summary email."""
    total_students = len(consolidated)
    ok_count   = len(sent_ok)
    fail_count = len(sent_fail)

    # Per-level (by assignment, matches what students actually see)
    lvl_counts = {lv: 0 for lv in LEVEL_ORDER}
    total_assignments = 0
    for s in consolidated:
        for a in s["assignments"]:
            lvl_counts[a["reminder_level"]] += 1
            total_assignments += 1

    # Per-course / batch breakdown
    batch_rows = defaultdict(lambda: {"students": 0, "1st": 0, "2nd": 0,
                                      "final": 0, "missed": 0})
    for s in consolidated:
        for a in s["assignments"]:
            key = a["class_name"] or "(Unknown Course)"
            batch_rows[key]["students"] += 1
            batch_rows[key][a["reminder_level"]] += 1

    # Failure list
    fail_items_html = ""
    if sent_fail:
        li_html = "".join(
            f"<li><b>{(s.get('student_name') or '—')}</b> "
            f"&lt;{s.get('student_email') or '—'}&gt; — "
            f"{s.get('reminder_label') or '—'} "
            f"({s.get('assignment_count', 0)} assignment(s))</li>"
            for s in sent_fail
        )
        fail_items_html = (
            "<h3 style='color:#B71C1C;margin-top:24px;'>❌ Failed sends "
            f"({len(sent_fail)})</h3>"
            f"<ul style='font-size:13px;color:#333;'>{li_html}</ul>"
        )

    # Per-batch table rows
    batch_tbl_rows = ""
    for course in sorted(batch_rows.keys()):
        r = batch_rows[course]
        batch_tbl_rows += (
            "<tr>"
            f"<td style='padding:6px 10px;border:1px solid #CFD8DC;'>{course}</td>"
            f"<td style='padding:6px 10px;border:1px solid #CFD8DC;text-align:center;'>{r['students']}</td>"
            f"<td style='padding:6px 10px;border:1px solid #CFD8DC;text-align:center;color:#1565C0;'>{r['1st']}</td>"
            f"<td style='padding:6px 10px;border:1px solid #CFD8DC;text-align:center;color:#E65100;'>{r['2nd']}</td>"
            f"<td style='padding:6px 10px;border:1px solid #CFD8DC;text-align:center;color:#B71C1C;'>{r['final']}</td>"
            f"<td style='padding:6px 10px;border:1px solid #CFD8DC;text-align:center;color:#791F1F;'>{r['missed']}</td>"
            "</tr>"
        )

    # PDF list
    pdf_list_html = ""
    if pdfs:
        li_html = "".join(
            f"<li><b>{p['course']}</b> — {p['assignment']}  "
            f"<span style='color:#666;'>({p['student_count']} student(s); "
            f"{len(p['pdf_bytes']):,} bytes)</span></li>"
            for p in pdfs
        )
        pdf_list_html = (
            "<h3 style='color:#1F3864;margin-top:24px;'>📎 PDF reports attached</h3>"
            f"<ul style='font-size:13px;color:#333;'>{li_html}</ul>"
        )

    html = f"""\
<html><body style="font-family:Segoe UI,Arial,sans-serif;background:#ECEFF1;padding:24px;">
  <div style="max-width:760px;margin:0 auto;background:#FFFFFF;border-radius:8px;
              padding:28px 32px;box-shadow:0 2px 8px rgba(0,0,0,0.06);">

    <h2 style="color:#1F3864;margin:0 0 4px 0;">
      IntelliBI — Assignment Reminder Run Summary
    </h2>
    <p style="color:#555;margin:0 0 18px 0;font-size:13px;">
      {today.strftime('%A, %d %B %Y')}  &nbsp;|&nbsp;
      Generated {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}
    </p>

    <table cellspacing="0" cellpadding="0" style="border-collapse:collapse;
           width:100%;font-size:13px;margin-bottom:16px;">
      <tr>
        <td style="padding:10px 14px;background:#E8F5E9;border-left:4px solid #2E7D32;">
          <b>✓ Sent successfully:</b> {ok_count} / {total_students}
        </td>
      </tr>
      <tr>
        <td style="padding:10px 14px;background:#FFEBEE;border-left:4px solid #C62828;">
          <b>✗ Failed:</b> {fail_count} / {total_students}
        </td>
      </tr>
      <tr>
        <td style="padding:10px 14px;background:#E3F2FD;border-left:4px solid #1565C0;">
          <b>Total pending assignments:</b> {total_assignments}
          &nbsp;|&nbsp; 1st: <b>{lvl_counts['1st']}</b>
          &nbsp;|&nbsp; 2nd: <b style="color:#E65100;">{lvl_counts['2nd']}</b>
          &nbsp;|&nbsp; Final: <b style="color:#B71C1C;">{lvl_counts['final']}</b>
          &nbsp;|&nbsp; Missed: <b style="color:#791F1F;">{lvl_counts['missed']}</b>
        </td>
      </tr>
    </table>

    <h3 style="color:#1F3864;margin-top:24px;">Per-course breakdown</h3>
    <table cellspacing="0" cellpadding="0" style="border-collapse:collapse;
           width:100%;font-size:13px;">
      <thead>
        <tr style="background:#1F3864;color:#FFFFFF;">
          <th style="padding:8px 10px;border:1px solid #1F3864;text-align:left;">Course / Batch</th>
          <th style="padding:8px 10px;border:1px solid #1F3864;">Pending</th>
          <th style="padding:8px 10px;border:1px solid #1F3864;">1st</th>
          <th style="padding:8px 10px;border:1px solid #1F3864;">2nd</th>
          <th style="padding:8px 10px;border:1px solid #1F3864;">Final</th>
          <th style="padding:8px 10px;border:1px solid #1F3864;">Missed</th>
        </tr>
      </thead>
      <tbody>
        {batch_tbl_rows}
      </tbody>
    </table>

    {pdf_list_html}
    {fail_items_html}

    <hr style="border:none;border-top:1px solid #E0E0E0;margin:24px 0 14px 0;">
    <p style="color:#555;font-size:12px;margin:0;">
      This is an automated summary of the daily assignment reminder run.<br>
      Student notifications were sent individually (one recipient per message)
      to comply with email provider limits.
    </p>
    <p style="color:#1F3864;font-size:13px;margin:14px 0 0 0;">
      Thanks &amp; Regards,<br>
      <b>IntelliBI</b><br>
      📞 <a href="tel:+917020629915" style="color:#1F3864;text-decoration:none;">7020629915</a>
    </p>
  </div>
</body></html>"""
    return html


def send_staff_summary_email(consolidated: list,
                             sent_ok: list,
                             sent_fail: list,
                             pdfs: list,
                             today: date,
                             dry_run: bool = False,
                             extra_pdfs: list = None) -> bool:
    """
    Send ONE consolidated summary email to the staff distro at the end of the
    run. Replaces the per-student Cc/Bcc so each student email is a 1-recipient
    send (keeps us under Gmail's daily quota & anti-spam thresholds).

    `extra_pdfs` — additional PDF dicts (same shape as `pdfs`) to attach to
    the summary email but NOT upload to Drive. Used for the standalone Email
    Status Report so the staff get it in their inbox without it polluting
    the Google Drive reports folder.
    """
    if not STAFF_SUMMARY_RECIPIENTS:
        print("[Summary] No STAFF_SUMMARY_RECIPIENTS configured — skipping.")
        return False

    extra_pdfs = extra_pdfs or []
    all_attachments = list(pdfs or []) + list(extra_pdfs)

    total   = len(consolidated)
    ok      = len(sent_ok)
    fail    = len(sent_fail)
    subject = (f"[IntelliBI] Reminder Run — {today.strftime('%d-%b-%Y')} "
               f"• {ok}/{total} sent"
               + (f" • {fail} failed" if fail else ""))

    html = _build_staff_summary_html(consolidated, sent_ok, sent_fail,
                                     all_attachments, today)

    if dry_run:
        # Local preview writing is DISABLED (Drive-only policy); just print a
        # short header. No HTML file is written to disk.
        preview_path = ""

        # Short, highly-visible box (no 200-line HTML dump)
        sep = "═" * 76
        print(f"\n{sep}")
        print(f"  📧  [DRY-RUN] STAFF SUMMARY EMAIL PREVIEW")
        print(f"{sep}")
        print(f"  From      : {GMAIL_SENDER}")
        print(f"  To        : {', '.join(STAFF_SUMMARY_RECIPIENTS)}")
        print(f"  Subject   : {subject}")
        print(f"  Attach    : {len(all_attachments)} PDF(s)")
        for p in all_attachments:
            print(f"              • {p['filename']}  ({len(p['pdf_bytes']):,} bytes)")
        print(f"  Students  : {len(consolidated)} total  "
              f"(OK: {len(sent_ok)}, Failed: {len(sent_fail)})")
        if preview_path:
            print(f"  HTML file : {preview_path}")
            print(f"              ↑  Open this in a browser to see the "
                  f"rendered summary email.")
        print(f"{sep}\n")
        return True

    try:
        msg            = MIMEMultipart()
        msg["From"]    = GMAIL_SENDER
        msg["To"]      = ", ".join(STAFF_SUMMARY_RECIPIENTS)
        msg["Subject"] = subject
        msg.attach(MIMEText(html, "html"))

        # Attach PDFs (reminder PDFs + extra status PDF, if any)
        for p in all_attachments:
            try:
                att = MIMEApplication(p["pdf_bytes"], _subtype="pdf")
                att.add_header("Content-Disposition", "attachment",
                               filename=p["filename"])
                msg.attach(att)
            except Exception as ae:
                print(f"    [Summary] ⚠ Failed to attach {p.get('filename','?')}: {ae}")

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_SENDER, GMAIL_APP_PASS)
            server.sendmail(GMAIL_SENDER, STAFF_SUMMARY_RECIPIENTS, msg.as_string())
        print(f"[Summary] ✓ Staff summary email sent to "
              f"{len(STAFF_SUMMARY_RECIPIENTS)} recipient(s): "
              f"{', '.join(STAFF_SUMMARY_RECIPIENTS)}")
        return True
    except Exception as e:
        print(f"[Summary] ⚠ Failed to send staff summary: {e}")
        return False


def save_html_preview(consolidated: list, today: date) -> str:
    """Local HTML preview writing is DISABLED (Drive-only policy).

    Returns "" so the dry-run flow simply skips the local file. The on-screen
    dry-run summary still prints what would be sent.
    """
    return ""

    # --- Disabled: previously wrote a local HTML preview file ---
    try:
        out_dir = os.path.join(_SCRIPT_DIR, "_email_previews")
        os.makedirs(out_dir, exist_ok=True)
        fname = f"EmailPreview_{today.strftime('%Y-%m-%d')}.html"
        fpath = os.path.join(out_dir, fname)

        parts = [
            "<html><head><meta charset='utf-8'>",
            f"<title>IntelliBI Reminder Preview — {today.strftime('%d-%b-%Y')}</title>",
            "</head><body style='font-family:Segoe UI,Arial,sans-serif;background:#ECEFF1;padding:20px;'>",
            f"<h1 style='color:#1F3864;'>📧 IntelliBI Reminder Email Preview — "
            f"{today.strftime('%d-%b-%Y')}</h1>",
            f"<p><strong>{len(consolidated)}</strong> consolidated email(s) "
            f"would be sent.</p><hr>",
        ]
        for i, s in enumerate(consolidated, 1):
            parts.append(
                f"<h2 style='color:#1F3864;'>#{i} — {s['student_name']} "
                f"&lt;{s['student_email']}&gt; — {s['reminder_label']} "
                f"({s['assignment_count']} assignment(s))</h2>"
            )
            parts.append(f"<p><em>Subject:</em> {_build_subject_line(s)}</p>")
            parts.append(_build_email_html(s))
            parts.append("<hr>")
        parts.append("</body></html>")
        with open(fpath, "w", encoding="utf-8") as f:
            f.write("\n".join(parts))
        return fpath
    except Exception as e:
        print(f"  [Preview] ⚠ Could not write HTML preview: {e}")
        return ""


# ─────────────────────────────────────────────────────────────────────────────
#  PER-ASSIGNMENT PDF GENERATION
#    One PDF per (Course, Assignment) pair.
#    Filename: Assignment_[CourseName]_[AssignmentsName]_Reminders_YYYY-MM-DD.pdf
# ─────────────────────────────────────────────────────────────────────────────

import re

def _sanitize_filename_part(s: str, max_len: int = 50) -> str:
    """Make a string safe for use inside a filename (Windows + *nix friendly)."""
    if not s:
        return "Unknown"
    # Replace path-unsafe chars with space, then collapse whitespace → underscore
    cleaned = re.sub(r'[\\/:*?"<>|\t\r\n]+', " ", str(s))
    # Drop other punctuation that looks ugly in filenames
    cleaned = re.sub(r"[,;()\[\]{}!@#$%^&+=`~']", " ", cleaned)
    # Collapse whitespace
    cleaned = re.sub(r"\s+", "_", cleaned.strip())
    # Remove any residual non-ascii (keep letters / digits / _ / - / .)
    cleaned = re.sub(r"[^A-Za-z0-9_\-.]+", "", cleaned)
    if not cleaned:
        return "Unknown"
    return cleaned[:max_len].rstrip("_-.") or "Unknown"


def _group_rows_by_course_assignment(consolidated: list) -> dict:
    """
    Expand consolidated (per-student) reminders into a mapping:
        (class_name, assessment_title, assessment_id) → list of row dicts
    Each row dict represents ONE student pending for that assignment.
    """
    groups = defaultdict(list)
    for s in consolidated:
        for a in s["assignments"]:
            key = (
                a["class_name"] or "(Unknown Course)",
                a["assessment_title"] or "(Untitled Assignment)",
                a["assessment_id"] or "",
            )
            groups[key].append({
                "student_name":   s["student_name"],
                "student_email":  s["student_email"],
                "student_phone":  s["student_phone"],
                "class_name":     a["class_name"],
                "class_subject":  a["class_subject"],
                "assessment_title": a["assessment_title"],
                "assessment_id":  a["assessment_id"],
                "maximum_marks":  a["maximum_marks"],
                "assigned_date_str":   a.get("assigned_date_str", ""),
                "days_since_assigned": a.get("days_since_assigned"),
                "deadline_str":   a["deadline_str"],
                "deadline":       a["deadline"],
                "days_remaining": a["days_remaining"],
                "reminder_level": a["reminder_level"],
                "reminder_label": a["reminder_label"],
            })
    return groups


def _build_single_assignment_pdf(course_name: str,
                                 assignment_title: str,
                                 rows: list,
                                 today: date) -> bytes:
    """Build a single PDF for one (course, assignment) pair."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units     import mm
    from reportlab.lib           import colors
    from reportlab.platypus      import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=12*mm, bottomMargin=12*mm,
        title=f"IntelliBI Reminder — {course_name} — {assignment_title}",
        author="IntelliBI Innovations Technologies",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"], fontSize=15,
        textColor=colors.HexColor("#1F3864"), spaceAfter=4, alignment=1)
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"], fontSize=10,
        textColor=colors.HexColor("#555555"), alignment=1, spaceAfter=8)
    small_style = ParagraphStyle(
        "Small", parent=styles["Normal"], fontSize=9, leading=11)
    cell_style = ParagraphStyle(
        "Cell", parent=styles["Normal"], fontSize=8, leading=10)
    header_cell_style = ParagraphStyle(
        "HeaderCell", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=colors.white, fontName="Helvetica-Bold", alignment=1)

    story = []

    # ── Header ────────────────────────────────────────────────────────────────
    story.append(Paragraph(
        "IntelliBI Assignment Submission — Reminder Report", title_style))
    story.append(Paragraph(
        f"Date: {today.strftime('%d-%b-%Y (%A)')}  |  "
        f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}",
        sub_style))

    # ── Assignment meta table ─────────────────────────────────────────────────
    # Derive details from the first row (all rows in this PDF share them)
    r0 = rows[0]
    assigned_meta_txt = _assigned_cell_text(
        r0.get("days_since_assigned"),
        r0.get("assigned_date_str", ""),
    ) or "—"
    meta_data = [
        ["Course / Batch",   course_name],
        ["Assignment",       assignment_title],
        ["Class Duration",   r0["class_subject"] or "—"],
        ["Maximum Marks",    r0["maximum_marks"] or "—"],
        ["Assigned Date",    assigned_meta_txt],
        ["Deadline",         r0["deadline_str"] or "—"],
        ["Pending Students", str(len(rows))],
    ]
    meta_table = Table(meta_data, colWidths=[45*mm, 210*mm], hAlign="LEFT")
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#1F3864")),
        ("TEXTCOLOR",  (0, 0), (0, -1), colors.white),
        ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
        ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#F5F7FA")),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.grey),
        ("LEFTPADDING",(0, 0), (-1, -1), 8),
        ("RIGHTPADDING",(0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 8))

    # ── Reminder level breakdown ──────────────────────────────────────────────
    bcount = {lv: 0 for lv in LEVEL_ORDER}
    for r in rows:
        bcount[r["reminder_level"]] += 1
    story.append(Paragraph(
        f"<b>Reminder Breakdown:</b> &nbsp; "
        f"1st Reminder: <b>{bcount['1st']}</b> &nbsp;|&nbsp; "
        f"2nd Reminder: <b>{bcount['2nd']}</b> &nbsp;|&nbsp; "
        f"Final Reminder: <b>{bcount['final']}</b> &nbsp;|&nbsp; "
        f"Missed Deadline: <b>{bcount['missed']}</b>",
        small_style))
    story.append(Spacer(1, 6))

    # ── Students table ────────────────────────────────────────────────────────
    # Sort students by urgency (Final → 2nd → 1st), then by name
    rows_sorted = sorted(
        rows,
        key=lambda r: (-LEVEL_PRIORITY[r["reminder_level"]],
                       (r["student_name"] or "").lower()),
    )

    header = [
        Paragraph("#", header_cell_style),
        Paragraph("Student Name", header_cell_style),
        Paragraph("Email", header_cell_style),
        Paragraph("Phone", header_cell_style),
        Paragraph("Deadline", header_cell_style),
        Paragraph("Reminder Number", header_cell_style),
    ]
    data = [header]
    row_bgs = []

    for i, r in enumerate(rows_sorted, 1):
        data.append([
            Paragraph(str(i), cell_style),
            Paragraph(r["student_name"] or "—", cell_style),
            Paragraph(r["student_email"] or "—", cell_style),
            Paragraph(r["student_phone"] or "—", cell_style),
            Paragraph(r["deadline_str"] or "—", cell_style),
            Paragraph(r["reminder_label"] or "—", cell_style),
        ])
        if r["reminder_level"] == "missed":
            row_bgs.append(colors.HexColor("#F7C1C1"))
        elif r["reminder_level"] == "final":
            row_bgs.append(colors.HexColor("#FFEBEE"))
        elif r["reminder_level"] == "2nd":
            row_bgs.append(colors.HexColor("#FFF3E0"))
        else:
            row_bgs.append(colors.HexColor("#E3F2FD"))

    # Landscape A4 usable width ≈ 273mm
    # #  Name  Email  Phone  Deadline  Reminder
    col_widths = [10*mm, 55*mm, 80*mm, 35*mm, 40*mm, 40*mm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")

    ts_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",      (0, 0), (0, -1), "CENTER"),
        ("ALIGN",      (4, 0), (5, -1), "CENTER"),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.grey),
        ("LEFTPADDING",(0, 0), (-1, -1), 5),
        ("RIGHTPADDING",(0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]
    for ri, bg in enumerate(row_bgs, start=1):
        ts_cmds.append(("BACKGROUND", (0, ri), (-1, ri), bg))

    tbl.setStyle(TableStyle(ts_cmds))
    story.append(tbl)
    story.append(Spacer(1, 8))

    # Footer note
    story.append(Paragraph(
        "<font size=7 color='#888888'>Automated report — "
        "IntelliBI Innovations Technologies</font>",
        styles["Normal"]))

    doc.build(story)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


def generate_assignment_pdfs(consolidated: list, today: date) -> list:
    """
    Generate one PDF per (course, assignment) pair.
    Returns a list of dicts:
        {"filename": ..., "pdf_bytes": ..., "course": ..., "assignment": ...,
         "student_count": ..., "reminder_counts": {...}}
    """
    groups = _group_rows_by_course_assignment(consolidated)
    if not groups:
        return []

    outputs = []
    date_str = today.strftime("%Y-%m-%d")

    # Sort keys for stable, pleasant order
    for (course, assignment, _aid) in sorted(groups.keys()):
        rows = groups[(course, assignment, _aid)]
        pdf_bytes = _build_single_assignment_pdf(course, assignment, rows, today)

        course_safe     = _sanitize_filename_part(course, max_len=40)
        assignment_safe = _sanitize_filename_part(assignment, max_len=60)
        filename = f"Assignment_{course_safe}_{assignment_safe}_Reminders_{date_str}.pdf"

        rcnt = {lv: 0 for lv in LEVEL_ORDER}
        for r in rows:
            rcnt[r["reminder_level"]] += 1

        outputs.append({
            "filename":        filename,
            "pdf_bytes":       pdf_bytes,
            "course":          course,
            "assignment":      assignment,
            "student_count":   len(rows),
            "reminder_counts": rcnt,
        })
    return outputs


# ─────────────────────────────────────────────────────────────────────────────
#  EMAIL STATUS PDF REPORT
#  A SEPARATE, standalone PDF containing the email delivery status for every
#  student in this run. Attached to the staff summary email only (not uploaded
#  to Drive). The existing per-assignment reminder PDFs are untouched.
# ─────────────────────────────────────────────────────────────────────────────

def build_email_status_pdf(consolidated: list,
                           email_status: dict,
                           sent_ok: list,
                           sent_fail: list,
                           today: date,
                           mode_label: str = "LIVE") -> bytes:
    """
    Build a single PDF showing the per-student email delivery status for this
    run. Returns the PDF bytes.

    Rows are grouped by course and sorted by urgency (Final → 2nd → 1st),
    then by student name, mirroring the reminder PDFs for visual consistency.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units     import mm
    from reportlab.lib           import colors
    from reportlab.platypus      import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, KeepTogether)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=12*mm, bottomMargin=12*mm,
        title=f"IntelliBI Email Delivery Status — {today.strftime('%d-%b-%Y')}",
        author="IntelliBI Innovations Technologies",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"], fontSize=15,
        textColor=colors.HexColor("#1F3864"), spaceAfter=4, alignment=1)
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"], fontSize=10,
        textColor=colors.HexColor("#555555"), alignment=1, spaceAfter=8)
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"], fontSize=11,
        textColor=colors.HexColor("#1F3864"), spaceBefore=8, spaceAfter=4)
    small_style = ParagraphStyle(
        "Small", parent=styles["Normal"], fontSize=9, leading=11)
    cell_style = ParagraphStyle(
        "Cell", parent=styles["Normal"], fontSize=8, leading=10)
    header_cell_style = ParagraphStyle(
        "HeaderCell", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=colors.white, fontName="Helvetica-Bold", alignment=1)

    # ── Status styles ─────────────────────────────────────────────────────────
    _STATUS_STYLE = {
        "Sent":    ("✓ Sent",    "#2E7D32"),
        "Failed":  ("✗ Failed",  "#C62828"),
        "Preview": ("Preview",   "#616161"),
        "Skipped": ("Skipped",   "#E65100"),
    }

    def _fmt_status_cell(status_str: str):
        label, hex_col = _STATUS_STYLE.get(
            status_str, (status_str or "—", "#616161"))
        return Paragraph(
            f'<font color="{hex_col}"><b>{label}</b></font>',
            cell_style,
        )

    story = []

    # ── Header ────────────────────────────────────────────────────────────────
    story.append(Paragraph(
        "IntelliBI — Email Delivery Status Report", title_style))
    story.append(Paragraph(
        f"Date: {today.strftime('%d-%b-%Y (%A)')}  |  "
        f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}  |  "
        f"Mode: {mode_label}",
        sub_style))

    # ── Summary stats box ─────────────────────────────────────────────────────
    total    = len(consolidated)
    n_ok     = sum(1 for s in consolidated
                   if email_status.get(s["student_email"]) == "Sent")
    n_fail   = sum(1 for s in consolidated
                   if email_status.get(s["student_email"]) == "Failed")
    n_prev   = sum(1 for s in consolidated
                   if email_status.get(s["student_email"]) == "Preview")
    n_skip   = sum(1 for s in consolidated
                   if email_status.get(s["student_email"]) == "Skipped")

    # Per-level counts (by assignment, matches reminder PDFs)
    lvl_counts = {lv: 0 for lv in LEVEL_ORDER}
    for s in consolidated:
        for a in s["assignments"]:
            lvl_counts[a["reminder_level"]] += 1

    stats_data = [
        [
            Paragraph("Total Students", header_cell_style),
            Paragraph("✓ Sent",         header_cell_style),
            Paragraph("✗ Failed",       header_cell_style),
            Paragraph("Preview",        header_cell_style),
            Paragraph("Skipped",        header_cell_style),
            Paragraph("1st Rem.",       header_cell_style),
            Paragraph("2nd Rem.",       header_cell_style),
            Paragraph("Final Rem.",     header_cell_style),
            Paragraph("Missed",         header_cell_style),
        ],
        [
            Paragraph(f"<b>{total}</b>", cell_style),
            Paragraph(f'<font color="#2E7D32"><b>{n_ok}</b></font>', cell_style),
            Paragraph(f'<font color="#C62828"><b>{n_fail}</b></font>', cell_style),
            Paragraph(f'<font color="#616161"><b>{n_prev}</b></font>', cell_style),
            Paragraph(f'<font color="#E65100"><b>{n_skip}</b></font>', cell_style),
            Paragraph(f'<font color="#1565C0"><b>{lvl_counts["1st"]}</b></font>', cell_style),
            Paragraph(f'<font color="#E65100"><b>{lvl_counts["2nd"]}</b></font>', cell_style),
            Paragraph(f'<font color="#B71C1C"><b>{lvl_counts["final"]}</b></font>', cell_style),
            Paragraph(f'<font color="#791F1F"><b>{lvl_counts["missed"]}</b></font>', cell_style),
        ],
    ]
    stats_col = 30*mm
    stats_table = Table(
        stats_data,
        colWidths=[stats_col]*9,
        hAlign="LEFT",
    )
    stats_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.grey),
        ("LEFTPADDING",(0, 0), (-1, -1), 4),
        ("RIGHTPADDING",(0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F5F7FA")),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 10))

    # ── Per-course status tables ──────────────────────────────────────────────
    # Group consolidated students by primary course (class_name of the most
    # urgent assignment). Each student appears exactly once.
    per_course = defaultdict(list)
    for s in consolidated:
        primary_course = (s.get("assignments") or [{}])[0].get("class_name") \
            or "(Unknown Course)"
        per_course[primary_course].append(s)

    status_of = lambda s: email_status.get(s["student_email"], "—")

    for course in sorted(per_course.keys()):
        students = per_course[course]

        # Sort: urgency → name
        students_sorted = sorted(
            students,
            key=lambda s: (
                -LEVEL_PRIORITY.get(s["reminder_level"], 0),
                (s.get("student_name") or "").lower(),
            ),
        )

        # Section header with per-course counts
        c_ok   = sum(1 for s in students_sorted if status_of(s) == "Sent")
        c_fail = sum(1 for s in students_sorted if status_of(s) == "Failed")
        story.append(Paragraph(
            f"{course}  "
            f"<font size=9 color='#555555'>"
            f"— {len(students_sorted)} student(s)"
            f"  |  ✓ {c_ok}   ✗ {c_fail}"
            f"</font>",
            section_style))

        header = [
            Paragraph("#", header_cell_style),
            Paragraph("Student Name", header_cell_style),
            Paragraph("Email", header_cell_style),
            Paragraph("Phone", header_cell_style),
            Paragraph("Reminder Number", header_cell_style),
            Paragraph("Pending", header_cell_style),
            Paragraph("Email Status", header_cell_style),
        ]
        data = [header]
        row_bgs = []
        for i, s in enumerate(students_sorted, 1):
            data.append([
                Paragraph(str(i), cell_style),
                Paragraph(s.get("student_name") or "—", cell_style),
                Paragraph(s.get("student_email") or "—", cell_style),
                Paragraph(s.get("student_phone") or "—", cell_style),
                Paragraph(s.get("reminder_label") or "—", cell_style),
                Paragraph(str(s.get("assignment_count") or 0), cell_style),
                _fmt_status_cell(status_of(s)),
            ])
            if s["reminder_level"] == "missed":
                row_bgs.append(colors.HexColor("#F7C1C1"))
            elif s["reminder_level"] == "final":
                row_bgs.append(colors.HexColor("#FFEBEE"))
            elif s["reminder_level"] == "2nd":
                row_bgs.append(colors.HexColor("#FFF3E0"))
            else:
                row_bgs.append(colors.HexColor("#E3F2FD"))

        # Landscape A4 usable width ≈ 273mm
        # #  Name  Email  Phone  Reminder  Pending  Status
        col_widths = [10*mm, 52*mm, 78*mm, 30*mm, 38*mm, 22*mm, 38*mm]
        tbl = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")

        ts_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN",      (0, 0), (0, -1), "CENTER"),
            ("ALIGN",      (4, 0), (6, -1), "CENTER"),
            ("GRID",       (0, 0), (-1, -1), 0.4, colors.grey),
            ("LEFTPADDING",(0, 0), (-1, -1), 5),
            ("RIGHTPADDING",(0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ]
        for ri, bg in enumerate(row_bgs, start=1):
            ts_cmds.append(("BACKGROUND", (0, ri), (-1, ri), bg))

        tbl.setStyle(TableStyle(ts_cmds))
        story.append(tbl)
        story.append(Spacer(1, 6))

    # ── Footer note ───────────────────────────────────────────────────────────
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "<font size=7 color='#888888'>Automated email delivery status report — "
        "IntelliBI Innovations Technologies</font>",
        styles["Normal"]))

    doc.build(story)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


def generate_email_status_pdf_entry(consolidated: list,
                                    email_status: dict,
                                    sent_ok: list,
                                    sent_fail: list,
                                    today: date,
                                    mode_label: str = "LIVE") -> dict:
    """
    Build the email-status PDF and return a `{filename, pdf_bytes, ...}` dict
    in the same shape as `generate_assignment_pdfs()` so it can share the
    attachment / logging code paths.
    """
    pdf_bytes = build_email_status_pdf(
        consolidated, email_status, sent_ok, sent_fail, today, mode_label)
    date_str = today.strftime("%Y-%m-%d")
    return {
        "filename":      f"EmailStatus_Report_{date_str}.pdf",
        "pdf_bytes":     pdf_bytes,
        "course":        "(all courses)",
        "assignment":    "Email Status Report",
        "student_count": len(consolidated),
        "reminder_counts": {lv: 0 for lv in LEVEL_ORDER},
    }


# ─────────────────────────────────────────────────────────────────────────────
#  GOOGLE DRIVE UPLOAD (for PDF)
# ─────────────────────────────────────────────────────────────────────────────

def _get_service_account_email() -> str:
    """Read the service account email from the JSON key file (for diagnostics)."""
    try:
        import json
        with open(SERVICE_ACCOUNT_FILE, "r", encoding="utf-8") as fh:
            return json.load(fh).get("client_email", "")
    except Exception:
        return ""


def _build_drive_client(impersonate_email: str = ""):
    """
    Build a Drive v3 client using the service account.
    If impersonate_email is given, try domain-wide delegation;
    fall back to non-delegated client on failure.
    Returns (drive_client, auth_mode_str).
    """
    from googleapiclient.discovery import build as gdrive_build
    from google.oauth2             import service_account as _sa

    base_creds = _sa.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=["https://www.googleapis.com/auth/drive"],
    )

    if impersonate_email:
        try:
            delegated = base_creds.with_subject(impersonate_email)
            drive = gdrive_build("drive", "v3", credentials=delegated,
                                 cache_discovery=False)
            return drive, f"impersonating {impersonate_email}"
        except Exception as e:
            print(f"[Drive] ⚠ Impersonation of {impersonate_email} failed: {e}")
            print(f"[Drive]   Falling back to direct service-account auth ...")

    drive = gdrive_build("drive", "v3", credentials=base_creds,
                         cache_discovery=False)
    return drive, "direct service account"


def _drive_folder_probe(drive, folder_id: str) -> dict:
    """Probe the target folder for accessibility. Returns metadata or {}. """
    try:
        meta = drive.files().get(
            fileId=folder_id,
            fields="id,name,mimeType,driveId,webViewLink,capabilities",
            supportsAllDrives=True,
        ).execute()
        return meta
    except Exception as e:
        print(f"[Drive] ⚠ Cannot access folder {folder_id}: {e}")
        return {}


def upload_pdf_to_drive(filename: str, pdf_bytes: bytes,
                        drive_client=None) -> str:
    """
    Upload a single PDF to the configured Drive folder.
    Returns webViewLink on success, '' on failure.
    If drive_client is provided, it's reused (helps avoid re-auth for batch upload).
    """
    import traceback
    from googleapiclient.http import MediaIoBaseUpload
    from googleapiclient.errors import HttpError

    sa_email = _get_service_account_email()

    try:
        if drive_client is None:
            drive_client, auth_mode = _build_drive_client(
                impersonate_email="info@intellibiinnovationstechnologies.in",
            )
            print(f"[Drive] Using auth: {auth_mode}")
            if sa_email:
                print(f"[Drive] Service account: {sa_email}")

        # Delete any existing file with the same name so we always upload fresh
        try:
            existing = drive_client.files().list(
                q=(f"'{PDF_DRIVE_FOLDER_ID}' in parents and "
                   f"name='{filename}' and trashed=false"),
                fields="files(id,name)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            ).execute().get("files", [])
            for f in existing:
                try:
                    drive_client.files().delete(
                        fileId=f["id"], supportsAllDrives=True,
                    ).execute()
                except Exception as de:
                    print(f"[Drive]   (could not delete old copy {f['id']}: {de})")
        except HttpError as he:
            print(f"[Drive] ⚠ Could not list existing files: {he}")

        # Actual upload
        buf = io.BytesIO(pdf_bytes)
        media = MediaIoBaseUpload(buf, mimetype="application/pdf",
                                  resumable=False)
        up = drive_client.files().create(
            body={"name": filename, "parents": [PDF_DRIVE_FOLDER_ID]},
            media_body=media,
            fields="id,webViewLink,parents",
            supportsAllDrives=True,
        ).execute()
        link = up.get("webViewLink", "")
        fid  = up.get("id", "")
        print(f"[Drive] ✓ Uploaded: {filename}  (fileId={fid})")
        if link:
            print(f"[Drive]   Link: {link}")
        return link

    except HttpError as he:
        print(f"[Drive] ⚠ HttpError while uploading '{filename}': "
              f"{he.resp.status if he.resp else '?'} — {he}")
        print(f"[Drive]   → Most likely cause: the target folder is not shared "
              f"with the service account or impersonated user.")
        if sa_email:
            print(f"[Drive]   → FIX: In Drive, open the folder → Share → add "
                  f"'{sa_email}' as Editor.")
        print(f"[Drive]   → Folder URL: https://drive.google.com/drive/folders/"
              f"{PDF_DRIVE_FOLDER_ID}")
        return ""
    except Exception as e:
        print(f"[Drive] ⚠ PDF upload failed for '{filename}': "
              f"{type(e).__name__}: {e}")
        traceback.print_exc()
        return ""


def upload_all_pdfs_to_drive(pdfs: list) -> list:
    """
    Upload multiple PDFs efficiently (reuse one Drive client).
    Performs an up-front folder probe so failures surface clearly.
    Returns list of (filename, link) tuples.
    """
    if not pdfs:
        print("[Drive] Nothing to upload.")
        return []

    sa_email = _get_service_account_email()
    drive_client, auth_mode = _build_drive_client(
        impersonate_email="info@intellibiinnovationstechnologies.in",
    )
    print(f"[Drive] Using auth: {auth_mode}")
    if sa_email:
        print(f"[Drive] Service account: {sa_email}")
    print(f"[Drive] Target folder ID: {PDF_DRIVE_FOLDER_ID}")
    print(f"[Drive] Target folder URL: "
          f"https://drive.google.com/drive/folders/{PDF_DRIVE_FOLDER_ID}")

    # ── Probe the folder up-front so failures are LOUD ───────────────────────
    meta = _drive_folder_probe(drive_client, PDF_DRIVE_FOLDER_ID)
    if not meta:
        print(f"[Drive] ⚠ Could not access target folder — aborting upload.")
        if sa_email:
            print(f"[Drive]   → FIX: Open the folder in Drive and share it with "
                  f"'{sa_email}' as Editor.")
        return []
    print(f"[Drive] ✓ Folder accessible: name='{meta.get('name','?')}', "
          f"mimeType='{meta.get('mimeType','?')}'"
          + (f", driveId={meta['driveId']}" if meta.get("driveId") else ""))
    caps = meta.get("capabilities", {})
    if not caps.get("canAddChildren", True):
        print(f"[Drive] ⚠ Service account cannot add children to this folder "
              f"(capabilities.canAddChildren=false). Upload will fail.")
        if sa_email:
            print(f"[Drive]   → FIX: Share the folder with '{sa_email}' as Editor.")
        return []

    # ── Upload each PDF ───────────────────────────────────────────────────────
    results = []
    for p in pdfs:
        link = upload_pdf_to_drive(p["filename"], p["pdf_bytes"],
                                   drive_client=drive_client)
        results.append((p["filename"], link))
    return results


def save_pdf_local(filename: str, pdf_bytes: bytes) -> str:
    """Local saving is DISABLED — reports are delivered to Google Drive only.

    Kept as a no-op so existing call sites stay intact. PDFs live only in
    memory and are pushed to Drive via upload_all_pdfs_to_drive().
    """
    return ""


# ─────────────────────────────────────────────────────────────────────────────
#  CONSOLE SUMMARY REPORT
# ─────────────────────────────────────────────────────────────────────────────

def print_summary_report(consolidated: list, sent_ok: list, sent_fail: list):
    """
    Print a detailed batch/subject-wise console summary.
    Counts are based on individual pending assignments (not just students).
    """
    sep  = "=" * 78
    line = "-" * 78

    print(f"\n{sep}")
    print(f"  ASSIGNMENT SUBMISSION EMAIL REMINDER — SUMMARY REPORT")
    print(f"  Generated: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}")
    print(f"{sep}\n")

    total_students = len(consolidated)
    ok_students    = len(sent_ok)
    fail_students  = len(sent_fail)

    total_items = sum(s["assignment_count"] for s in consolidated)
    ok_items    = sum(s["assignment_count"] for s in sent_ok)
    fail_items  = sum(s["assignment_count"] for s in sent_fail)

    print(f"  Students to notify         : {total_students}")
    print(f"  Emails sent successfully   : {ok_students}")
    print(f"  Emails failed              : {fail_students}")
    print(f"  Total pending assignments  : {total_items}  "
          f"(sent: {ok_items}, failed: {fail_items})")
    print(f"{line}")

    # ── Per reminder level (by assignment count, not student count) ───────────
    level_counts = {lv: {"total": 0, "sent": 0, "failed": 0} for lv in LEVEL_ORDER}
    for s in consolidated:
        for a in s["assignments"]:
            level_counts[a["reminder_level"]]["total"] += 1
    for s in sent_ok:
        for a in s["assignments"]:
            level_counts[a["reminder_level"]]["sent"] += 1
    for s in sent_fail:
        for a in s["assignments"]:
            level_counts[a["reminder_level"]]["failed"] += 1

    print(f"\n  {'Reminder Level':<20} {'Total':>8} {'Sent':>8} {'Failed':>8}")
    print(f"  {'-'*20} {'-'*8} {'-'*8} {'-'*8}")
    for lv in LEVEL_ORDER:
        c = level_counts[lv]
        print(f"  {LEVEL_LABEL[lv]:<20} {c['total']:>8} {c['sent']:>8} {c['failed']:>8}")

    # ── Batch (class_name) wise breakdown ─────────────────────────────────────
    print(f"\n{line}")
    print(f"  BATCH-WISE BREAKDOWN")
    print(f"{line}")

    batch_data = defaultdict(lambda: {"1st": 0, "2nd": 0, "final": 0,
                                      "missed": 0, "total": 0})
    for s in sent_ok:
        for a in s["assignments"]:
            batch = a["class_name"] or "(Unknown)"
            batch_data[batch][a["reminder_level"]] += 1
            batch_data[batch]["total"] += 1

    if batch_data:
        print(f"\n  {'Batch / Course':<34} {'1st':>6} {'2nd':>6} {'Final':>6} "
              f"{'Missed':>6} {'Total':>6}")
        print(f"  {'-'*34} {'-'*6} {'-'*6} {'-'*6} {'-'*6} {'-'*6}")
        for batch in sorted(batch_data.keys()):
            c = batch_data[batch]
            name_display = (batch[:32] + "..") if len(batch) > 34 else batch
            print(f"  {name_display:<34} {c['1st']:>6} {c['2nd']:>6} "
                  f"{c['final']:>6} {c['missed']:>6} {c['total']:>6}")
    else:
        print("  (No emails sent)")

    # ── Subject (class_subject) wise breakdown ────────────────────────────────
    print(f"\n{line}")
    print(f"  SUBJECT-WISE BREAKDOWN")
    print(f"{line}")

    subj_data = defaultdict(lambda: {"1st": 0, "2nd": 0, "final": 0,
                                     "missed": 0, "total": 0})
    for s in sent_ok:
        for a in s["assignments"]:
            subj = a["class_subject"] or "(Unknown)"
            subj_data[subj][a["reminder_level"]] += 1
            subj_data[subj]["total"] += 1

    if subj_data:
        print(f"\n  {'Subject / Duration':<34} {'1st':>6} {'2nd':>6} {'Final':>6} "
              f"{'Missed':>6} {'Total':>6}")
        print(f"  {'-'*34} {'-'*6} {'-'*6} {'-'*6} {'-'*6} {'-'*6}")
        for subj in sorted(subj_data.keys()):
            c = subj_data[subj]
            name_display = (subj[:32] + "..") if len(subj) > 34 else subj
            print(f"  {name_display:<34} {c['1st']:>6} {c['2nd']:>6} "
                  f"{c['final']:>6} {c['missed']:>6} {c['total']:>6}")
    else:
        print("  (No emails sent)")

    # ── Per-student detail ────────────────────────────────────────────────────
    print(f"\n{line}")
    print(f"  STUDENT-WISE DETAIL (consolidated emails)")
    print(f"{line}")

    if sent_ok:
        print(f"\n  {'#':>4}  {'Student Name':<25} {'Email':<32} "
              f"{'#Asgn':>6}  {'Top Reminder':<15}")
        print(f"  {'-'*4}  {'-'*25} {'-'*32} {'-'*6}  {'-'*15}")
        for i, s in enumerate(sent_ok, 1):
            sname = (s["student_name"][:23] + "..") if len(s["student_name"]) > 25 else s["student_name"]
            email = (s["student_email"][:30] + "..") if len(s["student_email"]) > 32 else s["student_email"]
            print(f"  {i:>4}  {sname:<25} {email:<32} "
                  f"{s['assignment_count']:>6}  {s['reminder_label']:<15}")
    else:
        print("  (No emails sent)")

    # ── Failed emails ─────────────────────────────────────────────────────────
    if sent_fail:
        print(f"\n{line}")
        print(f"  FAILED EMAILS")
        print(f"{line}")
        for i, s in enumerate(sent_fail, 1):
            print(f"  {i}. {s['student_name']} ({s['student_email']}) — "
                  f"{s['assignment_count']} assignment(s) — {s['reminder_label']}")

    print(f"\n{sep}")
    print(f"  End of Report")
    print(f"{sep}\n")


# ─────────────────────────────────────────────────────────────────────────────
#  OVERDUE SUBMISSIONS EXCEL REPORT
#    Lists every student whose assignment is OVERDUE: submission_status =
#    "Not Submitted" AND submission_deadline already in the past. This is
#    INDEPENDENT of the daily reminder window (2/1/0/-1 days) — it reports the
#    full overdue backlog. Reuses the Submissions data already read in main(),
#    so no extra Google Sheet read is needed. Saved as a formatted .xlsx next
#    to this script.
# ─────────────────────────────────────────────────────────────────────────────

def collect_overdue(subs_df, today: date,
                    include_all_not_submitted: bool = False) -> list:
    """Return overdue (or all not-submitted) row dicts, most-overdue first."""
    records = []
    if subs_df is None or subs_df.empty:
        return records

    for _, row in subs_df.iterrows():
        status = str(row.get("submission_status", "")).strip()
        if status.lower() != "not submitted":
            continue

        deadline = _parse_ist_date(row.get("submission_deadline", ""))
        if deadline is None:
            continue

        days_overdue = (today - deadline).days   # positive => overdue
        if not include_all_not_submitted and days_overdue <= 0:
            continue   # deadline not yet passed

        assigned = (_parse_ist_date(row.get("submission_start_date", ""))
                    or _parse_ist_date(row.get("assigned_date", ""))
                    or _parse_ist_date(row.get("assignment_start_date", "")))

        records.append({
            "student_name":     str(row.get("student_name", "") or "Student"),
            "student_email":    str(row.get("student_email", "") or "").strip(),
            "student_phone":    str(row.get("student_phone", "") or ""),
            "class_name":       str(row.get("class_name", "") or ""),
            "class_subject":    str(row.get("class_subject", "") or ""),
            "assessment_title": str(row.get("assessment_title", "") or ""),
            "maximum_marks":    str(row.get("maximum_marks", "") or ""),
            "assigned_str":     assigned.strftime("%d-%b-%Y") if assigned else "",
            "deadline_str":     deadline.strftime("%d-%b-%Y"),
            "days_overdue":     days_overdue,
        })

    records.sort(key=lambda r: (-r["days_overdue"], r["student_name"].lower()))
    return records


def build_overdue_workbook(records: list, today: date,
                           include_all_not_submitted: bool = False):
    """Build an openpyxl Workbook for the overdue submissions report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT_NAME = "Arial"
    NAVY = "1F3864"; NAVY_LIGHT = "D9E1F2"
    RED_FILL = "F7C1C1"; AMBER_FILL = "FCE4D6"; WHITE = "FFFFFF"

    wb = Workbook()
    ws = wb.active
    ws.title = "Overdue Submissions"

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    scope = ("All not-submitted assignments" if include_all_not_submitted
             else "Overdue assignments only")

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "IntelliBI — Overdue Assignment Submissions Report"
    c.font = Font(name=FONT_NAME, bold=True, size=14, color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = (f"Date: {today.strftime('%d-%b-%Y (%A)')}   |   "
               f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}   |   "
               f"Scope: {scope}   |   Total: {len(records)}")
    c.font = Font(name=FONT_NAME, italic=True, size=9, color="555555")

    headers = ["#", "Student Name", "Student Email", "Student Phone",
               "Course / Batch", "Subject / Duration", "Assignment",
               "Assigned Date", "Deadline", "Days Overdue", "Max Marks"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border

    first_data = header_row + 1
    for i, r in enumerate(records, 1):
        rr = first_data + i - 1
        values = [i, r["student_name"], r["student_email"], r["student_phone"],
                  r["class_name"], r["class_subject"], r["assessment_title"],
                  r["assigned_str"], r["deadline_str"], r["days_overdue"],
                  r["maximum_marks"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=rr, column=col, value=v)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 8, 9, 10, 11) else "left",
                vertical="center")
        fill = RED_FILL if r["days_overdue"] >= 3 else AMBER_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=fill)

    last_data = first_data + len(records) - 1 if records else first_data
    ws.freeze_panes = ws.cell(row=first_data, column=1)
    ws.auto_filter.ref = f"A{header_row}:K{max(last_data, header_row)}"

    widths = [5, 24, 32, 16, 26, 20, 30, 14, 14, 13, 11]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    if not records:
        ws.merge_cells(f"A{first_data}:K{first_data}")
        cc = ws[f"A{first_data}"]
        cc.value = "No overdue submissions found for this date."
        cc.font = Font(name=FONT_NAME, italic=True, size=11, color="2E7D32")
        cc.alignment = Alignment(horizontal="center")

    # ── Summary by course (live COUNTIF / SUM formulas) ───────────────────────
    ws2 = wb.create_sheet("Summary by Course")
    ws2.merge_cells("A1:B1")
    t2 = ws2["A1"]
    t2.value = "Overdue Submissions — Summary by Course / Batch"
    t2.font = Font(name=FONT_NAME, bold=True, size=12, color=NAVY)
    for col, h in enumerate(["Course / Batch", "Overdue Count"], 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    courses = sorted({r["class_name"] or "(Unknown Course)" for r in records})
    course_range = (f"'Overdue Submissions'!$E${first_data}:"
                    f"$E${max(last_data, first_data)}")
    srow = 4
    for course in courses:
        ws2.cell(row=srow, column=1,
                 value=course).font = Font(name=FONT_NAME, size=10)
        cnt = ws2.cell(row=srow, column=2)
        cnt.value = f'=COUNTIF({course_range},A{srow})'
        cnt.font = Font(name=FONT_NAME, size=10)
        cnt.alignment = Alignment(horizontal="center")
        for col in (1, 2):
            ws2.cell(row=srow, column=col).border = border
        srow += 1

    ws2.cell(row=srow, column=1,
             value="TOTAL").font = Font(name=FONT_NAME, bold=True)
    tot = ws2.cell(row=srow, column=2)
    tot.value = f"=SUM(B4:B{srow - 1})" if courses else 0
    tot.font = Font(name=FONT_NAME, bold=True)
    tot.alignment = Alignment(horizontal="center")
    for col in (1, 2):
        ws2.cell(row=srow, column=col).fill = PatternFill("solid",
                                                          fgColor=NAVY_LIGHT)
        ws2.cell(row=srow, column=col).border = border

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 16
    return wb


def generate_overdue_report(subs_df, today: date,
                            include_all_not_submitted: bool = False,
                            out_path: str = "") -> str:
    """Build + save the overdue submissions Excel report. Returns saved path."""
    records = collect_overdue(subs_df, today, include_all_not_submitted)
    scope = "all not-submitted" if include_all_not_submitted else "overdue only"
    print(f"\n[Overdue] Building Excel report ({scope}) — {len(records)} row(s) ...")
    try:
        wb = build_overdue_workbook(records, today, include_all_not_submitted)
    except ModuleNotFoundError as mne:
        missing = mne.name or str(mne)
        print(f"[Overdue] ⚠ Generation failed: missing Python package '{missing}'.")
        print(f"[Overdue]   → FIX: pip install {missing}")
        return ""
    except Exception as e:
        import traceback
        print(f"[Overdue] ⚠ Generation failed: {type(e).__name__}: {e}")
        traceback.print_exc()
        return ""

    if not out_path:
        out_path = os.path.join(
            _SCRIPT_DIR,
            f"OverdueSubmissions_Report_{today.strftime('%Y-%m-%d')}.xlsx")
    try:
        wb.save(out_path)
        print(f"[Overdue] ✓ Saved: {out_path}")
    except Exception as e:
        print(f"[Overdue] ⚠ Could not save report: {e}")
        return ""
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # ── Execution is controlled by the RUN CONFIGURATION variables at the top of
    #    this file, replacing the previous command-line arguments. They are
    #    gathered into a lightweight namespace so the logic below is unchanged.
    from types import SimpleNamespace
    args = SimpleNamespace(
        dry_run=dry_run,
        no_email=not send_email,
        no_pdf=not generate_pdf,
        no_overdue_report=not include_overdue_report,
        overdue_all=overdue_all,
        date=report_date,
    )

    today = (datetime.strptime(args.date, "%Y-%m-%d").date()
             if args.date else date.today())

    sep = "=" * 72
    mode = ("DRY-RUN" if args.dry_run else
            ("NO-EMAIL" if args.no_email else "LIVE"))
    print(f"\n{sep}")
    print(f"  IntelliBI Assignment Submission — Email Reminder")
    print(f"  Date    : {today.strftime('%d-%b-%Y')} ({today.strftime('%A')})")
    print(f"  Mode    : {mode}")
    print(f"{sep}\n")

    # ── Auth & data ───────────────────────────────────────────────────────────
    from utils import get_sheets_service
    service = get_sheets_service(SERVICE_ACCOUNT_FILE)

    print("[Data] Reading Submissions tab ...")
    subs_df = read_sheet_df(service, SUBMISSION_SHEET_ID, SUBMISSIONS_TAB)
    print(f"[Data] Total rows: {len(subs_df)}")

    if subs_df.empty:
        print("[Main] No submissions data found. Exiting.")
        return

    # ── Overdue submissions Excel report ──────────────────────────────────────
    # Built from the data already loaded above and is INDEPENDENT of today's
    # reminder window, so it runs even when no reminders are due today.
    if not args.no_overdue_report:
        generate_overdue_report(subs_df, today,
                                include_all_not_submitted=args.overdue_all)

    # ── Find pending students with nearing deadlines ──────────────────────────
    records = find_pending_reminders(subs_df, today)
    records = _dedup_records(records)
    consolidated = consolidate_by_student(records)

    n_students = len(consolidated)
    n_items    = sum(s["assignment_count"] for s in consolidated)
    print(f"[Reminders] {n_students} student(s) to notify — "
          f"{n_items} pending assignment(s) total")

    # Per-level breakdown (by assignment, so counts match what students see)
    lvl_counts = {lv: 0 for lv in LEVEL_ORDER}
    for s in consolidated:
        for a in s["assignments"]:
            lvl_counts[a["reminder_level"]] += 1
    for lv in LEVEL_ORDER:
        if lvl_counts[lv]:
            print(f"  {LEVEL_LABEL[lv]:<16}: {lvl_counts[lv]}")

    if not consolidated:
        print("[Main] No pending submissions match today's reminder criteria.")
        print("[Main] Nothing to do. Exiting.")
        return

    # ── Restrict reminder emails to ACTIVE students ───────────────────────────
    # Load is_candidate_active (Y/N) from the Students tab; every student is
    # validated in the send loop below before any email is sent.
    print("[Active-Check] Reading student active-status from Students tab ...")
    active_status_map = load_active_status_map(service)

    # ── Send emails (one consolidated per student) ────────────────────────────
    sent_ok   = []
    sent_fail = []
    skipped_inactive = []   # students excluded because they are not active

    skip_email = args.no_email

    print(f"\n[Email] {'Previewing' if args.dry_run else ('Skipping' if skip_email else 'Sending')} "
          f"{n_students} consolidated email(s) ...")

    # Throttling applies only to LIVE sends (not dry-run, not --no-email)
    throttled = (not args.dry_run) and (not skip_email)
    if throttled:
        est_secs = (
            n_students * INTER_EMAIL_DELAY_SEC
            + max(0, (n_students - 1) // EMAIL_BATCH_SIZE) * EMAIL_BATCH_PAUSE_SEC
        )
        est_min = est_secs / 60
        print(f"[Email] Throttle: {INTER_EMAIL_DELAY_SEC}s between sends, "
              f"{EMAIL_BATCH_PAUSE_SEC}s pause every {EMAIL_BATCH_SIZE} emails "
              f"(~{est_min:.1f} min total)")

    live_sent_count = 0   # count of LIVE sends completed (used for throttling)

    for i, s in enumerate(consolidated, 1):
        print(f"  [{i:>3}/{n_students}] {s['reminder_label']:>14} → "
              f"{s['student_name']:<25} ({s['student_email']}) — "
              f"{s['assignment_count']} assignment(s)")

        # Active-student gate — no email is sent unless is_candidate_active == 'Y'.
        active, reason = is_student_active(s, active_status_map)
        if not active:
            print(f"      [Active-Check] SKIP — no email sent: {reason}")
            skipped_inactive.append(s)
            continue

        if skip_email:
            sent_ok.append(s)
            continue

        success = send_reminder_email(s, dry_run=args.dry_run)
        if success:
            sent_ok.append(s)
        else:
            sent_fail.append(s)

        # ── Throttle live sends ──────────────────────────────────────────────
        if throttled:
            live_sent_count += 1
            is_last = (i == n_students)
            if not is_last:
                # Long pause every EMAIL_BATCH_SIZE emails
                if live_sent_count % EMAIL_BATCH_SIZE == 0:
                    print(f"    [Throttle] Batch of {EMAIL_BATCH_SIZE} done — "
                          f"pausing {EMAIL_BATCH_PAUSE_SEC}s before next batch ...")
                    time.sleep(EMAIL_BATCH_PAUSE_SEC)
                else:
                    time.sleep(INTER_EMAIL_DELAY_SEC)

    # Dry-run: also write a combined HTML preview file so it's easy to eyeball
    if args.dry_run:
        preview_path = save_html_preview(consolidated, today)
        if preview_path:
            print(f"\n[Preview] ✓ Combined HTML preview: {preview_path}")

    # ── Generate + upload per-assignment PDFs ─────────────────────────────────
    #    One PDF per (Course, Assignment) pair.
    #    Filename: Assignment_[CourseName]_[AssignmentsName]_Reminders_YYYY-MM-DD.pdf
    #    NOTE: PDFs are always generated (even in --dry-run) but are NEVER
    #          written to local disk — they are delivered to Google Drive only.
    #          The Google Drive upload is skipped during --dry-run.
    # ── Build per-student email-status map used by the standalone Email
    #    Status PDF report (attached to the staff summary email). The sending
    #    loop is already done at this point, so we know exactly which students
    #    succeeded, failed, were skipped, or were in dry-run preview mode.
    email_status = {}
    if args.dry_run:
        for s in consolidated:
            email_status[s["student_email"]] = "Preview"
    elif skip_email:
        for s in consolidated:
            email_status[s["student_email"]] = "Skipped"
    else:
        for s in sent_ok:
            email_status[s["student_email"]] = "Sent"
        for s in sent_fail:
            email_status[s["student_email"]] = "Failed"

    # Students excluded by the active-status gate are reported as "Skipped"
    # regardless of run mode (they were never sent an email).
    for s in skipped_inactive:
        email_status[s["student_email"]] = "Skipped"

    pdfs = []   # Populated below; used again by the staff summary email.
    if not args.no_pdf:
        mode_tag = "(DRY-RUN: local-only)" if args.dry_run else ""
        print(f"\n[PDF] Generating per-assignment reminder PDF(s) {mode_tag}...")
        try:
            pdfs = generate_assignment_pdfs(consolidated, today)
            if not pdfs:
                print("[PDF] No pending (course, assignment) groups — nothing to build.")
            else:
                print(f"[PDF] ✓ Built {len(pdfs)} PDF file(s):")
                for p in pdfs:
                    rc = p["reminder_counts"]
                    print(f"       • {p['filename']}  "
                          f"({p['student_count']} student(s); "
                          f"1st={rc['1st']}, 2nd={rc['2nd']}, Final={rc['final']}; "
                          f"{len(p['pdf_bytes']):,} bytes)")

                if args.dry_run:
                    print(f"[PDF] [DRY-RUN] {len(pdfs)} PDF(s) built in memory — "
                          f"Google Drive upload skipped (no local copies are written).")
                else:
                    print(f"\n[PDF] Uploading {len(pdfs)} PDF(s) to Google Drive ...")
                    results = upload_all_pdfs_to_drive(pdfs)
                    ok   = sum(1 for _, link in results if link)
                    fail = sum(1 for _, link in results if not link)
                    print(f"[Drive] Summary: {ok} uploaded, {fail} failed")
        except ModuleNotFoundError as mne:
            missing = mne.name or str(mne)
            print(f"[PDF] ⚠ Generation failed: missing Python package '{missing}'.")
            print(f"[PDF]   → FIX: Install it with:")
            print(f"[PDF]        pip install {missing}")
            if missing == "reportlab":
                print(f"[PDF]     (reportlab is required to build the batch-wise PDFs.)")
        except Exception as e:
            import traceback
            print(f"[PDF] ⚠ Generation failed: {type(e).__name__}: {e}")
            traceback.print_exc()
    else:
        print(f"\n[PDF] Skipped (--no-pdf).")

    # ── Standalone Email Status PDF (summary-email attachment only) ───────────
    # This PDF is intentionally SEPARATE from the per-assignment reminder PDFs:
    #   • It is NOT uploaded to Google Drive.
    #   • It is ONLY attached to the end-of-run staff summary email.
    #   • Existing reminder PDFs remain untouched (no status column added).
    status_pdf_entry = None
    if consolidated:
        try:
            mode_label = ("DRY-RUN" if args.dry_run else
                          ("NO-EMAIL" if skip_email else "LIVE"))
            print(f"\n[StatusPDF] Building standalone Email Status Report ({mode_label})...")
            status_pdf_entry = generate_email_status_pdf_entry(
                consolidated, email_status, sent_ok, sent_fail, today, mode_label)
            save_pdf_local(status_pdf_entry["filename"], status_pdf_entry["pdf_bytes"])
            print(f"[StatusPDF] ✓ Built {status_pdf_entry['filename']} "
                  f"({len(status_pdf_entry['pdf_bytes']):,} bytes) — "
                  f"will be attached to the staff summary email only.")
        except ModuleNotFoundError as mne:
            missing = mne.name or str(mne)
            print(f"[StatusPDF] ⚠ Generation failed: missing Python package '{missing}'.")
            print(f"[StatusPDF]   → FIX: pip install {missing}")
        except Exception as e:
            import traceback
            print(f"[StatusPDF] ⚠ Generation failed: {type(e).__name__}: {e}")
            traceback.print_exc()

    # ── Staff summary email (one combined mail, after all student sends) ──────
    # This REPLACES the per-student Cc/Bcc, so each student email counts as a
    # single-recipient send against Gmail's daily quota.
    #
    # Skip the summary if we were in dry-run (preview only) or --no-email
    # (explicit skip), and also if there's literally nothing to report.
    if consolidated:
        print(f"\n[Summary] {'Previewing' if args.dry_run else ('Skipping (--no-email)' if skip_email else 'Sending')} "
              f"staff summary email ...")
        if skip_email and not args.dry_run:
            print("[Summary] Staff summary email skipped because --no-email was set.")
        else:
            send_staff_summary_email(
                consolidated, sent_ok, sent_fail, pdfs, today,
                dry_run=args.dry_run,
                extra_pdfs=[status_pdf_entry] if status_pdf_entry else None,
            )

    # ── Summary report (console) ──────────────────────────────────────────────
    print_summary_report(consolidated, sent_ok, sent_fail)


if __name__ == "__main__":
    main()